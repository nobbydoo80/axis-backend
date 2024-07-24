"""tasks.py: Django checklist"""


import datetime
import os
import time


import dateutil.parser
from celery import shared_task
from celery.utils.log import get_task_logger
from django.db.models import Q
from django.db.models.signals import post_save
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils.timezone import now
from localflavor.us.us_states import STATES_NORMALIZED
from openpyxl.utils.exceptions import InvalidFileException

from axis.community.models import Community
from axis.company.models import Company
from axis.customer_aps.aps_calculator import APSInputException
from axis.customer_aps.aps_calculator.calculator import APSCalculator
from axis.eep_program.models import EEPProgram
from axis.filehandling.log_storage import LogStorage
from axis.filehandling.models import ResultObjectLog
from axis.filehandling.utils import XLSXParser, get_physical_file
from axis.floorplan.models import Floorplan
from axis.geographic.models import ClimateZone, County, City
from axis.home.models import EEPProgramHomeStatus, Home
from axis.home.tasks import (
    update_home_states,
    update_home_stats,
    certify_single_home,
    certify_sampleset,
)
from axis.qa.messages import QASubdivisionGatingCertification
from axis.sampleset import strings as sampleset_strings
from axis.sampleset.models import SampleSet
from axis.sampleset.strings import MISSING_SUBDIVISION
from axis.sampleset.utils import inspect_for_sampleset, validate_bulk_sampleset_configuration
from axis.scheduling.models import ConstructionStage, ConstructionStatus
from axis.subdivision.models import Subdivision, EEPProgramSubdivisionStatus
from . import strings
from .forms import BULK_UPLOAD_EQUIV_MAP
from .models import Question, QuestionChoice, CheckList, Section
from .utils import (
    validate_checklist_spreadsheet_format,
    validate_checklist_sheets,
    ExcelChecklist,
    get_answered_questions_from_data,
    answer_questions_for_home,
    set_annotations_for_home,
    MAX_NUM_CHOICES,
    VALID_AFFIRMATIVE_BULK_FILL_RESPONSES,
)

__author__ = "Steven Klass"
__date__ = "6/25/12 1:14 PM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

from ..core.validators import represents_integer

logger = get_task_logger(__name__)
User = get_user_model()


@shared_task
def process_checklist(company_id, user_id, result_object_id, **kwargs):
    """Validates the upload of a document containing a new checklist template."""
    # This does not process a completed checklist for a home, but rather a document detailing the
    # creation of a new checklist altogether.

    from axis.filehandling.models import AsynchronousProcessedDocument

    log = kwargs.get("log", logger)

    result_object = AsynchronousProcessedDocument.objects.get(id=result_object_id)
    result_log = ResultObjectLog(result_object_id=result_object.id)

    if None in [company_id, user_id, result_object_id]:
        lerrors = "Missing items {0}".format(", ".join([company_id, user_id, result_object_id]))
        result_log.update(errors=lerrors, raise_errors=True)
        return

    user = User.objects.get(id=user_id)
    company = result_object.company
    document = result_object.document

    try:
        filename, notes = get_physical_file(document.name)
        result_log.update(result_list=notes, raise_errors=True)
    except AttributeError:
        filename, notes = document, ["Using document string {0}".format(document)]
        result_log.update(warnings=notes, raise_errors=True)

    log.info(
        strings.USER_SUBMITTED_DOCUMENT_FOR_PROCESSING.format(
            user=user,
            document=filename,
            task=result_object.task_name,
            task_id=result_object.task_id,
        )
    )
    #
    # -- Here is where the real work begins --
    #

    # Validate our assumptions about the file's format, available sheets, columns, etc.
    validate_checklist_spreadsheet_format(filename, result_log)
    if len(result_log.get_errors()):
        result_log.update(raise_errors=True)
        return

    # Validate each sheet's format requirements and reference integrity.
    validate_checklist_sheets(filename, result_log, company)
    if len(result_log.get_errors()):
        result_log.update(raise_errors=True)
        return

    # Begin actual import of the data
    choice_dict = {}
    section_dict = {}
    question_dict = {}

    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Checklist")
    xlsx.get_columns(uniq_column="Name")
    result_log.update(latest="Processing sheet Checklist")
    result = xlsx.get_results_dictionary_list(set_lower=True, add_row_number=False)[0]

    if "public" in result and result["public"]:
        result["public"] = True
    else:
        result["public"] = False

    result["group"] = company.group

    checklist_obj, create = CheckList.objects.get_or_create(**result)
    checklist_obj.save()
    msg = "{} checklist {}".format("Created" if create else "Using existing", checklist_obj.name)
    result_log.update(info=msg)
    log.debug(msg)

    # Add our Choices
    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Choices")
    xlsx.get_columns(uniq_column="Choice")
    result_log.update(latest="Finalizing Choices")
    for result in xlsx.get_results_dictionary_list(set_lower=True, add_row_number=False):
        for item in [
            "is_considered_failure",
            "email_required",
            "comment_required",
            "display_as_failure",
            "photo_required",
            "document_required",
        ]:
            if item in result and result[item]:
                result[item] = True
            else:
                result[item] = False
        if result["choice_order"] is None:
            result["choice_order"] = 0
        result["object"], result["create"] = QuestionChoice.objects.get_or_create(**result)
        choice_dict[result["choice"]] = result

    added = sum(1 for v in choice_dict.values() if v["create"])
    updated = sum(1 for v in choice_dict.values() if not v["create"])

    msg = strings.ADDING_UPDATING_TYPE.format(added=added, updated=updated, type="Choices")
    result_log.update(info=msg)
    log.info(msg)

    # Add our Questions
    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Questions")
    xlsx.get_columns(uniq_column="Question")
    result_log.update(latest="Finalizing Questions")
    for result in xlsx.get_results_dictionary_list(set_lower=True, add_row_number=False):
        entry = result.copy()
        del entry["section"]
        for x in range(1, MAX_NUM_CHOICES + 1):
            if "choice{}".format(x) in entry:
                del entry["choice{}".format(x)]
        if "climate_zone" in entry and entry["climate_zone"] is not None:
            try:
                entry["climate_zone"] = ClimateZone.objects.get_by_code(entry["climate_zone"])
            except ClimateZone.DoesNotExist:
                warning = "Cannot find climate zone {} - skipping".format(entry["climate_zone"])
                result_log.update(warnings=[warning])

        entry["minimum_value"] = entry.pop("Minimum", None)
        entry["maximum_value"] = entry.pop("Maximum", None)
        entry["is_optional"] = entry.pop("Optional", False)
        if entry["type"] not in ["float", "integer"]:
            del entry["minimum_value"]
            del entry["maximum_value"]
        if "allow_bulk_fill" in entry.keys():
            allow_bulk_fill = False
            if str(entry["allow_bulk_fill"]).lower() in VALID_AFFIRMATIVE_BULK_FILL_RESPONSES:
                allow_bulk_fill = True
            entry["allow_bulk_fill"] = allow_bulk_fill
        result["object"], result["create"] = Question.objects.get_or_create(**entry)

        for choice in range(1, MAX_NUM_CHOICES + 1):
            key = "choice{}".format(choice)
            if key not in result:
                continue
            if result["type"] == "multiple-choice" and result[key]:
                result["object"].question_choice.add(choice_dict[result[key]]["object"])
            result.pop(key)
        question_dict[result["question"]] = result

        checklist_obj.questions.add(result["object"])

    added = sum(1 for v in question_dict.values() if v["create"])
    updated = sum(1 for v in question_dict.values() if not v["create"])
    msg = strings.ADDING_UPDATING_TYPE.format(added=added, updated=updated, type="Questions")
    result_log.update(info=msg)
    log.info(msg)

    # Add in our Sections
    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Sections")
    xlsx.get_columns(uniq_column="Name")
    result_log.update(latest="Finalizing Sections")
    for result in xlsx.get_results_dictionary_list(set_lower=True, add_row_number=False):
        result["checklist"] = checklist_obj
        result["object"], result["create"] = Section.objects.get_or_create(**result)
        section_dict[result["name"]] = result
        for question, value in question_dict.items():
            if value["section"] == result["name"]:
                result["object"].questions.add(value["object"])
    added = sum(1 for v in section_dict.values() if v["create"])
    updated = sum(1 for v in section_dict.values() if not v["create"])
    result_log.update(
        info=strings.ADDING_UPDATING_TYPE.format(added=added, updated=updated, type="Sections")
    )

    #
    # -- End of real work --
    #

    msg = strings.COMPLETED_PROCESSING.format(document=result_object.filename(), company=company)
    log.info(msg)
    return msg


@shared_task(time_limit=60 * 60)
def bulk_checklist_process(
    company_id,
    user_id,
    result_object_id,
    only_bulk_home_processing=False,
    overwrite_old_answers=False,
    **kwargs,
):
    """
    Processes a template XLSX where each row of content is a separate home, and the large number of
    columns represent the home details, homestatus details, checklist answers, and completed
    annotation data.

    results lists:
        results         - all results grabbed from the XLS.
        final_results   - subset of results that are not already certified,
                            has a builder and subdivision,
                            and has a city, county, and state.
        to_be_certified - subset of final_results that have no errors and a certification date.
    """
    from axis.filehandling.models import AsynchronousProcessedDocument
    from .collection.excel import BulkExcelChecklistCollector

    log = kwargs.get("log", logger)
    app_log = LogStorage(model_id=result_object_id)

    result_object = AsynchronousProcessedDocument.objects.get(id=result_object_id)

    if overwrite_old_answers:
        app_log.debug(strings.ALLOW_OVERWRITE)

    if None in [company_id, user_id, result_object_id]:
        missing = ", ".join([x for x in [company_id, user_id, result_object_id] if not x])
        # TODO: Ask Steve about likeliness of this occuring,
        #  what message should we put here for user.
        app_log.error(strings.MISSING_CORE.format(items=missing))
        return

    user = User.objects.get(id=user_id)
    company = result_object.company
    document = result_object.document

    filename = get_physical_file(document.name, log=app_log)

    app_log.debug(
        strings.SUBMIT_MSG.format(
            user=user.get_full_name(),
            document=os.path.basename(filename),
            task_id=result_object.task_id,
        )
    )

    #
    # -- Here is where the real work begins --
    #

    # Open the file and do a quick scan to make sure that the obvious things are in place.
    try:
        xlsx = ExcelChecklist(filename=filename, uniq_header_column="lot_number")
    except TypeError:
        # TODO: error (is not an excel document)
        app_log.error(strings.BAD_XLSX.format(document=os.path.basename(filename)))
        return

    try:
        xlsx.load_workbook_and_sheet()
    except AttributeError as err:
        # TODO: error (excel cannot load)
        app_log.error(strings.XLS_LOAD_ERROR.format(document=os.path.basename(filename), error=err))
        return
    except InvalidFileException:
        # TODO: error (is not an excel document)
        app_log.error(strings.BAD_XLSX.format(document=os.path.basename(filename)))
        return

    try:
        headers = xlsx.get_columns(equivalence_map=BULK_UPLOAD_EQUIV_MAP)
    except SystemError as err:
        app_log.error(
            strings.XLS_BAD_COLUMNS.format(document=os.path.basename(filename), error=err)
        )
        return
    except UnicodeDecodeError:
        app_log.error(strings.BAD_XLSX.format(document=os.path.basename(filename)))
        return

    if not len(headers):
        app_log.error(strings.NO_HEADERS.format(document=os.path.basename(filename)))
        return

    base_headers = xlsx.get_base_headers()
    for header in [
        "community",
        "subdivision_builder_name",
        "street_line2",
        "city",
        "state",
        "builder_org",
        "construction_stage",
        "sample_set",
        "certification_date",
        "floorplan",
        "is_multi_family",
        "alt_name",
    ]:
        if header in base_headers:
            base_headers.pop(base_headers.index(header))
    if only_bulk_home_processing:
        for header in ["rating_type", "is_multi_family", "alt_name"]:
            if header in base_headers:
                base_headers.pop(base_headers.index(header))

    if "rating_type" in base_headers:
        app_log.info(strings.DEPRECATED_HEADER_RATING_TYPE)

    base_headers_lower = [x.lower() for x in base_headers]
    headers_lower = base_headers_lower
    if len(headers):
        headers_lower = [x.lower() for x in headers if x]
    if not set(base_headers_lower).issubset(set(headers_lower)):
        missing = ", ".join(list(set(base_headers_lower) - set(headers_lower)))
        app_log.error(
            strings.XLS_MISSING_COLUMNS.format(document=os.path.basename(filename), missing=missing)
        )
        return

    results = xlsx.get_results_dictionary_list(
        equivalence_map=BULK_UPLOAD_EQUIV_MAP, set_lower=True
    )
    if not results:
        app_log.error(strings.XLS_NO_RESULTING_DATA.format(document=os.path.basename(filename)))
        return

    sampling_elegibility_dict = {}

    eep_name = None
    app_log.debug(strings.STAGE_ONE_COMPLETE)

    # ----------------------------------------------------------------------------------------------
    # STAGE TWO - Validation
    # ----------------------------------------------------------------------------------------------

    total_homes = len(results)
    final_results = []
    link_string = "<a target='_blank' href='{url}'>{object}</a>"
    for result in results:
        app_log.set_context(row=result["row_number"])
        stat = "{}/{}".format(results.index(result) + 1, len(results))
        app_log.debug(strings.STAGE_TWO_PCT_INCREMENT.format(stat=stat))
        # default all flags to false.
        app_log.set_flags(
            home_updated=False,
            home_created=False,
            home_already_certified=False,
            certification_date=False,
            home_certified=False,
            has_row_flags=True,
        )
        links = {}
        sampleset_tracking = None  # Set for row if sampling columns are in use
        collector = None  # Set once the row's program is identified

        # We just want to validate a couple things..
        if not result.get("eep_program") and eep_name:
            result["eep_program"] = eep_name

        state = None
        if result.get("state"):
            try:
                state = STATES_NORMALIZED[result.get("state").lower()]
            except (KeyError, AttributeError):
                app_log.error(strings.INVALID_US_STATE.format(state=result.get("state")))

        county = County.objects.verify(
            name=result.get("county"), state=state, ignore_missing=True, log=app_log
        )

        city = City.objects.verify(
            name=result.get("city"),
            county=county,
            state=state,
            info_only=True,
            ignore_missing=only_bulk_home_processing,
            log=app_log,
        )

        builder = None
        if result.get("builder_org"):
            builder = Company.objects.verify_existence_for_company(
                result.get("builder_org"), company, log=app_log
            )

        community = Community.objects.verify_existence_for_company(
            result.get("community"), company, ignore_missing=True, log=app_log
        )

        subdivision = None
        if result.get("subdivision") or result.get("subdivision_builder_name"):
            subdivision = Subdivision.objects.verify_for_company(
                name=result.get("subdivision"),
                builder_name=result.get("subdivision_builder_name"),
                company=company,
                community=community,
                builder=builder,
                log=app_log,
            )

        if subdivision:
            links["Subdivision"] = link_string.format(
                url=subdivision.get_absolute_url(), object=subdivision
            )
            app_log.set_flags(links=links)

        if not builder:
            if subdivision:
                builder = subdivision.builder_org
            else:
                app_log.error(strings.NO_BUILDER_NO_SUBDIVISION)
        else:
            if subdivision and builder != subdivision.builder_org:
                app_log.warning(
                    strings.BUILDER_SUBDIVISION_BUILDER_MISMATCH.format(
                        builder=builder, sub_builder=subdivision.builder_org
                    )
                )
                builder = subdivision.builder_org

        certification_date = None
        if result.get("certification_date"):
            app_log.set_flags(certification_date=True)
            if not isinstance(result.get("certification_date"), datetime.datetime):
                try:
                    cert = dateutil.parser.parse(result.get("certification_date")).replace(
                        tzinfo=datetime.timezone.utc
                    )
                except Exception:
                    app_log.error(
                        strings.INVALID_CERT_DATE.format(date=result.get("certification_date"))
                    )
                    certification_date = None
                else:
                    certification_date = cert
            else:
                certification_date = result.get("certification_date")

        eep_program = EEPProgram.objects.verify_for_company(
            result.get("eep_program"),
            company,
            subdivision,
            ignore_missing=only_bulk_home_processing,
            log=app_log,
        )

        use_legacy = not eep_program or eep_program.collection_request is None
        if not use_legacy:
            context = {
                "user": user,
            }
            collector = BulkExcelChecklistCollector(eep_program.collection_request, **context)

        questions = None
        annotations = None
        payloads = None
        if eep_program:
            if collector:
                # Clean these inputs, stage for saving.
                # Note that we always extend because we will be generating lists on every pass and
                # those do not replace the existing list, just get tacked on.
                payloads = collector.create_payloads_for_row_inputs(result)

                # FIXME: [djinco] collector.clean() is too broken when using a list
                payloads = [collector.clean_payload(payload) for payload in payloads]

            # For legacy and annotations--the input-collection setup doesn't handle the annotations
            questions, annotations = get_answered_questions_from_data(
                result, eep_program, ignore_missing=only_bulk_home_processing, log=app_log
            )

        # Determine Sampling info
        # We place this high up in the parsing process to provide this info to the results as soon
        # as possible.  If there are problems looking up other data about the row, this will still
        # appear on the report.
        sampling_detail_tuple = inspect_for_sampleset(
            **dict(
                result,
                **{
                    "company": company,
                    "subdivision": subdivision,
                    "builder_org": builder,
                    "questions": questions,
                },
            )
        )

        if sampling_detail_tuple.is_test_home and eep_program:
            # Will just do the sampling check
            EEPProgram.objects.verify_for_company(
                result.get("eep_program"),
                company,
                subdivision,
                ignore_missing=only_bulk_home_processing,
                is_test_home=sampling_detail_tuple.is_test_home,
                log=app_log,
            )

        floorplan_ignore = True if certification_date is None else False
        if certification_date and eep_program:
            floorplan_ignore = not eep_program.requires_floorplan()
            if not floorplan_ignore and result.get("floorplan") is None:
                app_log.warning(
                    "Certification date was defined and program requires "
                    "a floorplan - Missing floorplan will cause processing to stop"
                )

        floorplan, eep_program = Floorplan.objects.verify_for_company(
            name=result.get("floorplan"),
            company=company,
            subdivision=subdivision,
            builder=builder,
            eep_program=eep_program,
            ignore_missing=floorplan_ignore,
            certification_date=certification_date,
            user=user,
            return_eep=True,
            log=app_log,
            warn_on_missing_model_file=True,
        )

        sub_stat = EEPProgramSubdivisionStatus.objects.verify_for_company(
            subdivision=subdivision, eep_program=eep_program, company=company, log=app_log
        )

        if sampling_detail_tuple.use_sampling:
            if sampling_detail_tuple.url:
                sampleset_href = """<a href="{}">{}</a>""".format(
                    sampling_detail_tuple.url, sampling_detail_tuple.repr
                )
            else:
                sampleset_href = sampling_detail_tuple.repr

            if not subdivision:
                app_log.error(MISSING_SUBDIVISION)

            sampleset_tracking = sampling_elegibility_dict.setdefault(
                sampling_detail_tuple.given_name,
                {
                    "sampleset": sampling_detail_tuple.sampleset,
                    "types": [sampling_detail_tuple.is_test_home],
                    "homes": {},
                    "href": sampleset_href,
                },
            )
            app_log.info(strings.USING_SAMPLESET.format(sampleset=sampling_detail_tuple.given_name))

            if sampling_detail_tuple.sampleset:
                links["Sample Set"] = link_string.format(
                    **{
                        "url": sampling_detail_tuple.url,
                        "object": sampling_detail_tuple.sampleset,
                    }
                )
                app_log.set_flags(links=links)

                is_compatible, report = sampling_detail_tuple.sampleset.is_compatible(
                    subdivision=subdivision, builder_org=builder, return_report=True
                )
                if not report["subdivision"]["is_compatible"]:
                    if hasattr(report["subdivision"]["value"], "model"):
                        app_log.error(
                            sampleset_strings.MULTIPLE_METROS_IN_USE.format(
                                metros=", ".join(
                                    list(str(o) for o in report["subdivision"]["value"])
                                )
                            )
                        )
                    else:
                        if (
                            subdivision
                            and subdivision.metro != report["subdivision"]["value"].metro
                        ):
                            app_log.error(
                                sampleset_strings.MISMATCHED_METROS_SUBDIVISIONS.format(
                                    subdivision=subdivision,
                                    url=subdivision.get_absolute_url(),
                                    metro=subdivision.metro,
                                    ss_metro=report["subdivision"]["value"].metro,
                                )
                            )
                        else:
                            app_log.error(
                                sampleset_strings.INCOMPATIBLE_SUBDIVISION.format(
                                    sampleset=sampling_detail_tuple.given_name,
                                    subdivision=subdivision,
                                    url=report["subdivision"]["value"].get_absolute_url(),
                                )
                            )
                if not report["builder_org"]["is_compatible"]:
                    app_log.error(
                        sampleset_strings.INCOMPATIBLE_BUILDER_ORG.format(
                            home_builder=builder, sampleset_builder=report["builder_org"]["value"]
                        )
                    )
            else:
                if (
                    sampling_detail_tuple.name_type == "uuid"
                    and SampleSet.objects.filter(uuid=sampling_detail_tuple.given_name).exists()
                ):
                    msg = sampleset_strings.SAMPLESET_UUID_CONFLICT
                    app_log.error(msg.format(sampleset=sampling_detail_tuple.given_name))

        # Normalize geo variables
        try:
            if subdivision:
                city = city or subdivision.city
                county = county or subdivision.county
                state = state or subdivision.state
            else:
                city = city
                county = county or city.county
                state = state or county.state if county else None
        except AttributeError:
            pass
        if state is None:
            log.error(strings.NO_US_STATE)
        if None in [city, county, state]:
            continue

        # Errors have already been raised at this point..
        if {None} == {builder, subdivision}:
            continue

        def get_expected_string(value):
            value = value if value is not None else ""
            if not isinstance(value, str):
                if represents_integer(str(value)):
                    value = str(int(value))
                else:
                    value = str(value)
            return value

        lot_number = get_expected_string(result.get("lot_number"))
        zipcode = get_expected_string(result.get("zipcode"))

        home = Home.objects.verify_for_user(
            lot_number=lot_number,
            street_line1=result.get("street_line1"),
            street_line2=result.get("street_line2"),
            city=city,
            county=county,
            state=state,
            zipcode=zipcode,
            is_multi_family=result.get("is_multi_family"),
            alt_name=result.get("alt_name"),
            subdivision=subdivision,
            builder=builder,
            bulk_uploaded=True,
            user=user,
            log=app_log,
        )

        app_log.set_flags(home_updated=bool(home))
        if home:
            links["Home"] = link_string.format(
                url=home.get_absolute_url(), object=" ".join(["{}".format(home), str(home.id)])
            )
            app_log.set_flags(links=links)

        if eep_program and eep_program.owner.slug == "aps":
            # This should be enough to get us a preliminary program
            try:
                calculator = APSCalculator(
                    us_state=state,
                    electric_utility="aps",
                    floorplan=floorplan,
                    program=eep_program.slug,
                    notify_on_program_change=False,
                )
            except APSInputException as issues:
                for issue in issues[1:]:
                    if not app_log.has_message(issue):
                        app_log.error(issue)
            else:
                if not calculator.errors:
                    eep_program = calculator.eep_program
                else:
                    for issue in calculator.errors:
                        if not app_log.has_message(issue):
                            app_log.error(issue)
                    for issue in calculator.warnings:
                        if not app_log.has_message(issue):
                            app_log.warning(issue)

        # Add the home info to the sampleset tracking that it wants to be a part of
        if sampling_detail_tuple.use_sampling:
            home_str = ",".join(
                [
                    lot_number,
                    str(result.get("street_line1")),
                    str(city),
                    zipcode,
                ]
            )
            sampleset_tracking["homes"][home_str] = home
            sampleset_tracking["types"].append(sampling_detail_tuple.is_test_home)

        rater_of_record = None
        if result.get("rater_of_record"):
            rater_of_record_lookup = result.get("rater_of_record")
            try:
                lookup = Q(id=rater_of_record_lookup) | Q(rater_id=rater_of_record_lookup)
                rater_of_record = user.company.users.get(lookup)
            except User.DoesNotExist:
                warning_message = 'Could not find Rater of Record with RESNET RTIN or Axis ID "%s"'
                app_log.warning(warning_message, rater_of_record_lookup)
            except ValueError:
                warning_message = (
                    'Could not find Rater of Record with provided "%s". '
                    "Please provide an Axis ID or RESNET RTIN."
                )
                app_log.warning(warning_message, rater_of_record_lookup)
            else:
                app_log.info("Using %s as rater of record", rater_of_record)

        home_stat = EEPProgramHomeStatus.objects.verify_for_company(
            home=home,
            eep_program=eep_program,
            floorplan=floorplan,
            company=company,
            user=user,
            is_billable=result.get("is_billable"),
            ignore_missing=only_bulk_home_processing,
            ignore_missing_floorplan=floorplan_ignore,
            overwrite_old_answers=overwrite_old_answers,
            rater_of_record=rater_of_record,
            log=app_log,
        )

        if home_stat and home_stat.certification_date and home_stat.state == "complete":
            # TODO: why are we allowing them to overwrite answers for certified homes?
            if overwrite_old_answers is False:
                _url = reverse("home:view", kwargs={"pk": home.id})
                app_log.info(
                    strings.ALREADY_COMPLETE.format(
                        program=home_stat.eep_program, home=home_stat.home, url=_url
                    )
                )
                # Proof that the home exists, it has a certification date, and has been certified.
                app_log.set_flags(
                    home_updated=True, certification_date=True, home_already_certified=True
                )
                continue

        construction_stage = None
        if result.get("construction_stage"):
            stages = ConstructionStage.objects.filter_by_user(
                user=user, name__iexact=result.get("construction_stage")
            )
            if stages.count() != 1:
                valid_stages = ConstructionStage.objects.filter_by_user(user).values_list(
                    "name", flat=True
                )
                app_log.error(
                    strings.UNIDENTIFIABLE_CONSTRUCTION_STAGE.format(
                        stage=result.get("construction_stage"), valid_stages=", ".join(valid_stages)
                    )
                )
            else:
                construction_stage = stages[0]
                if home:
                    current = home.get_current_stage(user)
                    if current and construction_stage.order < current.stage.order:
                        app_log.error(
                            strings.CONSTRUCTION_STAGE_ORDER.format(
                                stage=current, new_stage=construction_stage
                            )
                        )

            if certification_date:
                if construction_stage.order < 100:
                    app_log.warning(strings.INVALID_CONSTRUCTION_STAGE)

        if subdivision and subdivision.is_multi_family:
            if not result.get("is_multi_family"):
                app_log.error(
                    strings.HOME_MULTIFAMLY_MISMATCH,
                    {
                        "url": subdivision.get_absolute_url(),
                        "subdivision": subdivision,
                    },
                )

        # If the sampleset already exists in the system, we can do an early check to verify that
        # there is space for the new homestatus to be added to it.
        # We can't provide this error sooner because we need to know that the homestatus itself
        # doesn't yet exist in the system.
        if sampling_detail_tuple.sampleset and home_stat is None:
            if sampling_detail_tuple.sampleset.is_full():
                app_log.error(
                    strings.ADDING_HOME_EXCEEDS_SAMPLESET.format(
                        **{
                            "sampleset": sampling_detail_tuple.repr,
                            "url": sampling_detail_tuple.url,
                        }
                    )
                )

        app_log.set_flags(links=links)

        final_results.append(
            {
                "collector": collector,
                "input_payloads": payloads,
                "lot_number": lot_number,
                "zipcode": zipcode,
                "state": state,
                "county": county,
                "city": city,
                "builder": builder,
                "community": community,
                "subdivision": subdivision,
                "eep_program": eep_program,
                "floorplan": floorplan,
                "result": result,
                "row_number": result["row_number"],
                "questions": questions,
                "annotations": annotations,
                "home": home,
                "home_stat": home_stat,
                "construction_stage": construction_stage,
                "certification_date": certification_date,
                "sub_stat": sub_stat,
                "floorplan_ignore": floorplan_ignore,
                # 'sampleset': sampleset,
                # 'sampleset_str': sampleset_str,
                "sampling_detail_tuple": sampling_detail_tuple,
                "upload_to_registry": True if result.get("upload_to_registry") == "Yes" else False,
                "rater_of_record": rater_of_record,
            }
        )

    # ==============================================================================================
    # STAGE 2.5 - validate sampleset
    # ==============================================================================================

    for ss_config in sampling_elegibility_dict.values():
        report = validate_bulk_sampleset_configuration(ss_config)
        for level, messages in report.items():
            for message in messages:
                getattr(app_log, level)(message)

    if app_log.has_errors:
        app_log.update_model(throttle_seconds=None)
        return

    app_log.set_context(row=None)
    # --------
    # TODO: info message (Does the user need to know this?)
    app_log.info(strings.STAGE_TWO_COMPLETE)

    # ----------------------------------------------------------------------------------------------
    # STAGE THREE - setting homes
    # ----------------------------------------------------------------------------------------------

    to_be_certified = []
    samplesets_to_be_certified = []

    for item in final_results:
        stat = "{}/{}".format(final_results.index(item) + 1, len(final_results))
        app_log.set_context(row=None)
        app_log.debug("%s Processing home - Starting", stat)
        app_log.set_context(row=item.get("row_number"))
        result = item.get("result")
        lot_number = item.get("lot_number")
        zipcode = item.get("zipcode")
        state = item.get("state")
        city = item.get("city")
        county = item.get("county")
        builder = item.get("builder")
        subdivision = item.get("subdivision")
        home = item.get("home")
        construction_stage = item.get("construction_stage")
        certification_date = item.get("certification_date")
        floorplan_ignore = item.get("floorplan_ignore")
        eep_program = item.get("eep_program")
        sampling_detail_tuple = item.get("sampling_detail_tuple")
        rater_of_record = item.get("rater_of_record")
        payloads = item.get("input_payloads")
        links = {}
        created_sampleset = None  # Stays None if detailtuple.sampleset is available

        collector = item["collector"]

        if not home:
            home, created = Home.objects.verify_and_create_for_user(
                lot_number=lot_number,
                street_line1=result.get("street_line1"),
                street_line2=result.get("street_line2"),
                city=city,
                county=county,
                state=state,
                zipcode=zipcode,
                is_multi_family=result.get("is_multi_family"),
                alt_name=result.get("alt_name"),
                subdivision=subdivision,
                builder=builder,
                bulk_uploaded=True,
                user=user,
                log=app_log,
                create=True,
            )
            # answer to whether the home was created
            app_log.set_flags(home_created=created)

            if subdivision:
                sub_cos = list(subdivision.relationships.all().values_list("company_id", flat=True))
                home_cos = list(home.relationships.all().values_list("company_id", flat=True))
                if set(sub_cos) != set(home_cos):
                    log.info(
                        strings.UPDATE_SUBDIVISION_RELATIONSHIPS_DIFFERENT_FROM_HOME.format(
                            stat=stat
                        )
                    )
                    subdivision.save()

        if home:
            links["Home"] = link_string.format(
                url=home.get_absolute_url(), object=" ".join(["{}".format(home), str(home.id)])
            )
            app_log.set_flags(links=links)

            item.update({"home": home})

        if subdivision:
            links["Subdivision"] = link_string.format(
                url=subdivision.get_absolute_url(), object=subdivision
            )
            app_log.set_flags(links=links)

        if sampling_detail_tuple.use_sampling:
            if sampling_detail_tuple.sampleset:
                app_log.info(
                    sampleset_strings.FOUND_SAMPLESET.format(
                        **{
                            "url": sampling_detail_tuple.url,
                            "sampleset": sampling_detail_tuple.sampleset,
                        }
                    )
                )
                links["Sample Set"] = link_string.format(
                    **{
                        "url": sampling_detail_tuple.url,
                        "object": sampling_detail_tuple.sampleset,
                    }
                )
                app_log.set_flags(links=links)
                item.update({"sample_set": sampling_detail_tuple.sampleset})
            else:
                app_log.debug("{} Processing home {} - Setting Sampleset".format(stat, home))
                created_sampleset, create = SampleSet.objects.get_or_create(
                    **{
                        # Create with uuid or with alt_name, depending on what was given
                        sampling_detail_tuple.name_type: sampling_detail_tuple.given_name,
                        "owner": company,
                    }
                )
                links["Sample Set"] = link_string.format(
                    **{
                        "url": created_sampleset.get_absolute_url(),
                        "object": created_sampleset,
                    }
                )
                app_log.set_flags(links=links)
                app_log.info(
                    sampleset_strings.CREATED_SAMPLESET.format(
                        url=created_sampleset.get_absolute_url(), sampleset=created_sampleset
                    )
                )

        floorplan = item.get("floorplan")
        home_stat = item.get("home_stat")
        if not home_stat:
            app_log.debug("%s Processing home %s - Setting House Status", stat, home)
            home_stat = EEPProgramHomeStatus.objects.verify_and_create_for_user(
                home=home,
                eep_program=eep_program,
                floorplan=floorplan,
                company=company,
                user=user,
                is_billable=result.get("is_billable"),
                ignore_missing=only_bulk_home_processing,
                ignore_missing_floorplan=floorplan_ignore,
                rater_of_record=rater_of_record,
                create=True,
                log=app_log,
            )
            item.update({"home_stat": home_stat})

            # Sumthin busted, error already reported via verify_and_create_for_user
            if not home_stat:
                continue

            app_log.set_flags(program_added=True)

            if floorplan and not home_stat.floorplans.filter(id=home_stat.floorplan.id).exists():
                home_stat.floorplans.add(home_stat.floorplan)
                update_home_states(
                    eepprogramhomestatus_id=home_stat.id, user_id=user.id, log=app_log
                )
                home_stat.validate_references()

            if subdivision and subdivision.is_multi_family:
                from axis.qa.models import QARequirement

                qastatus = subdivision.qastatus_set.filter(
                    requirement__eep_program=eep_program
                ).first()
                try:
                    qa_company = home.relationships.get(company__company_type="qa").company
                except Exception:
                    qa_company = None
                single_provider = len(home_stat.get_providers()) == 1
                available_req = (
                    QARequirement.objects.filter_for_add(subdivision, user)
                    .filter(eep_program=eep_program, type="file")
                    .first()
                )

                if qa_company and single_provider and available_req and not qastatus:
                    qastatus = subdivision.qastatus_set.create(
                        **{
                            "owner": qa_company,
                            "requirement": available_req,
                        }
                    )
                    url = subdivision.get_absolute_url() + "#/tabs/qa"
                    context = {
                        "subdivision": "{}".format(subdivision),
                        "url": url,
                        "program": available_req.eep_program.name,
                    }
                    QASubdivisionGatingCertification(url=url).send(
                        context=context,
                        company=qa_company,
                    )
                if qastatus:
                    item["qa_status"] = qastatus
        else:
            save_home_stat = False
            if rater_of_record:
                home_stat.rater_of_record = rater_of_record
                save_home_stat = True
                app_log.info("Rater of Record updated to {}".format(home_stat.rater_of_record))

            if floorplan:
                if not home_stat.floorplan:
                    home_stat.floorplan = floorplan
                    save_home_stat = True
                    app_log.info(
                        "Assigning Floorplan %s to %s with no previous Floorplan",
                        floorplan,
                        home_stat,
                    )
                else:
                    if home_stat.floorplan != floorplan:
                        app_log.warning(
                            "Previous Floorplan %s already exists, not attaching %s",
                            home_stat.floorplan,
                            floorplan,
                        )

            if save_home_stat:
                home_stat.save()

        # Ensure collector is pointed at the current homestatus's collection_request
        if eep_program.collection_request and not home_stat.collection_request:
            home_stat.set_collection_from_program()
            home_stat = EEPProgramHomeStatus.objects.get(id=home_stat.id)
        if home_stat.collection_request:
            collector.collection_request = home_stat.collection_request

        # Add the homestatus to the sampleset
        if sampling_detail_tuple.use_sampling:
            sampleset_obj = sampling_detail_tuple.sampleset or created_sampleset
            preexisting_sampleset = SampleSet.objects.exclude(id=sampleset_obj.id).filter(
                samplesethomestatus__home_status=home_stat
            )
            if preexisting_sampleset:
                if preexisting_sampleset[0].confirm_date:
                    msg = sampleset_strings.CANNOT_MOVE_HOME_FROM_CERTIFIED_SAMPLESET
                    app_log.info(
                        msg,
                        {"old_sampleset": preexisting_sampleset, "new_sampleset": sampleset_obj},
                    )
                else:
                    home_stat.remove_from_sampleset()
            sampleset_obj.add_home_status(
                home_stat, is_test_home=sampling_detail_tuple.is_test_home
            )

        sub_stat = item.get("sub_stat")
        if not sub_stat:
            app_log.debug("%s Processing home %s - Setting Subdivision Status", stat, home)
            sub_stat = EEPProgramSubdivisionStatus.objects.verify_and_create_for_company(
                subdivision=subdivision,
                eep_program=eep_program,
                company=company,
                create=True,
                log=app_log,
            )

        if home_stat.certification_date is not None:
            app_log.set_flags(home_already_certified=True)
            app_log.info(
                strings.SKIPPING_CERTIFIED_HOME.format(
                    url=home_stat.home.get_absolute_url(), home=home_stat.home
                )
            )
            continue

        answer_date = now()
        if construction_stage is not None:
            try:
                ConstructionStatus.objects.get_or_create(
                    stage=construction_stage,
                    home=home,
                    company=company,
                    defaults={
                        "start_date": answer_date,
                    },
                )
            except ConstructionStatus.MultipleObjectsReturned:
                app_log.warning(
                    f"Multiple construction stages for home {home.id} exist for stage "
                    f"{construction_stage} and company {company}"
                )

        if certification_date:
            app_log.debug("Using certification date for the answer date")
            answer_date = certification_date

        questions = item.get("questions")

        app_log.debug("%s Processing home %s - Setting Answers", stat, home)

        if collector:
            for payload in payloads:
                collector.store(**payload)
        else:
            answer_questions_for_home(
                questions=questions,
                home_stat=home_stat,
                answer_date=answer_date,
                company=company,
                user=user,
                ignore_missing=only_bulk_home_processing,
                overwrite_old_answers=overwrite_old_answers,
                log=app_log,
            )

        # FIXME: This message gets added in 'answer_questions_for_home'. Currently not known why.
        # Something to do with ConstructionStatus.
        app_log.remove_message("Setting the home construction stage")

        app_log.debug("%s Processing home %s - Answers have been set", stat, home)

        annotations = item.get("annotations")
        if annotations:
            assert isinstance(
                annotations, list
            ), "Bizzare we didn't get a list for our annotations?"

            app_log.debug("%s Processing home %s - Setting annotations", stat, home)
            set_annotations_for_home(
                annotations=annotations, home_stat=home_stat, user=user, log=app_log
            )

        # Update all the stats
        if home_stat.pct_complete < 99.9:
            update_home_stats(eepprogramhomestatus_id=home_stat.id, log=app_log)
            home_stat = EEPProgramHomeStatus.objects.get(id=home_stat.id)
        if home_stat.state != "complete":
            update_home_states(eepprogramhomestatus_id=home_stat.id, user_id=user.id, log=app_log)
            home_stat = EEPProgramHomeStatus.objects.get(id=home_stat.id)

        home_stat.validate_references()

        app_log.debug("%s Processing home %s - States and Stats have been updated", stat, home)
        app_log.set_flags(links=links)

        errors = home_stat.report_eligibility_for_certification()

        for error in errors:
            app_log.warning(error)

        can_certify = home_stat.can_user_certify(user, perform_eligiblity_check=False)
        if not can_certify:
            app_log.info(
                strings.UNABLE_TO_CERTIFY_INVALID_USER.format(
                    user=user.get_full_name(),
                    url=home_stat.home.get_absolute_url(),
                    home=home_stat.home,
                )
            )

        if not len(errors) and (certification_date and can_certify):
            if sampling_detail_tuple.use_sampling:
                samplesets_to_be_certified.append(
                    {
                        "row_number": item["row_number"],
                        "home_status": home_stat,
                        "certification_date": certification_date,
                        "sampleset": sampling_detail_tuple.sampleset or created_sampleset,
                        "upload_to_registry": item["upload_to_registry"],
                    }
                )
            else:
                to_be_certified.append(
                    {
                        "row_number": item["row_number"],
                        "home_status": home_stat,
                        "certification_date": certification_date,
                        "upload_to_registry": item["upload_to_registry"],
                    }
                )

    app_log.debug(strings.STAGE_THREE_COMPLETE)

    # ----------------------------------------------------------------------------------------------
    # STAGE FOUR - Certifications
    # ----------------------------------------------------------------------------------------------

    total_certified = 0
    num_items = len(to_be_certified)
    for i, item in enumerate(to_be_certified):
        app_log.set_context(row=item["row_number"])
        app_log.debug(
            "%s/%s Processing certification %s - Setting Certification Date", i + 1, num_items, home
        )
        errors = certify_single_home(
            user,
            item["home_status"],
            item["certification_date"],
            bulk_upload=True,
            upload_to_registry=item["upload_to_registry"],
        )

        for error in errors:
            app_log.error(strings.CERTIFY_ERROR.format(error=error))

        if not errors:
            item["home_status"] = EEPProgramHomeStatus.objects.get(id=item["home_status"].id)
            assert item["home_status"].certification_date is not None, "no errors but no cert date"
            total_certified += 1
            url = item["home_status"].home.get_absolute_url()
            app_log.info(
                strings.CERTIFIED_HOME.format(
                    url=url, home=item["home_status"].home, date=certification_date
                )
            )

    processed_samplesets = set()
    for i, item in enumerate(samplesets_to_be_certified):
        app_log.set_context(row=item["row_number"])
        if item["sampleset"] in processed_samplesets:
            app_log.info(
                strings.ALREADY_HANDLED.format(
                    **{
                        "home": item["home_status"].home,
                        "home_url": item["home_status"].home.get_absolute_url(),
                        "sampleset": item["sampleset"],
                        "sampleset_url": item["sampleset"].get_absolute_url(),
                    }
                )
            )
            continue

        app_log.debug("Certifing sampleset {!r}".format(item["sampleset"]))
        report = certify_sampleset(
            user,
            item["sampleset"],
            item["certification_date"],
            bulk_upload=True,
            upload_to_registry=item["upload_to_registry"],
        )

        # Sampleset messages can show up in General, instead of cramming them all into an arbitrary
        # home's logging row.
        app_log.set_context(row=None)
        for level, messages in report.items():
            for message in messages:
                getattr(app_log, level)(message)
        processed_samplesets.add(item["sampleset"])

    app_log.set_context(row=None)

    # ----------------------------------------------------------------------------------------------
    # STAGE FIVE - double check
    # ----------------------------------------------------------------------------------------------
    for item in final_results:
        if not item["home_stat"]:
            continue

        row = item["row_number"]
        app_log.set_context(row=row)

        # Check that the home was certified
        certification_date = item["certification_date"]
        stat = EEPProgramHomeStatus.objects.get(id=item["home_stat"].id)
        certified = bool(certification_date and stat.certification_date)

        if certified:
            app_log.remove_message(r"Unable to certify", use_current_context=True)
            app_log.remove_message(
                sampleset_strings.NO_TEST_HOMES.format(sampleset=".*?"), use_current_context=True
            )
            app_log.remove_message(
                r"unprovided questions? from its sampleset", use_current_context=True
            )

        flags = {"home_certified": certified}
        app_log.set_flags(**flags)

        # Go back and check homes with samplesets.
        # If there is a test home in their sampleset, remove the note saying there's not.
        sample_set = item.get("sample_set")
        if sample_set:
            if sample_set.samplesethomestatus_set.filter(is_test_home=True).count():
                app_log.remove_message("Sampled House does not have a Test Home")

    #
    # -- End of real work --
    #
    app_log.set_context(row=None)

    try:
        elapsed = app_log.storage[-1].created - app_log.storage[0].created
    except (IndexError, AttributeError):
        log.warning("No elapsed time")
        elapsed = 1e-100

    if total_certified > 0:
        msg = (
            "Completed processing {document} for {company} {total_certified}/{total_homes}"
            " certified. Total Time: {elapsed:.2f}s Avg: {avg:.2f}s/home"
        )
    else:
        msg = (
            "Completed processing {total_homes} homes in {document} for {company}. "
            "Total Time: {elapsed:.2f}s Avg: {avg:.2f}s/home"
        )
    avg = elapsed / float(total_homes)
    app_log.debug(
        msg.format(
            document=result_object.filename(),
            company=company,
            total_certified=total_certified,
            total_homes=total_homes,
            elapsed=elapsed,
            avg=avg,
        )
    )
    app_log.update_model(throttle_seconds=None)
    return msg.format(
        document=result_object.filename(),
        company=company,
        total_certified=total_certified,
        total_homes=total_homes,
        elapsed=elapsed,
        avg=avg,
    )


@shared_task(ignore_results=True)
def confirm_answers(user_id, answer_ids, home_href, sampleset_href=""):
    from axis.home.signals import update_stats_on_answer
    from axis.checklist.models import Answer

    log = logger

    # evaluate any ValuesListQuerySets
    answer_ids = list(answer_ids)

    if not len(answer_ids):
        return

    # Grab the instances we need
    user = User.objects.get(id=user_id)
    answers = Answer.objects.filter(id__in=answer_ids)

    # helper for confirming answers
    def confirm_answer(answer):
        answer.confirmed = True
        answer.save()

    # Disconnect signal
    post_save.disconnect(update_stats_on_answer, sender=Answer)

    # Do work
    current_time = None
    answer_count = len(answer_ids)
    for i, answer in enumerate(answers):
        if not current_time or time.time() - current_time > 5:
            current_time = time.time()
            log.debug(
                "Locking %s/%s answers%s on %s", i + 1, answer_count, sampleset_href, home_href
            )
        confirm_answer(answer)

    modified_dates = set(
        Answer.objects.filter(id__in=answer_ids).values_list("modified_date", flat=True)
    )
    Answer.history.model.objects.filter(modified_date__in=modified_dates, id__in=answer_ids).update(
        history_user=user
    )

    # Reconnect signal
    post_save.connect(update_stats_on_answer, sender=Answer)
