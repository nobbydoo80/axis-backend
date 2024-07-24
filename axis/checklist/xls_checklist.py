"""home_checklist.py: Django checklist"""


import datetime
import logging
import os
import random
import re
import shutil
import tempfile
import textwrap
import time
from tempfile import gettempdir

from django.utils.text import slugify
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.packaging.core import DocumentProperties
from openpyxl.styles import colors, fills, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

__author__ = "Steven Klass"
__date__ = "9/5/13 9:42 PM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

log = logging.getLogger(__name__)

# This is a list of cell / value pairs where if the found value is a value listed below ingore it.

valid_field_err = "'{value}' is not valid {type} field for {field} ({cell})"
minimum_value_err = (
    "'{value}' is less than minimum allowed value ({min_value}) field for {field} ({cell})"
)
maximum_value_err = (
    "'{value}' is greater than maximum allowed value ({max_value}) field for {field} ({cell})"
)

LOGO = os.path.abspath(os.path.dirname(__file__) + "/../core/static/images/Logo_Only_48.png")


class XLSItem(object):
    """This is one item"""

    def __init__(self, *args, **kwargs):
        self.name = kwargs.get("name", None)
        self.slug = kwargs.get("slug", slugify(self.name))
        self.value = kwargs.get("value", None)

        self.cell = kwargs.get("cell", None)
        self.choices = kwargs.get("choices", [])
        self.choice_map = kwargs.get("choice_map", None)
        self.choice_allow_any = kwargs.get("choice_allow_other", False)
        self.validate_choice_map_against_choices = kwargs.get("choice_map_only", False)

        self.required = kwargs.get("required", False)
        self.minimum_value = kwargs.get("minimum_value", None)
        self.maximum_value = kwargs.get("maximum_value", None)
        self.input_validator = kwargs.get("input_validator", False)
        self.attr_name = kwargs.get("attr_name", re.sub(r"-", "_", self.slug))

        assert self.slug is not None, "You must input a slug"
        assert self.name is not None, "You must input a name"

        for k, v in kwargs.items():
            if k not in self.__dict__.keys():
                setattr(self, k, v)

    def __repr__(self):
        value = self.value if self.value else "--"
        return "{name} ({cell}): '{value}'".format(name=self.name, cell=self.cell, value=value)

    def validate_choice(self, value):
        if self.choice_map and isinstance(self.choice_map, dict):
            for key, values in self.choice_map.items():
                if "{}".format(value).lower() in [
                    k.lower() for k in ["{}".format(v) for v in values] + [key]
                ]:
                    value = key if key != "__NONE__" else None
                    if not self.validate_choice_map_against_choices:
                        self.value = value
                        return
            if not self.validate_choice_map_against_choices:
                self.value = None
                err = (
                    "'{value}' is not in the available choices for field {field} ({cell}).  "
                    "Available choices are: {choices}"
                )
                _choices = self.choices

                # log.warning("REMOVE ME - NO validate_choice_map_against_choices - %r not in %f", value, _choices)

                if (
                    isinstance(self.choices, (list, tuple))
                    and len(self.choices)
                    and isinstance(self.choices[0], (list, tuple))
                ):
                    _choices = [x[0] for x in self.choices]

                raise ValueError(
                    err.format(
                        value=value, choices=", ".join(_choices), field=self.name, cell=self.cell
                    )
                )

        if len(self.choices) and isinstance(self.choices[0], (list, tuple)):
            self.value = next(
                (x[1] for x in self.choices if x[0].lower() == "{}".format(value).lower()), None
            )
        else:
            self.value = next(
                (x for x in self.choices if x.lower() == "{}".format(value).lower()), None
            )

        if self.value is None:
            err = (
                "'{value}' is not in the available choices for field {field} ({cell}).  "
                "Available choices are: {choices}"
            )

            # log.warning("REMOVE ME - %r not in %f", value, self.choices)

            _choices = self.choices
            if (
                isinstance(self.choices, (list, tuple))
                and len(self.choices)
                and isinstance(self.choices[0], (list, tuple))
            ):
                _choices = [x[0] for x in self.choices]
            raise ValueError(
                err.format(
                    value=value, choices=", ".join(_choices), field=self.name, cell=self.cell
                )
            )

    def validate_basestring(self, value):
        if not isinstance(value, str):
            try:
                value = "{value}".format(value=value)
            except ValueError:
                raise ValueError(
                    valid_field_err.format(
                        type="string", value=self.value, field=self.name, cell=self.cell
                    )
                )
        self.value = value

    def validate_min_value(self, value):
        if self.minimum_value:
            if float(value) < self.minimum_value:
                err = minimum_value_err.format(
                    min_value=self.minimum_value, value=self.value, field=self.name, cell=self.cell
                )
                raise ValueError(err)

    def validate_max_value(self, value):
        if self.maximum_value:
            if float(value) > self.maximum_value:
                err = maximum_value_err.format(
                    max_value=self.maximum_value, value=self.value, field=self.name, cell=self.cell
                )
                raise ValueError(err)

    def validate_float(self, value, is_pct=False):
        raise_error = False
        if not isinstance(value, float):
            try:
                f = float(value)
            except (TypeError, ValueError):
                raise_error = True
        try:
            m = re.search(r"[+-]?(\d+(\.\d*)?|\.\d+)([eE][+-]?\d+)?", str(value))
            if not m:
                raise_error = True
        except ValueError:
            raise_error = True

        if raise_error:
            raise ValueError(
                valid_field_err.format(
                    type="decimal", value=self.value, field=self.name, cell=self.cell
                )
            )

        value = float(value)
        if is_pct:
            if hasattr(self, "number_format_code") and "%" in self.number_format_code:
                if log.isEnabledFor(logging.DEBUG):
                    log.debug("Upconverting!! %s", value)
                value = float(value) * 100
            else:
                if log.isEnabledFor(logging.DEBUG):
                    log.debug("Leaving alone!! %s", value)
                value = float(value)

        self.validate_min_value(value)
        self.validate_max_value(value)

        self.value = value

    def validate_integer(self, value):
        raise_error = False
        if not isinstance(value, int):
            try:
                int(value)
            except ValueError:
                raise_error = True
        try:
            m = re.search(r"^\d+$", str(value))
            if not m:
                raise_error = True
        except ValueError:
            raise_error = True

        if raise_error:
            raise ValueError(
                valid_field_err.format(
                    type="whole number", value=self.value, field=self.name, cell=self.cell
                )
            )

        self.validate_min_value(value)
        self.validate_max_value(value)
        self.value = float(value)

    def validate_value(self, value=None, skip_required=False):
        if value is None:
            value = getattr(self, "value", None)

        try:
            value = value.strip()
            value = value.replace("\xa0", " ")
        except (AttributeError, UnicodeDecodeError, TypeError):
            pass

        if self.required and value in [None, ""] and skip_required is False:
            raise AttributeError(
                "Missing value on required field {} ({})".format(self.name, self.cell)
            )
        if value:
            if len(self.choices) or self.input_validator == "multiple-choice":
                self.validate_choice(value)
            elif self.input_validator in [float, "float"]:
                is_pct = hasattr(self, "number_format") and "%" in self.number_format
                self.validate_float(value, is_pct=is_pct)
            elif self.input_validator in [int, "integer", "int"]:
                self.validate_integer(value)
            elif self.input_validator in [str, "open"]:
                self.validate_basestring(value)
            else:
                log.warning("Unsupported %r for cell %s", self.input_validator, self.cell)

    def generate_dummy_value(self, keep_existing=True, only_required=False):
        if hasattr(self, "value") and self.value is not None and keep_existing:
            return
        required = self.required if hasattr(self, "required") else True
        if required and only_required:
            return
        if hasattr(self, "object_model"):
            use_existing = self.use_existing if hasattr(self, "use_existing") else False
            if use_existing:
                ModelObj = self.object_model.objects.all()[0]
                self.value = getattr(ModelObj, self.object_attribute)
            else:
                self.value = ".".join([self.object_type, self.object_attribute])
        elif hasattr(self, "choices") and len(self.choices):
            choice = random.choice(self.choices)
            self.value = choice[1] if isinstance(choice, (list, tuple)) else choice
        elif self.input_validator in [str, "open"]:
            self.value = self.slug
        elif self.input_validator in [int, "integer", "int"]:
            min_v = self.minimum_value if self.minimum_value else 0
            max_v = self.maximum_value if self.maximum_value else 100
            self.value = random.randint(min_v, max_v)
        elif self.input_validator in [float, "float"]:
            min_v = self.minimum_value if self.minimum_value else 0
            max_v = self.maximum_value if self.maximum_value else 100
            self.value = round(random.uniform(min_v, max_v), 2)
            if "pct" in self.slug:
                max_v = self.maximum_value if self.maximum_value else 1
                self.value = round(random.uniform(min_v, max_v), 2)
        elif self.input_validator == datetime.date:
            self.value = datetime.date(
                random.randint(2010, 2013), random.randint(1, 12), random.randint(1, 28)
            )
        elif self.input_validator == datetime.datetime:
            self.value = datetime.datetime(
                random.randint(2010, 2013),
                random.randint(1, 12),
                random.randint(1, 28),
                random.randint(0, 23),
                random.randint(0, 59),
            )
        else:
            log.warning("Unsupported %s for cell %s", self.input_validator, self.cell)
        log.debug("Setting default value for (%s) %s : %s ", self.cell, self.name, self.value)


class XLSChecklist(object):
    """This is a generic XLS Parser tailored to get data for populating a home"""

    def __init__(self, *args, **kwargs):
        self.filename = kwargs.get("filename", None)
        self.template = kwargs.get("template", None)
        self.sheet_name = kwargs.get("sheet_name", None)
        self.output_prefix = kwargs.get("output_prefix", "Axis_")

        self.creator = kwargs.get("creator", "Axis")
        self.title = kwargs.get("title", "Axis Document")
        self.subject = kwargs.get("subject", "Axis Generated Document")
        self.description = kwargs.get("description", "Axis Generated Document")

        self.user = kwargs.get("user", None)

        self.input_data = []
        self.input_clean = []
        self.row = 1
        self.column = 1

        self.log = kwargs.get("log")
        if self.log is None:
            self.log = logging.getLogger("XLSChecklist")

            def _xxx(*args, **kw):
                pass

            self.log.set_context = _xxx
            self.log.update_model = _xxx
        if kwargs.get("loglevel"):
            self.log.setLevel(kwargs.pop("loglevel"))
        self.start_time = kwargs.get("start_time", time.time())

    def set_cell_title_style(self, cell):
        cell.font = Font(name="Arial", size=18, bold=True)

    def set_cell_default_style(self, cell, **kwargs):
        cell.font = Font(name="Arial", size=12, bold=False)
        for key, value in kwargs.items():
            setattr(cell, key, value)

    def set_cell_bold_style(self, cell, **kwargs):
        cell.font = Font(name="Arial", size=12, bold=True)
        for key, value in kwargs.items():
            setattr(cell, key, value)

    def set_cell_header_style(self, cell, **kwargs):
        cell.font = Font(name="Arial", size=12, bold=True, color=colors.WHITE)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color="FF808080")

        for key, value in kwargs.items():
            setattr(cell, key, value)

    def set_cell_italic_small_style(self, cell):
        cell.font = Font(name="Arial", size=10, italic=True)

    def set_cell_large_style(self, cell):
        cell.font = Font(name="Arial", size=14, bold=True)

    def get_absolute_anchor(self, image, top, left, width=None, height=None):
        from openpyxl.drawing.spreadsheet_drawing import (
            AbsoluteAnchor,
            pixels_to_EMU,
            XDRPoint2D,
            XDRPositiveSize2D,
        )

        position = XDRPoint2D(pixels_to_EMU(left), pixels_to_EMU(top))
        width = width if width else image.width
        height = height if height else image.height
        size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
        return AbsoluteAnchor(pos=position, ext=size)

    def add_logo(self, workbook, sheet):
        image = Image(LOGO)
        sheet.add_image(image, anchor=self.get_absolute_anchor(image, 0, 10))

    def get_cell_positions(self, elements):
        """This is used to move the position to the starting position of a data element
        This should be a dictionary name: (column, row) elements."""
        pass

    def set_header_data(self, attributes):
        """Returns a list of items which simply a value and a cell"""
        self.header_data = []
        return self.header_data

    def add_value_set(self, add_header=False, cell_position_method=None):
        """Simply add a value_set"""
        # Create a fresh set of data with a copy of our baseline elements.
        kwargs = self.data_object_kwargs.copy()
        kwargs["cell_item"] = self.cell_item
        data_object = self.data_object(**kwargs)

        if add_header:
            self.header = self.set_header_data(data_object.attributes)

        positions = self.get_cell_positions(data_object.attributes)

        missing_positions = []
        for element in data_object.attributes:
            assert not isinstance(element, dict), "Expecting an object"
            if element.cell is None:
                try:
                    position = positions.get(element.name)
                except (KeyError, AttributeError):
                    if not element.value:
                        missing_positions.append(element.name)
                    continue
                element.cell = "{}".format(position)
        if not len(self.data):
            log.debug(
                "Assigned %s cell positions", len([x for x in data_object.attributes if x.cell])
            )
        if len(missing_positions):
            log.debug(
                "Unable to find %s elements because no position "
                "was found for them - either add them to this map or "
                "remove them %s",
                len(missing_positions),
                ",".join(missing_positions),
            )
        self.data.append(data_object)
        return data_object

    def get_template(self):
        return None

    def open_workbook_and_get_sheet(self, filename=None, sheet_name=None):
        """Open the worksheet"""
        if sheet_name is None:
            sheet_name = self.sheet_name
        if filename is None:
            filename = self.filename

        assert sheet_name is not None, "You need a valid sheet"

        workbook = load_workbook(filename=filename, data_only=True)
        if sheet_name not in workbook.sheetnames:
            if len(workbook.sheetnames) == 1:
                sheet_name = workbook.sheetnames[0]
                msg = "Using sheet name {sheet_name}".format(sheet_name=sheet_name)
                try:
                    if not self.log.has_message(msg, "WARNING"):
                        self.log.warning(msg)
                except AttributeError:
                    self.log.warning(msg)
            else:
                raise NameError(
                    "Sheet name specified '{}'does not exist {}".format(
                        sheet_name, ",".join(workbook.sheetnames)
                    )
                )
        return workbook[sheet_name]

    def pre_save(self, workbook, sheet):
        """This is to do something before the save"""
        pass

    def properties(self):
        """Set the document properties"""
        props = DocumentProperties()
        props.creator = self.creator
        props.title = self.creator
        props.subject = self.subject
        props.description = self.description
        return props

    def _set_valid_choices(self, sheet, cell, choices, show_dropdown=True):
        """Set the validation"""
        if isinstance(choices[0], (list, tuple)):
            valid_choices = [x[0] for x in choices]
        else:
            valid_choices = choices

        for choice in valid_choices:
            if not all(ord(c) < 128 for c in choice):
                log.error("Issue with %s in %s", choice, valid_choices)
                return

        # TODO - FIXME - Why can't I go past 10 items
        if len(valid_choices) > 9:
            valid_choices = valid_choices[:9] + [valid_choices[9][:22]]
        if sum([len(i) for i in valid_choices]) > 245:
            log.warning("FIXME - Why can't I go past 246 items..")

        formula = '"{choices}"'.format(choices=",".join(valid_choices))
        log.debug("Adding cell %s choices for cell %s", len(valid_choices), cell)
        data_validation = DataValidation(
            type="list", formula1=formula, allow_blank=True, showDropDown=True
        )
        sheet.add_data_validation(data_validation)
        data_validation.add(cell)
        data_validation.error = "Your entry is not in the list"
        data_validation.errorTitle = "Invalid Entry"
        data_validation.prompt = "Please select from the list"
        data_validation.promptTitle = "List Selection"

    def _set_validation(self, sheet, cell, **kwargs):
        data_validation = DataValidation(**kwargs)
        sheet.add_data_validation(data_validation)
        return data_validation.add(cell)

    def add_validation_whole_less_than_or_equal(self, sheet, cell, number):
        return self._set_validation(
            sheet, cell, type="whole", operator="lessThanOrEqual", formula1="{}".format(number)
        )

    def add_validation_whole_greater_than_or_equal(self, sheet, cell, number):
        return self._set_validation(
            sheet, cell, type="whole", operator="greaterThanOrEqual", formula1="{}".format(number)
        )

    def add_validation_whole_between(self, sheet, cell, minimum, maximum):
        log.debug("Adding whole validation to %s %s :%s", cell, minimum, maximum)
        return self._set_validation(
            sheet,
            cell,
            type="whole",
            operator="between",
            formula1="{}".format(minimum),
            formula2="{}".format(maximum),
        )

    def add_validation_decimal_less_than_or_equal(self, sheet, cell, number):
        return self._set_validation(
            sheet,
            cell,
            type="decimal",
            operator="lessThanOrEqual",
            formula1="{}".format(round(float(number), 2)),
        )

    def add_validation_decimal_greater_than_or_equal(self, sheet, cell, number):
        return self._set_validation(
            sheet,
            cell,
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="{}".format(round(float(number), 2)),
        )

    def add_validation_decimal_between(self, sheet, cell, minimum, maximum):
        """Sets whole number validation range"""
        log.debug("Adding decimal validation to %s %s:%s", cell, minimum, maximum)
        return self._set_validation(
            sheet,
            cell,
            type="decimal",
            operator="between",
            formula1="{}".format(round(float(minimum), 2)),
            formula2="{}".format(round(float(maximum), 2)),
        )

    def smart_trunctate(self, content, max_length=100, suffix="..."):
        if content is None:
            return
        if len(content) <= max_length:
            return content
        return textwrap.wrap(content, max_length - len(suffix))[0] + suffix

    def write(self, sheet_name=None, output=None):
        """Write this out"""
        if output is None:
            _, output = tempfile.mkstemp(suffix=".xlsx", prefix=self.output_prefix)
        if sheet_name is None:
            sheet_name = self.sheet_name

        assert output, "You need an output file"
        assert sheet_name is not None, "You need a valid sheet"

        if os.path.isfile(output):
            shutil.move(
                output,
                os.path.join(gettempdir(), os.path.basename(output)),
            )

        self.template = self.template if self.template else self.get_template()
        if self.template:
            shutil.copyfile(self.template, output)
            workbook = load_workbook(filename=output, data_only=True)
            sheet = workbook.get_sheet_by_name(self.sheet_name)
        else:
            workbook = Workbook()
            sheet = workbook.create_sheet(index=0, title=self.sheet_name)
        msg = "Adding %s data set%s to %s"
        log.debug(msg, len(self.data), "s" if len(self.data) > 1 else "", output)
        if hasattr(self, "header_data") and len(self.header_data):
            for item in self.header_data:
                cell = sheet.cell(item.cell)
                cell.value = item.value
                style_name = "set_cell_{}_style".format(getattr(item, "cell_style", "default"))
                try:
                    method = getattr(self, style_name)
                    method(cell)
                except AttributeError:
                    log.warning("Unable to find style name for %s", style_name)

        for data_set in self.data:
            for item in data_set.attributes:
                if hasattr(item, "required") and item.required:
                    if not (hasattr(item, "value") and item.value):
                        log.warning("Missing value on required field {%s", item.name)
                if not hasattr(item, "cell") or item.cell is None:
                    log.error("No value found for cell '%s'", item.name)
                    continue
                cell = sheet.cell(item.cell)
                cell.value = item.value
                if hasattr(item, "choices") and len(item.choices):
                    self._set_valid_choices(sheet, cell, item.choices)
                style_name = "set_cell_{}_style".format(getattr(item, "cell_style", "default"))
                try:
                    method = getattr(self, style_name)
                    method(cell)
                except AttributeError:
                    log.warning("Unable to find style name for %s", style_name)

        msg = "Completed adding %s items to %s"
        log.debug(msg, len([y for x in self.data for y in x.attributes]), output)
        workbook.properties = self.properties()
        self.pre_save(workbook, sheet)
        workbook.save(output)
        log.debug("Successfully wrote %s.", output)
        return output

    def generate_dummy_data(self, qty=1, only_required=False):
        """Generate dummy data"""
        log.debug("generating %s data sets", qty)
        for item in range(qty):
            self.add_value_set(add_header=True)
            data_set = self.data[-1]
            for item in data_set.attributes:
                item.generate_dummy_value(keep_existing=False, only_required=only_required)
        log.debug("Generated %s data sets", qty)

    def get_non_data_for_cell(self, cell):
        """This is just a map of values which can be considered None"""
        return None

    def read(self, sheet_name=None):
        assert not len(self.data), "You cannot combine a read with existing data"

        counter = 0
        sheet = self.open_workbook_and_get_sheet(sheet_name=sheet_name)

        while True:
            if counter == 100:
                break
            self.add_value_set()
            data_set = self.data[-1]
            missing_positions = []
            for item in data_set.attributes:
                item.value = sheet.cell(item.cell).value
            prior_cells = []
            if len(self.data) > 1:
                prior_cells = set([getattr(x, "cell", None) for x in self.data[-2].attributes])

            values = set([getattr(x, "value", None) for x in data_set.attributes])
            cells = set([getattr(x, "cell", None) for x in data_set.attributes])
            if values == {None} or cells == prior_cells:
                self.data = self.data[:-1]
                break

            if len(missing_positions):
                log.warning(
                    "Unable to find %s elements because no position " "was found for them %s",
                    len(missing_positions),
                    ",".join(missing_positions),
                )
            counter += 1

        if not len(self.data):
            self.log.error("No data was found in %s", self.filename)
        else:
            self.log.info("Identified %s datasets", len(self.data))

    def validate_data(self):
        pass

    def process_data(self, log=None):
        pass

    def _gather_nondata_validations(self, sheet_name=None):
        """Helper method used to create our validations"""
        validations = []
        sheet = self.open_workbook_and_get_sheet(sheet_name=sheet_name)
        for row in range(1, 60):
            for col in range(1, 60):
                cell_label = "{letter}{row}".format(letter=get_column_letter(col), row=row)
                value = sheet.cell(cell_label).value
                if value:
                    validations.append((cell_label, value))
        # pprint.pprint(validations)

    def update_task_progress(
        self, state="STARTED", steps=None, step=0, current=1, total=1, **kwargs
    ):
        """
        Update an asynchronous tasks state. Primarily used for send data to the client.
        :param state: string of custom state to set Task to.
        :param steps: list of processing steps
        :param step: int current step. 0 based
        :param current: int current progress point in total
        :param total: int total number of progress points
        """
        if hasattr(self, "task") and self.task:
            steps = steps or ["processing", "writing"]

            current_step = steps[step]

            meta = {}
            for s in steps:
                if s == current_step:
                    meta[s] = {"current": current, "total": total}
                elif steps.index(s) > steps.index(current_step):
                    # takes place after current step
                    meta[s] = {"current": 0, "total": 1}
                else:
                    # takes place before current step
                    meta[s] = {"current": 1, "total": 1}

                meta.update(kwargs)
                self.task.update_state(state=state, meta=meta)
