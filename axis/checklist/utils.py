import datetime
import logging
import os.path
import random
import re
import time
from collections import namedtuple, defaultdict, OrderedDict
from operator import itemgetter

from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import MultipleObjectsReturned
from django.db.models.signals import post_save, post_delete
from django.forms import ValidationError
from django.urls import reverse
from django.utils import formats
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.packaging.core import DocumentProperties
from openpyxl.styles import (
    colors,
    fills,
    Font,
    Alignment,
    PatternFill,
    Protection,
    Border,
    Side,
    Color,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation

from axis.annotation.messages import NWESHMeetsOrBeatsAnsweredNo
from axis.company.models import Company
from axis.core.utils import values_to_dict
from axis.eep_program.models import EEPProgram
from axis.filehandling.utils import XLSXParser
from axis.home.signals import update_stats_on_answer
from axis.home.tasks import update_home_stats, update_home_states
from axis.scheduling.models import ConstructionStatus, ConstructionStage
from . import strings
from .collection.excel import BulkExcelChecklistCollector
from .models import Question, Answer, TYPE_CHOICES, CheckList, QAAnswer

__author__ = "Steven Klass"
__date__ = "2/9/12 1:16 PM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

from axis.annotation.utils import validate_annotation_by_name

log = logging.getLogger(__name__)
User = get_user_model()

LOGO = os.path.abspath(os.path.dirname(__file__) + "/../core/static/images/Logo_Only_48.png")
ETO_LOGO = os.path.abspath(
    os.path.dirname(__file__) + "/../customer_eto/static/images/ET_EPS_Logo_White_Horz.png"
)

RANDOM_COMMENT = (
    "Lorem ipsum dolor sit amet, consectetur adipiscing elit. Sed elementum "
    "interdum tincidunt. Pellentesque leo lacus, tincidunt in pretium vel, "
    "interdum vitae nunc."
)

ANSWER_PRIORITY_MAP = {
    "must correct": 1,
    "not applicable": 2,
    "rater verified": 3,
    "builder verified": 4,
    "no": 1,
    "yes": 2,
}

EFL_ANSWER_MARK = "x"  # case insensitive mark expected for EFL answer

MAX_NUM_CHOICES = 25  # Arbitrary limit to the number of chosen responses on a spreadsheet
VALID_AFFIRMATIVE_BULK_FILL_RESPONSES = ["1", "yes", "true"]
VALID_BULK_FILL_RESPONSES = VALID_AFFIRMATIVE_BULK_FILL_RESPONSES + ["0", "no", "false", "", "none"]
IS_OPTIONAL_VALUE_MAP = {
    "yes": True,
    "y": True,
    "x": True,
    "no": False,
    "n": False,
    "": False,
}

REQUIRED_CELL_COLOR = "A5B8E1"
OPTIONAL_CELL_COLOR = "D0D9EE"
GREY_FONT_COLOR = "999999"
WHITE_FONT_COLOR = "FFFFFF"
NON_STORED_CELL_COLOR = "FFCCCC"

REQUIRED_COLUMNS = [
    "street_line1",
    "city",
    "state",
    "zipcode",
    "builder_org",
    "eep_program",
    "is_multi_family",
]

RandomAnswerTuple = namedtuple("RandomAnswerTuple", ["answer", "comment"])
AnsweredQuestionResultTuple = namedtuple(
    "AnsweredQuestionResultTuple", ["questions", "annotations"]
)


def get_random_answer_for_question(question, pct_error=1.5, seed=None):
    """
    Generates an automatic answer for a question based on its type.

    Returns a 2-namedtuple of (answer, comment), where the comment is sometimes ``None``.

    """

    # Sometimes generate a comment
    comment = None
    if random.randint(0, 100) > 95:
        comment = "{} {}".format(seed, RANDOM_COMMENT) if seed else RANDOM_COMMENT

    result = None
    if question.type == "open":
        result = RandomAnswerTuple("Open Answer", comment)
        if seed:
            result = RandomAnswerTuple("{} Open".format(seed), comment)
    elif question.type == "float":
        result = RandomAnswerTuple(3.14, comment)
    elif question.type in ["int", "integer"]:
        result = RandomAnswerTuple(random.randint(0, 9999), comment)
    elif question.type == "datetime":
        result = RandomAnswerTuple(now(), comment)
    elif question.type == "date":
        result = RandomAnswerTuple(datetime.date.today(), comment)
    elif question.type == "multiple-choice":
        # Generates a random choice list of mixed success/failure options and selects one.
        failures = list(question.question_choice.filter(is_considered_failure=True))
        passing = list(question.question_choice.filter(is_considered_failure=False))

        choice_list = []
        if pct_error and len(failures):
            choice_list = failures * int(float(pct_error) / len(failures))

        choice_list += passing * int((100 - len(choice_list)) / len(passing))
        choice = random.choice(choice_list)

        # Make sure comment exists where required.
        if choice.comment_required:
            comment = RANDOM_COMMENT
        result = RandomAnswerTuple(choice.choice, comment)
    elif question.type == "csv":
        result = RandomAnswerTuple("1,2,3", comment)
    elif question.type == "kvfloatcsv":
        result = RandomAnswerTuple("X:1,2,3 Y:1,2,3", comment)
    else:
        raise TypeError("Unknown question type {}!".format(question.type))
    return result


def build_questionanswer_dict(question_queryset, answer_queryset):
    """Returns a dictionary of question ids to their answer ids (or None if unanswered)."""
    question_id_list = list(question_queryset.values_list("id", flat=True))
    answers = answer_queryset.filter(question__id__in=question_id_list).values("id", "question__id")
    data = values_to_dict(answers, "question__id", "id")
    missing_answers = set(question_id_list) - set(data.keys())
    data.update({}.fromkeys(missing_answers))
    return data


def validate_answer(
    question,
    answer,
    comment=None,
    override_comment_requirement=False,
    log=None,
    ajax_error_reporting=False,
):
    """
    Generates a dictionary of normalized data given data from an input upload.

    ValueError can be raised for input data that doesn't completely validate.  The error message
    will contain HTML tags with links to the question that didn't pass inspection.

    Returns a dictionary containing the keys:

        answer, type, is_considered_failure, email_required, comment, error, choice_priority

    Note that the ``error`` key is None when validation is successful.

    """
    log = log if log else logging.getLogger(__name__)

    validator = question.get_validator_for_type(
        **{
            "override_comment_requirement": override_comment_requirement,
            "comment": comment,
            "log": log,
        }
    )
    try:
        answer = validator(answer)
    except ValidationError as e:
        log.warning(
            "Failed to validate %(answer)r: %(messages)r",
            {"answer": answer, "messages": ", ".join(e.messages)},
        )
        if ajax_error_reporting:
            raise ValueError(", ".join(e.messages))
        raise ValueError(
            "Failed to validate %(answer)r: %(messages)r"
            % {
                "answer": answer,
                "messages": ", ".join(e.messages),
            }
        )

    choice_priority = 10
    is_considered_failure, display_as_failure, email_required = False, False, False
    document_required, photo_required = False, False
    if question.type == "multiple-choice":
        choice = question.question_choice.get(choice=answer)
        if choice.is_considered_failure:
            choice_priority = 0
        else:
            choice_priority = ANSWER_PRIORITY_MAP.get(answer.lower())
        email_required = choice.email_required
        is_considered_failure = choice.is_considered_failure
        display_as_failure = choice.display_as_failure
        document_required = choice.document_required
        photo_required = choice.photo_required

    return {
        "answer": answer,
        "type": question.type,
        "is_considered_failure": is_considered_failure,
        "display_as_failure": display_as_failure,
        "email_required": email_required,
        "comment": comment,
        "error": None,
        "choice_priority": choice_priority,
        "document_required": document_required,
        "photo_required": photo_required,
    }


def get_answered_questions_from_data(data, eep_program, ignore_missing=False, log=None):
    """
    Correlates string questions with valid answers, which are the keys and values of ``data``.

    Returns a 2-namedtuple of (questions, annotations), where ``questions`` is a dictionary, and
    ``messages`` is a 3-namedtuple of (errors, warnings, info) of messages that came up during
    execution.

    The ``questions`` dictionary is a mapping of Question objects to a list of validated answers.
    Each such validated answer is itself a dictionary of the return value from
    ``validate_answer()``, which is essentially a valid kwargs dict for a call to get_or_create().
    """
    log = log if log else logging.getLogger(__name__)

    assert isinstance(data, dict), "Datatype must be a dictionary"
    assert isinstance(eep_program, EEPProgram), "Program must be EEProgram"

    questions = defaultdict(list)
    annotations = []

    question_pool = Question.objects.filter_by_eep(eep_program)

    for potential_question, potential_answer in data.items():
        if potential_question is None or potential_answer is None:
            continue

        question_start = None
        choice = None

        # Splits apart the question prefix from the actual question content
        # Example: "CC R1.1 - Test - Comment"
        bits = potential_question.rsplit(" - ", 1)
        if len(bits) == 1:
            question_start = bits[0]
        else:
            question_start, choice = bits
            if question_start.lower() == "annotation":
                annotation_data = validate_annotation_by_name(
                    eep_program, choice, potential_answer, log
                )
                if annotation_data:
                    annotations.append(annotation_data)
                else:
                    log.warning(
                        "%r was not a valid annotation name for the program." % question_start
                    )
                continue
            elif choice and choice.lower() in ["comment", "comments"]:
                continue

        # Find the text question's corresponding object
        try:
            question = question_pool.get(question__istartswith=question_start)
        except Question.DoesNotExist:
            continue
        except Question.MultipleObjectsReturned:
            # Try again with a space on the end to narrow the pool
            try:
                question = question_pool.get(question__istartswith="{} ".format(question_start))
            except Question.MultipleObjectsReturned:
                err = 'Many questions found for "{}" - skipped'
                log.error(err.format(question_start))
                continue
            except Question.DoesNotExist:
                # The first query returned many, but the second returned none.
                # Try again with a period on the end instead of a space.
                try:
                    question = question_pool.get(
                        question__istartswith="{}. ".format(question_start)
                    )  # A bug with OpenPyXL # FIXME: Document what the bug is
                except Question.DoesNotExist:
                    if ignore_missing:
                        issue = 'Unable to find a question which matches "%s" - skipped (%s)'
                        log.info(issue, potential_question, question_start)
                    continue

        # NOTE: Custom EFL handling
        # Here is where we will need to handle multiple answers when EFL gives us many
        # checkboxes we can use the fact that we have "choice".
        # Iterate through them all and find any others.

        if not choice:
            choice = potential_answer
        else:
            # EFL Style answers - only accept it if this is an "x".
            # Note, however, that if the choice happens to be a "YES" in the "YES" column, we let
            # it through as if it was an "x".
            if question.type == "multiple-choice" and potential_answer.lower() != EFL_ANSWER_MARK:
                # Choice is something other than the column name (which is a synonym for "x")
                if potential_answer.lower() != choice.lower():
                    log.warning(
                        "Skipping unintended multiple-choice answer '%(answer)s' - please leave "
                        " this blank or use '%(mark)s' to indicate this as a choice.",
                        {"answer": potential_answer, "mark": EFL_ANSWER_MARK},
                    )
                    continue

        # Pull the 'comment'/'comments' key from the data.
        comment = None
        comment_attr_name = question_start + " - comment"
        comment = data.get(comment_attr_name)
        if not comment:
            comment = data.get(comment_attr_name + "s")

        # Try to validate the answer within the question's context.
        try:
            answer_data = validate_answer(
                question, choice, comment, override_comment_requirement=True, log=log
            )
        except ValueError as e:
            log.warning("Error with answer - %s", e)
            continue
        else:
            if answer_data.get("answer") in [None, "", []]:
                log.info(
                    "Answer '%(choice)s' will not process as no value was found", {"choice": choice}
                )
                continue

            questions[question].append(answer_data)

    # ---- End of for loop

    for question, answers in questions.items():
        failing = [x for x in answers if x.get("is_considered_failure")]
        passing = [x for x in answers if not x.get("is_considered_failure")]
        if len(failing) > 1:
            log.error("Multiple failing answers for question %s is not allowed", question.pk)
        if len(passing) > 1:
            sorted_answers = sorted(passing, key=itemgetter("choice_priority"))
            log.debug(
                "Multiple passing answers is not recommended; accepted '%(answer)s' "
                "as final answer",
                {"answer": sorted_answers[-1]["answer"]},
            )
            questions[question] = failing + [sorted_answers[-1]]
    log.debug("Validated %s/%s questions", len(questions.keys()), question_pool.count())
    return AnsweredQuestionResultTuple(questions=questions, annotations=annotations)


def answer_questions_for_home(
    user,
    questions,
    home_stat,
    company,
    answer_date,
    log=None,
    ignore_missing=False,
    overwrite_old_answers=False,
):
    from axis.company.models import COMPANY_MODELS
    from axis.home.models import EEPProgramHomeStatus

    log = log if log else logging.getLogger(__name__)

    assert isinstance(company, COMPANY_MODELS), "Company must be of type Company"
    assert isinstance(
        home_stat, EEPProgramHomeStatus
    ), "home_stat must of type EEPProgramHomeStatus"
    assert isinstance(user, User), "User must be of type User"
    assert isinstance(questions, (type(None), dict)), "Questions must be a dict"
    assert isinstance(
        answer_date, (type(None), datetime.datetime)
    ), "If used answer_date must be a of type datetime.datetime"

    home_status_companies = home_stat.home.homestatuses.filter(
        eep_program__is_qa_program=True
    ).values_list("company__id", flat=True)
    answer_model = QAAnswer if company.id in home_status_companies else Answer

    log.debug("Using Model %s" % answer_model.__name__)

    if overwrite_old_answers:
        log.debug("Allowing overwrites for answers.")

    home = home_stat.home
    _url = reverse("home:view", kwargs={"pk": home.id})
    home_url = "<a href='{}'>{}</a>".format(_url, home)

    if questions is None or not len(questions.keys()):
        if ignore_missing:
            return None
        log.info("No answers provided for home %s", home_url)
        return None
    if home_stat.certification_date:
        info_already_certified = "Home %s has already been certified."
        log.info(info_already_certified, home_url)
        return None

    stage_updated = False
    new_answers, answer_ids, reused_answers, overwritten_answers = [], [], [], []

    # Disconnect HomeStatusUpdate - We want to update the states when we are done..
    post_save.disconnect(update_stats_on_answer, sender=answer_model)
    if overwrite_old_answers:
        post_delete.disconnect(update_stats_on_answer, sender=answer_model)

    ctime = None
    for question, answers in questions.items():
        if not ctime or time.time() - ctime > 3:
            ctime = time.time()
            debug_questions_process = "%s/%s questions have been processed for %s"
            log.debug(
                debug_questions_process,
                list(questions.keys()).index(question) + 1,
                len(questions.keys()),
                home_url,
            )

        if not len(answers):
            continue

        # Sort the answers, get all failures, and the last passing answer..
        sorted_answers = sorted(answers, key=lambda k: k["choice_priority"])
        assert len(sorted_answers) <= 2, "At most one passing and one failing is allowed"

        existing_answers = answer_model.objects.filter(home=home, question=question).order_by("-id")

        # Skip examination of new answers unless we're in overwrite mode.
        if not overwrite_old_answers or (existing_answers and not sorted_answers):
            if len(existing_answers):
                existing_answer = existing_answers[0]
                # msg = strings.REUSE_ANSWER
                # log.debug(msg.format(answer=existing_answer, question=question.question[:14]))
                reused_answers.append(existing_answer)
                continue

        for answer_dict in sorted_answers:
            if not stage_updated:
                info_stage_updated = "Setting the home construction stage to {stage}"
                stage = ConstructionStage.objects.get(name="Started", is_public=True, order=1)
                try:
                    status, create = ConstructionStatus.objects.get_or_create(
                        stage=stage,
                        home=home,
                        company=company,
                        defaults=dict(start_date=answer_date),
                    )
                except MultipleObjectsReturned:
                    _stats = ConstructionStatus.objects.filter(
                        stage=stage, home=home, company=company
                    )
                    for idx, _stat in enumerate(list(_stats.all())):
                        if idx == 0:
                            continue
                        _stat.delete()
                log.info(info_stage_updated.format(stage=stage.name))
                stage_updated = True

            # If the answer is new, then it means it's not one of the existing answers, since
            # we fetched the new answer via get_or_create().
            # In this situation, we will remove old answers in favor of the new one(s).
            if overwrite_old_answers and existing_answers:
                existing_answer = existing_answers[0]
                msg = strings.OVERWRITE_ANSWER
                ques_str = "{}...".format(question.question[:14])
                log.debug(
                    msg.format(
                        answer=existing_answer.answer,
                        question=ques_str,
                        new_answer=answer_dict.get("answer"),
                    )
                )
                overwritten_answers.append(existing_answer)
                existing_answers.delete()

            answer_datetime = datetime.datetime.combine(answer_date, datetime.time()).replace(
                tzinfo=datetime.timezone.utc
            )

            answer_obj, created = answer_model.objects.get_or_create(
                user=user,
                question=question,
                home=home,
                answer=answer_dict.get("answer"),
                is_considered_failure=answer_dict.get("is_considered_failure", False),
                display_as_failure=answer_dict.get("display_as_failure", False),
                type=question.type,
                defaults=dict(
                    comment=answer_dict.get("comment"),
                    created_date=answer_datetime,
                    confirmed=False,
                    bulk_uploaded=True,
                ),
            )

            if created:
                new_answers.append(answer_obj)
                historical_answer = answer_obj.history.first()
                if historical_answer and historical_answer.history_user is None:
                    historical_answer.history_user = user
                    historical_answer.save()

            if answer_dict.get("is_considered_failure"):
                log.info("Review of identified failure for '%s' set", answer_obj)
                answer_obj.failure_is_reviewed = True
                answer_obj.save()
                historical_answer = answer_obj.history.first()
                historical_answer.history_user = user
                historical_answer.save()
            answer_ids.append(answer_obj.id)

    # Now update the stats
    post_save.connect(update_stats_on_answer, sender=answer_model)
    if overwrite_old_answers:
        post_delete.connect(update_stats_on_answer, sender=answer_model)

    if len(new_answers) or len(overwritten_answers):
        if home_stat.pct_complete < 99.9:
            update_home_stats(eepprogramhomestatus_id=home_stat.id, log=log)
        if home_stat.state != "complete":
            update_home_states(eepprogramhomestatus_id=home_stat.id, user_id=user.id, log=log)

    home_stat.validate_references()

    msg = strings.BULK_ANSWERING_SUMMARY
    _url = reverse("home:view", kwargs={"pk": home_stat.home.id})
    href = '<a href="{}">{}</a>'.format(_url, home_stat.home)
    log.info(
        msg.format(
            created=len(new_answers),
            overwritten=len(overwritten_answers),
            reused=len(reused_answers),
            home=href,
        )
    )
    return answer_model.objects.filter(id__in=answer_ids)


def set_annotations_for_home(annotations, home_stat, user, log=None):
    log = log if log else logging.getLogger(__name__)
    objects = []
    total = 0
    changed = 0
    content_type = ContentType.objects.get_for_model(home_stat)
    for i, data in enumerate(annotations):
        obj, created = home_stat.annotations.get_or_create(
            content_type=content_type,
            object_id=home_stat.id,
            user=user,
            type=data["type"],
            defaults={
                "content": data["content"],
            },
        )
        if created:
            total += 1
        else:
            changed += 1
            obj.content = data["content"]
            obj.save()
        objects.append(obj)

        if data["type"].slug == "beat-annual-fuel-usage" and obj.content.lower() == "no":
            context = {"url": home_stat.get_absolute_url(), "text": "{}".format(home_stat)}
            NWESHMeetsOrBeatsAnsweredNo(url=home_stat.get_absolute_url()).send(
                user=user, context=context
            )

    if total:
        log.info("Added %d annotations to program status.", total)
    if changed:
        log.info("Changed %d annotations to program status.", changed)

    return objects


def validate_checklist_spreadsheet_format(filename, result_log):
    """
    Inspects the spreadsheet file to make sure it resembles the type of file we want to put through
    deeper processing.

    """

    EXPECTED_SHEET_HEADERS = [
        # format: ('Sheet name', 'Unique header name'),
        ("Choices", "Choice"),
        ("Questions", "Question"),
        ("Sections", "Name"),
        ("Checklist", "Name"),
    ]

    COLUMNS_WE_CARE_ABOUT = {
        "Choices": {
            "Choice",
            "Choice_order",
            "Is_Considered_Failure",
            "Email_Required",
            "Comment_Required",
        },
        "Questions": {
            "Section",
            "Type",
            "Question",
            "Description",
            "Help_URL",
            "Climate_Zone",
            "Unit",
            "Choice1",
            "Choice2",
            "Choice3",
            "Choice4",
            "Choice5",
            "Choice6",
        },
        "Sections": {"Name", "Priority", "Description"},
        "Checklist": {"Name", "Public", "Description"},
    }

    for sheet, uniq_header_column in EXPECTED_SHEET_HEADERS:
        result_log.update(latest="Validating sheet {}".format(sheet))
        try:
            xlsx = XLSXParser(filename=filename, uniq_header_column=uniq_header_column)
        except TypeError:
            err = "This does not appear to be a Microsoft XLSX document"
            result_log.update(errors=err, raise_errors=True)
            return
        try:
            xlsx.load_workbook_and_sheet(sheet_name=sheet)
        except AttributeError:
            err = "Unable to load the workbook with sheet Name - '{}'".format(sheet)
            result_log.update(errors=err)
            continue
        except InvalidFileException:
            err = "Unable to load the file - Is this an 'xlsx' file?"
            result_log.update(errors=err)
            continue

        headers = set(xlsx.get_columns(uniq_column=uniq_header_column))
        if not len(headers):
            err = (
                "Unable to find header row in sheet {} XLSX File. Looking for a column which "
                'exactly matches with "{}"'
            ).format(sheet, uniq_header_column)
            result_log.update(errors=err)
            continue

        if not COLUMNS_WE_CARE_ABOUT[sheet].issubset(headers):
            missing = ", ".join(list(COLUMNS_WE_CARE_ABOUT[sheet] - headers))
            err = "XLSX sheet {} is missing the following columns: {}".format(sheet, missing)
            result_log.update(errors=err)
            continue

        results = xlsx.get_results_dictionary_list()
        if not results:
            err = "There are no results on sheet {}".format(sheet)
            result_log.update(errors=err)
            continue

        log.info("Sheet %s looks good - %s results", sheet, len(results))


def validate_checklist_sheets(filename, result_log, company):
    """Validates that the various required sheets have valid data structures on them."""

    validate_checklist_sheet_Checklist(filename, result_log)

    section_list = validate_checklist_sheet_Sections(filename, result_log)
    choice_list = validate_checklist_sheet_Choices(filename, result_log)

    validate_checklist_sheet_Questions(filename, result_log, section_list, choice_list, company)


def validate_checklist_sheet_Checklist(filename, result_log):
    """Validates that only one checklist exists on the sheet and that it's not already used."""
    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Checklist")
    xlsx.get_columns(uniq_column="Name")
    results = xlsx.get_results_dictionary_list()
    if len(results) > 1:
        err = "You may only have one checklist listed per sheet. You provided {}".format(
            len(results)
        )
        result_log.update(errors=err)
    checklist = results[0]

    if CheckList.objects.filter(name=checklist["Name"]).count() > 1:
        err = (
            "More than one checklists with this name exists - I can't figure out what to do "
            "with this."
        )
        result_log.update(errors=err)


def validate_checklist_sheet_Sections(filename, result_log):
    """Validates that section names are unique.  Returns the list of names."""
    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Sections")
    xlsx.get_columns(uniq_column="Name")
    result_log.update(latest="Processing sheet Sections")

    section_list = []
    for result in xlsx.get_results_dictionary_list():
        name = result["Name"]
        if name in section_list:
            result_log.update(errors="Two sections given with the same name {}".format(name))
        else:
            section_list.append(name)
    return section_list


def validate_checklist_sheet_Choices(filename, result_log):
    """Validates that choice names are unique.  Returns the list of names."""
    xlsx = XLSXParser(filename=filename)
    xlsx.load_workbook_and_sheet(sheet_name="Choices")
    xlsx.get_columns(uniq_column="Choice")
    result_log.update(latest="Processing sheet Choices")

    choice_list = []
    for result in xlsx.get_results_dictionary_list(as_string=True):
        choice = result["Choice"]
        if choice in choice_list:
            result_log.update(errors="Two Choices given with the same name {}".format(choice))
        else:
            choice_list.append(choice)
    return choice_list


def validate_checklist_sheet_Questions(filename, result_log, section_list, choice_list, company):
    """Validates that question responses target correct sections and choices."""

    xlsx = XLSXParser(filename=filename)
    result_log.update(latest="Processing sheet Questions")
    xlsx.load_workbook_and_sheet(sheet_name="Questions")
    xlsx.get_columns(uniq_column="Question")

    for result in xlsx.get_results_dictionary_list():
        section = result["Section"]
        question = result["Question"]
        row_number = result["row_number"]
        is_optional = result.get("Optional", None)
        minimum_value = result.get("Minimum", None)
        maximum_value = result.get("Maximum", None)

        if section not in section_list:
            result_log.update(
                errors=(
                    'Section name "{}" not found in sections for Question ' '"{}" on row {}'
                ).format(section, question, row_number)
            )

        # Check that all specified choices exist on the Choices list
        for choice in range(1, MAX_NUM_CHOICES + 1):
            key = "Choice{}".format(choice)
            if key in result and result[key] and (result[key] not in choice_list):
                errors = (
                    'Choice name "{}" not found in Choices tab for Question "{}" on' " row {}"
                ).format(result[key], question, row_number)
                result_log.update(errors=errors)

        if result["Type"] not in dict(TYPE_CHOICES):
            errors = 'Question Type "{}" is not valid on row {}'.format(result["Type"], row_number)
            result_log.update(errors=errors)

        if minimum_value:
            if result["Type"] not in ["integer", "float"]:
                warning = 'Minimum value not applicable for "{type}"-type question'.format(
                    type=result["Type"]
                )
                result_log.update(warnings=warning)
            else:
                try:
                    float(minimum_value)
                except (TypeError, ValueError):
                    errors = "Invalid minimum value; should be a whole number or decimal"
                    result_log.update(errors=errors)

        if maximum_value:
            if result["Type"] not in ["integer", "float"]:
                warning = 'Maximum value not applicable for "{type}"-type question'.format(
                    type=result["Type"]
                )
                result_log.update(warnings=warning)
            else:
                try:
                    float(maximum_value)
                except (TypeError, ValueError):
                    errors = "Invalid maximum value; should be a whole number or decimal"
                    result_log.update(errors=errors)

        if minimum_value and maximum_value:
            try:
                if float(minimum_value) > float(maximum_value):
                    errors = "Maximum value is smaller than the minimum value."
                    result_log.update(errors=errors)
            except:
                pass

        if is_optional and is_optional.lower().strip() not in IS_OPTIONAL_VALUE_MAP:
            errors = 'If provided, "Optional" column should be one of: Yes, No, y, n'
            result_log.update(errors=errors)

        lowercase_values_data = dict((k.lower(), v) for k, v in result.items())
        if "allow_bulk_fill" in lowercase_values_data:
            bulk_fill = lowercase_values_data["allow_bulk_fill"]
            if str(bulk_fill).lower() not in VALID_BULK_FILL_RESPONSES:
                errors = 'Unknown Allow Bulk Fill Answer "{}" - must be "Yes" or "No"'
                result_log.update(errors=errors.format(bulk_fill))


class ExcelChecklist(XLSXParser):
    def __init__(self, *args, **kwargs):
        super(ExcelChecklist, self).__init__(*args, **kwargs)
        self.company_id = kwargs.get("company_id", None)
        self.company = None
        self.eep_program_id = kwargs.get("eep_program_id", None)
        self.eep_program = None
        self.editable_cells = []
        self.collector = None

        if kwargs.get("user_id", None):
            self.user = User.objects.get(id=kwargs.get("user_id", None))
        if self.company_id:
            self.company = Company.objects.get(id=self.company_id)
        if self.eep_program_id:
            self.eep_program = EEPProgram.objects.get(id=self.eep_program_id)

            collection_request = self.eep_program.collection_request
            if collection_request:
                context = {
                    "user": self.user,
                }
                self.collector = BulkExcelChecklistCollector(collection_request, **context)

    def set_cell_title_style(self, cell):
        cell.font = Font(name="Arial", size=18, bold=True)

    def set_cell_default_style(self, cell):
        cell.font = Font(name="Arial", size=12)

    def set_cell_header_style(self, cell, **kwargs):
        cell.font = Font(name="Arial", size=12, bold=True, color=colors.WHITE)
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color="FF808080")

    def set_cell_italic_small_style(self, cell):
        cell.font = Font(name="Arial", size=10, italic=True)

    def set_cell_large_style(self, cell):
        cell.font = Font(name="Arial", size=14, bold=True)

    def set_cell_center_style(self, cell):
        cell.font = Font(name="Arial", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def set_basic_optional_style(self, cell, editable=True):
        self.set_cell_default_style(cell)
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=OPTIONAL_CELL_COLOR)
        if editable and hasattr(self, "editable_cells"):
            self.editable_cells.append(cell.coordinate)

    def set_basic_unstored_data_style(self, cell, editable=True):
        self.set_cell_default_style(cell)
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=NON_STORED_CELL_COLOR)
        if editable and hasattr(self, "editable_cells"):
            self.editable_cells.append(cell.coordinate)

    def set_basic_required_style(self, cell, editable=True):
        self.set_cell_default_style(cell)
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=REQUIRED_CELL_COLOR)
        if editable and hasattr(self, "editable_cells"):
            self.editable_cells.append(cell.coordinate)

    def set_thin_bordered_style(self, cell):
        cell.border = Border(
            bottom=Side(border_style="thin", color=Color(rgb="FF000000")),
            top=Side(border_style="thin", color=Color(rgb="FF000000")),
            left=Side(border_style="thin", color=Color(rgb="FF000000")),
            right=Side(border_style="thin", color=Color(rgb="FF000000")),
        )

    def lock_cells(self, sheet_obj):
        if not hasattr(self, "editable_cells"):
            log.error("No editable cells")

        sheet_obj.protection.sheet = True
        for cell in self.editable_cells:
            cell_obj = sheet_obj[cell]
            cell_obj.protection = Protection(locked=False)

    def get_base_headers(self):
        """Columns in this method are treated as required."""
        return [
            "lot_number",
            "street_line1",
            "street_line2",
            "is_multi_family",
            "city",
            "state",
            "zipcode",
            "subdivision",
            "subdivision_builder_name",
            "community",
            "builder_org",
            "floorplan",
            "eep_program",
            "sample_set",
            "construction_stage",
            "certification_date",
        ]

    def get_registry_headers(self, eep_program):
        fields = ["rater_of_record"]
        if eep_program.require_input_data:
            fields.append("upload_to_registry")
        return fields

    def get_question_headers(self, breakout_choices, add_builder_verified):
        headers = []
        questions = self.get_questions()
        question_dict = self.get_questions_dict(questions)
        for question in questions:
            if breakout_choices and question.type == "multiple-choice":
                choices = question.question_choice.all().values_list("choice", flat=True)
                for choice in choices:
                    headers.append(question_dict[question]["short_name"] + " - {}".format(choice))
                if add_builder_verified and "Rater Verified" in choices:
                    headers.append(question_dict[question]["short_name"] + " - Builder Verified")
            else:
                headers.append(question_dict[question]["short_name"])
            headers.append(question_dict[question]["short_name"] + " - Comment")
        return headers

    def get_annotation_headers(self, company, eep_program):
        """Returns the column headers that denote annotations required for certification."""
        annotations = eep_program.required_annotation_types.all()
        return [("Annotation - %s" % a.name) for a in annotations]

    def get_questions(self, eep_program=None, home=None):
        """Get Questions"""
        questions = Question.objects.none()
        if eep_program is None:
            eep_program = self.eep_program
        if eep_program:
            questions = Question.objects.filter_by_eep(eep_program)
        return questions

    def get_min_length_for_unique_items(self, items):
        min_length = 1
        while True:
            shortest = [" ".join(q.split()[0:min_length]) for q in items]
            if len(list(set(shortest))) == len(items):
                break
            min_length += 1
        return min_length

    def get_questions_dict(self, questions):
        """This gets the question tuple"""
        question_dict = OrderedDict()
        for question in questions:
            question_dict[question] = {
                "type": question.type,
                "choices": dict(),
                "short_name": question.question,
            }
            if question.type == "multiple-choice":
                choice_dict = dict()
                choices = question.question_choice.all().values_list("choice", flat=True)
                for choice in choices:
                    ques = question_dict[question]["short_name"]
                    choice_dict[choice] = ques + " - " + choice
                question_dict[question]["choices"] = choice_dict
        return question_dict

    def get_absolute_anchor(self, image, top, left):
        from openpyxl.drawing.spreadsheet_drawing import (
            AbsoluteAnchor,
            pixels_to_EMU,
            XDRPoint2D,
            XDRPositiveSize2D,
        )

        position = XDRPoint2D(pixels_to_EMU(left), pixels_to_EMU(top))
        size = XDRPositiveSize2D(pixels_to_EMU(image.width), pixels_to_EMU(image.height))
        return AbsoluteAnchor(pos=position, ext=size)

    def add_logo(self, workbook, sheet, eep_program):
        image = Image(LOGO)
        sheet.add_image(image, anchor=self.get_absolute_anchor(image, 0, 10))

        if eep_program.id and "eto" in eep_program.slug:
            eto_image = Image(ETO_LOGO)
            eto_image.width = 207
            eto_image.height = 75
            sheet.add_image(
                eto_image, anchor="{letter}{row}".format(letter=get_column_letter(8), row=2)
            )

    def properties(self):
        """Set the document properties"""
        props = DocumentProperties()
        props.creator = "Axis"
        props.title = "{} Bulk Checklist Template".format(self.eep_program)
        props.subject = "Axis Checklist Tempalte"
        props.description = "A bulk checklist template for {}".format(self.eep_program)
        return props

    def find_reference_sheet_list(self, reference_sheet_name, choices):
        try:
            ref_sheet_obj = self.workbook[reference_sheet_name]
        except KeyError:
            ref_sheet_obj = self.workbook.create_sheet(reference_sheet_name)

        found = False
        for col in range(1, 1000):
            ref_cell_start, ref_cell_end = None, None
            for row in range(1, 20000):
                found = False
                ref_cell_start = "{letter}{row}".format(letter=get_column_letter(col), row=row)
                value = ref_sheet_obj[ref_cell_start].value
                # print("Looking at %s %s" % (ref_cell_start, value))
                if not value:
                    # print("Good we are in.. %s " % ref_cell_start)
                    ref_start_column, ref_start_row = get_column_letter(col), row
                    for idx, option in enumerate(choices, start=row):
                        ref_cell_end = "{letter}{row}".format(
                            letter=get_column_letter(col), row=idx
                        )
                        found = ref_sheet_obj[ref_cell_end].value is None
                        # print(" Sub review..  %s %s" % (ref_cell_end, ref_sheet_obj[ref_cell_end].value))
                        if not found:
                            break
                if found:
                    break

            if found:
                log.info(
                    "Escaping we have what we need %s:%s for %s",
                    ref_cell_start,
                    ref_cell_end,
                    "|".join(choices),
                )
                return ref_start_column, ref_start_row

        return None, None

    def is_valid_choices(self, choices):
        formula = '"{choices}"'.format(choices=",".join(choices))
        for choice in choices:
            if not all(31 < ord(c) < 128 for c in choice):
                return False
                # bad_actors = [c for c in choice if ord(c) >= 128]
                # raise IndexError("You need to provide a replace the characters %r with something in ASCII %s - Need to use reference list" % bad_actors, choice)
            if "," in choice or '"' in choice:
                return False
                # raise IndexError("You cannot use a comma, or double-quote in any choice option %r - Need to use reference list" % choice)
        if len(formula) > 257:  # Tested the shit out this..
            # See https://support.office.com/en-us/article/excel-specifications-and-limits-1672b34d-7043-467e-8e27-269d656771c3
            return False
            # log.warning("You need to provide a reference sheet and column when you have more than 257 chars you have %d %s", len(formula), choices)
            # raise IndexError("You need to provide a reference sheet and column when you have more than 257 chars")
        return True

    def _set_valid_choices(self, sheet, cell, choices, formula=None, show_dropdown=True):
        """Set the validation"""

        if formula:
            if len(formula) >= 255:
                raise IndexError(
                    "You cannot have a formula > 255 chars (%d) - you need to use a reference cell"
                    % len(formula)
                )
            if len(formula) >= 225:
                log.warning(
                    "You're formula is very close to exceeding 255 (%d) chars consider using a reference cell"
                    % len(formula)
                )
        else:
            if self.is_valid_choices(choices):
                formula = '"{choices}"'.format(choices=",".join(choices))
            else:
                column, start_row = self.find_reference_sheet_list("Reference", choices)
                if column and start_row:
                    ref_sheet_obj = self.workbook["Reference"]
                    for row, choice in enumerate(choices, start=start_row):
                        ref_cell_obj = ref_sheet_obj["{}{}".format(column, row)]
                        ref_cell_obj.value = choice
                formula = "={sheet}!${column}${start_row}:${column}${row}".format(
                    sheet="Reference", column=column, start_row=start_row, row=row
                )

        log.info("Formula for %s is %s", cell.coordinate, formula)

        data_validation = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=not show_dropdown,
            error="Your entry is not in the list",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
            promptTitle="List Selection",
        )

        sheet.add_data_validation(data_validation)
        data_validation.add(cell)
        return formula

    def _set_validation_decimal_between(self, sheet, cell, minimum, maximum):
        """Sets whole number validation range"""
        log.debug("Adding decimal validation to %s %s:%s", cell, minimum, maximum)
        data_validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="{}".format(minimum),
            formula2="{}".format(maximum),
        )

        sheet.add_data_validation(data_validation)
        data_validation.add(cell)

    def get_bulk_headers(
        self, company, eep_program, breakout_choices=False, add_builder_verified=False
    ):
        """Return the headers for answerable content."""
        headers = self.get_base_headers()
        headers.extend(self.get_registry_headers(eep_program))
        if self.collector:
            headers.extend(self.collector.get_xls_instrument_headers())
        else:
            headers.extend(self.get_question_headers(breakout_choices, add_builder_verified))
        headers.extend(self.get_annotation_headers(company, eep_program))
        return headers

    def translate_sample_type(self, name):
        """Get the rating type"""
        if name in ["confirmed", "Confirmed Rating"]:
            return 0
        elif name in ["sample test", "Sampled Test House"]:
            return 2
        elif name in ["sampled", "Sampled House"]:
            return 2
        raise ValueError("Unknown Sample Type {}".format(name))

    def get_reverse_map(self, input_key):
        from axis.checklist.forms import BULK_UPLOAD_EQUIV_MAP

        for value, key in BULK_UPLOAD_EQUIV_MAP:
            if key == input_key:
                return value
        return input_key

    def create_bulk_checklist(
        self,
        company,
        eep_program,
        home=None,
        filename=None,
        breakout_choices=False,
        lock_and_hide=True,
    ):
        """This will create a bulk upload template for download"""

        if filename is None:
            filename = self.filename
        else:
            self.filename = filename

        if eep_program is None and self.eep_program:
            log.info("Using class defined eep_program")
            eep_program = self.eep_program

        # Update collector settings
        if self.collector:
            self.collector.split_choices = breakout_choices

        questions = self.get_questions(eep_program=eep_program, home=home)

        headers = self.get_bulk_headers(company, eep_program, breakout_choices=breakout_choices)
        questions = self.get_questions()
        question_dict = self.get_questions_dict(questions)
        annotations = eep_program.required_annotation_types.all()

        self.workbook = Workbook()
        self.sheet = self.workbook.worksheets[0]
        self.sheet.page_setup.orientation = self.sheet.ORIENTATION_LANDSCAPE

        title = eep_program.name
        # This regex is taken from openpyxl.worksheet in _set_title()
        title = re.sub(r"[\\*?:/\[\]]+", "-", title)
        if len(title) > 28:
            self.sheet.title = title[:28] + "..."
        else:
            self.sheet.title = title[:28]

        today = formats.date_format(datetime.date.today(), "SHORT_DATE_FORMAT")
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        cell = self.sheet.cell(row=1, column=2)
        self.row += 1
        cell.value = "Bulk Home Checklist Upload Template"
        if eep_program:
            cell.value = "{} Bulk Checklist Template".format(eep_program)
        elif home:
            cell.value = "{} Bulk Checklist Template".format(home.get_addr())
        self.set_cell_title_style(cell)

        cell = self.sheet.cell(row=2, column=2)
        cell.value = "Ran by user: {} on {}".format(self.user.get_full_name(), today)
        self.set_cell_italic_small_style(cell)

        self.row += 2
        cell = self.sheet.cell(row=self.row, column=1)
        cell.value = "Notes:"
        self.set_cell_default_style(cell)
        cell.font = Font(name="Arial", size=12, bold=True)

        cell = self.sheet.cell(row=self.row - 1, column=8)
        cell.value = "Required"
        self.set_basic_required_style(cell, editable=False)

        cell = self.sheet.cell(row=self.row, column=8)
        cell.value = "Optional"
        self.set_basic_optional_style(cell, editable=False)

        notes = [
            " • Street Line 2 is used for unit/apt numbers",
            " • If Subdivision Builder Name is given it takes precedence over Subdivision",
            " • Is Multi-Family is used for multi-family complexes",
        ]
        self.row += 1
        for note in notes:
            self.row += 1
            cell = self.sheet.cell(row=self.row, column=1)
            cell.value = note
            self.set_cell_default_style(cell)

        eto2019_equipment_width = 42.33 * 3

        column_widths = {
            "lot_number": 8,
            "street_line1": 20,
            "street_line2": 20,
            "is_multi_family": 10,
            "city": 15,
            "state": 8,
            "zipcode": 10,
            "subdivision": 20,
            "subdivision_builder_name": 25,
            "community": 20,
            "builder_org": 20,
            "floorplan": 20,
            "eep_program": 30,
            "sample_set": 20,
            "construction_stage": 25,
            "certification_date": 25,
            "rater_of_record": 20,
            "upload_to_registry": 20,
            # FIXME: This is not sustainable.  We need to at least name the cells after measure
            "Select the furnace": eto2019_equipment_width,
            "Select the heat pump used for space conditioning": eto2019_equipment_width,
            "Select the water heater": eto2019_equipment_width,
            "Select the Refrigerator": eto2019_equipment_width,
            "Select the Dishwasher": eto2019_equipment_width,
            "Select the Clothes Washer": eto2019_equipment_width,
            "Select the Clothes Dryer": eto2019_equipment_width,
            "Select the Balanced Ventilation": eto2019_equipment_width,
        }

        self.row += 2

        questions_by_short_name = {
            info["short_name"]: question for question, info in question_dict.items()
        }
        choices_by_short_name = {}  # Filled in during first loop, referenced in second

        # Modify header column cells based on the instrument/question characteristics
        for index, column in enumerate(headers):
            col = get_column_letter(index + 1)
            cell = self.sheet["{}{}".format(col, self.row)]
            cell.value = self.get_reverse_map(headers[index])

            self.set_cell_header_style(cell)

            instrument = None
            question = None
            if self.collector:
                if not column.endswith(" - Comment"):
                    instrument = self.collector.resolve_instrument(headers[index], _raise=False)
            else:
                question = questions_by_short_name.get(headers[index])

            if (instrument or question) and not breakout_choices:
                # Add comment to header so full text is always available
                if self.collector:
                    full_text = instrument.description
                else:
                    full_text = question.description
                if full_text and len(full_text):
                    cell.comment = Comment(full_text, "Axis")

                if self.collector:
                    choices = self.collector.get_instrument_choices(instrument)
                    type_descriptor = self.collector.get_type_display(instrument)
                else:
                    choices = question.question_choice.all().values_list("choice", flat=True)
                    type_descriptor = question.get_type_display()

                if len(choices):
                    choices_by_short_name[headers[index]] = choices

                tcell = self.sheet["{}{}".format(col, self.row - 1)]
                tcell.value = type_descriptor
                self.set_cell_center_style(tcell)
                self.sheet.column_dimensions[col].width = 25

            elif column.endswith(" - Comment"):
                self.sheet.column_dimensions[col].width = 20

            elif column.startswith("Annotation - "):
                annotation = annotations.get(name=re.sub(r"Annotation - ", "", column))
                # tcell = self.sheet['{}{}'.format(col, self.row+1)]
                # choices = annotation.valid_multiplechoice_values.split(",")
                # if len(choices):
                #     self._set_valid_choices(self.sheet, tcell, choices)

                value = annotation.get_data_type_display()
                if annotation.is_required:
                    value += " (Required)"

                tcell = self.sheet["{}{}".format(col, self.row - 1)]
                tcell.value = value
                self.set_cell_center_style(tcell)

                self.sheet.column_dimensions[col].width = 25

            # Custom super-headers over field-based columns with no Question associated with them
            elif column in ["construction_stage"]:
                tcell = self.sheet["{}{}".format(col, self.row - 1)]
                tcell.value = "Multiple Choice"  # Made to look like 'Multiple Choice' Question type
                self.set_cell_center_style(tcell)

            # Override column width
            width = column_widths.get(column)
            if width:
                self.sheet.column_dimensions[col].width = width
            else:
                # Try and calculate a reasonable width if the question is long, so that the header
                # row isn't enormously tall.  (Keep the default width=25 if guess result is small.)
                guessed_width = len(column) * 6.8  # Crude average (495pt char width / 72)
                guessed_width = guessed_width / 6  # Crude conversion from pixels to Excel width
                guessed_width = guessed_width / 3  # Allow 3 lines of text wrapping
                self.sheet.column_dimensions[col].width = max(25, guessed_width)

        self.row += 1

        const_stages = list(
            ConstructionStage.objects.filter_by_company(self.user.company).values_list(
                "name", flat=True
            )
        )
        eep_choices = list(
            EEPProgram.objects.filter_by_user(self.user).values_list("name", flat=True)
        )

        # Set data row characteristics
        for row in range(50):  # We'll increment the self.row manually in the loop
            for index, column in enumerate(headers):
                col = get_column_letter(index + 1)
                cell = self.sheet["{}{}".format(col, self.row)]

                if column == "eep_program":
                    self._set_valid_choices(self.sheet, cell, eep_choices)
                    cell.value = self.eep_program.name

                if column in ["is_multi_family", "upload_to_registry"]:
                    choices = ["Yes", "No"]
                    self._set_valid_choices(self.sheet, cell, choices)

                if column in ["construction_stage"]:
                    self._set_valid_choices(self.sheet, cell, const_stages)

                if column in REQUIRED_COLUMNS:
                    self.set_basic_required_style(cell)
                else:
                    self.set_basic_optional_style(cell)

                instrument = None
                question = None
                if self.collector:
                    if not column.endswith(" - Comment"):  # Resolving isn't necessary for Comment
                        instrument = self.collector.resolve_instrument(headers[index], _raise=False)
                else:
                    question = questions_by_short_name.get(headers[index])
                if (instrument or question) and not breakout_choices:
                    # Choices
                    _choices = choices_by_short_name.get(headers[index])
                    choices, formula = _choices, None
                    if row:
                        choices, formula = None, _choices

                    if choices or formula:
                        formula = self._set_valid_choices(
                            sheet=self.sheet, cell=cell, choices=choices, formula=formula
                        )
                        choices_by_short_name[headers[index]] = formula

                    # Constraints
                    constraints = {
                        "min_value": None,
                        "max_value": None,
                    }
                    if self.collector:
                        method = self.collector.get_method(instrument)
                        constraints.update(method.get_constraints())
                    else:
                        constraints.update(
                            {
                                "min_value": question.minimum_value,
                                "max_value": question.maximum_value,
                            }
                        )
                    if constraints["min_value"] and constraints["max_value"]:
                        self._set_validation_decimal_between(
                            self.sheet,
                            cell,
                            *[
                                constraints["min_value"],
                                constraints["max_value"],
                            ],
                        )

                    # Optional markers
                    if self.collector:
                        is_optional = not instrument.response_policy.required
                    else:
                        is_optional = question.is_optional
                    if is_optional:
                        self.set_basic_optional_style(cell)
                    else:
                        self.set_basic_required_style(cell)

                if column.startswith("Annotation - "):
                    annotation = annotations.get(name=re.sub(r"Annotation - ", "", column))
                    choices = annotation.get_valid_multiplechoice_values()
                    if len(choices) and choices != [""]:
                        self._set_valid_choices(self.sheet, cell, choices)

                    if annotation.is_required:
                        self.set_basic_required_style(cell)
                    else:
                        self.set_basic_optional_style(cell)

                self.set_thin_bordered_style(cell)
            self.row += 1

        self.workbook.properties = self.properties()
        self.add_logo(self.workbook, self.sheet, eep_program)
        if lock_and_hide:
            self.lock_cells(self.sheet)

            try:
                references = self.workbook["Reference"]
            except KeyError:
                pass
            else:
                references.sheet_state = "hidden"
                self.lock_cells(references)

        self.workbook.save(filename=filename.name)
