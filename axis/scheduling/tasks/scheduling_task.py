"""scheduling_task.py: """
import io

from celery import shared_task
from django.core.files.base import ContentFile
from django.utils import formats, timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from axis.core.reports import AxisReportFormatter
from axis.scheduling.models import Task

__author__ = "Artem Hruzd"
__date__ = "01/24/2020 13:05"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]


@shared_task(bind=True)
def scheduling_task_report_task(self, asynchronous_process_document_id, user_id, task_ids=None):
    from axis.filehandling.models import AsynchronousProcessedDocument
    from axis.filehandling.log_storage import LogStorage
    from django.contrib.auth import get_user_model

    if not task_ids:
        task_ids = []

    task_meta = {
        "processing": {"current": 0, "total": 1},
        "writing": {"current": 0, "total": 2},
    }

    self.update_state(state="STARTED", meta=task_meta)

    tasks = Task.objects.filter(id__in=task_ids).prefetch_related("assignees")

    user = get_user_model().objects.get(id=user_id)

    app_log = LogStorage(model_id=asynchronous_process_document_id)

    asynchronous_process_document = AsynchronousProcessedDocument.objects.get(
        id=asynchronous_process_document_id
    )
    asynchronous_process_document.task_id = self.request.id
    asynchronous_process_document.task_name = self.name
    asynchronous_process_document.save()

    app_log.info(
        "{user} requested scheduling task report task [{task_id}]".format(
            user=user.get_full_name(), task_id=asynchronous_process_document.task_id
        )
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    app_log.debug("Starting to write report")

    today = formats.date_format(timezone.now().today(), "SHORT_DATE_FORMAT")
    axis_report_formatter = AxisReportFormatter(user=user)
    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="Accreditation list")

    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    cell = sheet.cell(row=1, column=2, value="Scheduled Tasks report")
    axis_report_formatter.set_cell_title_style(cell)

    cell = sheet.cell(
        row=2,
        column=2,
        value="Ran by user: {} on {}".format(user.get_full_name(), today),
    )
    axis_report_formatter.set_cell_italic_small_style(cell)

    axis_report_formatter.add_logo(sheet=sheet)

    row, column = 4, 1

    labels = [
        {"text": "Assignees", "col_width": 25},
        {"text": "Task", "col_width": 35},
        {"text": "Location", "col_width": 30},
        {"text": "Date&Time", "col_width": 25},
        {"text": "Status", "col_width": 35},
        {"text": "Note", "col_width": 15},
    ]

    row += 1
    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    task_meta["writing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    row += 1
    for task in tasks:
        names = [u.get_full_name() for u in task.assignees.all()]
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=1), ", ".join(names))

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=2), task.task_type)

        if task.home_id:
            axis_report_formatter.format_str_cell(sheet.cell(row=row, column=3), task.home)

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=4), task.datetime)

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=5), task.get_status_display()
        )

        axis_report_formatter.format_date_cell(sheet.cell(row=row, column=6), task.note)

        row += 1

    app_log.debug("Saving report")

    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    asynchronous_process_document.document.save(
        "Scheduled-Tasks-Report.xlsx", ContentFile(virtual_workbook.getvalue())
    )
    app_log.debug("Done saving")

    task_meta["writing"]["current"] = 2
    self.update_state(state="DONE", meta=task_meta)
    app_log.update_model(throttle_seconds=None)
