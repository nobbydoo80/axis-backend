__author__ = "Artem Hruzd"
__date__ = "03/16/2020 23:41"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import io
import logging
from datetime import datetime

import django_auto_prefetching
from django.apps import apps
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_extensions.mixins import NestedViewSetMixin

from axis.company.api_v3 import COMPANY_SEARCH_FIELDS, COMPANY_ORDERING_FIELDS
from axis.company.api_v3.filters import (
    AltNameFilter,
    AffiliationsPreferencesFilter,
    SponsoringPreferencesFilter,
    CompanyFilter,
    CompanyAccessFilter,
    CompanyRoleFilter,
)
from axis.company.api_v3.mixins import CompanyCountiesMixin
from axis.company.api_v3.permissions import (
    NestedIsCompanyMemberPermission,
    NestedCompanyHasAdminMembersPermission,
    NestedCompanyUpdatePermission,
    AltNameUpdatePermission,
    SponsorPreferencesUpdatePermission,
    CompanyUpdatePermission,
)
from axis.company.api_v3.serializers import (
    AltNameSerializer,
    NestedAltNameSerializer,
    AffiliationPreferencesSerializer,
    SponsoringPreferencesSerializer,
    UpdateSponsoringPreferencesSerializer,
    CompanySerializer,
    CopyResourcesToOtherCompanySerializer,
    CopiedResourcesToOtherCompanyResponseSerializer,
    SponsorPreferencesSerializer,
    CompanyFlatListSerializer,
    CompanyInfoSerializer,
    CompanyAccessSerializer,
    CompanyRoleSerializer,
    ChangeCompanySerializer,
)
from axis.company.models import Company, AltName, SponsorPreferences, CompanyAccess, CompanyRole
from axis.core.api_v3.filters import AxisSearchFilter, AxisOrderingFilter, AxisFilterBackend
from axis.core.api_v3.mixins import RecentlyViewedMixin
from axis.core.api_v3.permissions import (
    IsAdminUserOrSuperUserPermission,
    IsUserIsCompanyAdminPermission,
)
from axis.core.api_v3.viewsets.history import NestedHistoryViewSet
from axis.core.reports import AxisReportFormatter
from axis.customer_hirl.api_v3.permissions import HIRLCompanyMemberPermission
from axis.customer_hirl.models import BuilderAgreement
from axis.filehandling.api_v3.viewsets import NestedCustomerDocumentViewSet
from axis.relationship.api_v3.viewsets import NestedRelationshipViewSet

log = logging.getLogger(__name__)
User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")
customer_eto_app = apps.get_app_config("customer_eto")


class CompanyViewSet(RecentlyViewedMixin, CompanyCountiesMixin, viewsets.ModelViewSet):
    """
    Contains actions for all company types
    retrieve:
        Get company by ID


        Returns a Company
    list:
        Get all companies


        Returns all companies available for user
    create:
        Create a new Company


        Returns company
    partial_update:
        Update one or more fields on an existing builder organization


        Returns updated builder organization
    update:
        Update builder organization


        Returns updated builder organization
    delete:
        Delete builder organization and all related objects


        Delete builder organization and all related objects
    unattached:
        Inverse of main endpoint that filter objects by user.


        It’s the objects you can add relationships to,
        so that the start appearing in the main endpoint.
        “unattached” objects are safe to view, they’re just not officially connected to you yet
    counties:
        Get or update counties


        Get or update counties
    customer_hirl_clients_export:
        Returns XL sheet of all Builders/Developers/Architect/CommunityOwner companies related to Customer HIRL.
        Requesting user must be a member of HIRL.
    view:
        Add to recently viewed list


        Add Company to Recently viewed list
    recently_viewed:
        Get recently viewed list


        Get recently viewed Companies
    copy_resources_to_other_company:
        Allows copy related objects from one company to others


        Copy Client Agreement, Company COIs from one company to others

    change_company:
        Change company for current user


        Change company for current user by providing CompanyAccess
    """

    model = Company
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    filterset_class = CompanyFilter
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = COMPANY_SEARCH_FIELDS
    ordering_fields = COMPANY_ORDERING_FIELDS
    ordering = [
        "name",
    ]

    def get_queryset(self):
        queryset = super(CompanyViewSet, self).get_queryset()

        try:
            if not self.request.user.company:
                return queryset.none()
        except AttributeError:
            return queryset.none()

        if self.action == "customer_eto_rater_list":
            eto_company = Company.objects.get(slug=customer_eto_app.CUSTOMER_SLUG)
            queryset = Company.objects.filter_by_company(
                company=eto_company, include_self=False
            ).filter(company_type=Company.RATER_COMPANY_TYPE)

            return queryset

        queryset = queryset.filter(is_active=True).annotate(
            total_users=Count("companyaccess", distinct=True),
            total_company_admin_users=Count(
                "companyaccess",
                filter=Q(companyaccess__roles__slug=CompanyRole.IS_COMPANY_ADMIN),
                distinct=True,
            ),
            # customer_hirl specific
            active_customer_hirl_builder_agreements_count=Count(
                "customer_hirl_enrolled_agreements",
                filter=Q(customer_hirl_enrolled_agreements__state=BuilderAgreement.COUNTERSIGNED),
            ),
            active_coi_document_count=Count(
                "coi_documents",
                filter=Q(coi_documents__expiration_date__gt=timezone.now()),
            ),
        )
        return django_auto_prefetching.prefetch(queryset, self.get_serializer_class())

    @property
    def permission_classes(self):
        permission_classes = (IsAuthenticated,)
        if self.action in ["list", "retrieve"]:
            permission_classes = (IsAuthenticated,)
        elif self.action in [
            "create",
        ]:
            permission_classes = (
                IsAuthenticated,
                IsUserIsCompanyAdminPermission | IsAdminUserOrSuperUserPermission,
            )
        elif self.action in ["update", "partial_update"]:
            permission_classes = (CompanyUpdatePermission,)
        elif self.action in ["destroy"]:
            permission_classes = (IsAdminUserOrSuperUserPermission,)
        elif self.action == "counties":
            if self.request.method == "PATCH":
                permission_classes = (CompanyUpdatePermission,)
            else:
                permission_classes = (IsAuthenticated,)
        elif self.action == "customer_hirl_builder_organizations_export":
            permission_classes = (HIRLCompanyMemberPermission,)
        return permission_classes

    @action(detail=False, methods=["get"], serializer_class=CompanyFlatListSerializer)
    def flat_list(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    @action(detail=False, methods=["post"], serializer_class=CopyResourcesToOtherCompanySerializer)
    def copy_resources_to_other_company(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.save()
        response_serializer = CopiedResourcesToOtherCompanyResponseSerializer(data=data)
        response_serializer.is_valid()
        return Response(response_serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"])
    def customer_hirl_clients_export(self, request, *args, **kwargs):
        hirl_company = customer_hirl_app.get_customer_hirl_provider_organization()
        companies = (
            self.queryset.filter_by_company(company=hirl_company)
            .filter(hirlcompanyclient__isnull=False)
            .select_related("hirlcompanyclient", "city", "city__county")
        )

        axis_report_formatter = AxisReportFormatter(user=request.user)
        workbook = Workbook()
        sheet = workbook.create_sheet(index=0, title="Customers")

        row = 1
        column = 1
        labels = [
            {"text": "ID", "col_width": 25},
            {"text": "CustomerNumber", "col_width": 35},
            {"text": "Company", "col_width": 30},
            {"text": "AddressL1", "col_width": 25},
            {"text": "AddressL2", "col_width": 35},
            {"text": "City", "col_width": 35},
            {"text": "StateAbbr", "col_width": 15},
            {"text": "State", "col_width": 15},
            {"text": "Zip", "col_width": 15},
        ]

        for label in labels:
            cell = sheet.cell(row=row, column=column, value=label["text"])
            axis_report_formatter.set_cell_header_style(cell)
            sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
            column += 1

        row = 2
        for company in companies:
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=1), f"B{company.hirlcompanyclient.id:05}"
            )
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=2),
                f"B{company.hirlcompanyclient.id:05}",
            )
            axis_report_formatter.format_str_cell(sheet.cell(row=row, column=3), company.name)
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=4), company.street_line1 or ""
            )
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=5), company.street_line2 or ""
            )
            try:
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=6), company.city.name
                )
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=7), company.city.state
                )
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=8), company.city.county.get_state_display()
                )
            except AttributeError:
                pass

            try:
                sheet.cell(row=row, column=9, value="{:0>5}".format(company.zipcode))
            except TypeError:
                pass
            row += 1

        today = timezone.now().today().strftime("%Y%m%d")
        virtual_workbook = io.BytesIO()
        workbook.save(virtual_workbook)
        filename = "Builders{}.xlsx".format(today)

        response = HttpResponse(
            content=virtual_workbook.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename={}".format(filename)
        return response

    @action(detail=False, methods=["get"])
    def customer_hirl_verifiers_export(self, request, *args, **kwargs):
        customer_hirl = Company.objects.get(slug=customer_hirl_app.CUSTOMER_SLUG)
        raters = (
            self.queryset.filter_by_company(company=customer_hirl)
            .filter(hirlraterorganization__isnull=False)
            .values_list("id", flat=True)
            .distinct()
        )

        hirl_user_ids = (
            User.objects.filter(company__in=raters, accreditations__isnull=False)
            .select_related("company", "company__city", "company__city__county")
            .select_related("hirlrateruser")
            .prefetch_related("accreditations")
            .order_by("hirlrateruser__hirl_id", "-accreditations__date_last")
            .values_list("hirlrateruser__hirl_id", flat=True)
            .distinct()
        )
        # perform distinct by hirlrateruser__hirl_id field
        users = User.objects.filter(hirlrateruser__hirl_id__in=hirl_user_ids).distinct()

        axis_report_formatter = AxisReportFormatter(user=request.user)
        workbook = Workbook()
        sheet = workbook.create_sheet(index=0, title="Customers")

        row = 1
        column = 1
        labels = [
            {"text": "ID", "col_width": 25},
            {"text": "CustomerNumber", "col_width": 35},
            {"text": "AssignedVerifierID", "col_width": 35},
            {"text": "FirstName", "col_width": 35},
            {"text": "LastName", "col_width": 35},
            {"text": "Company", "col_width": 30},
            {"text": "AddressL1", "col_width": 25},
            {"text": "AddressL2", "col_width": 35},
            {"text": "City", "col_width": 35},
            {"text": "StateAbbr", "col_width": 15},
            {"text": "State", "col_width": 15},
            {"text": "Zip", "col_width": 15},
            {"text": "Email", "col_width": 15},
            {"text": "Accredited", "col_width": 15},
        ]

        for label in labels:
            cell = sheet.cell(row=row, column=column, value=label["text"])
            axis_report_formatter.set_cell_header_style(cell)
            sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
            column += 1

        row = 2
        for user in users:
            address = self._get_address_component(user)

            accredited = "FALSE"
            if (
                user.accreditations.annotate_expiration_date()
                .filter(expiration_date__gt=timezone.now())
                .exists()
            ):
                accredited = "TRUE"

            hirlrateruser = getattr(user, "hirlrateruser", None)

            if not hirlrateruser:
                log.debug("User {} has missing HIRLRaterUser. Skipping.".format(user))
                continue

            assigned_verifier_id = getattr(hirlrateruser, "assigned_verifier_id", None)

            if not assigned_verifier_id:
                log.debug("User {} has missing assigned_verifier_id. Skipping.".format(user))
                continue

            axis_report_formatter.format_integer_cell(
                sheet.cell(row=row, column=1), hirlrateruser.hirl_id
            )
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=2), "A{}".format(assigned_verifier_id[1:])
            )
            sheet.cell(row=row, column=3, value="{}".format(assigned_verifier_id))
            axis_report_formatter.format_str_cell(sheet.cell(row=row, column=4), user.first_name)
            axis_report_formatter.format_str_cell(sheet.cell(row=row, column=5), user.last_name)
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=6), "{}, {}".format(user.last_name, user.first_name)
            )
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=7), address["street_line1"] or ""
            )
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=8), address["street_line2"] or ""
            )

            try:
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=9), address["city"] or ""
                )
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=10), address["state_abbr"] or ""
                )
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=11), address["state"] or ""
                )
            except AttributeError:
                log.debug("Missing address details for user {}".format(user))

            if address["zipcode"]:
                zipcode = "{:0>5}".format(address["zipcode"])
            else:
                zipcode = ""
            sheet.cell(row=row, column=12, value=zipcode)
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=13), hirlrateruser.email
            )
            axis_report_formatter.format_str_cell(sheet.cell(row=row, column=14), accredited)
            row += 1

        today = datetime.today().strftime("%Y%m%d")

        virtual_workbook = io.BytesIO()
        workbook.save(virtual_workbook)

        filename = "Verifiers{}.xlsx".format(today)

        response = HttpResponse(
            content=virtual_workbook.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename={}".format(filename)
        return response

    def _get_address_component(self, user):
        address = {
            "street_line1": "",
            "street_line2": "",
            "city": "",
            "state_abbr": "",
            "state": "",
            "zipcode": "",
        }
        if user.mailing_geocode is None:
            address["street_line1"] = user.company.street_line1
            address["street_line2"] = user.company.street_line2
            address["zipcode"] = user.company.zipcode
            if user.company.city:
                address["city"] = user.company.city.name
                address["state_abbr"] = None
                address["state"] = None
                if user.company.city.county:
                    address["state_abbr"] = user.company.city.state
                    address["state"] = user.company.city.county.get_state_display()
        else:
            address["street_line1"] = user.mailing_geocode.raw_street_line1
            address["street_line2"] = user.mailing_geocode.raw_street_line2
            address["zipcode"] = user.mailing_geocode.raw_zipcode
            if user.mailing_geocode.raw_city:
                address["city"] = user.mailing_geocode.raw_city.name
                address["state_abbr"] = None
                address["state"] = None
                if user.mailing_geocode.raw_city.county:
                    address["state_abbr"] = user.mailing_geocode.raw_city.state
                    address["state"] = user.mailing_geocode.raw_city.county.get_state_display()

        return address

    @action(detail=False, methods=["get"], serializer_class=CompanyInfoSerializer)
    def customer_eto_rater_list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @swagger_auto_schema(request_body=ChangeCompanySerializer, responses={"200": CompanySerializer})
    @action(detail=False, methods=["post"], serializer_class=ChangeCompanySerializer)
    def change_company(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        company = serializer.save()
        company_serializer = CompanySerializer(instance=company)
        return Response(company_serializer.data)


class CompanyNestedRelationshipViewSet(NestedRelationshipViewSet):
    """
    list:
        Get relationships


        Returns all relationships
    create:
        Create a relationship between two companies


        Returns relationship
    """

    ct_model = Company


class AltNameViewSet(
    viewsets.GenericViewSet,
    viewsets.mixins.RetrieveModelMixin,
    viewsets.mixins.UpdateModelMixin,
    viewsets.mixins.DestroyModelMixin,
):
    """
    retrieve:
        Get alternative name by ID


        Returns alternative name
    partial_update:
        Update one or more fields on an existing alternative name


        returns updated alternative name
    update:
        Update alternative name


        returns updated alternative name
    delete:
        Delete alternative name


        Delete alternative name
    """

    model = AltName
    queryset = model.objects.all()
    filterset_class = AltNameFilter
    serializer_class = AltNameSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = ["company__name", "name"]
    ordering_fields = ["company", "name"]

    @property
    def permission_classes(self):
        permission_classes = (IsAuthenticated,)
        if self.action in ["update", "partial_update", "delete"]:
            permission_classes = (AltNameUpdatePermission,)
        return permission_classes

    def get_queryset(self):
        queryset = super(AltNameViewSet, self).get_queryset()
        queryset = queryset.filter_by_user(user=self.request.user)
        return django_auto_prefetching.prefetch(queryset, self.get_serializer_class())


class NestedAltNameViewSet(
    NestedViewSetMixin,
    viewsets.GenericViewSet,
    viewsets.mixins.CreateModelMixin,
    viewsets.mixins.ListModelMixin,
):
    """
    Nested Alternate company names endpoint
    list:
        Get all alternate names


        Returns all alternate names
    create:
        Create alternate name

        Create alternate name for company
    """

    model = AltName
    queryset = model.objects.all()
    filterset_class = AltNameFilter
    serializer_class = NestedAltNameSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = ["company__name", "name"]
    ordering_fields = ["company", "name"]

    @property
    def permission_classes(self):
        permission_classes = (IsAuthenticated,)
        if self.action in [
            "create",
        ]:
            permission_classes = (NestedCompanyUpdatePermission,)
        return permission_classes

    def perform_create(self, serializer):
        serializer.save(company_id=self.get_parents_query_dict().get("company_id"))


class NestedAffiliationsViewSet(
    NestedViewSetMixin, viewsets.GenericViewSet, viewsets.mixins.ListModelMixin
):
    """
    list:
        Get affiliation preferences for company


        Returns all affiliation preferences(SponsorPreferences) for company
    """

    model = SponsorPreferences
    permission_classes = (
        IsAuthenticated,
        IsAdminUserOrSuperUserPermission
        | NestedIsCompanyMemberPermission
        | ~NestedCompanyHasAdminMembersPermission,
    )
    queryset = model.objects.all()
    filterset_class = AffiliationsPreferencesFilter
    serializer_class = AffiliationPreferencesSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = ["sponsor__name"]
    ordering_fields = ["id", "sponsor"]

    def get_queryset(self):
        qs = super(NestedAffiliationsViewSet, self).get_queryset()
        if self.request.user.is_superuser:
            return qs
        if not hasattr(self.request.user, "company"):
            return qs.none()
        return qs.filter(
            Q(sponsored_company=self.request.user.company) | Q(sponsor=self.request.user.company)
        )


class SponsoringViewSet(
    viewsets.GenericViewSet,
    viewsets.mixins.CreateModelMixin,
    viewsets.mixins.RetrieveModelMixin,
    viewsets.mixins.UpdateModelMixin,
):
    """
    retrieve:
        Get sponsoring preference by ID


        Returns sponsoring preference
    partial_update:
        Update one or more fields on an existing sponsoring preference.


        returns updated sponsoring preference
    update:
        Update sponsoring preference.


        returns updated sponsoring preference
    """

    model = SponsorPreferences
    queryset = model.objects.all()
    filterset_class = SponsoringPreferencesFilter

    @property
    def permission_classes(self):
        if self.action == "create":
            return (IsAuthenticated, HIRLCompanyMemberPermission | IsAdminUserOrSuperUserPermission)
        if self.action == "update":
            return (SponsorPreferencesUpdatePermission,)
        return (IsAuthenticated,)

    def get_serializer_class(self):
        if self.action == "create":
            return SponsorPreferencesSerializer
        if self.action == "update":
            return UpdateSponsoringPreferencesSerializer
        return SponsorPreferencesSerializer

    def get_queryset(self):
        qs = super(SponsoringViewSet, self).get_queryset()
        if self.request.user.is_superuser:
            return qs
        if not hasattr(self.request.user, "company"):
            return qs.none()
        return qs.filter(sponsor=self.request.user.company)


class NestedSponsoringViewSet(
    NestedViewSetMixin, viewsets.GenericViewSet, viewsets.mixins.ListModelMixin
):
    """
    list:
        Get sponsoring preferences for company


        Returns all sponsoring preferences(SponsorPreferences) for company
    """

    model = SponsorPreferences
    permission_classes = (
        IsAuthenticated,
        IsAdminUserOrSuperUserPermission
        | NestedIsCompanyMemberPermission
        | ~NestedCompanyHasAdminMembersPermission,
    )
    queryset = model.objects.all()
    filterset_class = SponsoringPreferencesFilter
    serializer_class = SponsoringPreferencesSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = [
        "sponsored_company__name",
    ]
    ordering_fields = [
        "id",
        "sponsored_company",
    ]

    def get_queryset(self):
        qs = super(NestedSponsoringViewSet, self).get_queryset()
        if self.request.user.is_superuser:
            return qs
        if not hasattr(self.request.user, "company"):
            return qs.none()
        return qs.filter(sponsor=self.request.user.company)


class CompanyNestedCustomerDocumentViewSet(NestedCustomerDocumentViewSet):
    """
    Because of a bug in legacy code we save and query not
    actual ContentType of Company where documents should
    be attached, but self.request.user.company
    """

    def get_queryset(self):
        return (
            super(CompanyNestedCustomerDocumentViewSet, self)
            .get_queryset()
            .filter(content_type=ContentType.objects.get_for_model(self.request.user.company))
            .filter_by_user(self.request.user, include_public=True)
            .order_by("-pk")
        )

    def perform_create(self, serializer):
        serializer.save(
            content_type=ContentType.objects.get_for_model(self.request.user.company),
            object_id=self.get_parents_query_dict().get("object_id"),
            company=self.request.user.company,
        )


class CompanyNestedHistoryViewSet(NestedHistoryViewSet):
    model = Company.history.model
    queryset = model.objects.all()
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    blacklist_edit_fields = [
        "id",
        "company_ptr",
        "history_user",
        "history_type",
        "history_date",
        "history_id",
        "history_change_reason",
        "geocode_response",
        "shipping_geocode",
        "place",
        "counties",
    ]
    blacklist_create_fields = blacklist_edit_fields


class CompanyAccessViewSet(viewsets.ModelViewSet):
    model = CompanyAccess
    queryset = model.objects.all()
    filterset_class = CompanyAccessFilter
    serializer_class = CompanyAccessSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = [
        "company__id",
        "company__name",
        "user__first_name",
        "user__last_name",
        "user__username",
    ]
    ordering_fields = [
        "id",
        "company__id",
        "company__name",
        "user_id",
        "user__first_name",
        "user__last_name",
        "roles__name",
        "user__is_active",
        "user__is_approved",
    ]

    @property
    def permission_classes(self):
        if self.action in ["create", "update", "partial_update", "delete"]:
            return (IsAuthenticated, IsAdminUserOrSuperUserPermission)
        return (IsAuthenticated,)

    def get_queryset(self):
        queryset = super(CompanyAccessViewSet, self).get_queryset()
        # allow superusers to edit any record
        if not (self.request.user.is_authenticated and self.request.user.is_superuser):
            queryset = queryset.filter_by_user(user=self.request.user)
        return django_auto_prefetching.prefetch(queryset, self.get_serializer_class())


class NestedCompanyAccessViewSet(
    NestedViewSetMixin,
    viewsets.GenericViewSet,
    viewsets.mixins.CreateModelMixin,
    viewsets.mixins.ListModelMixin,
):
    """
    Nested Alternate company names endpoint
    list:
        Get all Company Access


        Returns all Company Access
    create:
        Create Company Access

        Create Company Access for company
    """

    model = CompanyAccess
    queryset = model.objects.all()
    filterset_class = CompanyAccessFilter
    serializer_class = CompanyAccessSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = ["company__name", "user__first_name", "user__last_name"]
    ordering_fields = [
        "id",
        "company__id",
        "company__name",
        "user_id",
        "user__first_name",
        "user__last_name",
        "roles__name",
    ]

    @property
    def permission_classes(self):
        permission_classes = (IsAuthenticated,)
        if self.action in [
            "create",
        ]:
            permission_classes = (NestedCompanyUpdatePermission,)
        return permission_classes

    def perform_create(self, serializer):
        serializer.save(company_id=self.get_parents_query_dict().get("company_id"))

    def get_queryset(self):
        queryset = super(NestedCompanyAccessViewSet, self).get_queryset()
        return django_auto_prefetching.prefetch(queryset, self.get_serializer_class())


class CompanyRoleViewSet(viewsets.ModelViewSet):
    model = CompanyRole
    queryset = model.objects.all()
    filterset_class = CompanyRoleFilter
    serializer_class = CompanyRoleSerializer
    filter_backends = [AxisFilterBackend, AxisSearchFilter, AxisOrderingFilter]
    search_fields = [
        "name",
        "slug",
    ]
    ordering_fields = [
        "id",
        "name",
    ]

    @property
    def permission_classes(self):
        if self.action in ["create", "update", "partial_update", "delete"]:
            return (IsAuthenticated, IsAdminUserOrSuperUserPermission)
        return (IsAuthenticated,)

    def get_queryset(self):
        qs = super(CompanyRoleViewSet, self).get_queryset()
        qs = qs.filter_by_user(user=self.request.user)
        return django_auto_prefetching.prefetch(qs, self.get_serializer_class())
