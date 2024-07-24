"""company.py: """

__author__ = "Artem Hruzd"
__date__ = "11/02/2020 11:58"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import copy
import io
from io import BytesIO
from unittest import skipIf

from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_user_model
from django.urls import reverse_lazy
from openpyxl import load_workbook
from rest_framework import status

from axis.company.models import Company, CompanyRole
from axis.company.tests.factories import (
    builder_organization_factory,
    company_role_factory,
    company_access_factory,
)
from axis.company.tests.factories import provider_organization_factory
from axis.company.tests.factories import rater_organization_factory
from axis.company.tests.mixins import CompaniesAndUsersTestMixin
from axis.core.models import RecentlyViewed
from axis.core.tests.factories import (
    provider_user_factory,
    builder_user_factory,
    rater_user_factory,
)
from axis.core.tests.testcases import ApiV3Tests
from axis.customer_hirl.models import HIRLRaterOrganization, HIRLRaterUser
from axis.customer_hirl.tests.factories import builder_agreement_factory, coi_document_factory
from axis.geographic.models import City
from axis.geographic.tests.factories import county_factory
from axis.relationship.models import Relationship
from axis.user_management.models import Accreditation
from axis.user_management.tests.factories import accreditation_factory

User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


class TestCompanyViewSet(ApiV3Tests):
    def test_list(self):
        list_url = "{url}?is_attached=attached&company_type={company_type}".format(
            url=reverse_lazy("api_v3:companies-list"), company_type=Company.BUILDER_COMPANY_TYPE
        )
        builder_organization_factory()
        builder_organization_factory()
        builder_organization_factory()
        user = provider_user_factory()
        builder_organizations = Company.objects.filter_by_user(
            user=user, include_self=True, company_type=Company.BUILDER_COMPANY_TYPE
        )
        kwargs = dict(url=list_url, user=user)
        data = self.list(**kwargs)

        self.assertEqual(len(data), builder_organizations.count())

    def test_retrieve(self):
        builder_organization = builder_organization_factory()

        provider_user = provider_user_factory(is_company_admin=True)
        builder_user = builder_user_factory(is_company_admin=True)
        superuser_builder = builder_user_factory(is_superuser=True)

        detail_url = reverse_lazy("api_v3:companies-detail", args=(builder_organization.id,))

        self.assertEqual(builder_user.company.relationships.count(), 0)

        with self.subTest("Retrieve builder company as builder"):
            data = self.retrieve(
                url=detail_url,
                user=builder_user,
                expected_status=status.HTTP_200_OK,
            )
            self.assertEqual(data["id"], builder_organization.id)

        with self.subTest("Retrieve builder company as provider"):
            data = self.retrieve(
                url=detail_url,
                user=provider_user,
                expected_status=status.HTTP_200_OK,
            )
            self.assertEqual(data["id"], builder_organization.id)

        with self.subTest("Retrieve builder company as superuser builder"):
            data = self.retrieve(
                url=detail_url,
                user=superuser_builder,
                expected_status=status.HTTP_200_OK,
            )
            self.assertEqual(data["id"], builder_organization.id)

    def test_create(self):
        """
        All users except builders allowed to create builder organization
        and modify basic attributes
        """
        create_url = reverse_lazy("api_v3:companies-list")
        provider_user = provider_user_factory(is_company_admin=True)
        builder_user = builder_user_factory(is_company_admin=True)
        superuser_builder = builder_user_factory(is_superuser=True)

        city = City.objects.first()
        self.assertEqual(builder_user.company.relationships.count(), 0)

        with self.subTest("Create builder company as builder"):
            self.create(
                url=create_url,
                user=builder_user,
                data=dict(
                    name="Test Builder Create by {}".format(builder_user.company),
                    company_type=Company.BUILDER_COMPANY_TYPE,
                    city=city.pk,
                    is_active=True,
                ),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )

        with self.subTest("Create builder company as provider"):
            obj = self.create(
                url=create_url,
                user=provider_user,
                data=dict(
                    name="Test Builder Create by {}".format(provider_user.company),
                    company_type=Company.BUILDER_COMPANY_TYPE,
                    city=city.pk,
                    is_active=True,
                ),
                expected_status=status.HTTP_201_CREATED,
            )
            self.assertEqual(obj["company_type"], Company.BUILDER_COMPANY_TYPE)
            self.assertEqual(obj["is_active"], True)
            builder_organization = Company.objects.get(id=obj["id"])
            self.assertEqual(provider_user.company.relationships.count(), 1)
            self.assertEqual(builder_organization.relationships.count(), 0)

        with self.subTest("Create builder company as superuser builder"):
            obj = self.create(
                url=create_url,
                user=superuser_builder,
                data=dict(
                    name="Test Builder Create by {}".format(superuser_builder.company),
                    company_type=Company.BUILDER_COMPANY_TYPE,
                    city=city.pk,
                    is_active=True,
                ),
                expected_status=status.HTTP_201_CREATED,
            )
            self.assertEqual(obj["company_type"], Company.BUILDER_COMPANY_TYPE)
            self.assertEqual(obj["is_active"], True)
            builder_organization = Company.objects.get(id=obj["id"])
            self.assertEqual(provider_user.company.relationships.count(), 1)
            self.assertEqual(builder_organization.relationships.count(), 0)

    def test_update_as_company_admin_member(self):
        """
        Edit company as Company Admin member of this company
        """
        builder_organization = builder_organization_factory()
        builder_user = provider_user_factory(is_company_admin=True, company=builder_organization)
        update_url = reverse_lazy("api_v3:companies-detail", args=(builder_organization.id,))
        obj = self.update(
            user=builder_user,
            url=update_url,
            data=dict(
                pk=builder_organization.pk,
                name="Changed name",
                street_line1="street_line1",
                address_override=False,
            ),
            partial=True,
        )
        self.assertEqual(obj["name"], "Changed name")
        self.assertEqual(obj["street_line1"], "street_line1")
        self.assertEqual(obj["address_override"], False)

    def test_update_as_company_admin_but_not_a_member(self):
        builder_organization = builder_organization_factory()
        builder_user_factory(is_company_admin=True, company=builder_organization)

        rater_user = rater_user_factory(is_company_admin=True)

        update_url = reverse_lazy("api_v3:companies-detail", args=(builder_organization.id,))
        self.update(
            user=rater_user,
            url=update_url,
            partial=True,
            data=dict(pk=builder_organization.pk, name="Changed name"),
            expected_status=status.HTTP_403_FORBIDDEN,
        )
        builder_organization.refresh_from_db()
        self.assertNotEqual(builder_organization.name, "Changed name")

    def test_update_as_common_user_member(self):
        """
        Edit company as member of this company is not allowed
        """
        builder_organization = builder_organization_factory()
        builder_user = builder_user_factory(is_company_admin=False, company=builder_organization)

        update_url = reverse_lazy("api_v3:companies-detail", args=(builder_organization.id,))
        self.update(
            user=builder_user,
            url=update_url,
            partial=True,
            data=dict(pk=builder_organization.pk, name="Changed name"),
            expected_status=status.HTTP_403_FORBIDDEN,
        )
        builder_organization.refresh_from_db()
        self.assertNotEqual(builder_organization.name, "Changed name")

    def test_update_as_superuser(self):
        """
        Edit company as superuser
        """
        builder_organization = builder_organization_factory()
        builder_user_factory(is_company_admin=True, company=builder_organization)

        super_user = rater_user_factory(is_superuser=True)

        update_url = reverse_lazy("api_v3:companies-detail", args=(builder_organization.id,))
        self.update(
            user=super_user,
            url=update_url,
            partial=True,
            data=dict(pk=builder_organization.pk, name="Changed name"),
            expected_status=status.HTTP_200_OK,
        )
        builder_organization.refresh_from_db()
        self.assertEqual(builder_organization.name, "Changed name")

    def test_update_counties_as_company_admin_member(self):
        """
        Update company counties as company admin member
        """
        builder_organization = builder_organization_factory()
        builder_user = builder_user_factory(is_company_admin=True, company=builder_organization)
        builder_organization.counties.all().delete()
        county = county_factory()

        url = reverse_lazy("api_v3:companies-counties", args=(builder_organization.pk,))
        data = self.update(
            user=builder_user,
            url=url,
            partial=True,
            data=dict(
                ids=[
                    county.id,
                ]
            ),
        )
        self.assertEqual(data[0]["id"], county.id)
        self.assertEqual(len(data), 1)

    def test_aggregated_counties_by_state(self):
        builder_organization = builder_organization_factory()
        builder_user = builder_user_factory(is_company_admin=True, company=builder_organization)

        url = reverse_lazy(
            "api_v3:companies-aggregated-counties-by-state",
            args=(builder_organization.pk,),
        )
        self.client.force_authenticate(user=builder_user)
        response = self.client.get(url, format="json")

        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["counties"], 1)
        self.assertEqual(response.data[0]["selected_counties"], 1)

    def test_update_counties_for_state_as_company_admin_member(self):
        builder_organization = builder_organization_factory()
        builder_user = builder_user_factory(is_company_admin=True, company=builder_organization)
        county = county_factory()
        builder_organization.counties.all().delete()

        url = reverse_lazy(
            "api_v3:companies-update-counties-for-state",
            args=(builder_organization.pk,),
        )
        data = self.update(
            user=builder_user,
            url=url,
            partial=True,
            data=dict(
                state=county.state,
                counties=[
                    county.id,
                ],
            ),
        )
        self.assertEqual(data[0]["id"], county.id)
        self.assertEqual(len(data), 1)

    def test_copy_resources_to_other_company(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        builder_organization = builder_organization_factory()
        builder_agreement = builder_agreement_factory(
            owner=hirl_company, company=builder_organization
        )
        coi = coi_document_factory(company=builder_organization)

        builder_organization2 = builder_organization_factory()

        url = reverse_lazy("api_v3:companies-copy-resources-to-other-company")
        self.client.force_authenticate(user=hirl_user)
        response = self.client.post(
            url,
            data={
                "company_from": builder_organization.id,
                "companies_to": [
                    builder_organization2.id,
                ],
                "copy_client_agreement": True,
                "copy_coi": True,
            },
            format="json",
        )
        self.assertGreater(len(response.data["client_agreements"]), 0)
        self.assertGreater(len(response.data["cois"]), 0)

        self.assertGreater(builder_organization2.coi_documents.count(), 0)
        self.assertGreater(builder_organization2.customer_hirl_enrolled_agreements.count(), 0)

        copied_builder_agreement = builder_organization2.customer_hirl_enrolled_agreements.first()
        self.assertEqual(copied_builder_agreement.date_created, builder_agreement.date_created)

        copied_coi = builder_organization2.coi_documents.first()

        self.assertEqual(copied_coi.created_at, coi.created_at)

    def test_move_resources_to_other_company(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        builder_organization = builder_organization_factory()
        builder_agreement_factory(owner=hirl_company, company=builder_organization)
        coi_document_factory(company=builder_organization)

        builder_organization2 = builder_organization_factory()

        url = reverse_lazy("api_v3:companies-copy-resources-to-other-company")
        self.client.force_authenticate(user=hirl_user)
        response = self.client.post(
            url,
            data={
                "company_from": builder_organization.id,
                "companies_to": [
                    builder_organization2.id,
                ],
                "move_client_agreement": True,
                "move_coi": True,
            },
            format="json",
        )
        self.assertGreater(len(response.data["client_agreements"]), 0)
        self.assertGreater(len(response.data["cois"]), 0)

        self.assertEqual(builder_organization.coi_documents.count(), 0)
        self.assertEqual(builder_organization.customer_hirl_enrolled_agreements.count(), 0)

        self.assertEqual(builder_organization2.coi_documents.count(), 1)
        self.assertEqual(builder_organization2.customer_hirl_enrolled_agreements.count(), 1)

    def test_do_not_allow_copy_and_move_resources_to_other_company(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        builder_organization = builder_organization_factory()
        builder_agreement_factory(owner=hirl_company, company=builder_organization)
        coi_document_factory(company=builder_organization)

        builder_organization2 = builder_organization_factory()

        url = reverse_lazy("api_v3:companies-copy-resources-to-other-company")
        self.client.force_authenticate(user=hirl_user)
        response = self.client.post(
            url,
            data={
                "company_from": builder_organization.id,
                "companies_to": [
                    builder_organization2.id,
                ],
                "move_client_agreement": True,
                "move_coi": True,
                "copy_client_agreement": True,
                "copy_coi": True,
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_customer_hirl_clients_export(self):
        builder_organization = builder_organization_factory()
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        Relationship.objects.create_mutual_relationships(hirl_company, builder_organization)

        list_url = reverse_lazy("api_v3:companies-customer-hirl-clients-export")
        self.client.force_authenticate(user=hirl_user)
        response = self.client.get(list_url, format="json")

        # check actual data in response
        wb = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = wb["Customers"]
        id_heading = sheet.cell(row=1, column=1).value  # ID
        id_data = sheet.cell(row=2, column=1).value
        customer_number = sheet.cell(row=2, column=2).value
        self.assertEqual(id_heading, "ID")
        self.assertEqual(id_data, f"B{builder_organization.hirlcompanyclient.id:05}")
        self.assertEqual(customer_number, f"B{builder_organization.hirlcompanyclient.id:05}")
        wb.close()

    @skipIf(
        settings.DATABASES["default"]["ENGINE"] != "django.db.backends.mysql",
        "Only can be run on MYSQL Engine",
    )
    def test_hirl_raters_export(self):
        rater_organization = rater_organization_factory()
        HIRLRaterOrganization.objects.create(
            rater_organization=rater_organization, data={}, hirl_id=567
        )
        rater_user = rater_user_factory(company=rater_organization)
        accreditation_factory(trainee=rater_user, name=Accreditation.NGBS_2020_NAME)
        hirl_rater_user = HIRLRaterUser.objects.create(
            user=rater_user, data={}, hirl_id=8910, assigned_verifier_id="000007"
        )
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        Relationship.objects.create_mutual_relationships(hirl_company, rater_organization)

        list_url = reverse_lazy("api_v3:companies-customer-hirl-verifiers-export")
        self.client.force_authenticate(user=hirl_user)
        response = self.client.get(list_url, format="json")

        wb = load_workbook(BytesIO(response.content), data_only=True)
        sheet = wb["Customers"]
        internal_id = sheet.cell(row=2, column=1).value
        cust_number = sheet.cell(row=2, column=2).value
        verifier_id = sheet.cell(row=2, column=3).value
        self.assertEqual(internal_id, 8910)
        self.assertEqual(cust_number, "A00007")
        self.assertEqual(verifier_id, "000007")

    def test_change_company_as_superuser(self):
        builder_organization = builder_organization_factory(name="Builder Company 1")
        superuser = builder_user_factory(company=builder_organization, is_superuser=True)
        builder_organization2 = builder_organization_factory(name="Builder Company 2")
        rater_organization = rater_organization_factory(name="Rater Company 1")

        is_company_admin_role = company_role_factory(
            name="Is Company Admin", slug=CompanyRole.IS_COMPANY_ADMIN
        )
        custom_role = company_role_factory(name="Custom Role")

        url = reverse_lazy("api_v3:companies-change-company")

        superuser_default_company_company_access = superuser.companyaccess_set.get(
            company=superuser.default_company
        )
        builder_organization2_company_access = company_access_factory(
            user=superuser, company=builder_organization2
        )
        rater_organization_company_access = company_access_factory(
            user=superuser, company=rater_organization, roles=[]
        )

        minimum_required_data = {
            "company_access": builder_organization2_company_access.id,
        }
        with self.subTest(
            "Test company_access = None",
        ):
            invalid_data = copy.deepcopy(minimum_required_data)
            invalid_data.pop("company_access")
            self.create(
                url=url,
                user=superuser,
                data=invalid_data,
                expected_status=status.HTTP_400_BAD_REQUEST,
            )

        with self.subTest(
            "Test change company for superuser to Builder Company 2",
        ):
            self.assertNotEqual(superuser.company, builder_organization2)
            self.assertEqual(superuser.default_company, builder_organization)

            data = self.create(
                url=url,
                user=superuser,
                data=minimum_required_data,
                expected_status=status.HTTP_200_OK,
            )
            self.assertEqual(data["id"], builder_organization2.id)

            superuser.refresh_from_db()

            self.assertEqual(superuser.company, builder_organization2)

        with self.subTest(
            "Test change company for superuser to Rater Company 1",
        ):
            self.assertNotEqual(superuser.company, rater_organization)
            self.assertEqual(superuser.default_company, builder_organization)

            request_data = {
                "company_access": rater_organization_company_access.id,
            }
            data = self.create(
                url=url,
                user=superuser,
                data=request_data,
                expected_status=status.HTTP_200_OK,
            )
            self.assertEqual(data["id"], rater_organization.id)

            superuser.refresh_from_db()

            self.assertEqual(superuser.company, rater_organization)

    def test_change_company(self):
        """
        Change user as common user
        """
        builder_organization = builder_organization_factory(name="Builder Company 1")
        user = builder_user_factory(company=builder_organization, is_superuser=False)
        builder_organization2 = builder_organization_factory(name="Builder Company 2")

        is_company_admin_role = company_role_factory(
            name="Is Company Admin", slug=CompanyRole.IS_COMPANY_ADMIN
        )
        custom_role = company_role_factory(name="Custom Role")

        url = reverse_lazy("api_v3:companies-change-company")

        superuser_default_company_company_access = user.companyaccess_set.get(
            company=user.default_company
        )
        builder_organization2_company_access = company_access_factory(
            user=user, company=builder_organization2, roles=[is_company_admin_role, custom_role]
        )

        minimum_required_data = {
            "company_access": builder_organization2_company_access.id,
        }

        with self.subTest(
            "Test change company to Builder Company 2",
        ):
            self.assertNotEqual(user.company, builder_organization2)
            self.assertEqual(user.default_company, builder_organization)

            data = self.create(
                url=url,
                user=user,
                data=minimum_required_data,
                expected_status=status.HTTP_200_OK,
            )
            self.assertEqual(data["id"], builder_organization2.id)

            user.refresh_from_db()

            self.assertEqual(user.company, builder_organization2)


class TestCompanyRecentlyViewedViewSet(CompaniesAndUsersTestMixin, ApiV3Tests):
    include_company_types = ["provider"]
    include_unrelated = False
    include_noperms = False

    def test_companies_recent_items_list(self):
        user = User.objects.filter(
            is_superuser=False, is_company_admin=True, company__company_type="provider"
        ).first()
        recently_viewed, _ = RecentlyViewed.objects.view(instance=user.company, by=user)

        list_url = reverse_lazy("api_v3:companies-recently-viewed")
        kwargs = dict(url=list_url, user=user)
        data = self.list(**kwargs)
        self.assertEqual(data[0]["id"], recently_viewed.id)

    def test_companies_recent_items_view(self):
        user = User.objects.filter(
            is_superuser=False, is_company_admin=True, company__company_type="provider"
        ).first()
        view_url = reverse_lazy("api_v3:companies-view", args=(user.company.pk,))
        kwargs = dict(url=view_url, user=user)
        _ = self.create(**kwargs)
        recently_viewed = RecentlyViewed.objects.first()
        self.assertIsNotNone(recently_viewed)
