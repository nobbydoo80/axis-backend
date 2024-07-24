"""training_tasks.py: """
import io

from celery import shared_task
from django.apps import apps
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.utils import timezone, formats
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from axis.core.reports import AxisReportFormatter
from axis.filehandling.log_storage import LogStorage
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.user_management.models import Training
from axis.user_management.states import TrainingStatusStates

__author__ = "Artem Hruzd"
__date__ = "12/10/2019 18:54"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

user_management_app = apps.get_app_config("user_management")


@shared_task
def training_status_expire_task():
    """Expire training status by checking training_date and TRAINING_CYCLE"""
    trainings = Training.objects.filter(
        trainingstatus__state=TrainingStatusStates.APPROVED
    ).prefetch_related("statuses")

    for training in trainings:
        training_date = training.training_date
        date_diff = timezone.now().date() - user_management_app.TRAINING_CYCLE
        if training_date < date_diff:
            for training_status in training.trainingstatus_set.all():
                training_status.expire()
                training_status.save()


@shared_task(bind=True)
def training_report_task(self, asynchronous_process_document_id, user_id, training_ids=None):
    if not training_ids:
        training_ids = []

    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 2}}

    self.update_state(state="STARTED", meta=task_meta)

    trainings = (
        Training.objects.filter(id__in=training_ids)
        .prefetch_related("trainee")
        .prefetch_related("statuses")
    )

    user = get_user_model().objects.get(id=user_id)

    app_log = LogStorage(model_id=asynchronous_process_document_id)

    asynchronous_process_document = AsynchronousProcessedDocument.objects.get(
        id=asynchronous_process_document_id
    )
    asynchronous_process_document.task_id = self.request.id
    asynchronous_process_document.task_name = self.name
    asynchronous_process_document.save()

    app_log.info(
        "{user} requested accreditation report task [{task_id}]".format(
            user=user.get_full_name(), task_id=asynchronous_process_document.task_id
        )
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    app_log.debug("Starting to write report")

    today = formats.date_format(timezone.now().today(), "SHORT_DATE_FORMAT")
    axis_report_formatter = AxisReportFormatter(user=user)
    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="Certification Metrics list")

    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    cell = sheet.cell(row=1, column=2, value="Training report")
    axis_report_formatter.set_cell_title_style(cell)

    cell = sheet.cell(
        row=2, column=2, value="Ran by user: {} on {}".format(user.get_full_name(), today)
    )
    axis_report_formatter.set_cell_italic_small_style(cell)

    axis_report_formatter.add_logo(sheet=sheet)

    row, column = 4, 1

    labels = [
        {"text": "Training name", "col_width": 25},
        {"text": "Training Company or Conference", "col_width": 25},
        {"text": "Training type", "col_width": 30},
        {"text": "Attendance", "col_width": 25},
        {"text": "Credit hours", "col_width": 35},
        {"text": "Trainee", "col_width": 35},
        {"text": "Training date", "col_width": 35},
    ]

    row += 1
    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    task_meta["writing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    row += 1
    for training in trainings:
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=1), training.name)
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=2), training.address)
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=3), training.get_training_type_display()
        )
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=4), training.get_attendance_type_display()
        )
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=5), training.credit_hours)
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=6), training.trainee.get_full_name()
        )
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=7), training.training_date)

        row += 1

    app_log.debug("Saving report")
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    asynchronous_process_document.document.save(
        "Training-Report.xlsx", ContentFile(virtual_workbook.getvalue())
    )
    app_log.debug("Done saving")

    task_meta["writing"]["current"] = 2
    self.update_state(state="DONE", meta=task_meta)
    app_log.update_model(throttle_seconds=None)
