"""inspection_grade_tasks.py: """

__author__ = "Artem Hruzd"
__date__ = "12/10/2019 18:54"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import datetime
import io
import os
from urllib.parse import quote

from celery import shared_task
from dateutil.relativedelta import relativedelta
from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import (
    Count,
    When,
    Case,
    IntegerField,
    Value,
    Q,
    Prefetch,
)
from django.utils import formats
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from axis.company.models import Company
from axis.core.reports import AxisReportFormatter
from axis.filehandling.log_storage import LogStorage
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.user_management.messages import InspectionGradeCustomerHIRLQuarterReportMessage
from axis.user_management.models import InspectionGrade

User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


@shared_task(bind=True)
def inspection_grade_report_task(
    self, asynchronous_process_document_id, user_id, inspection_grade_ids=None
):
    if not inspection_grade_ids:
        inspection_grade_ids = []

    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 2}}

    self.update_state(state="STARTED", meta=task_meta)

    inspectiong_grades = InspectionGrade.objects.filter(id__in=inspection_grade_ids).select_related(
        "qa_status",
        "qa_status__home_status",
        "qa_status__home_status__home",
        "user",
        "user__company",
        "approver",
    )

    user = get_user_model().objects.get(id=user_id)

    app_log = LogStorage(model_id=asynchronous_process_document_id)

    asynchronous_process_document = AsynchronousProcessedDocument.objects.get(
        id=asynchronous_process_document_id
    )
    asynchronous_process_document.task_id = self.request.id
    asynchronous_process_document.task_name = self.name
    asynchronous_process_document.save()

    app_log.info(
        "{user} requested inspection grade report task [{task_id}]".format(
            user=user.get_full_name(), task_id=asynchronous_process_document.task_id
        )
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    app_log.debug("Starting to write report")

    today = formats.date_format(timezone.now().today(), "SHORT_DATE_FORMAT")
    axis_report_formatter = AxisReportFormatter(user=user)
    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="Inspection Grades list")

    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    cell = sheet.cell(row=1, column=2, value="Inspection Grades report")
    axis_report_formatter.set_cell_title_style(cell)

    cell = sheet.cell(
        row=2, column=2, value="Ran by user: {} on {}".format(user.get_full_name(), today)
    )
    axis_report_formatter.set_cell_italic_small_style(cell)

    axis_report_formatter.add_logo(sheet=sheet)

    row, column = 4, 1

    labels = [
        {"text": "Name", "col_width": 25},
        {"text": "Company", "col_width": 25},
        {"text": "Date Grading", "col_width": 25},
        {"text": "Grade", "col_width": 35},
        {"text": "QA Type", "col_width": 35},
        {"text": "Address", "col_width": 120},
        {"text": "Reviewer", "col_width": 25},
        {"text": "Reviewer Notes", "col_width": 100},
    ]

    row += 1
    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    task_meta["writing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    row += 1
    for inspectiong_grade in inspectiong_grades:
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=1), inspectiong_grade.user.get_full_name()
        )
        company = getattr(inspectiong_grade.user, "company", None)
        if company:
            axis_report_formatter.format_str_cell(sheet.cell(row=row, column=2), company.name)
        axis_report_formatter.format_date_cell(
            sheet.cell(row=row, column=3), inspectiong_grade.graded_date
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=4), inspectiong_grade.display_grade()
        )

        if getattr(inspectiong_grade, "qa_status", None):
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=5),
                f"{inspectiong_grade.qa_status.requirement.get_type_display()}",
            )
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=6),
                f"{inspectiong_grade.qa_status.home_status.home.get_home_address_display()}",
            )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=7), inspectiong_grade.approver.get_full_name()
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=8), inspectiong_grade.notes
        )

        row += 1

    app_log.debug("Saving report")
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    asynchronous_process_document.document.save(
        "Inspection-Grades-Report.xlsx", ContentFile(virtual_workbook.getvalue())
    )
    app_log.debug("Done saving")

    task_meta["writing"]["current"] = 2
    self.update_state(state="DONE", meta=task_meta)
    app_log.update_model(throttle_seconds=None)


@shared_task(bind=True)
def customer_hirl_inspection_grade_quarter_report_task(self):
    today = datetime.datetime.now(datetime.timezone.utc)
    half_year_ago = today - relativedelta(months=6)

    sponsored_by_ngbs_companies = Company.objects.filter(
        sponsors__slug=customer_hirl_app.CUSTOMER_SLUG
    )

    prefetch_related_inspection_grade_qs = InspectionGrade.objects.filter(
        created_at__range=[half_year_ago, today]
    ).annotate(
        letter_grade_as_value=Case(
            When(letter_grade=InspectionGrade.A_GRADE, then=Value(5)),
            When(letter_grade=InspectionGrade.B_GRADE, then=Value(4)),
            When(letter_grade=InspectionGrade.C_GRADE, then=Value(3)),
            When(letter_grade=InspectionGrade.D_GRADE, then=Value(2)),
            When(letter_grade=InspectionGrade.F_GRADE, then=Value(1)),
            default=Value(0),
            output_field=IntegerField(),
        ),
    )

    ngbs_sponsored_users = (
        User.objects.filter(company__in=sponsored_by_ngbs_companies)
        .annotate(
            inspection_grade_total_count=Count(
                "inspectiongrade",
                filter=Q(inspectiongrade__created_at__range=[half_year_ago, today]),
            )
        )
        .filter(inspection_grade_total_count__gt=0)
        .prefetch_related(
            Prefetch("inspectiongrade_set", queryset=prefetch_related_inspection_grade_qs)
        )
    )

    for user in ngbs_sponsored_users:
        avg_grade = (
            sum(user.inspectiongrade_set.all().values_list("letter_grade_as_value", flat=True))
            / user.inspection_grade_total_count
        )
        url = (
            "/app/user_management/inspection_grades?search={verifier_name_escaped}"
            "&graded_date__gte={from_date_escaped}&graded_date__lte={to_date_escaped}".format(
                verifier_name_escaped=quote(user.get_full_name()),
                from_date_escaped=quote(half_year_ago.strftime("%m/%d/%Y")),
                to_date_escaped=quote(today.strftime("%m/%d/%Y")),
            )
        )

        message = InspectionGradeCustomerHIRLQuarterReportMessage(url=url)
        if avg_grade > 3:
            message.content = (
                "Congratulations! Throughout {from_date} to {to_date}, "
                "you consistently earned high performance marks from the NGBS Green Review team. Great job! "
                "<a href='{url}' "
                "target='_blank'>View your current NGBS Green 6-month grade average</a>"
            )
            message.email_content = os.path.join(
                settings.SITE_ROOT,
                "axis",
                "user_management",
                "templates",
                "inspection_grade",
                "customer_hirl_inspection_grade_quarter_report_email.html",
            )
        elif avg_grade == 3:
            message.content = (
                "<a href='{url}' target='_blank'>"
                "View your current NGBS Green 6-month grade average</a>"
            )
            message.email_content = os.path.join(
                settings.SITE_ROOT,
                "axis",
                "user_management",
                "templates",
                "inspection_grade",
                "customer_hirl_inspection_grade_quarter_report_for_c_grade_email.html",
            )
        else:
            message.content = (
                "<a href='{url}' target='_blank'>"
                "View your current NGBS Green 6-month grade average</a>"
            )
            message.email_content = os.path.join(
                settings.SITE_ROOT,
                "axis",
                "user_management",
                "templates",
                "inspection_grade",
                "customer_hirl_inspection_grade_quarter_report_for_f_grade_email.html",
            )

        message.send(
            user=user,
            context={
                "from_date": half_year_ago.strftime("%m-%d-%Y"),
                "to_date": today.strftime("%m-%d-%Y"),
                "url": url,
                "verifier": user,
            },
            url=url,
        )
