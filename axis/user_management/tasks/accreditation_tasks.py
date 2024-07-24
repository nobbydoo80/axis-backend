"""accreditation_tasks.py: """

__author__ = "Artem Hruzd"
__date__ = "12/10/2019 18:54"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import io

from celery import shared_task
from django.apps import apps
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.urls import reverse_lazy
from django.utils import formats, timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from axis.company.models import RaterOrganization, Company
from axis.core.reports import AxisReportFormatter
from axis.user_management.messages import AccreditationExpireWarningMessage
from axis.user_management.models import Accreditation

User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


@shared_task
def accreditation_status_expire_notification_warning_task(days_before_expire=60):
    """
    Provide a notification to administrators when an accreditation is within N days of expiring.
    We send notification in current day only, to prevent user spamming,
    because we do not save already sent messages
    :param days_before_expire: int
    """

    accreditations = (
        Accreditation.objects.exclude(state=Accreditation.INACTIVE_STATE)
        .annotate_expiration_date()
        .filter(
            expiration_date=(timezone.now() + timezone.timedelta(days=days_before_expire)).date()
        )
    )

    for accreditation in accreditations:
        url = reverse_lazy("profile:detail", kwargs={"pk": accreditation.trainee.pk})
        AccreditationExpireWarningMessage(url=url).send(
            company=accreditation.approver.company,
            context={
                "trainee": accreditation.trainee,
                "url": url,
                "days_before_expire": days_before_expire,
                "accreditation": accreditation,
            },
        )


@shared_task
def accreditation_status_expire_task():
    """
    Accreditation should transition to "Expired" based on a calculation of
    "Last Date" plus the "Accreditation Cycle" value.
    So if cycle is set to "Every 2 Years",
    the accreditation should transition to expired 2 years from the "Last Date"
    """

    accreditations = (
        Accreditation.objects.exclude(state=Accreditation.INACTIVE_STATE)
        .annotate_expiration_date()
        .filter(expiration_date__lt=timezone.now().date())
    )

    for accreditation in accreditations:
        accreditation.state = Accreditation.INACTIVE_STATE
        accreditation.save()


@shared_task(bind=True)
def accreditation_report_task(
    self, asynchronous_process_document_id, user_id, accreditation_ids=None
):
    from axis.filehandling.models import AsynchronousProcessedDocument
    from axis.filehandling.log_storage import LogStorage
    from django.contrib.auth import get_user_model
    from axis.user_management.models import Accreditation

    if not accreditation_ids:
        accreditation_ids = []

    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 2}}

    self.update_state(state="STARTED", meta=task_meta)

    accreditations = Accreditation.objects.filter(id__in=accreditation_ids).prefetch_related(
        "trainee"
    )

    user = get_user_model().objects.get(id=user_id)

    app_log = LogStorage(model_id=asynchronous_process_document_id)

    asynchronous_process_document = AsynchronousProcessedDocument.objects.get(
        id=asynchronous_process_document_id
    )
    asynchronous_process_document.task_id = self.request.id
    asynchronous_process_document.task_name = self.name
    asynchronous_process_document.save()

    app_log.info(
        "{user} requested accreditation report task [{task_id}]".format(
            user=user.get_full_name(), task_id=asynchronous_process_document.task_id
        )
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    app_log.debug("Starting to write report")

    today = formats.date_format(timezone.now().today(), "SHORT_DATE_FORMAT")
    axis_report_formatter = AxisReportFormatter(user=user)
    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="Accreditation list")

    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    cell = sheet.cell(row=1, column=2, value="Accreditation report")
    axis_report_formatter.set_cell_title_style(cell)

    cell = sheet.cell(
        row=2, column=2, value="Ran by user: {} on {}".format(user.get_full_name(), today)
    )
    axis_report_formatter.set_cell_italic_small_style(cell)

    axis_report_formatter.add_logo(sheet=sheet)

    row, column = 4, 1

    labels = [
        {"text": "Trainee", "col_width": 25},
        {"text": "Accreditation program", "col_width": 35},
        {"text": "Accreditation ID", "col_width": 30},
        {"text": "Role", "col_width": 25},
        {"text": "Accreditation status", "col_width": 35},
        {"text": "Accreditation cycle", "col_width": 35},
        {"text": "Initial accreditation date", "col_width": 15},
        {"text": "Most Recent Accreditation Date", "col_width": 15},
        {"text": "Date expired", "col_width": 15},
    ]

    row += 1
    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    task_meta["writing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    row += 1
    for accreditation in accreditations:
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=1), accreditation.trainee.get_full_name()
        )

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=2), accreditation.name)

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=3), accreditation.accreditation_id
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=4), accreditation.get_role_display()
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=5), accreditation.get_state_display()
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=6), accreditation.get_accreditation_cycle_display()
        )

        axis_report_formatter.format_date_cell(
            sheet.cell(row=row, column=7), accreditation.date_initial
        )

        axis_report_formatter.format_date_cell(
            sheet.cell(row=row, column=8), accreditation.date_last
        )

        axis_report_formatter.format_date_cell(
            sheet.cell(row=row, column=9), accreditation.get_expiration_date()
        )

        row += 1

    app_log.debug("Saving report")
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    asynchronous_process_document.document.save(
        "Accreditation-Report.xlsx", ContentFile(virtual_workbook.getvalue())
    )
    app_log.debug("Done saving")

    task_meta["writing"]["current"] = 2
    self.update_state(state="DONE", meta=task_meta)
    app_log.update_model(throttle_seconds=None)


@shared_task(bind=True)
def customer_hirl_accreditation_report_task(self, asynchronous_process_document_id, user_id):
    from axis.filehandling.models import AsynchronousProcessedDocument
    from axis.filehandling.log_storage import LogStorage
    from axis.user_management.models import Accreditation

    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 2}}

    self.update_state(state="STARTED", meta=task_meta)

    user = User.objects.get(id=user_id)

    verifiers = (
        User.objects.filter_by_user(user=user)
        .filter(company__company_type=Company.RATER_COMPANY_TYPE, is_active=True, is_approved=True)
        .prefetch_related("accreditations")
    )

    app_log = LogStorage(model_id=asynchronous_process_document_id)

    asynchronous_process_document = AsynchronousProcessedDocument.objects.get(
        id=asynchronous_process_document_id
    )
    asynchronous_process_document.task_id = self.request.id
    asynchronous_process_document.task_name = self.name
    asynchronous_process_document.save()

    app_log.info(
        "{user} requested accreditation report task [{task_id}]".format(
            user=user.get_full_name(), task_id=asynchronous_process_document.task_id
        )
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    app_log.debug("Starting to write report")

    today = formats.date_format(timezone.now().today(), "SHORT_DATE_FORMAT")
    axis_report_formatter = AxisReportFormatter(user=user)
    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="Verifier Accreditation list")

    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    cell = sheet.cell(row=1, column=2, value="Verifier Accreditation report")
    axis_report_formatter.set_cell_title_style(cell)

    cell = sheet.cell(
        row=2, column=2, value="Ran by user: {} on {}".format(user.get_full_name(), today)
    )
    axis_report_formatter.set_cell_italic_small_style(cell)

    axis_report_formatter.add_logo(sheet=sheet)

    row, column = 4, 1

    labels = [
        {"text": "Verifier", "col_width": 25},
        {"text": "Verifier Company", "col_width": 35},
        {"text": "Verifier Email", "col_width": 35},
        {"text": "Master Verifier", "col_width": 25},
        {"text": "WRI", "col_width": 25},
        {"text": "2020 NGBS", "col_width": 25},
        {"text": "2015 NGBS", "col_width": 25},
        {"text": "Green Field Rep", "col_width": 25},
        {"text": "Expiration Date", "col_width": 25},
    ]

    row += 1
    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    task_meta["writing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    row += 1

    accreditation_col_map = {
        Accreditation.MASTER_VERIFIER_NAME: 4,
        Accreditation.NGBS_WRI_VERIFIER_NAME: 5,
        Accreditation.NGBS_2020_NAME: 6,
        Accreditation.NGBS_2015_NAME: 7,
        Accreditation.NGBS_GREEN_FIELD_REP_NAME: 8,
    }

    for verifier in verifiers:
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=1), verifier.get_full_name()
        )
        if verifier.company:
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=2), verifier.company.name
            )
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=3), verifier.email)
        for accreditation in verifier.accreditations.all():
            column = accreditation_col_map.get(accreditation.name, None)
            if column:
                axis_report_formatter.format_date_cell(
                    sheet.cell(row=row, column=column),
                    accreditation.date_initial,
                )
                expiration_date = accreditation.get_expiration_date()
                if expiration_date:
                    axis_report_formatter.format_date_cell(
                        sheet.cell(row=row, column=9),
                        expiration_date,
                    )

        row += 1

    app_log.debug("Saving report")
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    asynchronous_process_document.document.save(
        f"Accreditation-Report{today}.xlsx", ContentFile(virtual_workbook.getvalue())
    )
    app_log.debug("Done saving")

    task_meta["writing"]["current"] = 2
    self.update_state(state="DONE", meta=task_meta)
    app_log.update_model(throttle_seconds=None)
