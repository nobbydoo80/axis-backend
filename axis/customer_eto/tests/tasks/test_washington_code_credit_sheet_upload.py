"""test_washington_code_credit_sheet_upload.py - Axis"""

__author__ = "Steven K"
__date__ = "10/26/21 15:08"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven K",
]

import io
import logging
import os
import shutil
import tempfile
import zipfile
from unittest import mock

from celery import states
from django.contrib.contenttypes.models import ContentType
from django.core.files import File
from openpyxl import load_workbook

from axis.annotation.models import Annotation
from axis.annotation.models import Type as AnnotationType
from axis.community.models import Community
from axis.company.models import Company, AltName
from axis.core.tests.client import AxisClient
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.geographic.models import County, City
from axis.geographic.tests.factories import county_factory
from axis.relationship.models import Relationship
from axis.subdivision.models import Subdivision

from ...tasks import WashingtonCodeCreditUploadTask
from ...tasks.washington_code_credit import WashingtonCodeCreditUploadHandler

from ...eep_programs.washington_code_credit import (
    BuildingEnvelope,
    AirLeakageControl,
    HighEfficiencyHVAC,
    HighEfficiencyHVACDistribution,
    DWHR,
    EfficientWaterHeating,
    RenewableEnergy,
    Appliances,
    WACCFuelType,
    ThermostatType,
    FireplaceType,
    FramingType,
    VentilationType,
    FurnaceLocation,
    DuctLocation,
)
from ...enumerations import YesNo
from ..program_checks.test_washington_code_credit import (
    WashingtonCodeCreditProgramBase,
)

log = logging.getLogger(__name__)

mock_worksheet_data = {
    "subdivision": {"name": None},
    "community": {"name": None},
    "home": {
        "lot_number": "6068403972",
        "street_line1": "5803 W Metaline Ave",
        "street_line2": "# 606",
        "zipcode": "99336",
        "city": "kennewick",
        "county": "benton",
    },
    "home_status": {"company": "RATER", "rater_of_record": "frank black"},
    "associations": {
        "builder": "BUILDER",
        "electric_utility": "Pacific Power",
        "gas_utility": "NW Natural Gas",
        "hvac": None,
        "rater": "RATER",
    },
    "annotations": [
        {"type": "wcc-envelope-option", "content": BuildingEnvelope.OPTION_1p7.value},
        {"type": "wcc-air-leakage-option", "content": AirLeakageControl.OPTION_2p4.value},
        {"type": "wcc-hvac-option", "content": HighEfficiencyHVAC.OPTION_3p3.value},
        {
            "type": "wcc-hvac-distribution-option",
            "content": HighEfficiencyHVACDistribution.OPTION_4p1.value,
        },
        {"type": "wcc-dwhr-option", "content": DWHR.OPTION_5p1.value},
        {
            "type": "wcc-water-heating-option",
            "content": EfficientWaterHeating.OPTION_5p2.value,
        },
        {
            "type": "wcc-renewable-electric-option",
            "content": RenewableEnergy.OPTION_6p1a.value,
        },
        {"type": "wcc-appliance-option", "content": Appliances.OPTION_7p1.value},
    ],
    "collected_input": [
        {"measure": "wcc-conditioned_floor_area", "content": 2000, "comment": "Some comment"},
        {"measure": "wcc-water_heating_fuel", "content": WACCFuelType.GAS.value},
        {"measure": "wcc-thermostat_type", "content": ThermostatType.ECOBEE_VOICE.value},
        {"measure": "wcc-fireplace_efficiency", "content": FireplaceType.NONE.value},
        {"measure": "wcc-wall_cavity_r_value", "content": 90},
        {"measure": "wcc-wall_continuous_r_value", "content": 40},
        {"measure": "wcc-framing_type", "content": FramingType.ADVANCED.value},
        {"measure": "wcc-window_u_value", "content": 0.25},
        {"measure": "wcc-window_shgc", "content": 0.30},
        {"measure": "wcc-floor_cavity_r_value", "content": 20},
        {"measure": "wcc-slab_perimeter_r_value", "content": 0},
        {"measure": "wcc-under_slab_r_value", "content": 0},
        {"measure": "wcc-ceiling_r_value", "content": 30},
        {"measure": "wcc-raised_heel", "content": YesNo.NO.value},
        {"measure": "wcc-air_leakage_ach", "content": 5.5},
        {"measure": "wcc-ventilation_type", "content": VentilationType.BALANCED.value},
        {"measure": "wcc-ventilation_brand", "content": "Ventilation Brand"},
        {"measure": "wcc-ventilation_model", "content": "Ventilation Model"},
        {"measure": "wcc-hrv_asre", "content": 89},
        {"measure": "wcc-furnace_brand", "content": "Furnace Brand"},
        {"measure": "wcc-furnace_model", "content": "Furnace Model"},
        {"measure": "wcc-furnace_afue", "content": 96},
        {"measure": "wcc-furnace_location", "content": FurnaceLocation.UNCONDITIONED_SPACE.value},
        {"measure": "wcc-duct_location", "content": DuctLocation.CONDITIONED_SPACE.value},
        {"measure": "wcc-duct_leakage", "content": 39},
        {"measure": "wcc-dwhr_installed", "content": YesNo.YES.value},
        {"measure": "wcc-water_heater_brand", "content": "Water Heater Brand"},
        {"measure": "wcc-water_heater_model", "content": "Water Heater Model"},
        {"measure": "wcc-gas_water_heater_uef", "content": 0.78},
        {"measure": "wcc-electric_water_heater_uef", "content": 2.9},
        {"measure": "wcc-total_ua_alternative", "content": 5},
    ],
}


def mock_extract_cell_data(*_args, **_kwargs) -> dict:
    return mock_worksheet_data


class WashingtonCodeCreditUploadViewTests(WashingtonCodeCreditProgramBase):
    client_class = AxisClient

    @classmethod
    def setUpTestData(cls):
        super(WashingtonCodeCreditUploadViewTests, cls).setUpTestData()
        from axis.core.tests.factories import rater_admin_factory

        cls.user = rater_admin_factory(
            company=cls.rater_company, first_name="frank", last_name="Black"
        )
        cls.rater_company.update_permissions()

        cls.document = os.path.join(
            os.path.dirname(__file__), "../../static/templates/washington_code_credit.xlsm"
        )

        cls.task_name = (
            "axis.customer_eto.tasks.washington_code_credit.WashingtonCodeCreditUploadTask"
        )

        with io.open(cls.document, "rb") as f:
            cls.result_object = AsynchronousProcessedDocument(
                company=cls.rater_company,
                download=True,
                task_name=cls.task_name,
                document=File(f, name=os.path.basename(cls.document)),
            )
            cls.result_object.save()
        cls.task_kwargs = {
            "result_object_id": cls.result_object.id,
            "company_id": cls.rater_company.id,
            "user_id": cls.user.id,
        }

    def test_workbook_read(self):
        """Verify that our document is readable.."""
        self.assertTrue(os.path.exists(self.document))
        wb = load_workbook(self.document, data_only=True, read_only=True)

        required_sheets = {
            "Step 1 - AXIS Home Details",
            "Step 2 - Select Code Credits",
            "Step 3 - Enter Specifications",
        }
        self.assertTrue(required_sheets.issubset(set(wb.sheetnames)))

    @mock.patch(
        "axis.customer_eto.tasks.washington_code_credit"
        ".WashingtonCodeCreditUploadHandler.extract_cell_data",
        side_effect=mock_extract_cell_data,
    )
    def test_ok_file(self, _mock):
        task = WashingtonCodeCreditUploadTask.apply_async(kwargs=self.task_kwargs)
        self.assertEqual(task.status, states.SUCCESS)

    def test_wrong_filetype(self):
        with io.open(__file__, "rb") as f:
            self.result_object.document = File(f, name=os.path.basename(self.document))
            self.result_object.save()

        self.assertRaises(
            zipfile.BadZipfile,
            WashingtonCodeCreditUploadTask.apply_async,
            kwargs=self.task_kwargs,
        )

    def test_wrong_sheets(self):
        document = os.path.join(
            os.path.dirname(__file__), "../../../checklist/static/templates/Single_Home_Upload.xlsx"
        )
        with io.open(document, "rb") as f:
            self.result_object.document = File(f, name=os.path.basename(self.document))
            self.result_object.save()

        self.assertRaises(
            IndexError,
            WashingtonCodeCreditUploadTask.apply_async,
            kwargs=self.task_kwargs,
        )

    def test_validate_associations(self):
        """Verify our associations"""
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        with self.subTest("Basic Passing"):
            result = hndlr.validate_associations(**mock_worksheet_data["associations"])
            self.assertEqual(result["builder"], Company.objects.get(id=self.builder_company.id))
            self.assertEqual(result["electric_utility"], Company.objects.get(id=self.pac_pwr.id))
            self.assertEqual(result["gas_utility"], Company.objects.get(id=self.nw_nat.id))
            self.assertEqual(result["rater"], Company.objects.get(id=self.rater_company.id))
            self.assertEqual(result["provider"], self.eep_program.certifiable_by.first())

        with self.subTest("Basic Alias"):
            AltName.objects.create(
                company=Company.objects.get(id=self.builder_company.id), name="FrodoCom Inc."
            )
            data = mock_worksheet_data["associations"].copy()
            data["builder"] = "FROdOcom"
            result = hndlr.validate_associations(**data)
            self.assertEqual(result["builder"], Company.objects.get(id=self.builder_company.id))

        with self.subTest("Provider Upload"):
            hndlr = WashingtonCodeCreditUploadHandler(
                result_object=self.result_object,
                company=self.provider_company,
                user=self.provider_user,
            )
            data = mock_worksheet_data["associations"].copy()
            data["rater"] = self.rater_company.name
            result = hndlr.validate_associations(**data)
            self.assertEqual(result["rater"], Company.objects.get(id=self.rater_company.id))
            self.assertEqual(result["provider"], self.eep_program.certifiable_by.first())

    def test_validate_community(self):
        """Verify our associations"""
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )

        with self.subTest("No Community"):
            result = hndlr.validate_community(**mock_worksheet_data["community"])
            self.assertIsNone(result)

        self.assertEqual(Community.objects.count(), 0)
        community = Community.objects.create(name="Test", city=self.city)
        with self.subTest("Community Exists no Relationship"):
            result = hndlr.validate_community(name="Test")
            self.assertIsNone(result)

        self.assertEqual(Community.objects.count(), 1)
        Relationship.objects.validate_or_create_relations_to_entity(community, self.rater_company)
        with self.subTest("Community passing"):
            result = hndlr.validate_community(name="Test")
            self.assertEqual(result, community)

    def test_validate_subdivision(self):
        """Verify our subdivision"""
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )

        with self.subTest("No Subdivision"):
            result, builder = hndlr.validate_subdivision(
                name=mock_worksheet_data["subdivision"]["name"],
                builder=Company.objects.get(id=self.builder_company.id),
            )
            self.assertIsNone(result)

        self.assertEqual(Subdivision.objects.count(), 0)
        subdivision = Subdivision.objects.create(
            name="Test", city=self.city, builder_org=self.builder_company
        )
        with self.subTest("Subdivision Exists no Relationship"):
            result, builder = hndlr.validate_subdivision(name="Test")
            self.assertIsNone(result)

        self.assertEqual(Subdivision.objects.count(), 1)
        Relationship.objects.validate_or_create_relations_to_entity(subdivision, self.rater_company)
        with self.subTest("Subdivision passing"):
            result, builder = hndlr.validate_subdivision(name="Test")
            self.assertEqual(result, subdivision)

    def test_validate_county(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )

        with self.subTest("No County"):
            result = hndlr.validate_county(name=None)
            self.assertIsNone(result)

        county = County.objects.filter(state="WA").first()
        with self.subTest("No County"):
            result = hndlr.validate_county(name=county.name.upper())
            self.assertEqual(result, county)

    def test_validate_city(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )

        with self.subTest("No City"):
            result = hndlr.validate_city(name=None)
            self.assertIsNone(result)

        city = City.objects.filter(county__state="WA").first()
        initial = City.objects.count()
        with self.subTest("Valid City Valid county"):
            result = hndlr.validate_city(name=city.name.lower(), county=city.county)
            self.assertEqual(result, city)
            self.assertEqual(city.county, result.county)
            self.assertEqual(City.objects.count(), initial)

        city = City.objects.filter(county__state="WA").first()
        with self.subTest("Valid City No county"):
            result = hndlr.validate_city(name=city.name.lower())
            self.assertEqual(result, city)
            self.assertEqual(city.county, result.county)
            self.assertEqual(City.objects.count(), initial)

    def test_validate_home(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        with self.subTest("Verify home already exists"):
            result = hndlr.validate_home(
                lot_number=self.home.lot_number,
                street_line1=self.home.street_line1,
                street_line2=self.home.street_line2,
                city=self.home.city,
                county=self.home.city.county,
                state=self.home.state,
                zipcode=self.home.zipcode,
                subdivision=self.home.subdivision,
                builder=self.builder_company,
            )
            self.assertEqual(result, self.home)

        with self.subTest("Verify home already exists minimum"):
            result = hndlr.validate_home(
                lot_number=self.home.lot_number,
                street_line1=self.home.street_line1,
                street_line2=self.home.street_line2,
                city=self.home.city,
                county=self.home.city.county,
                zipcode=self.home.zipcode,
                builder=self.builder_company,
            )
            self.assertEqual(result, self.home, hndlr.app_log.report_chronological())

        with self.subTest("New Home"):
            result = hndlr.validate_home(
                lot_number=self.home.lot_number + "-a",
                street_line1=self.home.street_line1,
                street_line2=self.home.street_line2,
                city=self.home.city,
                county=self.home.city.county,
                zipcode=self.home.zipcode,
                builder=self.builder_company,
            )
            self.assertEqual(type(result), dict, hndlr.app_log.report_chronological())

    def test_validate_home_status(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        with self.subTest("Verify project already exists"):
            result = hndlr.validate_home_status(
                home=self.home,
                company=hndlr.company,
                rater_of_record=hndlr.user.username,
            )
            self.assertEqual(result, self.home_status)

        with self.subTest("Verify project company string"):
            hndlr = WashingtonCodeCreditUploadHandler(
                result_object=self.result_object,
                company=self.provider_company,
                user=self.provider_user,
            )
            result = hndlr.validate_home_status(
                home=self.home,
                company=self.provider_company.name,
                rater_of_record=hndlr.user.username,
            )
            self.assertEqual(type(result), dict, hndlr.app_log.report_chronological())

    def test_annotations(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        with self.subTest("Verify new annotation"):
            annotations = mock_worksheet_data["annotations"]
            result = hndlr.validate_annotations(
                home_status=self.home_status,
                annotations=annotations,
            )
            self.assertEqual(len(result), len(annotations))
        with self.subTest("Wrong annotation"):
            # Deep copy is required.
            annotations = [x.copy() for x in mock_worksheet_data["annotations"]]
            annotations[-1]["type"] = "foo"
            result = hndlr.validate_annotations(
                home_status=self.home_status,
                annotations=annotations,
            )
            self.assertEqual(len(result), len(annotations) - 1)
            self.assertIn(
                "Annotation provided (foo) is not included in program",
                hndlr.app_log.report_chronological()[-1],
            )

        with self.subTest("Already Existing annotation same value no update"):
            annotation = Annotation.objects.create(
                type=AnnotationType.objects.get(slug="wcc-envelope-option"),
                content=BuildingEnvelope.OPTION_1p7.value,
                content_type=ContentType.objects.get_for_model(self.home_status),
                object_id=self.home_status.id,
            )
            annotations = mock_worksheet_data["annotations"]
            result = hndlr.validate_annotations(
                home_status=self.home_status,
                annotations=annotations,
            )
            self.assertNotIn("wcc-envelope-option", [x["type"] for x in result])
            self.assertEqual(len(result), len(annotations) - 1)
            Annotation.objects.filter(id=annotation.id).delete()

        with self.subTest("Already Existing annotation different value - updatable"):
            annotation = Annotation.objects.create(
                type=AnnotationType.objects.get(slug="wcc-air-leakage-option"),
                content=AirLeakageControl.OPTION_2p4,
                content_type=ContentType.objects.get_for_model(self.home_status),
                object_id=self.home_status.id,
            )
            annotations = mock_worksheet_data["annotations"]
            result = hndlr.validate_annotations(
                home_status=self.home_status,
                annotations=annotations,
            )
            self.assertIn("wcc-air-leakage-option", [x["type"] for x in result])
            self.assertEqual(len(result), len(annotations))
            Annotation.objects.filter(id=annotation.id).delete()

    def test_collected_input(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        with self.subTest("Verify new inputs"):
            collected_input = mock_worksheet_data["collected_input"]
            result = hndlr.validate_collected_input(
                home_status=self.home_status,
                eep_program=self.home_status.eep_program,
                collected_input=collected_input,
                user=self.user,
            )
            self.assertEqual(len(result), len(collected_input))

        with self.subTest("Already Existing response same value no update"):
            item = mock_worksheet_data["collected_input"][0]
            self.add_collected_input(
                value=item["content"], measure_id=item["measure"], home_status=self.home_status
            )
            collected_input = mock_worksheet_data["collected_input"]
            result = hndlr.validate_collected_input(
                home_status=self.home_status,
                eep_program=self.home_status.eep_program,
                collected_input=collected_input,
                user=self.user,
            )
            self.assertEqual(len(result), len(collected_input) - 1)
            self.assertIn("Skipping existing input", hndlr.app_log.report_chronological()[-1])
            self.remove_collected_input(item["measure"])

        with self.subTest("Already Existing response different value - updatable"):
            item = mock_worksheet_data["collected_input"][0]
            self.add_collected_input(
                value=2500, measure_id=item["measure"], home_status=self.home_status
            )
            collected_input = mock_worksheet_data["collected_input"]
            result = hndlr.validate_collected_input(
                home_status=self.home_status,
                eep_program=self.home_status.eep_program,
                collected_input=collected_input,
                user=self.user,
            )
            self.assertEqual(len(result), len(collected_input))

        with self.subTest("Response no longer required"):
            # We added Gas heat therefore electric uef is not needed.  Way cool
            item = mock_worksheet_data["collected_input"][1]
            self.add_collected_input(
                value=item["content"], measure_id=item["measure"], home_status=self.home_status
            )
            collected_input = mock_worksheet_data["collected_input"]
            result = hndlr.validate_collected_input(
                home_status=self.home_status,
                eep_program=self.home_status.eep_program,
                collected_input=collected_input,
                user=self.user,
            )
            self.assertIn("not required at this point", hndlr.app_log.report_chronological()[-1])
            # "2" the item we added and the no longer needed one.
            self.assertEqual(len(result), len(collected_input) - 2)
            self.remove_collected_input(item["measure"])

    def test_home_creation(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        self.assertEqual(type(cleaned_data["home"]), dict)
        input_data = cleaned_data["home"]
        result = hndlr.create({"home": input_data})
        self.assertIsNotNone(result["home"])
        home = result["home"]
        self.assertEqual(home.lot_number, input_data["lot_number"])
        self.assertEqual(home.street_line1, input_data["street_line1"])
        self.assertIn(home.street_line2, input_data["street_line2"])
        self.assertEqual(home.zipcode, input_data["zipcode"])
        self.assertEqual(home.city, input_data["city"])
        self.assertEqual(home.county, input_data["city"].county)
        self.assertEqual(home.state, input_data["state"])
        self.assertEqual(home.is_multi_family, False)
        self.assertEqual(home.history.count(), 1)
        self.assertEqual(home.history.get().history_user, hndlr.user)
        self.assertIsNotNone(home.geocode_response)
        self.assertIsNotNone(home.geocode_response.geocode)

    def test_home_status_creation(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        self.assertEqual(type(cleaned_data["home"]), dict)
        input_data = cleaned_data["home_status"]
        result = hndlr.create({"home": cleaned_data["home"], "home_status": input_data})
        self.assertIsNotNone(result["home_status"])
        home_status = result["home_status"]
        self.assertEqual(home_status.company, input_data["company"])
        self.assertEqual(home_status.home, result["home"])
        self.assertEqual(home_status.eep_program, self.eep_program)
        self.assertIsNone(home_status.floorplan)
        self.assertEqual(home_status.rater_of_record, input_data["rater_of_record"])
        self.assertEqual(home_status.rater_of_record, input_data["rater_of_record"])
        self.assertEqual(home_status.history.count(), 3)
        expected = set(home_status.history.all().values_list("history_user", flat=True))
        self.assertEqual(expected, {hndlr.user.id})

    def test_associations_creation(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        self.assertEqual(type(cleaned_data["associations"]), dict)

        result = hndlr.create(cleaned_data)
        self.assertIsNotNone(result["associations"])

        associations = result["associations"]
        self.assertEqual(associations["removed"], None)
        self.assertEqual(associations["added"].count(), 2)

    def test_annotation_creation(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        self.assertGreater(len(cleaned_data["annotations"]), 2)
        annotations = cleaned_data["annotations"][:]
        with self.subTest("Partial Annotations"):
            cleaned_data["annotations"] = annotations[:2]
            result = hndlr.create(cleaned_data)
            self.assertIsNotNone(result["annotations"])
            self.assertEqual(len(result["annotations"]), 2)
        with self.subTest("Full Some exist"):
            cleaned_data["annotations"] = annotations
            result = hndlr.create(cleaned_data)
            self.assertIsNotNone(result["annotations"])
            self.assertEqual(len(result["annotations"]), len(annotations) - 2)

    def test_input_collection_add(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        self.assertGreater(len(cleaned_data["collected_input"]), 2)
        collected_input = cleaned_data["collected_input"][:]

        with self.subTest("Partial Questions"):
            cleaned_data["collected_input"] = collected_input[:2]
            result = hndlr.create(cleaned_data)
            self.assertIsNotNone(result["collected_input"])
            self.assertEqual(len(result["collected_input"]["added"]), 2)
            self.assertIsNone(result["collected_input"]["removed"])
        with self.subTest("Remainder+  Questions"):
            cleaned_data["collected_input"] = collected_input
            result = hndlr.create(cleaned_data)
            self.assertIsNotNone(result["collected_input"])
            # Note this is 3 (in lieu pf 2)  b/c we added a fuel so the UEF
            # question b/c not needed and is excluded
            self.assertEqual(len(result["collected_input"]["added"]), len(collected_input) - 3)
            self.assertIsNone(result["collected_input"]["removed"])
        with self.subTest("Update  Question"):
            collected_input = collected_input[0].copy()
            collected_input["content"] = 4500
            cleaned_data["collected_input"] = [collected_input]
            result = hndlr.create(cleaned_data)
            self.assertIsNotNone(result["collected_input"])
            self.assertEqual(len(result["collected_input"]["added"]), 1)
            self.assertEqual(len(result["collected_input"]["removed"]), 1)

    def test_save(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        result = hndlr.create(cleaned_data)
        self.assertIn("home", result)
        self.assertIsNotNone(result["home"])
        self.assertIn("home_status", result)
        self.assertIsNotNone(result["home_status"])
        self.assertIn("associations", result)
        self.assertIsNotNone(result["associations"])
        self.assertIn("annotations", result)
        self.assertIsNotNone(result["annotations"])
        self.assertIn("collected_input", result)
        self.assertIsNotNone(result["collected_input"])

    def test_notify(self):
        hndlr = WashingtonCodeCreditUploadHandler(
            result_object=self.result_object, company=self.rater_company, user=self.user
        )
        cleaned_data = hndlr.validate_data(data=mock_worksheet_data)
        with self.subTest("First Notification"):
            data = hndlr.create(cleaned_data)
            result = hndlr.notify_parties(**data)
            self.assertEqual(result, True)
        with self.subTest("Subsequent No-Notification"):
            data = hndlr.create(cleaned_data)
            result = hndlr.notify_parties(**data)
            self.assertEqual(result, False)
