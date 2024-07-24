"""washington_code_credit.py - Axis"""

__author__ = "Steven K"
__date__ = "10/26/21 11:14"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven K",
]

import datetime
import io
import logging
import os
import tempfile
import typing
import zipfile
from functools import cached_property

from analytics.tests.test_api import Home, EEPProgramHomeStatus
from celery import Task, states
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from openpyxl import load_workbook, Workbook
from simple_history.utils import bulk_create_with_history

from axis.annotation.models import Annotation
from axis.annotation.utils import validate_annotations
from axis.checklist.collection.excel import ExcelChecklistCollector
from axis.community.models import Community
from axis.company.models import Company, BuilderOrganization
from axis.core.models import User
from axis.customer_eto.api_v3.serializers.reports.washington_code_credit import (
    WashingtonCodeCreditReportSerializer,
)
from axis.customer_eto.messages import WashingtonCodeCreditUploadComplete
from axis.customer_eto.models import FastTrackSubmission
from axis.customer_eto.reports.washington_code_credit import WashingtonCodeCreditReport
from axis.eep_program.models import EEPProgram
from axis.filehandling.log_storage import LogStorage
from axis.filehandling.models import AsynchronousProcessedDocument, CustomerDocument
from axis.geographic.models import County, City
from axis.relationship.utils import create_or_update_spanning_relationships
from axis.subdivision.models import Subdivision
from celeryapp import celery_app

log = logging.getLogger(__name__)

HOME_DETAILS_SHEET = "Step 1 - AXIS Home Details"
CODE_CREDITS_SHEET = "Step 2 - Select Code Credits"
SPECIFICATIONS_SHEET = "Step 3 - Enter Specifications"

HOME_DATA = {
    "lot_number": (HOME_DETAILS_SHEET, "J7"),
    "street_line1": (HOME_DETAILS_SHEET, "D6"),
    "street_line2": (HOME_DETAILS_SHEET, "D7"),
    "city": (HOME_DETAILS_SHEET, "D8"),
    "county": (HOME_DETAILS_SHEET, "D10"),
    "zipcode": (HOME_DETAILS_SHEET, "D11"),
}

COMMUNITY_DATA = {
    "name": (HOME_DETAILS_SHEET, "D12"),
}

SUBDIVISION_DATA = {
    "name": (HOME_DETAILS_SHEET, "D13"),
}

HOME_STATUS_DATA = {
    "company": (HOME_DETAILS_SHEET, "D17"),
    "rater_of_record": (HOME_DETAILS_SHEET, "D18"),
}

ASSOCIATIONS = {
    "builder": (HOME_DETAILS_SHEET, "D22"),
    "electric_utility": (HOME_DETAILS_SHEET, "D23"),
    "gas_utility": (HOME_DETAILS_SHEET, "D24"),
    "hvac": (HOME_DETAILS_SHEET, "D25"),
    "rater": (HOME_DETAILS_SHEET, "D27"),
}

COLLECTION_DATA = {
    "wcc-conditioned_floor_area": (CODE_CREDITS_SHEET, "C6"),
    "wcc-water_heating_fuel": (CODE_CREDITS_SHEET, "C7"),
    "wcc-thermostat_type": (CODE_CREDITS_SHEET, "C8"),
    "wcc-fireplace_efficiency": (CODE_CREDITS_SHEET, "C9"),
    "wcc-wall_cavity_r_value": (SPECIFICATIONS_SHEET, "E12"),
    "wcc-wall_continuous_r_value": (SPECIFICATIONS_SHEET, "E13"),
    "wcc-framing_type": (SPECIFICATIONS_SHEET, "E14"),
    "wcc-window_u_value": (SPECIFICATIONS_SHEET, "E15"),
    "wcc-window_shgc": (SPECIFICATIONS_SHEET, "E16"),
    "wcc-floor_cavity_r_value": (SPECIFICATIONS_SHEET, "E17"),
    "wcc-slab_perimeter_r_value": (SPECIFICATIONS_SHEET, "E18"),
    "wcc-under_slab_r_value": (SPECIFICATIONS_SHEET, "E19"),
    "wcc-ceiling_r_value": (SPECIFICATIONS_SHEET, "E20"),
    "wcc-raised_heel": (SPECIFICATIONS_SHEET, "E21"),
    "wcc-total_ua_alternative": (SPECIFICATIONS_SHEET, "E22"),
    "wcc-air_leakage_ach": (SPECIFICATIONS_SHEET, "E27"),
    "wcc-ventilation_type": (SPECIFICATIONS_SHEET, "E28"),
    "wcc-ventilation_brand": (SPECIFICATIONS_SHEET, "E29"),
    "wcc-ventilation_model": (SPECIFICATIONS_SHEET, "E30"),
    "wcc-hrv_asre": (SPECIFICATIONS_SHEET, "E31"),
    "wcc-furnace_brand": (SPECIFICATIONS_SHEET, "E36"),
    "wcc-furnace_model": (SPECIFICATIONS_SHEET, "E37"),
    "wcc-furnace_afue": (SPECIFICATIONS_SHEET, "E38"),
    "wcc-furnace_location": (SPECIFICATIONS_SHEET, "E43"),
    "wcc-duct_location": (SPECIFICATIONS_SHEET, "E44"),
    "wcc-duct_leakage": (SPECIFICATIONS_SHEET, "E45"),
    "wcc-dwhr_installed": (SPECIFICATIONS_SHEET, "E51"),
    "wcc-water_heater_brand": (SPECIFICATIONS_SHEET, "E52"),
    "wcc-water_heater_model": (SPECIFICATIONS_SHEET, "E53"),
    "wcc-gas_water_heater_uef": (SPECIFICATIONS_SHEET, "E54"),
    "wcc-electric_water_heater_uef": (SPECIFICATIONS_SHEET, "E55"),
}

ANNOTATIONS = {
    "wcc-envelope-option": (CODE_CREDITS_SHEET, "B13"),
    "wcc-air-leakage-option": (CODE_CREDITS_SHEET, "B14"),
    "wcc-hvac-option": (CODE_CREDITS_SHEET, "B15"),
    "wcc-hvac-distribution-option": (CODE_CREDITS_SHEET, "B16"),
    "wcc-dwhr-option": (CODE_CREDITS_SHEET, "B17"),
    "wcc-water-heating-option": (CODE_CREDITS_SHEET, "B18"),
    "wcc-renewable-electric-option": (CODE_CREDITS_SHEET, "B19"),
    "wcc-appliance-option": (CODE_CREDITS_SHEET, "B20"),
}

DEFAULT_PROVIDER_SLUG = "peci"


class WashingtonCodeCreditUploadHandler:
    def __init__(self, *_args, **kwargs):
        self.result_object = kwargs.get("result_object", None)
        self.company = kwargs.get("company", None)
        self.user = kwargs.get("user", None)
        self.app_log = kwargs.get("app_log", kwargs.get("log", None))
        if self.app_log is None and self.result_object:
            self.app_log = LogStorage(
                model_id=self.result_object.id,
                level=kwargs.get("log_level", logging.NOTSET),
            )
            self.set_flags(
                home_updated=False,
                home_created=False,
                program_added=False,
                home_already_certified=False,
                certification_date=False,
                home_certified=False,
                has_row_flags=False,
            )
        self.collector = None
        self.data_links = {}

    @cached_property
    def eep_program(self):
        return EEPProgram.objects.filter(slug="washington-code-credit").first()

    def extract_section(
        self,
        wb: Workbook,
        fields: dict,
        as_list: bool = False,
        retain_empty: bool = False,
        null_select_stmts: bool = False,
        key_name: str = "field",
    ) -> typing.Union[list, dict]:
        """Pulls data from the sheet"""
        results = []
        for field, (sheet, cell) in fields.items():
            value = wb[sheet][cell].value
            if null_select_stmts and value and "Select" in str(value):
                value = None
            if value is None and retain_empty is False:
                continue
            if wb[sheet][cell].number_format == "0%":  # This handles the percent to a real number
                value = int(round(value * 100))
            data = {key_name: field, "content": value}
            if hasattr(wb[sheet][cell], "comment") and wb[sheet][cell].comment:
                data["comment"] = wb[sheet][cell].comment
            results.append(data)
        if as_list:
            return results
        return {x[key_name]: x["content"] for x in results}

    def extract_cell_data(self, wb: Workbook = None):
        """One place to pull the data such that we can patch this later"""
        return {
            "subdivision": self.extract_section(wb, SUBDIVISION_DATA, retain_empty=True),
            "community": self.extract_section(wb, COMMUNITY_DATA, retain_empty=True),
            "home": self.extract_section(wb, HOME_DATA, retain_empty=True),
            "home_status": self.extract_section(wb, HOME_STATUS_DATA, retain_empty=True),
            "associations": self.extract_section(wb, ASSOCIATIONS, retain_empty=True),
            "collected_input": self.extract_section(
                wb,
                COLLECTION_DATA,
                retain_empty=False,
                null_select_stmts=True,
                as_list=True,
                key_name="measure",
            ),
            "annotations": self.extract_section(
                wb,
                ANNOTATIONS,
                retain_empty=False,
                null_select_stmts=True,
                as_list=True,
                key_name="type",
            ),
        }

    def set_flags(self, **kwargs):
        if hasattr(self.app_log, "set_flags"):
            self.app_log.set_flags(**kwargs)

    def add_data_link(self, label, url, string):
        self.data_links[label] = f"<a target='_blank' href='{url}'>{string}</a>"
        self.set_flags(**self.data_links)

    def extract_spreadsheet_data(self, result_object: AsynchronousProcessedDocument) -> dict:
        """Read the contents of the data"""

        byte_file = result_object.document.read()

        try:
            wb = load_workbook(ContentFile(byte_file), data_only=True, read_only=True)
            result_object.document.close()
        except zipfile.BadZipfile:
            result_object.document.close()
            msg = f"{result_object.document.name} does not appear to be a Microsoft Excel Document"
            self.app_log.error(msg)
            raise zipfile.BadZipfile(msg)

        required_sheets = {
            "Step 1 - AXIS Home Details",
            "Step 2 - Select Code Credits",
            "Step 3 - Enter Specifications",
        }
        if not required_sheets.issubset(set(wb.sheetnames)):
            msg = "This is not the correct xls document.  Expected sheets do not exist"
            self.app_log.error(msg)
            raise IndexError(msg)

        return self.extract_cell_data(wb)

    def validate_associations(
        self,
        builder: typing.Union[str, None] = None,
        electric_utility: typing.Union[str, None] = None,
        gas_utility: typing.Union[str, None] = None,
        hvac: typing.Union[str, None] = None,
        rater: typing.Union[str, None] = None,
    ) -> dict:
        builder_org = None
        if builder is not None:
            builder_org = Company.objects.verify_existence_for_company(
                name=builder,
                company=self.company,
                company_type="builder",
                log=self.app_log,
            )
        electric_utility_org = None
        if electric_utility:
            electric_utility_org = Company.objects.verify_existence_for_company(
                name=electric_utility,
                company=self.company,
                company_type="utility",
                log=self.app_log,
            )
        gas_utility_org = None
        if gas_utility:
            gas_utility_org = Company.objects.verify_existence_for_company(
                name=gas_utility,
                company=self.company,
                company_type="utility",
                log=self.app_log,
            )
        hvac_org = None
        if hvac:
            hvac_org = Company.objects.verify_existence_for_company(
                name=hvac, company=self.company, company_type="hvac", log=self.app_log
            )
        rater_org = None
        if self.company.company_type == "rater":
            rater_org = self.company
        elif self.company.company_type == "provider" and rater:
            rater_org = Company.objects.verify_existence_for_company(
                name=rater, company=self.company, company_type="rater", log=self.app_log
            )
        provider_org = None
        if self.company.company_type == "rater":
            provider_org = Company.objects.get(slug=DEFAULT_PROVIDER_SLUG)
        elif self.company.company_type == "provider":
            provider_org = self.user.company

        return {
            "builder": builder_org,
            "electric_utility": electric_utility_org,
            "gas_utility": gas_utility_org,
            "hvac": hvac_org,
            "rater": rater_org,
            "provider": provider_org,
        }

    def create_associations(
        self,
        home_status: EEPProgramHomeStatus,
        builder: typing.Union[Company, None] = None,
        rater: typing.Union[Company, None] = None,
        provider: typing.Union[Company, None] = None,
        electric_utility: typing.Union[Company, None] = None,
        gas_utility: typing.Union[Company, None] = None,
        hvac: typing.Union[Company, None] = None,
    ) -> dict:
        data = {
            "builder": builder,
            "rater": rater,
            "provider": provider,
            "electric_utility": electric_utility,
            "gas_utility": gas_utility,
            "hvac": hvac,
        }

        existing_relation_ids = home_status.home.relationships.all().values_list(
            "company", flat=True
        )
        existing_companies = Company.objects.filter(id__in=existing_relation_ids)

        to_add, to_remove = [], []
        for narrowed_company_type, company in data.items():
            company_type = narrowed_company_type
            if "utility" in narrowed_company_type:
                company_type = "utility"
            if company is None or company in existing_companies:
                continue
            # There can only be one of these
            if company_type in ["builder", "utilty", "hvac"]:
                kw = {
                    "company_type": company_type,
                }
                if narrowed_company_type == "electric_utility":
                    kw["utilityorganization__electricity_provider"] = True
                if narrowed_company_type == "gas_utility":
                    kw["utilityorganization__gas_provider"] = True
                if existing_companies.filter(**kw).count():
                    to_remove += [existing_companies.filter(**kw).values_list("id", flat=True)]
            to_add.append(company.id)

        result = {"added": None, "removed": None}
        if to_remove:
            self.app_log.info(f"Removing {len(to_remove)} associations to {home_status.home}")
            result["removed"] = Company.objects.filter(id__in=to_remove)
            home_status.home.relationships.filter(company_id__in=to_remove).delete()

        if to_add:
            self.app_log.info(f"Adding {len(to_add)} associations to {home_status.home}")
            result["added"] = Company.objects.filter(id__in=to_add)
            create_or_update_spanning_relationships(list(result["added"]), home_status)
            home_status.home._generate_utility_type_hints(None, None, discover=True)
        return result

    def validate_community(
        self,
        name: typing.Union[str, None] = None,
    ) -> typing.Union[str, Community]:
        community = None
        if name:
            community = Community.objects.verify_existence_for_company(
                name=name, company=self.company, log=self.app_log
            )
            if community:
                self.add_data_link("Community", community.get_absolute_url(), community)
        return community

    def validate_subdivision(
        self,
        name: typing.Union[str, None] = None,
        builder: typing.Union[Company, BuilderOrganization, None] = None,
        community: typing.Union[Community, None] = None,
    ) -> tuple:
        subdivision = None
        builder_org = None
        if name:
            subdivision = Subdivision.objects.verify_for_company(
                name=name,
                community=community,
                builder=builder,
                company=self.company,
                log=self.app_log,
            )
            if subdivision:
                builder_org = Company.objects.get(id=subdivision.builder_org.id)
                self.add_data_link("Subdivision", subdivision.get_absolute_url(), subdivision)
            else:
                builder_org = builder
        return subdivision, builder_org

    def validate_county(
        self, name: typing.Union[str, None] = None, state: str = "WA"
    ) -> typing.Union[str, County]:
        """Get a valid county"""
        county = None
        if name:
            county = County.objects.verify(name=name, state=state, log=self.app_log)
        return county

    def validate_city(
        self,
        name: typing.Union[str, None] = None,
        county: typing.Union[str, County] = None,
        state: str = "WA",
    ) -> typing.Union[str, City]:
        """Get a valid city"""
        city = None
        if name:
            city = City.objects.verify(
                name=name,
                county=county,
                state=state,
                log=self.app_log,
                ignore_missing=True,
            )
        return city

    def validate_home(
        self,
        lot_number: typing.Union[str, None] = None,
        street_line1: typing.Union[str, None] = None,
        street_line2: typing.Union[str, None] = None,
        zipcode: typing.Union[str, None] = None,
        state: str = "WA",
        city: typing.Union[City, None] = None,
        county: typing.Union[County, None] = None,
        subdivision: typing.Union[str, Subdivision] = None,
        builder: typing.Union[Company, BuilderOrganization, None] = None,
    ) -> typing.Union[Home, dict]:
        home = Home.objects.verify_for_user(
            lot_number=lot_number,
            street_line1=street_line1,
            street_line2=street_line2,
            city=city,
            county=county,
            state=state,
            zipcode=zipcode,
            is_multi_family=False,
            subdivision=subdivision,
            builder=builder,
            bulk_uploaded=False,
            user=self.user,
            log=self.app_log,
        )
        if home is None:
            return dict(
                lot_number=lot_number,
                street_line1=street_line1,
                street_line2=street_line2,
                city=city,
                county=county,
                state=state,
                zipcode=zipcode,
                is_multi_family=False,
                subdivision=subdivision,
                builder=builder,
                bulk_uploaded=False,
                user=self.user,
                log=self.app_log,
            )
        self.add_data_link("Home", home.get_absolute_url(), f"{home} [{home.id}]")
        return home

    def validate_home_status(
        self,
        home: typing.Union[Home, None],
        company: typing.Union[str, Company],
        rater_of_record: typing.Union[str, None] = None,
    ) -> typing.Union[EEPProgramHomeStatus, dict]:
        company_obj = company if isinstance(company, Company) else None
        if company_obj is None and self.company.company_type == "rater":
            company_obj = self.company
        if company_obj is None:
            if isinstance(company, str):
                # Priority on this
                company_obj = Company.objects.verify_existence_for_company(
                    name=company,
                    company=self.company,
                    company_type="rater",
                    log=self.app_log,
                    log_errors=False,
                    include_self=True,
                )
                if company_obj is None:
                    company_obj = Company.objects.verify_existence_for_company(
                        name=company,
                        company=self.company,
                        company_type="provider",
                        log=self.app_log,
                        log_errors=False,
                    )

        rater_o_record = None
        if company_obj is None:
            self.app_log.error("Unable to determine company bound to home status")
        elif rater_of_record:
            rater_o_record = User.objects.verify_rater_of_record_for_company(
                rater_of_record=rater_of_record, company=company_obj, log=self.app_log
            )
        if home:
            home_status = EEPProgramHomeStatus.objects.verify_for_company(
                home=home,
                eep_program=self.eep_program,
                floorplan=None,
                company=company_obj,
                user=self.user,
                ignore_missing=False,
                ignore_missing_floorplan=False,
                overwrite_old_answers=False,
                rater_of_record=rater_o_record,
                log=self.app_log,
            )
            if isinstance(home_status, EEPProgramHomeStatus):
                return home_status
        return dict(
            home=home,
            eep_program=self.eep_program,
            floorplan=None,
            company=company_obj,
            user=self.user,
            ignore_missing=False,
            ignore_missing_floorplan=False,
            overwrite_old_answers=False,
            rater_of_record=rater_o_record,
            log=self.app_log,
        )

    def validate_annotations(
        self,
        home_status: typing.Union[EEPProgramHomeStatus, None],
        annotations: typing.Union[list, None] = None,
        create: bool = False,
    ) -> list:
        annotations = annotations if annotations is not None else []

        existing_annotations, content_type = {}, None
        if home_status:
            existing_annotations = {
                annotation.type: annotation for annotation in home_status.annotations.all()
            }

        annotations = validate_annotations(
            eep_program=self.eep_program, annotations=annotations, log=self.app_log
        )

        results, to_add = [], []
        for item in annotations:
            if existing_annotations.get(item["type"]):
                if existing_annotations[item["type"]].content == item["content"]:
                    continue
                results.append({"type": item["type"].slug, "content": item["content"]})
                if create:
                    annotation = existing_annotations[item["type"]]
                    annotation.content = item["content"]
                    annotation.user = self.user
                    annotation.save()
                    if annotation.history.all()[0].history_user is None:
                        history = annotation.history.all()[0]
                        history.history_user = self.user
                        history.save()
            else:
                results.append({"type": item["type"].slug, "content": item["content"]})
                if home_status and create:
                    if content_type is None:
                        content_type = ContentType.objects.get_for_model(home_status)
                    to_add.append(
                        Annotation(
                            type=item["type"],
                            content=item["content"],
                            content_type=content_type,
                            object_id=home_status.pk,
                            user=self.user,
                        )
                    )
        if create and to_add:
            results = bulk_create_with_history(to_add, Annotation, default_user=self.user)
            self.app_log.info(f"Added {len(to_add)} Annotations for {home_status.home}")

        return results

    def create_annotations(
        self,
        home_status: EEPProgramHomeStatus,
        annotations: typing.Union[list, None] = None,
    ) -> list:
        return self.validate_annotations(
            home_status=home_status, annotations=annotations, create=True
        )

    def validate_collected_input(
        self,
        eep_program: EEPProgram,
        home_status: typing.Union[EEPProgramHomeStatus, None],
        collected_input: typing.Union[list, None] = None,
        answer_overwrite: bool = True,
        user: User = None,
        create: bool = False,
    ) -> list:
        collected_input = collected_input if collected_input is not None else []
        collection_request = eep_program.collection_request
        if home_status:
            collection_request = home_status.collection_request
            self.app_log.debug(
                "Using Collection Request on home status (%(home_status)s)",
                dict(home_status=home_status.pk),
            )
        context = {
            "user": user or self.user,
        }
        self.collector = ExcelChecklistCollector(collection_request, **context)
        self.collector.context["user_role"] = self.collector.get_user_role(self.user)
        instruments = self.collector.get_xls_instruments()
        instrument_lookup = {i.measure_id: i for i in instruments}
        results, to_add, to_remove = [], [], []
        for item in collected_input:
            measure, data, comment = (
                item["measure"],
                item["content"],
                item.get("comment"),
            )
            if measure not in instrument_lookup:
                self.app_log.info(f"Skipping {measure!r} it's not required at this point")
                continue
            instrument = instrument_lookup[measure]
            extra = {"comment": comment} if comment else {}
            payload = self.collector.make_payload(instrument, data, extra=extra)

            if self.collector.get_inputs(payload["instrument"]):
                existing = self.collector.get_inputs(payload["instrument"]).first()
                if existing.data["input"] == payload["data"]["input"]:
                    self.app_log.info("Skipping existing input {data!r} for {measure}")
                    continue
                elif answer_overwrite is False:
                    continue
                else:
                    to_remove.append(existing)
            try:
                payload = self.collector.clean_payload(payload)
            except ValidationError as err:
                self.app_log.error(f"Error with validating {measure} - {err}")
                continue
            to_add.append(payload)
            results.append(item)
        if create:
            results = {
                "added": [] if to_add else None,
                "removed": [] if to_remove else None,
            }
            for item in to_add:
                item["measure"] = item.pop("instrument").measure_id
                results["added"].append(self.collector.store(**item))
            for item in to_remove:
                results["removed"].append(
                    {
                        "id": item.id,
                        "data": item.data,
                        "measure": item.instrument.measure,
                    }
                )
                item.delete()

        return results

    def create_collected_input(
        self,
        eep_program: EEPProgram,
        home_status: EEPProgramHomeStatus,
        collected_input: typing.Union[list, None] = None,
        answer_overwrite: bool = True,
        user: User = None,
    ) -> list:
        return self.validate_collected_input(
            eep_program=eep_program,
            home_status=home_status,
            collected_input=collected_input,
            answer_overwrite=answer_overwrite,
            user=user,
            create=True,
        )

    def validate_data(self, data: dict) -> dict:
        """This will go through and validate each piece.  Order matters here.  The only things we
        will create is a home / home_status.  Everything else is expected to be here.  We want to
        collect errors so we don't have the user piece meal upload
        """

        associations = self.validate_associations(**data["associations"])
        community = self.validate_community(**data["community"])
        subdivision, builder = self.validate_subdivision(
            name=data["subdivision"]["name"],
            builder=associations["builder"],
            community=community,
        )
        if associations["builder"] is None and builder:
            associations["builder"] = builder

        county = self.validate_county(name=data["home"]["county"], state="WA")

        city = self.validate_city(
            name=data["home"]["city"],
            county=county,
        )
        if city and county is None:
            county = city.county

        home = self.validate_home(
            lot_number=data["home"]["lot_number"],
            street_line1=data["home"]["street_line1"],
            street_line2=data["home"]["street_line2"],
            zipcode=data["home"]["zipcode"],
            city=city,
            county=county,
            subdivision=subdivision,
            builder=associations["builder"],
        )

        home_status = self.validate_home_status(
            home=home if isinstance(home, Home) else None,
            company=data["home_status"]["company"] or associations["rater"],
            rater_of_record=data["home_status"]["rater_of_record"],
        )
        if isinstance(home_status, EEPProgramHomeStatus):
            company = home_status.company
        else:
            company = home_status.get("company")

        annotations = self.validate_annotations(
            home_status=home_status if isinstance(home_status, EEPProgramHomeStatus) else None,
            annotations=data["annotations"],
        )

        user = self.user
        if company and self.user.company != company:
            user = company.users.first()

        collected_input = self.validate_collected_input(
            eep_program=self.eep_program,
            home_status=home_status if isinstance(home_status, EEPProgramHomeStatus) else None,
            collected_input=data["collected_input"],
            user=user,
        )

        return {
            "associations": associations,
            "home": home,
            "home_status": home_status,
            "annotations": annotations,
            "collected_input": collected_input,
        }

    def create(self, validated_data: dict) -> dict:
        """Create the home"""

        result = {
            "home": None,
            "home_status": None,
            "associations": None,
            "annotations": None,
            "collected_input": None,
        }

        home = validated_data.get("home")
        if home is not None and isinstance(home, dict):
            home = Home.objects.verify_and_create_for_user(**home)[0]
            if home is None:
                self.app_log.error("Home was not created!?")
                return result
            self.add_data_link("Home", home.get_absolute_url(), f"{home} [{home.id}]")
            result["home"] = home

        home_status = validated_data.get("home_status")
        if home_status is not None and isinstance(home_status, dict):
            if home_status["home"] is None:
                home_status["home"] = home
            home_status = EEPProgramHomeStatus.objects.verify_and_create_for_user(**home_status)
            if home_status is None:
                self.app_log.error("Program was not added to the the home!?")
                return result
            self.set_flags(program_added=True)
            result["home_status"] = home_status

        associations = validated_data.get("associations")
        if associations is not None and isinstance(associations, dict):
            relationships = self.create_associations(home_status=home_status, **associations)
            result["associations"] = relationships

        _annotations = validated_data.get("annotations")
        if _annotations is not None and isinstance(_annotations, list):
            annotations = self.create_annotations(
                home_status=home_status,
                annotations=_annotations,
            )
            result["annotations"] = annotations

        _collected_input = validated_data.get("collected_input")
        if _collected_input is not None and isinstance(_collected_input, list):
            collected_input = self.create_collected_input(
                home_status=home_status,
                eep_program=self.eep_program,
                collected_input=_collected_input,
                user=self.user,
            )
            result["collected_input"] = collected_input

        return result

    def notify_parties(
        self,
        home: Home,
        home_status: EEPProgramHomeStatus,
        associations: dict,
        annotations: list,
        collected_input: dict,
    ) -> bool:
        # Remove all prior uploads

        has_new_associations = associations.get("added") and len(associations["added"])
        has_new_input = collected_input.get("added") and len(collected_input["added"])
        new_input_or_annotations = has_new_input or has_new_associations or len(annotations)

        if home is not None and new_input_or_annotations:
            existing = CustomerDocument.objects.filter(
                document__contains="washington_code_credit_upload_",
                content_type=ContentType.objects.get_for_model(home),
                object_id=home.pk,
            ).first()

            _now = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            _ext = os.path.splitext(self.result_object.document.name)[1]

            result_object = AsynchronousProcessedDocument.objects.get(id=self.result_object.id)
            CustomerDocument.objects.store(
                content_object=home,
                company=self.user.company,
                document=result_object.document.read(),
                filename=f"washington_code_credit_upload_{_now}{_ext}",
                filesize=result_object.document.size,
                type="document",
                pk=existing.pk if existing else None,
            )
            result_object.document.close()

            notification_user = self.user
            context = {
                "home": home,
                "url": home.get_absolute_url(),
            }
            WashingtonCodeCreditUploadComplete(url=context["url"]).send(
                context=context,
                user=notification_user,
            )
            return True
        return False


class WashingtonCodeCreditUploadTask(Task):
    ignore_result = False

    # store_errors_even_if_ignored = True
    time_limit = 60 * 60

    def run(self, result_object_id: int, company_id: int, user_id: int, *args, **kwargs):
        self.update_state(state=states.STARTED, meta={"stage": "initializing", "percentage": 1})
        log.info(f"Starting {self.__class__.__name__} ({self.request.id})")

        result_object = AsynchronousProcessedDocument.objects.get(id=result_object_id)
        handler = WashingtonCodeCreditUploadHandler(
            result_object=result_object,
            company=Company.objects.get(id=company_id),
            user=User.objects.get(id=user_id),
        )
        self.update_state(state=states.STARTED, meta={"stage": "initialized", "percentage": 5})
        try:
            data = handler.extract_spreadsheet_data(result_object)
        except ValueError as err:
            return self.retry(exc=err)

        self.update_state(state=states.STARTED, meta={"stage": "extracting data", "percentage": 10})
        cleaned_data = handler.validate_data(data)

        self.update_state(state=states.STARTED, meta={"stage": "validating data", "percentage": 40})
        if handler.app_log.has_errors:
            self.update_state(
                state=states.FAILURE,
                meta={"stage": "validating data", "percentage": 40},
            )
            handler.app_log.update_model(throttle_seconds=None)
            return

        final = handler.create(cleaned_data)
        self.update_state(state=states.STARTED, meta={"stage": "creating data", "percentage": 70})

        if handler.app_log.has_errors:
            self.update_state(
                state=states.FAILURE,
                meta={"stage": "validating data", "percentage": 70},
            )
            handler.app_log.update_model(throttle_seconds=None)
            return

        self.update_state(state=states.STARTED, meta={"stage": "creating data", "percentage": 90})
        handler.notify_parties(**final)

        handler.app_log.update_model(throttle_seconds=None)
        self.update_state(state=states.SUCCESS, meta={"stage": "Complete", "percentage": 100})


class WashingtonCodeCreditBulkReportTask(Task):
    ignore_result = False

    time_limit = 60 * 60

    def run(
        self,
        asynchronous_process_document_id: int,
        eep_program_home_status_ids: list = [],
        user_id: int = None,
    ):
        task_meta = {
            "processing": {"current": 0, "total": len(eep_program_home_status_ids)},
            "writing": {"current": 0, "total": 1},
        }
        self.update_state(state=states.STARTED, meta=task_meta)
        log.info(f"Starting {self.__class__.__name__} ({self.request.id})")

        user = User.objects.get(id=user_id)
        asynchronous_process_document = AsynchronousProcessedDocument.objects.get(
            id=asynchronous_process_document_id
        )

        app_log = LogStorage(model_id=asynchronous_process_document_id)

        with tempfile.SpooledTemporaryFile() as tmp:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as archive:
                for home_status_id in eep_program_home_status_ids:
                    try:
                        instance = FastTrackSubmission.objects.get(home_status=home_status_id)
                    except FastTrackSubmission.DoesNotExist:
                        msg = f"FastTrackSubmission instance is missing for home status id {home_status_id}"
                        app_log.error(msg)
                        app_log.update_model(throttle_seconds=None)
                        return

                    try:
                        serializer = WashingtonCodeCreditReportSerializer(instance=instance)
                    except ValueError:
                        msg = f"FastTrackSubmission instance is invlid for home status id {home_status_id}"
                        app_log.error(msg)
                        app_log.update_model(throttle_seconds=None)
                        return

                    virtual_workbook = io.BytesIO()

                    report = WashingtonCodeCreditReport()
                    report.build(response=virtual_workbook, user=user, data=serializer.data)

                    archive.writestr(
                        f"WCC_Compliance_Report_{home_status_id}.pdf",
                        virtual_workbook.getvalue(),
                    )
                    task_meta["processing"]["current"] += 1
                    self.update_state(state="Started", meta=task_meta)
            tmp.seek(0)
            app_log.info("Saving Report")
            filename = "WCC_Compliance_Reports.zip"
            asynchronous_process_document.document.save(filename, ContentFile(tmp.read()))

        app_log.info("Done")
        task_meta["writing"]["current"] = 1
        self.update_state(state="Done", meta=task_meta)
        app_log.update_model(throttle_seconds=None)


WashingtonCodeCreditUploadTask = celery_app.register_task(WashingtonCodeCreditUploadTask())
WashingtonCodeCreditBulkReportTask = celery_app.register_task(WashingtonCodeCreditBulkReportTask())
