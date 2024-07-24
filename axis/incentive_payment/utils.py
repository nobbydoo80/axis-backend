"""utils.py: Django incentive_payment"""


import logging
import datetime

from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ObjectDoesNotExist

from openpyxl import Workbook
from openpyxl.styles import colors, Font, Alignment, PatternFill
from openpyxl.styles.fills import fills
from openpyxl.utils import get_column_letter

from django.contrib.auth import get_user_model
from axis.annotation.models import Type
from axis.company.models import Company
from axis.incentive_payment.models import IncentivePaymentStatus, IPPItem
from axis.relationship.models import Relationship

__author__ = "Steven Klass"
__date__ = "4/16/13 9:41 AM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

log = logging.getLogger(__name__)
User = get_user_model()


def add_annotation_types(**kwargs):
    """This will add in default annotation types"""

    annotation_data = [
        {
            "name": "APS IPP Status Note",
            "data_type": "open",
            "description": "APS IPP Note",
            "content_types": [IncentivePaymentStatus],
            "is_unique": False,
            "slug": "aps-ipp-status-note",
        }
    ]
    log.debug("Adding in annotation data")
    for item in annotation_data:
        type, create = Type.objects.get_or_create(
            name=item["name"],
            description=item.get("description", ""),
            data_type=item.get("data_type"),
            is_unique=item.get("is_unique", True),
            valid_multiplechoice_values=item.get("valid_multiplechoice_values", ""),
        )
        for content_type in item.get("content_types"):
            content_type_obj = ContentType.objects.get_for_model(content_type)
            type.applicable_content_types.add(content_type_obj)
        type.save()
        company = Company.objects.get(name=kwargs.get("company", "APS"))
        Relationship.objects.validate_or_create_relations_to_entity(
            entity=type, direct_relation=company, implied_relations=[]
        )


def gather_cycle_time_metrics():
    stats = IncentivePaymentStatus.objects.filter(state="complete")
    print("Analyzed {} Programs".format(stats.count()))

    earliest = datetime.datetime.now(datetime.timezone.utc)

    deltas = {}
    delta_list = []
    for i in range(50):
        deltas[i] = 0
    numerator = 0
    denominator = 0
    for stat in stats:
        first = stat.get_state_transitions()[0]
        last = list(stat.get_state_transitions())[-1]

        if first.start_time < earliest:
            earliest = first.start_time

        delta = (last.start_time - first.start_time).days

        if delta > 0:
            delta_list.append(delta)
            try:
                deltas[delta] += 1
            except KeyError:
                deltas[delta] = 1
            numerator += delta
            denominator += 1

    print("Earliest Record: {}".format(earliest))
    print("Reduced to {} records".format(len(deltas)))

    print("Cycle Times")
    for k, v in deltas.items():
        print("{}, {}".format(k, v))

    for i in range(50):
        print("{}, ".format(i)),
    print("")
    for i in range(50):
        print("{}, ".format(deltas[i])),

    print("")

    def average(s):
        return sum(s) * 1.0 / len(s)

    avg = average(delta_list)
    print("Average", avg)

    variance = map(lambda x: (x - avg) ** 2, delta_list)
    print("Variance", variance)

    import math

    standard_deviation = math.sqrt(average(variance))
    print("Std Dev", standard_deviation)

    return delta_list


class CycleTimeMetrics:
    def __init__(self, *args, **kwargs):
        start_date = kwargs.get("start_date", None)
        if start_date is None:
            now = datetime.datetime.now(datetime.timezone.utc)
            start_date = datetime.datetime(now.year, 1, 1, tzinfo=datetime.timezone.utc)

        user = kwargs.get("user", None)
        if user is None:
            user = User.objects.get(username="rjohnson")

        self.earliest = now()
        self.delta_days = {}
        for i in range(50):
            self.delta_days[i] = 0
        self.deltas = []

        stat_ids = IPPItem.objects.filter(incentive_distribution__paid_date__gte=start_date)
        stat_ids = stat_ids.values_list("home_status", flat=True)
        incentive_stats = IncentivePaymentStatus.objects.filter_by_user(user)
        self.incentive_stats = incentive_stats.filter(state="complete", home_status__in=stat_ids)

    def set_cell_header_style(self, cell, **kwargs):
        cell.font = Font(name="Arial", size=14, bold=True, color=colors.WHITE)
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=colors.DARKRED)

        for key, value in kwargs.items():
            setattr(cell, key, value)

    def _set_cell_default_style(self, cell):
        cell.font = Font(name="Arial", size=12)

    def rec_getattr(self, obj, attr, default=None):
        """Get object's attribute. May use dot notation."""
        if "." not in attr:
            if "()" in attr:
                attr = attr.split("(")[0]
                return getattr(obj, attr, default)()
            return getattr(obj, attr, default)
        else:
            split_attr = attr.split(".")
            return self.rec_getattr(getattr(obj, split_attr[0]), ".".join(split_attr[1:]), default)

    def get_rater_paid_date(self, status):
        try:
            stat = IPPItem.objects.filter(
                home_status=status.home_status,
                incentive_distribution__customer__company_type="provider",
            )[0]
            return stat.incentive_distribution.paid_date
        except (ObjectDoesNotExist, IndexError):
            return "-"

    def get_builder_paid_date(self, status):
        try:
            stat = IPPItem.objects.filter(
                home_status=status.home_status,
                incentive_distribution__customer__company_type="builder",
            )[0]
            return stat.incentive_distribution.paid_date
        except (ObjectDoesNotExist, IndexError):
            return "-"

    def get_received_date(self, status):
        try:
            stat = list(status.get_state_transitions())[0]
            return stat.start_time
        except ObjectDoesNotExist:
            return "-"

    def get_completed_date(self, status):
        try:
            stat = list(status.get_state_transitions())[-1]
            return stat.start_time
        except ObjectDoesNotExist:
            return "-"

    def get_elapsed_days(self, status):
        received_date = self.get_received_date(status)
        completed_date = self.get_completed_date(status)
        if received_date < self.earliest:
            self.earliest = received_date

        delta = (completed_date - received_date).days

        # if delta > 0:
        self.deltas.append(delta)
        try:
            self.delta_days[delta] += 1
        except KeyError:
            self.delta_days[delta] = 1
        return delta

    def report(self, filename=None, display_results=False):
        if filename is None:
            filename = "ipp_cycle_time.xlsx"

        attrs = [
            ("Axis Home", "home_status.home"),
            ("Builder", "home_status.home.get_builder()"),
            ("Rater", "home_status.company"),
            ("EEP Program", "home_status.eep_program"),
            ("Certification Date", "home_status.certification_date"),
            ("Rater Paid Date", "get_rater_paid_date"),
            ("Builder Paid Date", "get_builder_paid_date"),
            ("Received Date", "get_received_date"),
            ("Completed Date", "get_completed_date"),
            ("Elapsed Days", "get_elapsed_days"),
        ]

        row = 1
        workbook = Workbook()
        sheet = workbook.create_sheet(index=0, title="Cycle Time Metrics")
        for col, (label, _attr) in enumerate(attrs):
            cell_label = "{letter}{row}".format(letter=get_column_letter(col + 1), row=row)
            cell = sheet[cell_label]
            cell.value = label
            self._set_cell_header_style(cell)
        row += 1

        print("Analyzed {} Programs".format(self.incentive_stats.count()))

        for object in self.incentive_stats.all().distinct():
            for col, (label, attr) in enumerate(attrs):
                cell = sheet.cell(column=attrs.index((label, attr)) + 1, row=row)
                if hasattr(self, attr):
                    method = getattr(self, attr)
                    value = method(object)
                else:
                    value = self.rec_getattr(object, attr, "-")
                if display_results:
                    print("  {} {} = {}".format(cell.address, attr, value))
                cell.value = "{}".format(value)
                self._set_cell_default_style(cell)
            # Last one is elasped days..
            row += 1

        row = 1
        sheet = workbook.create_sheet(index=0, title="Summary")
        cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
        cell.value = "Earliest Record:"
        cell = sheet["{letter}{row}".format(letter=get_column_letter(1 + 1), row=row)]
        cell.value = self.earliest
        row += 2

        cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
        cell.value = "Cycle Times"
        row += 1
        cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
        cell.value = "Days"
        cell = sheet["{letter}{row}".format(letter=get_column_letter(1 + 1), row=row)]
        cell.value = "Number of entries"
        row += 1
        for k, v in self.delta_days.items():
            cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
            cell.value = k
            cell = sheet["{letter}{row}".format(letter=get_column_letter(1 + 1), row=row)]
            cell.value = v
            row += 1
        row += 2

        def average(s):
            return sum(s) * 1.0 / len(s)

        avg = average(self.deltas)
        cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
        cell.value = "Average"
        cell = sheet["{letter}{row}".format(letter=get_column_letter(1 + 1), row=row)]
        cell.value = avg
        row += 2

        variance = map(lambda x: (x - avg) ** 2, self.deltas)
        cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
        cell.value = "Variance"
        cell = sheet["{letter}{row}".format(letter=get_column_letter(1 + 1), row=row)]
        cell.value = variance
        row += 2

        import math

        standard_deviation = math.sqrt(average(variance))
        variance = map(lambda x: (x - avg) ** 2, self.deltas)
        cell = sheet["{letter}{row}".format(letter=get_column_letter(0 + 1), row=row)]
        cell.value = "Std Deviation"
        cell = sheet["{letter}{row}".format(letter=get_column_letter(1 + 1), row=row)]
        cell.value = standard_deviation

        workbook.save(filename)

        return self.deltas
