"""export_data: Django builder_agreement"""

import datetime
import logging
import tempfile
from functools import partial

from django.contrib.auth import get_user_model
from django.utils import formats
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import colors, fills, Side, Border, Color, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from axis.builder_agreement.models import BuilderAgreement, os
from axis.builder_agreement.views.views import BuilderAgreementMixin
from axis.checklist.xls_checklist import XLSChecklist

__author__ = "Steven Klass"
__date__ = "8/21/15 1:30 PM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

log = logging.getLogger(__name__)
User = get_user_model()

LOGO = os.path.abspath(os.path.dirname(__file__) + "/../core/static/images/sponsor_aps.png")
AXIS_LOGO = os.path.abspath(os.path.dirname(__file__) + "/../core/static/images/Logo_Only_128.png")


class BuilderAgreementXLSExport(XLSChecklist, BuilderAgreementMixin):
    """Generate the Builder Agreement XLS Report"""

    def __init__(self, *_args, **kwargs):
        self.kwargs = kwargs
        self.model = BuilderAgreement
        self.creator = kwargs.get("creator", "Axis")
        self.title = kwargs.get("title", "Builder Information Export")
        self.subject = kwargs.get("subject", "Builder Information Export")
        self.description = kwargs.get("description", "Axis Generated Document")
        self.max_num = kwargs.get("max_num", None)
        self.hers_min = kwargs.get("hers_min", None)
        self.hers_max = kwargs.get("hers_max", None)

        kwargs = {
            "user_id": kwargs.get("user_id"),
            "log": kwargs.get("log"),
            "max_num": kwargs.get("max_num", None),
            "task": kwargs.get("task"),
            "reuse_storage": kwargs.get("reuse_storage", True),
        }

        self.user = User.objects.get(id=kwargs.get("user_id"))

        self.aps_sponsored = "aps" in self.user.company.sponsors.values_list("slug", flat=True)
        if not self.aps_sponsored and self.user.company.slug == "aps":
            self.aps_sponsored = True
        self.queryset = None
        super(BuilderAgreementXLSExport, self).__init__(**kwargs)

    def get_queryset(self):
        """Apply filters requested by UI."""
        if self.queryset is not None:
            return self.queryset
        self.user = User.objects.get(id=self.kwargs.get("user_id"))
        queryset = self.model.objects.filter_by_user(self.user)
        queryset = queryset.select_related("subdivision__community")
        for key, value in self.kwargs.items():
            try:
                if "," in value:
                    self.kwargs["key"] = ",".split(value)
            except TypeError:
                pass
        self.queryset = self.apply_filters(queryset, self.user, self.kwargs)
        return self.queryset

    def get_floorplan_data(self, builder_id, subdivision_id, company_id):
        """Pull out the floorplan data"""
        data = {"company": None, "floorplans": []}
        floorplans = self.model.objects.get_floorplans_for_user(
            self.user, company_id, builder_id, subdivision_id
        )

        if self.hers_min:
            floorplans = floorplans.filter(remrate_target__hers__score__gte=self.hers_min)
        if self.hers_max:
            floorplans = floorplans.filter(remrate_target__hers__score__lte=self.hers_max)

        if floorplans.count():
            log.debug("     Total Floorplans %s", floorplans.count())
            data["company"] = floorplans.first().owner
            data["floorplans"] = floorplans
        else:
            log.info(
                "Skipping Floorplans Builder: %s Subdivision: %s Company: %s found",
                builder_id,
                subdivision_id,
                company_id,
            )
        return data

    def get_subdivision_data(self, builder_id, subdivision_id):
        """Pull out the subdivision data"""
        from axis.subdivision.models import Subdivision

        subdivision = subdivision_id
        if subdivision_id:
            subdivision = Subdivision.objects.get(id=subdivision_id)
        agreement = BuilderAgreement.objects.get(subdivision=subdivision, builder_org_id=builder_id)
        data = {"companies": [], "subdivision": subdivision, "agreement": agreement}
        # pylint: disable=not-an-iterable
        for company in self.model.objects.get_raters_and_providers(builder_id, subdivision_id):
            log.debug("    Company %s", company)
            floorplans = self.get_floorplan_data(builder_id, subdivision_id, company.id)
            data["companies"].append(floorplans)
        return data

    def get_builder_data(self, builder_id):
        """Pull the builder data"""
        data = {"builder": None, "builder_id": builder_id, "subdivisions": [], "start_date": None}
        for agreement in self.queryset.filter(builder_org_id=builder_id).order_by("subdivision"):
            data["builder"] = agreement.builder_org
            subdivision_id = agreement.subdivision.id if agreement.subdivision else None
            subdivision = self.get_subdivision_data(builder_id, subdivision_id)
            if subdivision.get("companies"):
                log.debug("  Subdivision %s", agreement.subdivision)
                data["subdivisions"].append(subdivision)
                if subdivision.get("agreement") and subdivision.get("agreement").start_date:
                    data["start_date"] = data.get("start_date") or datetime.date(2025, 1, 1)
                    _date = subdivision.get("agreement").start_date < data["start_date"]
                    if not data["start_date"] or _date:
                        data["start_date"] = subdivision.get("agreement").start_date
            else:
                log.info(
                    "Skipping Builder: %s  Subdivision: %s not companies found",
                    builder_id,
                    subdivision_id,
                )
        return data

    def gather_data(self):
        """Gather the data set"""
        data = []
        queryset = self.get_queryset()
        values = queryset.values_list("builder_org", flat=True).order_by("builder_org__name")
        for builder_id in list(set(values)):
            builder = self.get_builder_data(builder_id)
            if builder.get("subdivisions"):
                log.debug(" Builder: %s", builder_id)
                data.append(builder)
            else:
                log.info("Skipping Builder %s not subdivisions found", builder_id)
        data = sorted(data, key=lambda x: x.get("builder").name, reverse=True)
        return data

    def set_no_border_row(self, row=None, cols=10, start_col=1, bold=False, **attrs):
        """Visual effects"""
        if row is None:
            row = self.row
        style = {
            "font": Font(name="Helvetica", size=12, bold=bold),
            "border": Border(
                bottom=Side(border_style="thin", color=Color(rgb="FF000000")),
                top=Side(border_style="thin", color=Color(rgb="FF000000")),
                left=Side(border_style="thin", color=Color(rgb="FF000000")),
                right=Side(border_style="thin", color=Color(rgb="FF000000")),
            ),
        }
        style.update(attrs)
        cell = self.sheet.cell(row=row, column=start_col)
        for k, v in style.items():
            setattr(cell, k, v)
        for col in range(start_col, cols):
            cell = self.sheet.cell(row=row, column=col)
            for k, v in style.items():
                setattr(cell, k, v)

    def set_cell_bold_style(self, cell, **kwargs):
        """Visual Bold Style"""
        cell.font = Font(name="Helvetica", size=12, bold=True)
        for key, value in kwargs.items():
            setattr(cell, key, value)

    def add_builder_header(self, builder, start_date):
        """Builder Header"""
        cell = self.sheet.cell(row=self.row, column=1)
        cell.value = "{}".format(builder)
        self.set_cell_bold_style(cell)
        self.row += 1
        cell = self.sheet.cell(row=self.row, column=1)
        try:
            text = "Builder Information Since {}".format(
                formats.date_format(start_date, "SHORT_DATE_FORMAT")
            )
        except AttributeError as err:
            log.error("Unable to get start date from %s - %s", start_date, err)
            text = "Builder Information"
        cell.value = text
        self.set_cell_bold_style(cell)
        self.row += 1
        cell = self.sheet.cell(row=self.row, column=1)
        today = formats.date_format(datetime.date.today(), "SHORT_DATE_FORMAT")
        cell.value = "Report Generated: {}".format(today)
        self.set_cell_bold_style(cell)
        self.row += 3

    def set_cell_header_style(self, cell, color=None, **kwargs):
        """General cell header"""
        if color is None:
            color = Color(rgb="666666")

        cell.font = Font("Helvetica", size=12, bold=True, color=colors.WHITE)
        cell.fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=color)

        for key, value in kwargs.items():
            setattr(cell, key, value)

    def add_floorplan_section(self, floorplans):  # noqa: MC0001
        """Floorplan section info"""
        vals = [
            "Name",
            "Number",
            "Simulation Project",
            "HERS No PV",
            "Software Version",
            "Smart Thermostat Count",
        ]
        column = 2

        color = Color(rgb="666666")
        if self.aps_sponsored:
            color = Color(rgb="F47B20")

        top_style = {
            "font": Font(name="Helvetica", size=12, bold=False, color=colors.WHITE),
            "alignment": Alignment(wrap_text=False, vertical="bottom"),
            "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color=color),
            "border": Border(
                # bottom=Side(border_style='thin', color=Color(rgb='FF000000')),
                top=Side(border_style="thin", color=Color(rgb="FF000000")),
                left=Side(border_style="thin", color=Color(rgb="FF000000")),
                right=Side(border_style="thin", color=Color(rgb="FF000000")),
            ),
        }

        for value in vals:
            cell = self.sheet.cell(row=self.row, column=column)
            cell.value = value

            for key, style_value in top_style.items():
                setattr(cell, key, style_value)

            column += 1
        self.sheet.row_dimensions[self.row].height = 16

        self.row += 1
        top_style.update(
            {
                "font": Font(name="Arial", size=11, bold=False),
                "border": Border(
                    # bottom=Side(border_style='thin', color=Color(rgb='FF000000')),
                    top=Side(border_style="thin", color=Color(rgb="FF000000")),
                    left=Side(border_style="thin", color=Color(rgb="FF000000")),
                    right=Side(border_style="thin", color=Color(rgb="FF000000")),
                ),
                "number_format": "General",
            }
        )

        for idx, floorplan in enumerate(floorplans):
            top_style.update(
                {
                    "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color=colors.WHITE),
                    "number_format": "General",
                }
            )

            if idx % 2:
                top_style.update(
                    {
                        "fill": PatternFill(
                            fill_type=fills.FILL_SOLID, start_color=Color(rgb="E6E6E6")
                        )
                    }
                )

            if floorplan == floorplans.last():
                top_style.update(
                    {
                        "border": Border(
                            bottom=Side(border_style="thin", color=Color(rgb="FF000000")),
                            top=Side(border_style="thin", color=Color(rgb="FF000000")),
                            left=Side(border_style="thin", color=Color(rgb="FF000000")),
                            right=Side(border_style="thin", color=Color(rgb="FF000000")),
                        )
                    }
                )

            cell = self.sheet.cell(row=self.row, column=2)
            cell.value = "{}".format(floorplan.name)
            for k, v in top_style.items():
                setattr(cell, k, v)

            cell = self.sheet.cell(row=self.row, column=3)
            cell.value = "{}".format(floorplan.number)
            for k, v in top_style.items():
                setattr(cell, k, v)

            # Prep number format based on current format
            number_top_style = top_style.copy()
            number_top_style.update({"number_format": "0.0"})

            # Dig up the right simulation data
            simulation_info = {
                "project_name": None,
                "hers_score": None,
                "engine_version": None,
            }
            if floorplan.remrate_target:
                try:
                    simulation_info["project_name"] = "{}".format(
                        floorplan.remrate_target.building.project.name
                    )
                except AttributeError:
                    pass
                try:
                    simulation_info["hers_score"] = float(
                        floorplan.remrate_target.energystar.energy_star_v3_pv_score
                    )
                except (AttributeError, TypeError):
                    try:
                        simulation_info["hers_score"] = float(
                            floorplan.remrate_target.energystar.energy_star_v2p5_pv_score
                        )
                    except (AttributeError, TypeError):
                        pass
                try:
                    simulation_info["engine_version"] = floorplan.remrate_target.version
                except AttributeError:
                    pass
            elif floorplan.ekotrope_houseplan:
                houseplan = floorplan.ekotrope_houseplan
                try:
                    simulation_info["project_name"] = "{}".format(houseplan.name)
                except AttributeError:
                    pass
                try:
                    simulation_info["hers_score"] = float(houseplan.analysis.data["hersScoreNoPv"])
                except (AttributeError, TypeError):
                    pass
                try:
                    engine_version = houseplan.project.data.get("algorithmVersion")
                    simulation_info["engine_version"] = engine_version
                    if engine_version is None:
                        log.error(
                            "Unable to get 'algorithmVersion' key from "
                            "houseplan.project.data houseplan id: %(id)s",
                            {"id": houseplan.id},
                        )
                except AttributeError:
                    pass

            cell = self.sheet.cell(row=self.row, column=4)
            cell.value = simulation_info["project_name"] or "-"
            for k, v in top_style.items():
                setattr(cell, k, v)

            cell = self.sheet.cell(row=self.row, column=5)
            if simulation_info["hers_score"]:
                cell.value = simulation_info["hers_score"]
                for k, v in number_top_style.items():
                    setattr(cell, k, v)
            else:
                cell.value = "-"
                for k, v in top_style.items():
                    setattr(cell, k, v)

            cell = self.sheet.cell(row=self.row, column=6)
            cell.value = simulation_info["engine_version"] or "-"
            for k, v in top_style.items():
                setattr(cell, k, v)

            cell = self.sheet.cell(row=self.row, column=7)
            cell.value = floorplan.get_approved_status().thermostat_qty
            for k, v in top_style.items():
                setattr(cell, k, v)

            self.row += 1

    def add_company_and_floorplan(self, subdivision, company, agreement, show_label):
        """Add company Data"""

        def set_cell_style(cell, style):
            """Set cell style"""
            for key, value in style.items():
                setattr(cell, key, value)

        style = {
            "font": Font(name="Helvetica", size=12, bold=True),
            "alignment": Alignment(wrap_text=True, vertical="center", shrinkToFit=False),
        }
        if show_label:
            cell = self.sheet.cell(row=self.row, column=1)
            try:
                cell.value = "{}".format(subdivision.name)
            except AttributeError:
                cell.value = "-"

            set_cell_style(cell, style)

        cell = self.sheet.cell(row=self.row, column=2)
        community = "-"
        if subdivision and subdivision.community:
            community = subdivision.community
        cell.value = "{}".format(community)
        set_cell_style(cell, style)

        cell = self.sheet.cell(row=self.row, column=3)
        alt_name = "-"
        if subdivision:
            alt_name = subdivision.builder_name if subdivision.builder_name else "-"
        cell.value = "{}".format(alt_name)
        set_cell_style(cell, style)

        cstyle = {
            "font": Font(name="Helvetica", size=12, bold=True),
            "alignment": Alignment(
                wrap_text=False, horizontal="center", vertical="center", shrinkToFit=True
            ),
        }

        cell = self.sheet.cell(row=self.row, column=4)
        try:
            start = formats.date_format(agreement.start_date, "SHORT_DATE_FORMAT")
        except AttributeError:
            start = "-"
        cell.value = "{}".format(start)
        set_cell_style(cell, cstyle)

        cell = self.sheet.cell(row=self.row, column=5)
        cell.value = "{}".format(agreement.total_lots)
        set_cell_style(cell, cstyle)

        cell = self.sheet.cell(row=self.row, column=6)
        cell.value = "{}".format(company.get("company"))
        set_cell_style(cell, cstyle)  # Shrink to fix EFL

        cell = self.sheet.cell(row=self.row, column=7)
        cell.value = "Yes" if subdivision.get_fuel_types(self.user) == "Electric" else "No"
        set_cell_style(cell, cstyle)

        cell = self.sheet.cell(row=self.row, column=8)
        if agreement.total_lots:
            cell.value = "{}/{}".format(agreement.lots_paid, agreement.total_lots)
        else:
            cell.value = "{}".format(agreement.lots_paid)
        set_cell_style(cell, cstyle)
        self.sheet.row_dimensions[self.row].height = 14 * 2

        dstyle = {
            "font": Font(name="Helvetica", size=12, bold=True),
            "alignment": Alignment(wrap_text=False, horizontal="left", vertical="bottom"),
        }

        self.row += 1
        cell = self.sheet.cell(row=self.row, column=1)
        cell.value = "Thermostat Eligibility"
        set_cell_style(cell, dstyle)
        self.sheet.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)

        cell = self.sheet.cell(row=self.row, column=3)
        cell.value = subdivision.get_aps_thermostat_option(pretty=True)
        set_cell_style(cell, dstyle)
        self.sheet.row_dimensions[self.row].height = 20
        self.sheet.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=8)

        self.row += 1
        cell = self.sheet.cell(row=self.row, column=1)
        cell.value = "Thermostat Models"
        set_cell_style(cell, dstyle)
        self.sheet.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)

        cell = self.sheet.cell(row=self.row, column=3)
        value = "-"
        if subdivision.get_aps_thermostat_option():
            value = ",".join(subdivision.aps_thermostat_option.get_models_pretty())
        cell.value = value
        set_cell_style(cell, dstyle)
        self.sheet.row_dimensions[self.row].height = 20
        self.sheet.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=8)

        self.row += 1
        self.add_floorplan_section(company.get("floorplans"))

    def add_subdivision_section(self, subdivision):
        """Add subdivision Section"""

        vals = [
            "Subdivision",
            "Community",
            "Subdivision Alt ID",
            "Start Date",
            "Total Lots",
            "Rater/Provider",
            "Electric Only",
            "Lots Paid/Total Lots",
        ]
        column = 1

        color = None
        if self.aps_sponsored:
            color = Color(rgb="001DC6")

        for value in vals:
            cell = self.sheet.cell(row=self.row, column=column)
            cell.value = value
            self.set_cell_header_style(cell, color=color)
            column += 1
        self.sheet.row_dimensions[self.row].height = 16

        self.row += 1
        for idx, company in enumerate(subdivision.get("companies")):
            self.add_company_and_floorplan(
                subdivision.get("subdivision"), company, subdivision.get("agreement"), not idx
            )

        self.row += 1

    def pre_save(self, workbook, sheet):
        """Pre-Save add logos"""
        if self.aps_sponsored:
            image = Image(LOGO)
            sheet.add_image(
                image, anchor="{letter}{row}".format(letter=get_column_letter(5), row=2)
            )
        else:
            image = Image(AXIS_LOGO)
            image.height = 100
            image.width = 100
            sheet.add_image(
                image, anchor="{letter}{row}".format(letter=get_column_letter(6), row=1)
            )

    def write(self, output=None):
        """Write this out"""

        queryset = self.get_queryset()
        self.update_task = partial(self.update_task_progress, total=queryset.count() + 1)
        count = 1
        self.update_task(current=count)
        count += 1

        if output is None:
            _, output = tempfile.mkstemp(suffix=".xlsx", prefix=self.output_prefix)

        workbook = Workbook()
        data = self.gather_data()
        self.update_task(current=2)

        for tab in data:
            self.row = 1
            self.sheet = workbook.create_sheet(index=0, title="{}".format(tab.get("builder"))[:30])
            self.sheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
            self.sheet.set_printer_settings("1", "landscape")
            self.add_builder_header(tab.get("builder"), tab.get("start_date"))
            for subdivision in tab.get("subdivisions"):
                self.add_subdivision_section(subdivision)
                self.update_task(current=count)
                count += 1
            self.sheet.column_dimensions["A"].width = 16
            self.sheet.column_dimensions["B"].width = 16
            self.sheet.column_dimensions["C"].width = 15
            self.sheet.column_dimensions["D"].width = 15
            self.sheet.column_dimensions["E"].width = 15
            self.sheet.column_dimensions["F"].width = 15
            self.sheet.column_dimensions["G"].width = 12
            self.sheet.column_dimensions["H"].width = 15
            self.sheet.column_dimensions["I"].width = 20
            self.sheet.column_dimensions["J"].width = 25
            self.pre_save(workbook, self.sheet)

        msg = "Completed adding %s items to %s"
        log.debug(msg, queryset.count(), output)
        workbook.properties = self.properties()
        workbook.save(output)
        self.update_task(current=count + 1)
        log.debug("Successfully wrote %s.", output)
        return output


def main(args):
    """Main Program -

    :param args: argparse.Namespace
    """
    logging.basicConfig(
        level=logging.INFO,
        datefmt="%H:%M:%S",
        stream=sys.stdout,
        format="%(asctime)s %(levelname)s [%(filename)s] (%(name)s) %(message)s",
    )

    args.verbose = 4 if args.verbose > 4 else args.verbose
    loglevel = 50 - args.verbose * 10
    log.setLevel(loglevel)

    # Tis is a large kwarg test
    # kwargs = {'provider_id': 9,
    #           'end': '08-01-2015',
    #           'city_id': 29515,
    #           'builder_id': 1,
    #           'user_id': 15,
    #           'hers_max': 100}

    # # This is a good None Subdivision Test
    # kwargs = {'builder_id': 26, 'user_id': 15L, 'hers_max': 100}
    #
    # # This is a good None Subdivision Test
    # kwargs = {'builder_id': 26, 'user_id': 15L, 'hers_max': 100}
    #
    # # Multiple Builders
    # kwargs = {'city_id': 29516, 'user_id': 15L}

    kwargs = {"builder_id": 21, "user_id": 15, "hers_min": 60, "hers_max": 63}

    obj = BuilderAgreementXLSExport(**kwargs)
    obj.write("/Users/sklass/Desktop/aps_builder.xlsx")


if __name__ == "__main__":
    import sys
    import argparse

    parser = argparse.ArgumentParser(description="")
    parser.add_argument(
        "-v",
        "--verbose",
        const=1,
        default=1,
        type=int,
        nargs="?",
        help="increase verbosity: 1=errors, 2=warnings, 3=info, 4=debug. "
        "No number means warning. Default is no verbosity.",
    )
    sys.exit(main(parser.parse_args()))
