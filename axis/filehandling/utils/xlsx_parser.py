import collections
import csv
import logging
import unicodedata
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

__author__ = "Steven Klass"
__date__ = "5/28/12 7:56 AM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

log = logging.getLogger(__name__)

CTRL_CHARS = dict.fromkeys(c for c in range(65536) if unicodedata.category(chr(c)).startswith("C"))


# pylint: disable
class XLSXParser(object):
    def __init__(self, *args, **kwargs):
        self.filename = kwargs.get("filename", None)
        self.sheet_name = kwargs.get("sheet_name", None)

        self.dialect = kwargs.get("dialect", csv.excel_tab)

        self.row = kwargs.get("row", 1)
        self.column = kwargs.get("column", 1)
        self.uniq_header_column = kwargs.get("uniq_header_column", None)
        self.sheet = None

    def load_workbook_and_sheet(self, sheet_name=None, data_only=True, read_only=True):
        log.debug("Loading {} {}".format(self.filename, type(self.filename)))
        self.workbook = load_workbook(
            filename=self.filename, data_only=data_only, read_only=read_only
        )
        if sheet_name:
            self.sheet_name = sheet_name
        if self.sheet_name:
            self.sheet = self.workbook[self.sheet_name]
            cell = self.sheet["A1"]  # Try to get a cell
        else:
            data_sheets = []
            for sheet in self.workbook.sheetnames:
                data_found = False
                sheet_obj = self.workbook[sheet]
                for row in range(1, 40):
                    if data_found:
                        break
                    for col in range(1, 20):
                        cell_name = "{}{}".format(get_column_letter(col), row)
                        cell = sheet_obj[cell_name]
                        if cell.value:
                            log.debug("Cell '%s' ans data", cell.value)
                            data_found = True
                            data_sheets.append(sheet)
                            break
            if not len(data_sheets):
                raise AttributeError("No data found in any sheets.")
            # if len(data_sheets) > 1:
            #     raise AttributeError, "Multiple Sheets with data: {}. Only one sheet may " \
            #                           "have data.".format(", ".join(data_sheets))
            # else:
            self.sheet = self.workbook[data_sheets[0]]
            cell = self.sheet["A1"]  # Try to get a cell

    def get_columns(self, uniq_column=None, equivalence_map=None):
        if equivalence_map:
            equivalence_map = dict(equivalence_map)
        if uniq_column is not None:
            self.uniq_header_column = uniq_column
        self.column_headers = []
        found = False
        for row in range(1, 40):
            cell = self.sheet["A{}".format(row)]
            if not cell.value:
                continue
            # log.debug("{}: {}, {}".format(row, cell.value, self.uniq_header_column))
            if self.uniq_header_column is None:
                found = True
            else:
                _values = [
                    self.sheet["{}{}".format(get_column_letter(col), row)] for col in range(1, 1000)
                ]
                values = []
                for col, x in enumerate(_values, start=1):
                    if x:
                        try:
                            values.append(x.value)
                        except TypeError:
                            log.warning(
                                "Unable to figure out value for cell: {}{} on {}".format(
                                    get_column_letter(col), row, self.filename
                                )
                            )
                            values.append(None)
                for value in values:
                    if value is None:
                        continue
                    if isinstance(value, str):
                        value = value.replace("\x00", "")
                        value = value.replace("\xa0", " ")

                    if equivalence_map and "{}".format(value).lower() in [
                        "{}".format(x).lower() for x in equivalence_map.keys()
                    ]:
                        eq_keys = [x.lower() for x in equivalence_map.keys()]
                        eq_values = [x for x in equivalence_map.values()]
                        idx = eq_keys.index(value.lower())
                        value = eq_values[idx]

                    if (
                        "{}".format(value).strip().lower()
                        == "{}".format(self.uniq_header_column).strip().lower()
                    ):
                        # log.debug("Found {} on row {}".format(self.uniq_header_column, row))
                        found = True
                        break
            if found:
                skipped = 0
                for col in range(1, 10000):
                    cell = self.sheet["{}{}".format(get_column_letter(col), row)]
                    if cell.value:
                        if skipped == 1:
                            log.warning(
                                "The column prior to %s was skipped..", get_column_letter(col)
                            )
                            skipped = 0
                        try:
                            value = "{}".format(cell.value).strip()
                        except UnicodeEncodeError:
                            value = cell.value
                        if equivalence_map and value.lower() in [
                            x.lower() for x in equivalence_map.keys()
                        ]:
                            eq_keys = [x.lower() for x in equivalence_map.keys()]
                            eq_values = [x for x in equivalence_map.values()]
                            idx = eq_keys.index(value.lower())
                            value = eq_values[idx]
                        self.column_headers.append(value)
                    else:
                        self.column_headers.append(None)
                        skipped += 1
                break

        # This chops the end None values off..
        found, final = False, []
        self.column_headers.reverse()
        for item in self.column_headers:
            if item is None and not found:
                continue
            found = True
            final.append(item)
        final.reverse()
        self.column_headers = final

        lower_headers = ["{}".format(x).lower() for x in self.column_headers if x]
        if len(set(lower_headers)) != len(lower_headers):
            bad = [x for x, y in collections.Counter(lower_headers).items() if y > 1]
            raise IndexError("You have multiple columns named {}".format(",".join(bad)))

        self.row = row + 1
        return self.column_headers

    def get_results_dictionary_list(
        self, set_lower=False, add_row_number=True, equivalence_map=None, as_string=False
    ):
        """This gets the results

        If an equivalence_map map is used it should be in the form of
            equivalence_map = ((foo, target) (foo2, target))
        This will map any found foo header to target dictname
        """
        # TODO: If there is no data besides the headers
        if equivalence_map:
            equivalence_map = dict(equivalence_map)
        results = []
        for row in range(self.row, 10000):
            if not self.sheet["A{}".format(row)].value and not self.sheet["B{}".format(row)].value:
                break
            record = {}
            for col_label in self.column_headers:
                col = get_column_letter(self.column_headers.index(col_label) + 1)
                if set_lower:
                    col_label = "{}".format(col_label).lower()
                if equivalence_map and col_label in equivalence_map.keys():
                    col_label = equivalence_map[col_label]

                # # A way to handle https://bitbucket.org/openpyxl/openpyxl/issue/343
                cell = self.sheet["{}{}".format(col, row)]
                try:
                    value = cell.value
                except TypeError:
                    if cell.is_date():
                        value = None
                    else:
                        raise
                else:
                    # We want to keep values that shouldn't be strings.
                    # But in the case we receive a unicode strings,
                    # we want to make sure it isn't littered with control
                    # characters that could affect queries down the line.
                    if isinstance(value, str):
                        value = value.translate(CTRL_CHARS)

                # log.debug("Column: {:<2} Row: {:<2} Label: {:<20} Value: {}".format(col, row,  col_label[0:20],value))
                if as_string:
                    if value:
                        value = "{}".format(value)
                    # elif value == 'None': value = None
                try:
                    value = value.replace("\xa0", " ")
                except (AttributeError, UnicodeDecodeError, TypeError):
                    pass
                try:
                    value = value.strip()
                except AttributeError:
                    pass
                record[col_label] = value
            if add_row_number:
                record["row_number"] = row
            results.append(record)
        return results
