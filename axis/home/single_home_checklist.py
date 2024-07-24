"""single_home_checklist.py: Django axis.home"""

import datetime
import logging
import os
import re
import shutil
import tempfile
from collections import defaultdict, OrderedDict
from itertools import chain, zip_longest

from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import FieldDoesNotExist
from localflavor.us.us_states import US_STATES
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.formula.tokenizer import TokenizerError
from openpyxl.styles import Font, PatternFill, fills, Border, Side, Color, Alignment, Protection
from openpyxl.utils import range_boundaries, get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from simple_history.utils import bulk_create_with_history

from axis.checklist.collection.collectors import get_user_role_for_homestatus
from axis.checklist.collection.excel import ExcelChecklistCollector
from axis.checklist.models import Question
from axis.checklist.strings import UNABLE_TO_CERTIFY_INVALID_USER, CERTIFY_ERROR, CERTIFIED_HOME
from axis.checklist.xls_checklist import XLSChecklist
from axis.eep_program.models import EEPProgram
from axis.qa.models import QARequirement, QAStatus, QANote
from axis.relationship.utils import create_or_update_spanning_relationships
from .models import Home, EEPProgramHomeStatus
from .tasks import update_home_stats, update_home_states, certify_single_home

TEMPLATE = os.path.abspath(
    os.path.dirname(os.path.realpath(__file__))
    + "/../checklist/static/templates/Single_Home_Upload.xlsx"
)

TEMPLATE_QA_ETO_2020 = os.path.abspath(
    os.path.dirname(os.path.realpath(__file__))
    + "/../checklist/static/templates/Single_Home_Upload_FieldQA_ETO_2020.xlsx"
)
AXIS_LOGO = os.path.abspath(os.path.dirname(__file__) + "/../core/static/images/Logo_Only_128.png")
ETO_LOGO = os.path.abspath(
    os.path.dirname(__file__) + "/../customer_eto/static/images/ET_EPS_Logo_White_Horz.png"
)

MEDIUM_GREY = "6D6D6D"
REQUIRED_CELL_COLOR = "A5B8E1"
OPTIONAL_CELL_COLOR = "FBFCFE"
NON_STORED_CELL_COLOR = "FFCCCC"
GREY_FONT_COLOR = "999999"
DARK_GREY_FONT_COLOR = "4C4C4C"
WHITE_FONT_COLOR = "FFFFFF"

__author__ = "Steven Klass"
__date__ = "11/8/17 15:35"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

log = logging.getLogger(__name__)
User = get_user_model()


class ChecklistWriter(object):
    def get_cell_from_label(self, sheet_name, label, row=None, raise_on_error=True):
        assert hasattr(self, "workbook"), "Need a workbook"
        try:
            defined_name = self.workbook.defined_names[label]
        except KeyError:
            if raise_on_error:
                raise
            return None
        try:
            if row:
                sn, ca = next(
                    ((x, y) for x, y in defined_name.destinations if x == sheet_name and row == row)
                )
            else:
                sn, ca = next(((x, y) for x, y in defined_name.destinations if x == sheet_name))
        except (StopIteration, TokenizerError):
            if raise_on_error:
                raise StopIteration("Unable to find %s in sheet %s" % (label, sheet_name))
            return None
        try:
            return self.workbook[sheet_name][ca.replace("$", "")]
        except IndexError as exc:
            if raise_on_error:
                raise exc
            return None

    def get_value_from_label(self, sheet_name, label, row=None, raise_on_error=True):
        cell_object = self.get_cell_from_label(sheet_name, label, row, raise_on_error)
        if cell_object:
            return cell_object.value

    def get_comment_from_label(self, sheet_name, label, row=None):
        cell_object = self.get_cell_from_label(sheet_name, label, row)
        if cell_object.comment:
            return cell_object.comment.text

    @property
    def basic_style(self):
        return {"font": Font(name="Calibri", size=12), "alignment": Alignment(horizontal="left")}

    @property
    def basic_white_style(self):
        color = Color(rgb="FFFFFF")
        if hasattr(self, "user") and self.user.is_superuser:
            color = Color(rgb="D0D0D0")
        return {"font": Font(name="Calibri", size=10, color=color)}

    @property
    def basic_optional_style(self):
        kwargs = self.basic_style
        kwargs["fill"] = PatternFill(fill_type=fills.FILL_SOLID, start_color=OPTIONAL_CELL_COLOR)
        return kwargs

    @property
    def basic_non_stored_style(self):
        kwargs = self.basic_style
        kwargs["fill"] = PatternFill(fill_type=fills.FILL_SOLID, start_color=NON_STORED_CELL_COLOR)
        return kwargs

    def set_thin_bordered_style(self, color=None):
        if color:
            color = Color(rgb=color)

        return Border(
            bottom=Side(border_style="thin", color=color),
            top=Side(border_style="thin", color=color),
            left=Side(border_style="thin", color=color),
            right=Side(border_style="thin", color=color),
        )

    @property
    def thin_bordered_style(self):
        kwargs = self.basic_style
        kwargs["border"] = self.set_thin_bordered_style(color="FF000000")
        return kwargs

    @property
    def basic_required_style(self):
        kwargs = self.basic_style
        kwargs["fill"] = PatternFill(fill_type=fills.FILL_SOLID, start_color=REQUIRED_CELL_COLOR)
        return kwargs

    def merge_and_set(self, cell_range_str, value, sheet_obj, style=None, **styles):
        styles.setdefault("alignment", Alignment(horizontal="left", wrap_text=True))
        sheet_obj.merge_cells(cell_range_str)
        cell_obj = sheet_obj[cell_range_str.split(":")[0]]
        self.set_text(cell_obj, value, sheet_obj, style=style)
        self.set_merged_bordered_style(sheet_obj, cell_obj, **styles)

    def set_merged_bordered_style(self, sheet_obj, cell_obj, **attrs):
        def get_merge_range(cell_coordinate):
            for merge_range in sheet_obj.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merge_range).upper())
                for cell_list in sheet_obj.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    for cell_unit in cell_list:
                        if cell_unit.coordinate == cell_coordinate:
                            return merge_range

        cell_range = get_merge_range(cell_obj.coordinate)

        # Don't want to double this up.
        if hasattr(self, "handled_merged_cells"):
            if cell_range in self.handled_merged_cells:
                return cell_range
            self.handled_merged_cells.append(cell_range)

        min_col, min_row, max_col, max_row = range_boundaries(str(cell_range).upper())
        for cell_list in sheet_obj.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell_obj in cell_list:
                style = self.thin_bordered_style
                style.update(attrs)
                for k, v in style.items():
                    setattr(cell_obj, k, v)
        return cell_range

    def set_style(self, sheet_obj, cell_obj, **attrs):
        if cell_obj.coordinate in sheet_obj.merged_cells:
            return self.set_merged_bordered_style(sheet_obj, cell_obj, **attrs)
        else:
            for k, v in attrs.items():
                setattr(cell_obj, k, v)
            return cell_obj.coordinate

    def set_border_range(self, cell_range, sheet_obj):
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_range).upper())
        rows = list(
            sheet_obj.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        )
        for row_idx, row in enumerate(rows, start=1):
            style = {}
            if row_idx == 1:
                style["top"] = Side(border_style="thick", color=Color(rgb="FF000000"))
            if row_idx == len(rows):
                style["bottom"] = Side(border_style="thick", color=Color(rgb="FF000000"))
            for col_idx, cell in enumerate(row, start=1):
                style = {}
                if row_idx == 1:
                    style["top"] = Side(border_style="thick", color=Color(rgb="FF000000"))
                elif cell.border.top:
                    style["top"] = cell.border.top

                if row_idx == len(rows):
                    style["bottom"] = Side(border_style="thick", color=Color(rgb="FF000000"))
                elif cell.border.bottom:
                    style["bottom"] = cell.border.bottom

                if col_idx == 1:
                    style["left"] = Side(border_style="thick", color=Color(rgb="FF000000"))
                elif cell.border.left:
                    style["left"] = cell.border.left

                if col_idx == len(row):
                    style["right"] = Side(border_style="thick", color=Color(rgb="FF000000"))
                elif cell.border.right:
                    style["right"] = cell.border.right

                if style:
                    cell.border = Border(**style)

    def lock_cells(self, sheet_obj):
        if not hasattr(self, "editable_cells"):
            log.error("No editable cells")

        sheet_obj.protection.sheet = True
        for cell in self.editable_cells:
            cell_obj = sheet_obj[cell]
            cell_obj.protection = Protection(locked=False)

    def add_validation(self, sheet_obj, cell_obj, data_validation):
        sheet_obj.add_data_validation(data_validation)
        if cell_obj.coordinate not in sheet_obj.merged_cells:
            data_validation.add(cell_obj)
        else:

            def get_merge_range(cell_coordinate):
                for merge_range in sheet_obj.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merge_range).upper())
                    for cell_list in sheet_obj.iter_rows(
                        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                    ):
                        for cell_unit in cell_list:
                            if cell_unit.coordinate == cell_coordinate:
                                return merge_range

            cell_range = get_merge_range(cell_obj.coordinate)
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range).upper())
            for cell_list in sheet_obj.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                for cell_obj in cell_list:
                    data_validation.add(cell_obj)

    def is_valid_choices(self, choices):
        formula = '"{choices}"'.format(choices=",".join(choices))
        for choice in choices:
            if not all(31 < ord(c) < 128 for c in choice):
                return False
                # bad_actors = [c for c in choice if ord(c) >= 128]
                # raise IndexError("You need to provide a replace the characters %r with something in ASCII %s - Need to use reference list" % bad_actors, choice)
            if "," in choice or '"' in choice:
                return False
                # raise IndexError("You cannot use a comma, or double-quote in any choice option %r - Need to use reference list" % choice)
        if len(formula) > 257:  # Tested the shit out this..
            # See https://support.office.com/en-us/article/excel-specifications-and-limits-1672b34d-7043-467e-8e27-269d656771c3
            return False
            # log.warning("You need to provide a reference sheet and column when you have more than 257 chars you have %d %s", len(formula), choices)
            # raise IndexError("You need to provide a reference sheet and column when you have more than 257 chars")
        return True

    def find_reference_sheet_list(self, reference_sheet_name, choices):
        ref_sheet_obj = self.workbook[reference_sheet_name]

        found = False
        for col in range(27, 1000):
            for row in range(1, 2000):
                found = False
                ref_cell_start = "{letter}{row}".format(letter=get_column_letter(col), row=row)
                value = ref_sheet_obj[ref_cell_start].value
                # print("Looking at %s %s" % (ref_cell_start, value))
                if not value:
                    # print("Good we are in.. %s " % ref_cell_start)
                    ref_start_column, ref_start_row = get_column_letter(col), row
                    for idx, option in enumerate(choices, start=row):
                        ref_cell_end = "{letter}{row}".format(
                            letter=get_column_letter(col), row=idx
                        )
                        found = ref_sheet_obj[ref_cell_end].value is None
                        # print(" Sub review..  %s %s" % (ref_cell_end, ref_sheet_obj[ref_cell_end].value))
                        if not found:
                            break
                if found:
                    break

            if found:
                # log.debug("Escaping we have what we need %s:%s for %s", ref_cell_start, ref_cell_end, "|".join(choices))
                return ref_start_column, ref_start_row

        return None, None

    def set_valid_choices(
        self,
        sheet_obj,
        cell_obj,
        choices,
        formula=None,
        reference_sheet_name=None,
        reference_column_letter=None,
        show_dropdown=True,
    ):
        """Set the validation"""
        assert isinstance(choices, (list, tuple))

        if formula:
            if len(formula) >= 255:
                raise IndexError(
                    "You cannot have a formula > 255 chars (%d) - you need to use a reference cell"
                    % len(formula)
                )
            if len(formula) >= 225:
                log.warning(
                    "You're formula is very close to exceeding 255 (%d) chars consider using a reference cell"
                    % len(formula)
                )
        elif reference_sheet_name and reference_column_letter:
            ref_sheet_obj = self.workbook[reference_sheet_name]
            for row, choice in enumerate(choices, start=2):
                ref_cell_obj = ref_sheet_obj["{}{}".format(reference_column_letter, row)]
                ref_cell_obj.value = choice
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet=reference_sheet_name, column=reference_column_letter, row=row
            )
        else:
            if self.is_valid_choices(choices):
                formula = '"{choices}"'.format(choices=",".join(choices))
            else:
                column, start_row = self.find_reference_sheet_list("Geography", choices)
                if column and start_row:
                    ref_sheet_obj = self.workbook["Geography"]
                    for row, choice in enumerate(choices, start=start_row):
                        ref_cell_obj = ref_sheet_obj["{}{}".format(column, row)]
                        ref_cell_obj.value = choice
                formula = "={sheet}!${column}${start_row}:${column}${row}".format(
                    sheet="Geography", column=column, start_row=start_row, row=row
                )

        log.debug("Formula for %s is %s", cell_obj.coordinate, formula)

        data_validation = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=not show_dropdown,
            error="Your entry is not in the list",
            errorTitle="Invalid Entry",
            prompt="Please select from the list",
            promptTitle="List Selection",
        )
        self.add_validation(sheet_obj, cell_obj, data_validation)

    def get_last_data_row_number(self, sheet_obj):
        return [x for x in sheet_obj.rows][-1][0].row

    def set_text(self, cell_obj, value, sheet_obj, style=None):
        cell_obj.value = value
        if style == "header":
            self.set_style(
                sheet_obj,
                cell_obj,
                **{
                    "font": Font(
                        name="Calibri", size=14, italic=True, color=Color(rgb=WHITE_FONT_COLOR)
                    ),
                    "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color=GREY_FONT_COLOR),
                },
            )


class HomeChecklistData(ChecklistWriter):
    def set_label_non_stored_value(self, sheet_obj, cell_label, value, row=None):
        cell_obj = self.get_cell_from_label(sheet_obj.title, cell_label, row)
        cell_obj.value = value
        self.set_style(sheet_obj, cell_obj, **self.basic_non_stored_style)
        if hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)
        return cell_obj

    def set_label_optional_value(self, sheet_obj, cell_label, value, row=None):
        cell_obj = self.get_cell_from_label(sheet_obj.title, cell_label, row)
        cell_obj.value = value
        self.set_style(sheet_obj, cell_obj, **self.basic_optional_style)
        if hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)
        return cell_obj

    def set_label_required_value(self, sheet_obj, cell_label, value, row=None):
        cell_obj = self.get_cell_from_label(sheet_obj.title, cell_label, row)
        cell_obj.value = value
        self.set_style(sheet_obj, cell_obj, **self.basic_required_style)
        if hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)
        return cell_obj

    def set_label_white_value(self, sheet_obj, cell_label, value, row=None):
        cell_obj = self.get_cell_from_label(sheet_obj.title, cell_label, row)
        cell_obj.value = value
        self.set_style(sheet_obj, cell_obj, **self.basic_white_style)
        return cell_obj

    def set_label_default_value(self, sheet_obj, cell_label, value, row=None, edit_flag=False):
        cell_obj = self.get_cell_from_label(sheet_obj.title, cell_label, row)
        cell_obj.value = value
        self.set_style(sheet_obj, cell_obj, **self.basic_style)
        if edit_flag and hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)
        return cell_obj

    def set_home__overwrite(self, sheet_obj, cell_label, value="No", row=None):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        self.set_valid_choices(sheet_obj, cell_obj, choices=["No", "New"], formula=None)

    def set_home_status_overwrite(self, sheet_obj, cell_label, value="No", row=None):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        self.set_valid_choices(sheet_obj, cell_obj, choices=["No", "Yes", "New"], formula=None)

    def set_answer__overwrite(self, sheet_obj, cell_label, value="No", row=None):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        self.set_valid_choices(sheet_obj, cell_obj, choices=["No", "Yes"], formula=None)

    def set_multi_family(self, sheet_obj, cell_label, value, row=None):
        value = "Yes" if value is True else "No"
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        self.set_valid_choices(sheet_obj, cell_obj, choices=["No", "Yes"], formula=None)

    def set_company_overwrite(self, sheet_obj, cell_label, value="No", row=None):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        self.set_valid_choices(sheet_obj, cell_obj, choices=["No", "Yes"], formula=None)

    def set_county(self, sheet_obj, cell_label, value, row=None, required=False):
        method = self.set_label_optional_value
        if required:
            method = self.set_label_required_value
        cell_obj = method(sheet_obj, cell_label, value, row)
        formula = "=OFFSET(Geography!$C$2,MATCH(home__state__name,Geography!$A$2:$A$3146,0)-1,0,COUNTIF(Geography!$A$2:$A$3146,home__state__name),1)"
        self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_state(self, sheet_obj, cell_label, value, row=None):
        state_dict = dict(US_STATES)
        value = "{}".format(state_dict.get(value)) if value else None
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)

        if hasattr(self, "user") and not self.user.is_superuser:
            from axis.geographic.models import County

            _states = list(County.objects.filter_by_user(self.user).values_list("state", flat=True))
            choices = filter(None, [state_dict.get(x) for x in _states] + [value])
        else:
            choices = state_dict.values()

        choices = sorted(["{}".format(x) for x in list(set(choices))])
        if len(choices) > 0:
            self.set_valid_choices(
                sheet_obj,
                cell_obj,
                choices,
                reference_sheet_name="Geography",
                reference_column_letter="E",
            )

    def set_zipcode(self, sheet_obj, cell_label, value, row):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        cell_obj.value = value
        data_validation = DataValidation(
            type="custom",
            operator="equal",
            formula1="=AND(ISNUMBER(home__zipcode),LEN(home__zipcode)=5)",
            errorStyle="stop",
            errorTitle="ZIPCode",
            error="Please enter a 5 digit ZIPCode",
        )
        sheet_obj.add_data_validation(data_validation)
        data_validation.add(cell_obj)
        cell_obj.number_format = "#####"
        return cell_obj

    def set_subdivision(self, sheet_obj, cell_label, value, row):
        cell_obj = self.set_label_optional_value(sheet_obj, cell_label, value, row)

        from axis.subdivision.models import Subdivision

        obj_lists = Subdivision.objects.all()
        if hasattr(self, "user"):
            obj_lists = Subdivision.objects.filter_by_user(self.user)
        obj_lists = list(obj_lists.values_list("name", flat=True).order_by("name"))

        reference_sheet_obj = self.workbook["Geography"]
        subdivision_column = "J"

        for row, (name) in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(subdivision_column, row)]
            ref_cell_obj.value = name

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=subdivision_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_community(self, sheet_obj, cell_label, value, row):
        cell_obj = self.set_label_optional_value(sheet_obj, cell_label, value, row)

        reference_sheet_obj = self.workbook["Geography"]
        community_column = "G"

        from axis.community.models import Community

        obj_lists = Community.objects.all()
        if hasattr(self, "user"):
            obj_lists = Community.objects.filter_by_user(self.user)
        obj_lists = list(obj_lists.values_list("name", flat=True).order_by("name"))

        for row, (name) in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(community_column, row)]
            ref_cell_obj.value = name

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=community_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_program(self, sheet_obj, cell_label, value, row, eep_program_id=None):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)

        reference_sheet_obj = self.workbook["Geography"]
        target_column = "L"

        from axis.eep_program.models import EEPProgram

        obj_lists = EEPProgram.objects.all()
        if hasattr(self, "user"):
            obj_lists = EEPProgram.objects.filter_by_user(self.user, visible_for_use=True)
        if eep_program_id:
            obj_lists = obj_lists.filter(id=eep_program_id)

        obj_lists = list(obj_lists.values_list("name", flat=True).order_by("name"))

        for row, (name) in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(target_column, row)]
            ref_cell_obj.value = name

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=target_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_company_type(
        self, sheet_obj, cell_label, value, row, target_column, company_type, utility_type=None
    ):
        value = value.name if value else None
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)

        reference_sheet_obj = self.workbook["Geography"]

        if isinstance(company_type, str):
            company_type = [company_type]

        query_args = {"company_type__in": company_type}
        if utility_type == "gas":
            query_args["utilityorganization__gas_provider"] = True
        elif utility_type == "electric":
            query_args["utilityorganization__electricity_provider"] = True

        from axis.company.models import Company

        obj_lists = Company.objects.filter(**query_args)
        if hasattr(self, "user"):
            obj_lists = Company.objects.filter_by_user(self.user, include_self=True).filter(
                **query_args
            )
            if self.user.company.company_type in company_type:
                obj_lists = Company.objects.filter(id=self.user.company_id)

        obj_lists = sorted(set(obj_lists.values_list("name", flat=True)), key=str.casefold)

        for row, (name) in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(target_column, row)]
            ref_cell_obj.value = name

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=target_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_home_status_company(self, sheet_obj, cell_label, value, row, eep_program):
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)

        reference_sheet_obj = self.workbook["Geography"]
        target_column = "N"

        from axis.company.models import Company

        company_types = ["qa", "provider"] if eep_program.is_qa_program else ["rater", "provider"]
        obj_lists = Company.objects.filter(company_type__in=company_types)
        if hasattr(self, "user"):
            obj_lists = Company.objects.filter_by_user(self.user).filter(
                company_type__in=company_types
            )
            if self.user.company.company_type == "rater":
                obj_lists = Company.objects.filter(id=self.user.company_id)
            if eep_program.is_qa_program:
                obj_lists = Company.objects.filter(
                    id__in=[self.user.company_id, eep_program.owner.id]
                )

        obj_lists = sorted(set(obj_lists.values_list("name", flat=True)), key=str.casefold)

        for row, (name) in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(target_column, row)]
            ref_cell_obj.value = name

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=target_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_home_status_rater(self, sheet_obj, cell_label, value, row, company_id):
        cell_obj = self.set_label_optional_value(sheet_obj, cell_label, value, row)

        reference_sheet_obj = self.workbook["Geography"]
        target_column = "P"

        obj_lists = User.objects.filter(company_id=company_id)
        obj_lists = sorted(["{}, {}".format(x.last_name, x.first_name) for x in obj_lists])

        for row, (name) in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(target_column, row)]
            ref_cell_obj.value = name

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=target_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_floorplan(self, sheet_obj, cell_label, value, row):
        def set_floorplan_name(name=None, type=None, pk=None, floorplan=None):
            if floorplan:
                name = floorplan.name
                type = floorplan.type
                pk = floorplan.pk
            if name:
                label = name
                if type:
                    label += " ({})".format(type.capitalize())
                if pk:
                    label += " [{}]".format(pk)
                return label

        cell_obj = self.set_label_optional_value(
            sheet_obj, cell_label, set_floorplan_name(floorplan=value), row
        )

        reference_sheet_obj = self.workbook["Geography"]
        target_column = "Q"

        from axis.floorplan.models import Floorplan

        obj_lists = Floorplan.objects.all()
        if hasattr(self, "user"):
            obj_lists = Floorplan.objects.filter_by_user(self.user)
        obj_lists = obj_lists.exclude(name__isnull=True)
        obj_lists = list(obj_lists.values_list("name", "type", "pk").order_by("-created_date"))[
            :200
        ]
        obj_lists.sort(key=lambda x: x[0])

        for row, data in enumerate(obj_lists, start=2):
            ref_cell_obj = reference_sheet_obj["{}{}".format(target_column, row)]
            ref_cell_obj.value = set_floorplan_name(*data)

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=target_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_simulation(self, sheet_obj, cell_label, value, row, company=None):
        from axis.remrate_data.strings import EXPORT_TYPES
        from axis.remrate_data.models import Simulation, DESIGN_MODELS

        def set_simulation_name(
            rating_number=None,
            building__filename=None,
            export_type=None,
            version=None,
            simulation_date=None,
            pk=None,
            simulation=None,
        ):
            if simulation:
                rating_number = simulation.rating_number
                version = simulation.version
                building__filename = simulation.building.filename
                export_type = simulation.export_type
                simulation_date = simulation.simulation_date
                pk = simulation.id

            if not pk:
                return None

            if rating_number:
                rating_number = rating_number
            elif building__filename:
                rating_number = building__filename
            else:
                rating_number = "Unknown"

            version = " (v{})".format(version) if version else ""
            exp_type = ""
            if export_type and export_type != 1:
                exp_type = " [{}]".format(dict(EXPORT_TYPES).get(export_type))
            simulation_date = simulation_date.strftime("%m/%d/%Y") if simulation_date else ""
            return "{} {}{}{} [{}]".format(rating_number, simulation_date, version, exp_type, pk)

        cell_obj = self.set_label_optional_value(
            sheet_obj, cell_label, set_simulation_name(simulation=value), row
        )

        reference_sheet_obj = self.workbook["Geography"]
        target_column = "R"

        query_args = {"export_type__in": [1] + DESIGN_MODELS}
        if company:
            query_args["company"] = company

        obj_lists = Simulation.objects.filter(**query_args)

        if hasattr(self, "user"):
            obj_lists = Simulation.objects.filter_by_user(self.user).filter(**query_args)

        vals = (
            "rating_number",
            "building__filename",
            "export_type",
            "version",
            "simulation_date",
            "id",
        )
        obj_lists = list(obj_lists.values_list(*vals).order_by("-simulation_date"))

        for row, data in enumerate(obj_lists, start=2):
            _value = set_simulation_name(*data)
            ref_cell_obj = reference_sheet_obj["{}{}".format(target_column, row)]
            ref_cell_obj.value = _value
            if row > 102:
                log.warning("Capping at 100")
                break

        if len(obj_lists) > 0:
            formula = "={sheet}!${column}$2:${column}${row}".format(
                sheet="Geography", column=target_column, row=row
            )
            self.set_valid_choices(sheet_obj, cell_obj, choices=[], formula=formula)

    def set_certification_date(self, sheet_obj, cell_label, value, row):
        value = value.strftime("%m/%d/%Y") if value else None
        cell_obj = self.set_label_optional_value(sheet_obj, cell_label, value, row)
        data_validation = DataValidation(type="date")
        sheet_obj.add_data_validation(data_validation)
        data_validation.add(cell_obj)
        cell_obj.number_format = "MM/DD/YYYY"

    @property
    def qa_state_choices(self):
        return OrderedDict(
            [
                ("complete", "Complete — Pass"),
                ("correction_required", "Correction Required"),
            ]
        )

    def set_qa_state(self, sheet_obj, cell_label, value, row):
        choices_lookup = dict(map(reversed, self.qa_state_choices.items()))
        cell_obj = self.set_label_required_value(sheet_obj, cell_label, value, row)
        self.set_valid_choices(sheet_obj, cell_obj, choices=choices_lookup.values(), formula=None)


class SingleHomeChecklist(XLSChecklist, HomeChecklistData):
    def __init__(self, *args, **kwargs):
        kwargs["creator"] = kwargs.get("creator", "Axis")
        kwargs["title"] = kwargs.get("title", "Single Home Checklist")
        kwargs["subject"] = kwargs.get("subject", "Single Home Checklist")
        kwargs["description"] = kwargs.get("description", "Single Home Checklist")
        kwargs["sheet_name"] = kwargs.get("sheet_name", "Home")
        kwargs["overwrite_old_answers"] = kwargs.get("overwrite_old_answers", False)
        kwargs["template"] = kwargs.get("template")
        self.overwrite_old_answers = kwargs["overwrite_old_answers"]
        self.handled_merged_cells = []
        self.data = []
        self.input_issues = []
        self.collector = None
        super(SingleHomeChecklist, self).__init__(*args, **kwargs)

    def read(self, filename=None):
        assert not len(self.data), "You cannot combine a read with existing data"

        if filename is None:
            filename = self.filename

        if not os.path.exists(filename):
            raise IOError("File does not exist - %s" % filename)

        self.workbook = load_workbook(filename=filename, data_only=True)

        home_ids = self.workbook.defined_names["home_id"]
        sheets = [sheet for sheet, cell in home_ids.destinations]

        if not len(sheets):
            raise IOError("This is not the right template - Need cell with a home_id defined name")

        # This imagines that each row has a representative
        if len(sheets) != len(set(sheets)):
            raise IOError("Bulk Upload not supported in this.")

        for i, sheet_name in enumerate(sheets):
            try:
                self.log.set_context(row=None)
            except AttributeError:
                pass

            sheet_obj = self.workbook[sheet_name]
            data = self.read_sections(sheet_obj=sheet_obj, sheet_name=sheet_name, sheet_index=i + 1)
            self.data.append(data)
        return self.data

    def read_sections(self, sheet_obj, sheet_name, sheet_index):
        self.data_links = {}

        link_string = "<a target='_blank' href='{url}'>{object}</a>"

        try:
            self.log.set_context(row=sheet_index)
            self.log.set_flags(
                home_updated=False,
                home_created=False,
                program_added=False,
                home_already_certified=False,
                certification_date=False,
                home_certified=False,
                has_row_flags=True,
                row_label=sheet_name,
            )
        except AttributeError:
            pass

        home_data = self.read_home_values(sheet_obj=self.workbook[sheet_name])
        if home_data["subdivision_url"]:
            self.data_links["Subdivision"] = link_string.format(
                url=home_data["subdivision_url"], object=home_data["subdivision_string"]
            )
        if home_data["community_url"]:
            self.data_links["Community"] = link_string.format(
                url=home_data["community_url"], object=home_data["community_string"]
            )
        if home_data["home_url"]:
            self.data_links["Home"] = link_string.format(
                url=home_data["home_url"], object=home_data["home_string"]
            )
        try:
            self.log.set_flags(links=self.data_links)
        except AttributeError:
            pass

        home_status_data = self.read_home_stat_values(
            sheet_obj=self.workbook[sheet_name], **home_data
        )
        try:
            self.log.set_flags(
                home_already_certified=home_status_data["is_certified"],
                certification_date=home_status_data["certification_date"],
            )
        except AttributeError:
            pass

        ans_overwrite = self.get_value_from_label(sheet_name=sheet_name, label="answers__overwrite")
        company_data = self.read_company_values(
            sheet_obj=self.workbook[sheet_name], **home_status_data
        )

        try:
            annotations_overwrite = self.get_value_from_label(
                sheet_name=sheet_name, label="annotations__overwrite"
            )
            annotations = self.read_annotation_values(
                sheet_obj=self.workbook[sheet_name], **home_status_data
            )
        except KeyError:
            annotations_overwrite = False
            annotations = []

        questions = self.read_question_values(
            sheet_obj=self.workbook[sheet_name], **home_status_data
        )
        field_data = self.read_field_qa_values(
            sheet_obj=self.workbook[sheet_name], **home_status_data
        )

        return {
            "home": home_data,
            "home_status": home_status_data,
            "associations": company_data,
            "questions": questions,
            "answer_overwrite": ans_overwrite,
            "annotations": annotations,
            "annotations_overwrite": annotations_overwrite,
            "field_qa": field_data,
        }

    def validate_home_data(
        self,
        create,
        overwrite,
        home,
        lot_number,
        street_line1,
        street_line2,
        city,
        county,
        state,
        zipcode,
        subdivision,
        is_multi_family,
        builder,
        sheet_idx,
        **data,
    ):
        link_string = "<a target='_blank' href='{url}'>{object}</a>"

        if overwrite == "New":
            self.log.debug("%s New home", "Create" if create else "Validate")

            if builder is None:
                self.log.error("New home without a builder is not allowed")
                return

            home, _cr = Home.objects.verify_and_create_for_user(
                lot_number=lot_number,
                street_line1=street_line1,
                street_line2=street_line2,
                city=city,
                county=county,
                state=state,
                zipcode=zipcode,
                subdivision=subdivision,
                builder=builder,
                is_multi_family=is_multi_family,
                create=create,
                user=self.user,
                log=self.log,
            )

            if create:
                if _cr:
                    self.log.set_flags(home_created=True)
                self.data_links["Home"] = link_string.format(
                    url=home.get_absolute_url(), object="{} [{}]".format(home, home.id)
                )
                self.log.set_flags(links=self.data_links)

                self.data[sheet_idx]["home"]["home"] = home
                self.data[sheet_idx]["home"]["home_id"] = home.id

        # TODO Support Yes?
        # elif overwrite == "Yes":
        #     self.log.debug("{} existing home {}".format("Update" if create else "Validate", home.id))
        #     changed_keys = []
        #     for k in ['lot_number', 'street_line1', 'street_line2', 'city', 'county', 'state', 'zipcode', 'is_multi_family', ]
        else:
            self.log.debug("Using existing Axis Home ID: %s", home.id)

    def uploaded_as_qa(self, eep_program, sheet_idx):
        if self.data[sheet_idx].get("upload_as_qa") is not None:
            return self.data[sheet_idx]["upload_as_qa"]

        if eep_program.is_qa_program:
            # This appears to be field QA
            if not EEPProgram.objects.filter_by_user(self.user).filter(id=eep_program.id).count():
                self.log.error(
                    "This appears to be a QA Field Checklist however you do not access to this"
                )
                raise ValueError(
                    "This appears to be a QA Field Checklist however you do not access to this"
                )
            qa_requirement = (
                QARequirement.objects.filter_by_user(self.user)
                .filter(eep_program=eep_program.get_rater_program(), type="field")
                .first()
            )
            if not qa_requirement:
                self.log.error(
                    "This appears to be a QA Field Checklist however you do not have any requirements"
                )
                raise ValueError(
                    "This appears to be a QA Field Checklist however you do not have any requirements"
                )
            self.data[sheet_idx]["upload_as_qa"] = True
        else:
            self.data[sheet_idx]["upload_as_qa"] = False
        log.debug("Setting upload_as_qa %r", self.data[sheet_idx]["upload_as_qa"])
        return self.data[sheet_idx]["upload_as_qa"]

    def validate_home_status_data(
        self,
        create,
        overwrite,
        home_status,
        home,
        eep_program,
        company,
        floorplan,
        simulation,
        rater,
        certification_date,
        sheet_idx,
        **data,
    ):
        if self.uploaded_as_qa(eep_program, sheet_idx) is not None:
            log.debug(
                "Working with%s QA Program" % ""
                if self.uploaded_as_qa(eep_program, sheet_idx)
                else "out"
            )

        if overwrite == "New" or home_status is None:
            self.log.debug(
                "{}{} Program{}".format(
                    "Create" if create else "Validate",
                    "ing" if home_status is None else " New",
                    " for home {}".format(home) if home else "",
                )
            )

            created = True
            stats = EEPProgramHomeStatus.objects.filter(
                home=home, company=company, eep_program=eep_program
            )
            home_status = stats.first()
            if home_status:
                self.log.info(
                    "Program %(program)s (%(home_status)s) already exists on home for %(company)s.",
                    dict(
                        program=eep_program,
                        home_status=home_status.id if home_status else "--",
                        company=company,
                    ),
                )
                created = False

            home_status = EEPProgramHomeStatus.objects.verify_and_create_for_user(
                company=company,
                eep_program=eep_program,
                floorplan=floorplan,
                remrate=simulation,
                rater_of_record=rater,
                home=home,
                create=create,
                ignore_missing_floorplan=True,
                user=self.user,
                log=self.log,
            )

            if home_status:
                if self.uploaded_as_qa(eep_program, sheet_idx):
                    if eep_program.is_qa_program:
                        log.debug("Setting uploaded QA HS %r", home_status)
                        self.data[sheet_idx]["home_status"]["home_status"] = home_status
                        self.data[sheet_idx]["home_status"]["home_status_id"] = home_status.id
                    else:
                        log.debug("Setting uploaded QA NON HS %r", home_status)
                        self.data[sheet_idx]["home_status"]["home_status_non_qa"] = home_status
                else:
                    log.debug("Setting uploaded Default HS %r", home_status)
                    self.data[sheet_idx]["home_status"]["home_status"] = home_status
                    self.data[sheet_idx]["home_status"]["home_status_id"] = home_status.id

            # Set flags and extra references for qa scenarios
            if create:
                if created:
                    self.log.set_flags(program_added=True)
                    if home_status and home_status.certification_date:
                        self.log.set_flags(
                            home_certified=True, certification_date=home_status.certification_date
                        )

        elif overwrite == "Yes":
            self.log.debug(
                "{} New Program home {}".format("Updating" if create else "Validate", home)
            )

            save_required = 0
            for key, value in [
                ("company", company),
                ("eep_program", eep_program),
                ("floorplan", floorplan),
                ("rater_of_record", rater),
            ]:
                if getattr(home_status, key) != value:
                    self.log.debug("Updating %s to new value %s", key, value)
                    save_required += 1
                    if create:
                        setattr(home_status, key, value)
            if save_required and create:
                self.log.info(
                    "Updating %d fields one the program for home %s",
                    save_required,
                    home_status.home,
                )
                home_status.save()
        else:
            self.log.debug("Using existing Axis Program ID: %s - %r", home_status.id, home_status)
            self.data[sheet_idx]["home_status"]["home_status"] = home_status
            self.data[sheet_idx]["home_status"]["home_status_id"] = home_status.id

        # ALWAYS ensure the home_status has its collection_request set, if one is available.
        if home_status:
            if eep_program.collection_request and home_status.collection_request is None:
                home_status.set_collection_from_program()

        if eep_program.is_qa_program:
            rater_program = eep_program.get_rater_program()
            rater_company = self.data[sheet_idx]["associations"]["rater"]
            self.log.debug(
                "Switching over to the rater program %s for %s", rater_program, rater_company
            )
            self.validate_home_status_data(
                create,
                overwrite,
                None,
                home,
                rater_program,
                rater_company,
                None,
                None,
                None,
                None,
                sheet_idx,
                **data,
            )

    def validate_field_qa(
        self, create, eep_program, home_status=None, qa_home_status=None, sheet_idx=None, **kwargs
    ):
        if not self.uploaded_as_qa(eep_program, sheet_idx):
            return

        qa_requirement = (
            QARequirement.objects.filter_by_user(self.user)
            .filter(eep_program=eep_program.get_rater_program(), type="field")
            .first()
        )

        qa_company = qa_requirement.qa_company
        self.data[sheet_idx]["home_status"]["qa_requirement"] = qa_requirement
        self.data[sheet_idx]["home_status"]["qa_company"] = qa_company

        new_state = kwargs["field_qa"].get("new_state")
        rater = kwargs.get("rater")

        choices_lookup = dict(map(reversed, self.qa_state_choices.items()))
        if new_state and new_state not in choices_lookup:
            self.log.error("QA State choice %r is not available", new_state)

        new_note = kwargs["field_qa"].get("new_note")

        if not create:
            return

        qastatus, _create = QAStatus.objects.get_or_create(
            **{
                "home_status": home_status,
                "owner": qa_company,
                "requirement": qa_requirement,
            }
        )

        self.log.info("%s QA for %s", "Creating" if _create else "Using Existing", home_status)

        if rater and qastatus.qa_designee != rater:
            self.log.info("Setting Verifier of record to %s" % rater)
            qastatus.qa_designee = rater

        self.data[sheet_idx]["home_status"]["qa_status"] = qastatus
        self.log.set_flags(qa_added=_create)

        if qastatus.state == "received":
            qastatus.make_transition("received_to_in_progress")

        if new_state:
            transitions = qastatus.get_state_info().possible_transitions
            transition_choices = [transition.get_name() for transition in transitions]
            transition_name = "{from_state}_to_{new_state}".format(
                from_state=qastatus.state, new_state=choices_lookup[new_state]
            )

            if qastatus.state != "received" and transition_name not in transition_choices:
                if qa_home_status.state != choices_lookup[new_state]:
                    self.log.error(
                        "'{new_state_display}' isn't a valid state change operation while "
                        "QA is '{from_state}'".format(
                            new_state_display=choices_lookup[new_state], from_state=qastatus.state
                        )
                    )
            if qastatus.state != choices_lookup[new_state]:
                self.log.info(
                    "Transition required from %(from_state)s for requested state: %(new_state)s",
                    dict(from_state=qastatus.state, new_state=choices_lookup[new_state]),
                )
                qastatus.make_transition(transition_name)

        if new_note:
            QANote.objects.create(
                **{
                    "user": self.user,
                    "qa_status": qastatus,
                    "content_type": ContentType.objects.get_for_model(EEPProgramHomeStatus),
                    "object_id": home_status.id,
                    "note": new_note,
                }
            )
            self.log.info("Added QA Note")

    def validate_relationship_data(self, create, home_status, sheet_idx, **data):
        from axis.company.models import Company

        if not home_status:
            return

        my_companies = Company.objects.filter_by_user(self.user).values_list("id", flat=True)

        for co_type in "builder", "rater", "provider", "electric_utility", "gas_utility", "hvac":
            if not data.get(co_type):
                continue

            all_rels = home_status.home.relationships.filter(
                company__company_type__in=co_type
            ).values_list("id", flat=True)

            kw = {
                "company_type": co_type if "utility" not in "co_type" else "utility",
                "id__in": list(set(my_companies).intersection(set(all_rels))),
            }
            if co_type == "electric_utility":
                kw["utilityorganization__electricity_provider"] = True
            elif co_type == "gas_utility":
                kw["utilityorganization__gas_provider"] = True

            if data["{}_overwrite".format(co_type)] in ["Yes"]:
                existing = Company.objects.filter(**kw).exclude(id=data["{}_id".format(co_type)])
                if existing.count():
                    if create:
                        self.log.info(
                            "Removing existing %s companies - %s",
                            co_type,
                            ", ".join(list(existing)),
                        )
                        # existing.delete()
                    else:
                        self.log.info("Existing %s companies found - %s" ", ".join(list(existing)))

                self.log.info("Adding %s %s", co_type, data[co_type])
                create_or_update_spanning_relationships(data[co_type], home_status)
            if data["{}_overwrite".format(co_type)] in ["New"]:
                existing = Company.objects.filter(**kw).filter(id=data["{}_id".format(co_type)])
                if not existing.count():
                    if create:
                        self.log.info("Adding %s %s", co_type, data[co_type])
                        create_or_update_spanning_relationships(data[co_type], home_status)
                    else:
                        self.log.info("%s %s is not attached to home", co_type, data[co_type])

    def validate_annotation_data(
        self, create, user, home_status, annotation_date, overwrite, annotation_data
    ):
        from axis.annotation.models import Annotation
        from axis.annotation.models import Type as AnnotationType

        if overwrite == "No":
            self.log.info("Not processing any existing annotations")

        existing_annotations = {}
        if home_status and home_status.pk:
            existing_annotations = {
                k.id: v for k, v in home_status.get_annotations_breakdown().items() if v
            }
        final = []
        for annotation in annotation_data:
            annotation_type = AnnotationType.objects.get(id=annotation["type_id"])
            if annotation_type.data_type == "multiple-choice":
                if annotation["content"] not in annotation_type.get_valid_multiplechoice_values():
                    log.error("Annotation {annotation['content']} is not valid choice skipping")
                    continue
            elif annotation_type.data_type == "int":
                if annotation["content"] != str(int(annotation["content"])):
                    log.error("Annotation {annotation['content']} is not valid integer skipping")
                    continue
            # Allow floats and open
            existing_annotation = None
            if annotation["type_id"] in existing_annotations:
                existing_annotation = existing_annotations[annotation["type_id"]]
                if overwrite == "No":
                    log.debug(
                        f"Skipping existing annotation - {annotation_type.slug} "
                        f"{annotation['content']} already exists"
                    )
                    continue

            if create:
                if existing_annotation:
                    existing_annotation.user_id = annotation["user_id"]
                    existing_annotation.content = annotation["content"]
                    existing_annotation.last_update = annotation_date
                    existing_annotation.save()
                    self.log.info(
                        f"Updated {annotation_type.slug} annotation for " f"{home_status.home}"
                    )
                    continue
                if home_status is None:
                    raise IOError("We should have a home status at this point")

                final.append(
                    Annotation(
                        type_id=annotation["type_id"],
                        content=annotation["content"],
                        content_type_id=annotation["content_type_id"],
                        object_id=home_status.pk,
                        user_id=user.pk,
                    )
                )

        if create and final:
            bulk_create_with_history(final, Annotation, default_user=user)
            self.log.info(f"Added {len(final)} Annotations for {home_status.home}")

    def validate_answers_data(
        self, create, home_status, company, answer_date, sheet_idx, overwrite, question_data
    ):
        from axis.checklist.utils import validate_answer, answer_questions_for_home

        if not self.use_legacy:
            if home_status:  # Could be None until a create=True run comes along
                self.collector.collection_request = home_status.collection_request

        if overwrite == "No":
            self.log.info("Not processing any existing answers")

        def _validate_legacy(question, data):
            try:
                return validate_answer(
                    question, log=self.log, answer=data["input"], comment=data.get("comment")
                )
            except Exception as e:
                log.error(
                    "Issue with question [%s] and answer %s", question, data["input"], extra=e
                )
            if data["input"] in [None, "", []]:
                question_href = '<a href="{}">{}[...]</a>'.format(
                    question.get_absolute_url(), "{}".format(question)[:50]
                )
                issue = "Question %s with answer '%s' will not process as no value was found"
                log.info(issue, question_href, data["input"])
            return None

        def _store_legacy():
            answer_questions_for_home(
                self.user,
                question_data,
                home_status,
                company,
                answer_date,
                log=self.log,
                ignore_missing=False,
                overwrite_old_answers=overwrite,
            )

        if not create:
            # Validate only
            inputs = defaultdict(list)
            for payload in question_data:
                if self.use_legacy:
                    payload["data"]["input"] = _validate_legacy(
                        payload["instrument"], payload["data"]
                    )
                else:
                    if self.collector.get_inputs(payload["instrument"]) and overwrite == "No":
                        log.debug(
                            "Skipping existing answer - %s",
                            "{}".format(payload["instrument"]).encode("utf-8"),
                        )
                        continue
                    payload = self.collector.clean_payload(payload)

                self.log.debug(
                    "Validated answer %r for %r", payload["data"]["input"], payload["instrument"]
                )
                inputs[payload["instrument"]].append(payload)

            self.data[sheet_idx]["question_data"] = inputs
            return

        # Store data
        if self.use_legacy:
            _store_legacy()
        else:
            if home_status:
                for info in chain(*question_data.values()):
                    info["measure"] = info.pop("instrument").measure_id
                    self.log.debug(
                        "Storing answer %r for %r",
                        info["data"]["input"],
                        "{}".format(info["measure"]).encode("utf-8"),
                    )
                    self.collector.store(**info)  # Already contains 'instrument' reference

    def validate_data(self):
        self.log.info("Validating Data")
        assert len(self.data), "No data to validate - forgot to read.."
        create = False
        for idx, sheet in enumerate(self.data):
            self.log.set_context(row=idx + 1)
            self.validate_home_data(
                create=create,
                builder=sheet["associations"].get("builder"),
                sheet_idx=idx,
                **sheet["home"],
            )
            self.validate_home_status_data(
                create=create, home=sheet["home"].get("home"), sheet_idx=idx, **sheet["home_status"]
            )
            self.validate_relationship_data(
                create=create,
                home_status=sheet["home_status"].get("home_status"),
                sheet_idx=idx,
                **sheet["associations"],
            )

            answer_date = datetime.datetime.today()
            if sheet["home_status"].get("certification_date"):
                answer_date = sheet["home_status"].get("certification_date")

            home_status = sheet["home_status"].get("home_status")
            log.debug("Validating data for %r", home_status if home_status else "new status")

            self.validate_annotation_data(
                create=create,
                user=self.user,
                home_status=home_status,
                overwrite=sheet["annotations_overwrite"],
                annotation_date=answer_date,
                annotation_data=sheet["annotations"],
            )

            self.validate_answers_data(
                create=create,
                home_status=home_status,
                company=self.user.company,
                overwrite=sheet["answer_overwrite"],
                answer_date=answer_date,
                sheet_idx=idx,
                question_data=sheet["questions"],
            )

            if self.uploaded_as_qa(sheet["home_status"].get("eep_program"), idx):
                self.validate_field_qa(
                    create=create,
                    eep_program=sheet["home_status"].get("eep_program"),
                    home_status=sheet["home_status"].get("home_status_non_qa"),
                    qa_home_status=sheet["home_status"].get("home_status"),
                    rater=sheet["home_status"].get("rater"),
                    sheet_idx=idx,
                    field_qa=sheet["field_qa"],
                )
        self.log.set_context(row=None)
        self.log.update_model(throttle_seconds=None)

    def process_data(self):
        self.log.info("Processing Data")
        assert len(self.data), "No data to validate - forgot to read.."
        create = True
        for idx, sheet in enumerate(self.data):
            self.log.set_context(row=idx + 1)
            self.validate_home_data(
                create=create,
                builder=sheet["associations"].get("builder"),
                sheet_idx=idx,
                **sheet["home"],
            )
            self.validate_home_status_data(
                create=create, home=sheet["home"].get("home"), sheet_idx=idx, **sheet["home_status"]
            )

            home_status = sheet["home_status"].get("home_status")

            log.debug("Processing data for %r", home_status)
            self.validate_relationship_data(
                create=create, home_status=home_status, sheet_idx=idx, **sheet["associations"]
            )
            answer_date = datetime.datetime.today()
            if sheet["home_status"].get("certification_date"):
                answer_date = sheet["home_status"].get("certification_date")

            if sheet["annotations"]:
                self.validate_annotation_data(
                    create=create,
                    user=self.user,
                    home_status=home_status,
                    overwrite=sheet["annotations_overwrite"],
                    annotation_date=answer_date,
                    annotation_data=sheet["annotations"],
                )

            if sheet["question_data"]:
                self.validate_answers_data(
                    create=create,
                    home_status=home_status,
                    company=self.user.company,
                    overwrite=sheet["answer_overwrite"],
                    answer_date=answer_date,
                    sheet_idx=idx,
                    question_data=sheet["question_data"],
                )

            if self.uploaded_as_qa(sheet["home_status"].get("eep_program"), idx):
                self.validate_field_qa(
                    create=create,
                    eep_program=sheet["home_status"].get("eep_program"),
                    home_status=sheet["home_status"].get("home_status_non_qa"),
                    qa_home_status=sheet["home_status"].get("home_status"),
                    rater=sheet["home_status"].get("rater"),
                    sheet_idx=idx,
                    field_qa=sheet["field_qa"],
                )

            if home_status:
                if home_status.pct_complete < 99.9:
                    update_home_stats(eepprogramhomestatus_id=home_status.id, log=self.log)
                    home_status = EEPProgramHomeStatus.objects.get(id=home_status.id)
                if home_status.state != "complete":
                    update_home_states(
                        eepprogramhomestatus_id=home_status.id, user_id=self.user.id, log=self.log
                    )
                    home_status = EEPProgramHomeStatus.objects.get(id=home_status.id)

                errors = home_status.report_eligibility_for_certification()
                home_status.validate_references()

                if errors:
                    for error in errors:
                        self.log.warning(error)
                    continue

                can_certify = home_status.can_user_certify(
                    self.user, perform_eligiblity_check=False
                )
                if not can_certify:
                    self.log.info(
                        UNABLE_TO_CERTIFY_INVALID_USER.format(
                            user=self.user.get_full_name(),
                            url=home_status.home.get_absolute_url(),
                            home=home_status.home,
                        )
                    )
                    continue

                errors = certify_single_home(
                    self.user, home_status, sheet["home_status"]["certification_date"]
                )

                for error in errors:
                    self.log.error(CERTIFY_ERROR.format(error=error))

                if not errors:
                    url = home_status.home.get_absolute_url()
                    self.log.info(
                        CERTIFIED_HOME.format(
                            url=url, home=home_status.home, date=home_status.certification_date
                        )
                    )

                self.log.set_flags(
                    home_certified=True, certification_date=home_status.certification_date
                )

        self.log.set_context(row=None)
        self.log.update_model(throttle_seconds=None)

    def clean_state(self, state_name):
        state_dict = dict([(x[0], "{}".format(x[1])) for x in US_STATES])
        reversed_state_dict = dict([("{}".format(x[1]), x[0]) for x in US_STATES])

        try:
            state = reversed_state_dict[state_name]
            state_name = state_dict[state]
        except KeyError:
            try:
                state_name = state_dict[state_name]
                state = reversed_state_dict[state_name]
            except KeyError:
                state_name = None
                state = None

        if hasattr(self, "user"):
            from axis.geographic.models import County

            _states = list(County.objects.filter_by_user(self.user).values_list("state", flat=True))
            if state not in _states:
                self.log.error("The state selected '%s' is not in your company territory", state)
        return state, state_name

    def clean_county(self, county_name, state, **kwargs):
        from axis.geographic.models import County

        counties = County.objects.all()
        if hasattr(self, "user"):
            counties = County.objects.filter_by_user(self.user)

        county, county_id = None, None
        if county_name:
            try:
                county = counties.get(name=county_name.strip(), state=state)
            except County.DoesNotExist:
                self.log.error("Unable to find a county '%s' in %s", county_name, state)
        if county:
            county_id = county.id
        else:
            county_id = None
        return county, county_id

    def clean_city(self, city_name, county, state, **kwargs):
        from axis.geographic.models import City

        city_kwargs = {"county__state": state}
        if county:
            city_kwargs["county"] = county
        if city_name:
            city_kwargs["name"] = city_name

        cities = City.objects.all()
        if hasattr(self, "user"):
            cities = City.objects.filter_by_user(self.user)

        city, city_id = None, None
        try:
            city = cities.get(**city_kwargs)
        except City.MultipleObjectsReturned:
            if county:
                self.log.error(
                    "Multiple cities found with the name '%s' in county %s %s",
                    city_name,
                    county.name,
                    state,
                )
            else:
                self.log.error(
                    "Multiple cities found with the name '%s' in %s - use the county to narrow this",
                    city_name,
                    state,
                )
        except City.DoesNotExist:
            if "county" in city_kwargs:
                self.log.error(
                    "No cities found with the name '%s' in county %s %s",
                    city_name,
                    county.name,
                    state,
                )
            else:
                self.log.error("No cities found with the name '%s' in %s", city_name, state)
        if city:
            city_id = city.id
        else:
            city_id = None
        return city, city_id

    def clean_home(self, home_id, overwrite, **kwargs):
        home = None

        homes = Home.objects.all()
        if hasattr(self, "user"):
            homes = Home.objects.filter_by_user(self.user)

        if home_id and overwrite in ["Yes", "No"]:
            try:
                home = homes.get(id=home_id)
            except Home.DoesNotExist:
                self.log.error("Home with Axis ID: '%s' does not exist", home_id)

        return home, home_id

    def clean_subdivision(self, subdivision_name, subdivision_id, city, community, **kwargs):
        from axis.subdivision.models import Subdivision

        subdivision, subdivision_id = None, None
        if not subdivision_name:
            return subdivision, subdivision_id

        subdivisions = Subdivision.objects.all()
        if hasattr(self, "user"):
            subdivisions = Subdivision.objects.filter_by_user(self.user)

        query_args = {"name": subdivision_name}
        if city:
            query_args["city"] = city
        if community:
            query_args["community"] = community
        try:
            subdivision = subdivisions.get(**query_args)
            subdivision_id = subdivision.id
        except Subdivision.DoesNotExist:
            msg = "Subdivision/MF Development %r does not exist" % subdivision_name
            if community:
                msg += " in community %s" % (community.name)
            if city:
                msg += f" in city {city}"
            self.log.error(msg)
        except Subdivision.MultipleObjectsReturned:
            msg = "Multiple Subdivisions/MF Developments found for '%s'" % subdivision_name
            if not city:
                msg += " please add a city"
                if not community:
                    msg += " or community"
            elif not community:
                msg += " please add a community"
            self.log.error(msg)
        return subdivision, subdivision_id

    def clean_community(self, community_name, community_id, city, **kwargs):
        from axis.community.models import Community

        community, community_id = None, None
        if not community_name:
            return community, community_id

        communities = Community.objects.all()
        if hasattr(self, "user"):
            communities = Community.objects.filter_by_user(self.user)

        query_args = {"name": community_name}
        if city:
            query_args["city"] = city
        try:
            community = communities.get(**query_args)
            community_id = community.id
        except Community.DoesNotExist:
            msg = "Community '%s' does not exist" % community_name
            if city:
                msg += f" in city {city}"
            self.log.error(msg)
        except Community.MultipleObjectsReturned:
            msg = "Multiple Communities found for %r" % community_name
            if not city:
                msg += " please add a city"
            self.log.error(msg)
        return community, community_id

    def clean_simulation(self, simulation_name, simulation_id, **kwargs):
        from axis.remrate_data.models import Simulation

        object, object_id = None, None
        if not simulation_name:
            return object, object_id

        objects = Simulation.objects.all()
        if hasattr(self, "user"):
            objects = Simulation.objects.filter_by_user(self.user)

        query_args = {"pk": re.search(r"\[(\d+)\]$", simulation_name).group(1)}

        try:
            object = objects.get(**query_args)
            object_id = object.id
        except Simulation.DoesNotExist:
            self.log.error("Simulation ID: '%s' does not exist", query_args["pk"])
        except Simulation.MultipleObjectsReturned:
            self.log.error("Multiple Simulations found for ID: '%s'", query_args["pk"])
        return object, object_id

    def clean_floorplan(self, floorplan_name, floorplan_id, **kwargs):
        from axis.floorplan.models import Floorplan

        object, object_id = None, None

        if not floorplan_name:
            return object, object_id

        objects = Floorplan.objects.all()
        if hasattr(self, "user"):
            objects = Floorplan.objects.filter_by_user(self.user)

        query_args = {"pk": re.search(r"\[(\d+)\]$", floorplan_name).group(1)}

        try:
            object = objects.get(**query_args)
            object_id = object.id
        except Floorplan.DoesNotExist:
            self.log.error("Floorplan ID: '%s' does not exist", query_args["pk"])
        except Floorplan.MultipleObjectsReturned:
            self.log.error("Multiple Floorplans found for ID: '%s'", query_args["pk"])
        return object, object_id

    def clean_eep_program(self, eep_program_name, eep_program_id, **kwargs):
        from axis.eep_program.models import EEPProgram

        object, object_id = None, None
        if not eep_program_name:
            return object, object_id

        objects = EEPProgram.objects.all()
        if hasattr(self, "user"):
            objects = EEPProgram.objects.filter_by_user(self.user, visible_for_use=True)

        query_args = {"name": eep_program_name}
        try:
            object = objects.get(**query_args)
            object_id = object.id
        except EEPProgram.DoesNotExist:
            self.log.error("Program '%s' does not exist", eep_program_name)
        except EEPProgram.MultipleObjectsReturned:
            self.log.error("Multiple Programs found for '%s'", eep_program_name)
        return object, object_id

    def clean_status_company(self, company_name, company_id, eep_program, **kwargs):
        from axis.company.models import Company

        object, object_id = None, None
        if not company_name:
            if eep_program.is_qa_program:
                company_name = self.user.company.name
            else:
                return object, object_id

        company_types = ["qa", "provider"] if eep_program.is_qa_program else ["rater", "provider"]

        objects = Company.objects.filter(company_type__in=company_types)
        if hasattr(self, "user"):
            objects = Company.objects.filter_by_user(self.user).filter(
                company_type__in=company_types
            )
            if self.user.company.company_type == "rater":
                objects = Company.objects.filter(id=self.user.company_id)
            if eep_program.is_qa_program:
                if company_id:
                    objects = Company.objects.filter(id=company_id)
                else:
                    objects = Company.objects.filter(
                        id__in=[self.user.company_id, eep_program.owner.id]
                    )

        query_args = {"name": company_name}
        try:
            object = objects.get(**query_args)
            object_id = object.id
        except Company.DoesNotExist:
            self.log.error("Company '%s' does not exist", company_name)
        except Company.MultipleObjectsReturned:
            self.log.error(f"Multiple Companies found for {company_name!r} ;)")
        return object, object_id

    def clean_company(self, company_name, company_id, overwrite, company_type, utility_type=None):
        from axis.company.models import Company

        object, object_id = None, None
        if not company_name:
            return object, object_id

        if isinstance(company_type, str):
            company_type = [company_type]

        if overwrite == "No" and company_id:
            try:
                return Company.objects.get(id=company_id), int(company_id)
            except Company.DoesNotExist:
                self.log.error(
                    "Company with ID '%s' does not exist - Specified do not overwrite", company_id
                )
                return object, object_id

        query_args = {"company_type__in": company_type}
        if utility_type == "gas":
            query_args["utilityorganization__gas_provider"] = True
        elif utility_type == "electric":
            query_args["utilityorganization__electricity_provider"] = True

        objects = Company.objects.all().filter(**query_args)
        if hasattr(self, "user"):
            objects = Company.objects.filter_by_user(self.user).filter(**query_args)
            if self.user.company.company_type in company_type:
                objects = Company.objects.filter(id=self.user.company_id)

        query_args = {"name": company_name}
        try:
            object = objects.get(**query_args)
            object_id = object.id
        except Company.DoesNotExist:
            self.log.error("Company '%s' does not exist", company_name)
        except Company.MultipleObjectsReturned:
            self.log.error("Multiple Companies found for '%s'", company_name)
        return object, object_id

    def clean_rater(self, rater_name, rater_id, company=None, **kwargs):
        from axis.company.models import Company

        object, object_id = None, None
        if not rater_name:
            return object, object_id

        last, first = rater_name.split(", ")[0], ", ".join(rater_name.split(", ")[1:])

        query_args = {"last_name": last, "first_name": first}
        if company:
            query_args["company"] = company

        objects = User.objects.all()
        if hasattr(self, "user"):
            objects = User.objects.filter_by_user(self.user)
            if self.user.company.company_type in ["rater", "provider"]:
                comps = Company.objects.filter_by_user(self.user, include_self=True).filter(
                    company_type__in=["rater", "provider"]
                )
                objects = User.objects.filter(company_id__in=comps.values_list("id", flat=True))
                objects = User.objects.filter(
                    id__in=list(objects.values_list("id", flat=True)) + [self.user.id]
                )
        try:
            object = objects.get(**query_args)
            object_id = object.id
        except User.DoesNotExist:
            self.log.error("User '%s' does not exist", rater_name)
        except User.MultipleObjectsReturned:
            self.log.error("Multiple Users found for '%s'", rater_name)
        return object, object_id

    def clean_home_status(self, home, home_status_id, overwrite, **kwargs):
        from axis.home.models import EEPProgramHomeStatus

        object, object_id = None, None
        objects = EEPProgramHomeStatus.objects.all()
        if hasattr(self, "user"):
            objects = EEPProgramHomeStatus.objects.filter_by_user(self.user)

        if home_status_id and overwrite in ["Yes", "No"]:
            try:
                object = objects.get(id=home_status_id, home=home)
                object_id = object.id
            except EEPProgramHomeStatus.DoesNotExist:
                self.log.error("Project with ID: '%s' does not exist", home_status_id)

        return object, object_id

    def clean_certification(self, certification_date, **kwargs):
        if certification_date is None:
            return None
        cert_date = None
        try:
            cert_date = datetime.datetime.strptime(certification_date, "%m/%d/%Y")
        except Exception:
            try:
                cert_date = datetime.datetime.strptime(certification_date, "%Y-%m-%d %H:%M:%S")
            except Exception:
                self.log.error("Unable to parse certification date from '%s'", certification_date)
        return cert_date

    def read_home_values(self, sheet_obj, row=None):
        data = dict(
            [
                (
                    "home_id",
                    self.get_value_from_label(sheet_name=sheet_obj.title, label="home_id", row=row),
                ),
                (
                    "overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__overwrite", row=row
                    ),
                ),
                (
                    "is_multi_family",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__is_multi_family", row=row
                    ),
                ),
                (
                    "lot_number",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__lot_number", row=row
                    ),
                ),
                (
                    "street_line1",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__street_line1", row=row
                    ),
                ),
                (
                    "street_line2",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__street_line2", row=row
                    ),
                ),
                (
                    "city_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__city__name", row=row
                    ),
                ),
                (
                    "city_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__city_id", row=row
                    ),
                ),
                (
                    "county_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__county__name", row=row
                    ),
                ),
                (
                    "county_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__county_id", row=row
                    ),
                ),
                (
                    "state",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__state", row=row
                    ),
                ),
                (
                    "state_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__state__name", row=row
                    ),
                ),
                (
                    "zipcode",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home__zipcode", row=row
                    ),
                ),
                (
                    "community_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="community__name", row=row
                    ),
                ),
                (
                    "community_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="community_id", row=row
                    ),
                ),
                (
                    "subdivision_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="subdivision__name", row=row
                    ),
                ),
                (
                    "subdivision_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="subdivision_id", row=row
                    ),
                ),
            ]
        )

        data["state"], data["state_name"] = self.clean_state(data["state_name"])
        data["county"], data["county_id"] = self.clean_county(**data)
        data["city"], data["city_id"] = self.clean_city(**data)
        data["home"], data["home_id"] = self.clean_home(**data)
        data["community"], data["community_id"] = self.clean_community(**data)
        data["subdivision"], data["subdivision_id"] = self.clean_subdivision(**data)

        if data["overwrite"] == "Yes" and not data["home_id"]:
            self.log.error('Home overwrite cannot be set to "Yes" without an Axis home ID')
        if data["overwrite"] == "New" and data["home_id"]:
            self.log.error('Home overwrite cannot be set to "New" when using an Axis home ID')
        if data["subdivision"] and data["community"]:
            if data["subdivision"].community != data["community"]:
                self.log.error(
                    'Community specified "%s" does not match Subdivision Community "%s"',
                    data["community"],
                    data["subdivision"].community,
                )

        data["home_url"] = (
            data["home"].get_absolute_url() if data["home_id"] and data["home"] else None
        )
        data["home_string"] = "{} [{}]".format(data["home"], data["home_id"])
        data["city_url"] = data["city"].get_absolute_url() if data["city_id"] else None
        data["city_string"] = "{}".format(data["city"])
        data["subdivision_url"] = (
            data["subdivision"].get_absolute_url() if data["subdivision_id"] else None
        )
        data["subdivision_string"] = "{}".format(data["subdivision"])
        data["community_url"] = (
            data["community"].get_absolute_url() if data["community_id"] else None
        )
        data["community_string"] = "{}".format(data["community"])

        return data

    def read_home_stat_values(self, sheet_obj, row=None, home=None, **kwargs):
        data = dict(
            [
                (
                    "home_status_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status_id", row=row
                    ),
                ),
                (
                    "overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__overwrite", row=row
                    ),
                ),
                (
                    "eep_program_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__program", row=row
                    ),
                ),
                (
                    "eep_program_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__program_id", row=row
                    ),
                ),
                (
                    "company_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__company", row=row
                    ),
                ),
                (
                    "company_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__company_id", row=row
                    ),
                ),
                (
                    "rater_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__rater", row=row
                    ),
                ),
                (
                    "rater_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_status__rater_id", row=row
                    ),
                ),
                (
                    "certification_date",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="certification_date", row=row
                    ),
                ),
                (
                    "floorplan_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title,
                        label="home_status__floorplan",
                        row=row,
                        raise_on_error=False,
                    ),
                ),
                (
                    "floorplan_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title,
                        label="home_status__floorplan_id",
                        row=row,
                        raise_on_error=False,
                    ),
                ),
                (
                    "simulation_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title,
                        label="home_status__simulation",
                        row=row,
                        raise_on_error=False,
                    ),
                ),
                (
                    "simulation_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title,
                        label="home_status__simulation_id",
                        row=row,
                        raise_on_error=False,
                    ),
                ),
            ]
        )

        if kwargs.get("overwrite") == "New" and data["overwrite"] != "New":
            self.log.info('Setting "New" on home status as it\'s a new home.')
            data["overwrite"] = "New"

        data["floorplan"], data["floorplan_id"] = self.clean_floorplan(**data)
        data["simulation"], data["simulation_id"] = self.clean_simulation(**data)

        data["eep_program"], data["eep_program_id"] = self.clean_eep_program(**data)
        data["company"], data["company_id"] = self.clean_status_company(**data)
        data["rater"], data["rater_id"] = self.clean_rater(**data)
        data["home_status"], data["home_status_id"] = self.clean_home_status(home, **data)
        data["certification_date"] = self.clean_certification(**data)

        data["is_certified"] = data["home_status"] and data["home_status"].certification_date

        if data["overwrite"] == "Yes" and not data["home_status_id"]:
            self.log.error(
                'Project overwrite cannot be set to "Yes" without an existing program on a home'
            )
        if data["overwrite"] == "New" and data["home_status_id"]:
            self.log.error('Project overwrite cannot be set to "New" when using program on a home')

        if data["overwrite"] in ["New", "Yes"] and data["is_certified"]:
            self.log.warning("Program on home is already certified modifications not allowed")

        return data

    def read_field_qa_values(
        self, sheet_obj, row=None, eep_program=None, home_status=None, **kwargs
    ):
        has_field_qa = self.get_cell_from_label(
            sheet_obj.title, "field_qa_overwrite", raise_on_error=False
        )

        qa_program = eep_program
        rater_program = qa_program.get_rater_program()

        qa_company = (
            home_status.company if home_status and home_status.company_id else self.user.company
        )

        try:
            QARequirement.objects.get(
                eep_program=rater_program, type="field", qa_company=qa_company
            )
        except QARequirement.DoesNotExist:
            log.info("QA Requirement does not exist")
            return {}

        if not has_field_qa:
            return {}

        return dict(
            [
                (
                    "overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="field_qa_overwrite", row=row
                    ),
                ),
                (
                    "new_state",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="field_qa__new_state", row=row
                    ),
                ),
                (
                    "new_note",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="field_qa__new_notes", row=row
                    ),
                ),
            ]
        )

    def read_company_values(self, sheet_obj, row=None, home_status=None, **kwargs):
        data = dict(
            [
                (
                    "builder_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_builder", row=row
                    ),
                ),
                (
                    "builder_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_builder_id", row=row
                    ),
                ),
                (
                    "builder_overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="home_builder_overwrite", row=row
                    ),
                ),
                (
                    "electric_utility_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="electric_utility", row=row
                    ),
                ),
                (
                    "electric_utility_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="electric_utility_id", row=row
                    ),
                ),
                (
                    "electric_utility_overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="electric_utility_overwrite", row=row
                    ),
                ),
                (
                    "gas_utility_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="gas_utility", row=row
                    ),
                ),
                (
                    "gas_utility_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="gas_utility_id", row=row
                    ),
                ),
                (
                    "gas_utility_overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="gas_utility_overwrite", row=row
                    ),
                ),
                (
                    "provider_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="provider_organization", row=row
                    ),
                ),
                (
                    "provider_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="provider_organization_id", row=row
                    ),
                ),
                (
                    "provider_overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="provider_organization_overwrite", row=row
                    ),
                ),
                (
                    "rater_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="rating_company", row=row
                    ),
                ),
                (
                    "rater_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="rating_company_id", row=row
                    ),
                ),
                (
                    "rater_overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="rating_company_overwrite", row=row
                    ),
                ),
                (
                    "hvac_name",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="hvac_company", row=row
                    ),
                ),
                (
                    "hvac_id",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="hvac_company_id", row=row
                    ),
                ),
                (
                    "hvac_overwrite",
                    self.get_value_from_label(
                        sheet_name=sheet_obj.title, label="hvac_company_overwrite", row=row
                    ),
                ),
            ]
        )

        data["builder"], data["builder_id"] = self.clean_company(
            company_name=data["builder_name"],
            company_id=data["builder_id"],
            overwrite=data["builder_overwrite"],
            company_type="builder",
        )

        data["electric_utility"], data["electric_utility_id"] = self.clean_company(
            company_name=data["electric_utility_name"],
            company_id=data["electric_utility_id"],
            overwrite=data["electric_utility_overwrite"],
            company_type="utility",
            utility_type="electric",
        )

        data["gas_utility"], data["gas_utility_id"] = self.clean_company(
            company_name=data["gas_utility_name"],
            company_id=data["gas_utility_id"],
            overwrite=data["gas_utility_overwrite"],
            company_type="utility",
            utility_type="gas",
        )

        data["provider"], data["provider_id"] = self.clean_company(
            company_name=data["provider_name"],
            company_id=data["provider_id"],
            overwrite=data["provider_overwrite"],
            company_type="provider",
        )

        company_type = home_status.company.company_type if home_status else "rater"
        data["rater"], data["rater_id"] = self.clean_company(
            company_name=data["rater_name"],
            company_id=data["rater_id"],
            overwrite=data["rater_overwrite"],
            company_type=company_type,
        )

        data["hvac"], data["hvac_id"] = self.clean_company(
            company_name=data["hvac_name"],
            company_id=data["hvac_id"],
            overwrite=data["hvac_overwrite"],
            company_type="hvac",
        )

        return data

    def read_question_values(
        self, sheet_obj, row=None, home_status=None, eep_program=None, **kwargs
    ):
        assert hasattr(self, "workbook"), "Need a workbook"

        prefix = r"^(measure|question)_"
        questions = [
            x.name
            for x in self.workbook.defined_names.definedName
            if re.match(prefix, x.name)
            and x.attr_text.startswith(sheet_obj.title)
            and (not row or re.search(r"[A-Z]\$%s$" % row, x.attr_text))
        ]

        collection_request = eep_program.collection_request
        if home_status:
            collection_request = home_status.collection_request
            self.log.debug(
                "Using Collection Request on home status (%(home_status)s)",
                dict(home_status=home_status.pk),
            )

        self.use_legacy = collection_request is None
        if self.use_legacy:
            self.log.info(
                "No collection_request for homestatus.  Treating data as a legacy checklist"
            )

        # Build Collector reguardless of legacy use, because it will power the payload object
        # generation that legacy has been updated to receive.
        context = {
            "user": self.user,
        }
        self.collector = ExcelChecklistCollector(collection_request, **context)

        if not self.use_legacy:
            instruments = self.collector.get_xls_instruments()
            instrument_lookup = {i.measure_id.replace("-", "_"): i for i in instruments}

        results = []
        for cell_name in questions:
            instrument = None
            measure = re.sub(prefix, "", cell_name)

            # FIXME: The single home field qa xlsx likely went out with a typo'd measure name, so
            # this will help us avoid the discrepancy.
            if measure == "equipment-primary-heating-type":
                measure = "primary-heating-equipment-type"

            if self.use_legacy:
                instrument = Question.objects.filter(
                    id=measure, checklist__eepprogram=eep_program
                ).first()
                instrument.collection_request = None
            elif measure in instrument_lookup:
                instrument = instrument_lookup[measure]

            if instrument is None:
                raise ValueError("Unexpected answer cell %r." % (measure,))

            data = self.get_value_from_label(sheet_name=sheet_obj.title, label=cell_name, row=row)

            if data is None:
                continue

            comment = self.get_comment_from_label(
                sheet_name=sheet_obj.title, label=cell_name, row=row
            )
            payload = self.collector.make_payload(instrument, data, extra={"comment": comment})

            results.append(payload)

        return results

    def read_annotation_values(
        self, sheet_obj, row=None, home_status=None, eep_program=None, **kwargs
    ):
        assert hasattr(self, "workbook"), "Need a workbook"
        prefix = r"^annotation_"
        annotations = [
            x.name
            for x in self.workbook.defined_names.definedName
            if re.match(prefix, x.name)
            and sheet_obj.title in x.attr_text
            and (not row or re.search(r"[A-Z]\$%s$" % row, x.attr_text))
        ]

        if home_status:
            self.log.debug(f"Using Home Status {home_status.id}", dict(home_status=home_status.pk))

        required_annotations = {}
        for annotation in eep_program.required_annotation_types.all():
            required_annotations[annotation.slug] = annotation
            required_annotations[re.sub("_", "-", annotation.slug)] = annotation

        results = []
        for cell_name in annotations:
            annotation = re.sub("_", "-", re.sub(prefix, "", cell_name))

            if annotation not in required_annotations:
                raise ValueError("Unexpected annotation cell %r." % (annotation,))

            data = self.get_value_from_label(sheet_name=sheet_obj.title, label=cell_name, row=row)
            if data is None:
                continue
            results.append(
                {
                    "type_id": required_annotations[annotation].id,
                    "content_type_id": ContentType.objects.get_for_model(EEPProgramHomeStatus).pk,
                    "content": data,
                    "user_id": self.user.pk,
                }
            )

        return results

    def write(
        self,
        home_status_id=None,
        home_id=None,
        program_id=None,
        output=None,
        lock_and_hide=True,
        return_workbook=False,
    ):
        """
        get all information given home_status and print to 640S excel.
        :returns: workbook, output
        to save to location:
            workbook, output = object.write()
            workbook.save(output)
        """
        from axis.home.models import EEPProgramHomeStatus

        home = None
        home_status = None
        eep_program = None
        collection_request = None

        # Resolve objects
        if home_status_id:
            home_status = EEPProgramHomeStatus.objects.get(id=home_status_id)
            home = home_status.home

            eep_program = home_status.eep_program
            role = get_user_role_for_homestatus(home_status, self.user)
            if role == "qa":
                eep_program = eep_program.get_qa_program()
                home_status = home.homestatuses.filter(eep_program=eep_program).first()
                log.info("Shifting to using QA program")

            collection_request = home_status.collection_request
            if (
                home_status.collection_request is None
                and home_status.eep_program.collection_request_id
            ):
                home_status.set_collection_from_program()

        elif home_id:
            home = Home.objects.get(id=home_id)
            _kw = {"home_id": home_id}
            if program_id:
                _kw = {"home_id": home_id, "eep_program_id": program_id}
            try:
                home_status = EEPProgramHomeStatus.objects.get(**_kw)
            except EEPProgramHomeStatus.DoesNotExist:
                home_status = EEPProgramHomeStatus()
            else:
                eep_program = home_status.eep_program

                # Check homestatus.id softly, it could be an unbound qa instance now
                if (
                    home_status.collection_request is None
                    and home_status.eep_program.collection_request_id
                ):
                    home_status.set_collection_from_program()

            collection_request = home_status.collection_request

        elif program_id:
            home = Home()
            home_status = EEPProgramHomeStatus()
            eep_program = EEPProgram.objects.get(id=program_id)
            collection_request = eep_program.collection_request

        # Setup
        if output is None:
            _, output = tempfile.mkstemp(suffix=".xlsx", prefix=self.output_prefix)
            self.log.debug("Look here for the file that was just generated %s", output)
        if os.path.isfile(output):
            shutil.move(output, os.path.join(tempfile.gettempdir(), os.path.basename(output)))

        # handle special templates
        if eep_program and not self.template:
            template_map = {
                "eto-2020-qa": TEMPLATE_QA_ETO_2020,
            }
            self.template = template_map.get(eep_program.slug, TEMPLATE)
        else:
            self.template = self.template if self.template else self.get_template()

        self.log.debug("Using template %s", self.template)
        shutil.copyfile(self.template, output)

        self.workbook = load_workbook(filename=output, keep_vba=False, data_only=True)
        sheet_obj = self.workbook[self.workbook.sheetnames[0]]
        self.editable_cells = []

        log.debug(
            "Writing Sections -- %(eep_program)r, %(home)r, %(home_status)r",
            {"eep_program": eep_program, "home_status": home_status, "home": home},
        )
        # Write
        self.write_sections(
            **{
                "sheet_obj": sheet_obj,
                "eep_program": eep_program,
                "home": home,
                "home_status": home_status,
                "collection_request": collection_request,
            }
        )

        # Cleanup
        for cell_range in sheet_obj.merged_cells.ranges:
            if cell_range in self.handled_merged_cells:
                continue
            cell_label = cell_range.coord.split(":")[0]
            self.set_merged_bordered_style(sheet_obj, sheet_obj[cell_label], **self.basic_style)
        self.pre_save(self.workbook, sheet_obj, lock_and_hide, eep_program)

        if return_workbook:
            return self.workbook

        self.workbook.save(output)
        return output

    def write_sections(self, **context):
        self.write_home_values(**context)
        self.write_home_status_values(**context)
        self.write_company_values(**context)
        self.write_annotations(**context)
        self.write_checklist(**context)
        self.write_field_qa_values(**context)
        self.write_extra(**context)

    def pre_save(self, workbook, sheet_obj, lock_and_hide=True, eep_program=None):
        if lock_and_hide:
            geo_sheet = self.workbook["Geography"]
            geo_sheet.sheet_state = "hidden"
            self.lock_cells(geo_sheet)

        image = Image(AXIS_LOGO)
        image.width = 100
        image.height = 100
        image.anchor = "B1"

        cell_obj = sheet_obj["D2"]
        cell_obj.value = "AXIS Single Home Checklist"
        self.set_style(sheet_obj, cell_obj, **{"font": Font(name="Calibri", size=14, bold=True)})

        cell_obj = sheet_obj["D4"]
        cell_obj.value = "Prepared specifically for {} of {} on {}".format(
            self.user, self.user.company, datetime.date.today()
        )
        self.set_style(
            sheet_obj,
            cell_obj,
            **{
                "font": Font(name="Calibri", size=10, italic=True, color=Color(rgb=GREY_FONT_COLOR))
            },
        )

        if lock_and_hide:
            self.lock_cells(sheet_obj)
            cells_needing_white = (
                "M9 M10 M11 M15 M16 M19 M20 M21 M22 M25 M26 M29 M30 M31 M32 M33 M34".split()
            )
            if self.template == TEMPLATE_QA_ETO_2020:
                cells_needing_white = (
                    "M10 M12 M13 M14 M18 M19 M22 M23 M24 M25 M28 M29 M30 M31 M32 M33 M34".split()
                )

            for cell in cells_needing_white:
                cell_obj = sheet_obj[cell]
                self.set_style(sheet_obj, cell_obj, **self.basic_white_style)

        sheet_obj.add_image(image)

        if eep_program.id and "eto" in eep_program.slug:
            sheet_obj.column_dimensions["J"].width = 3
            sheet_obj.column_dimensions["K"].width = 8
            eto_image = Image(ETO_LOGO)
            eto_image.width = 207
            eto_image.height = 75
            eto_image.anchor = "K1"

            sheet_obj.add_image(eto_image)

    def write_home_values(self, home, sheet_obj, row=None, **context):
        self.set_label_default_value(
            sheet_obj=sheet_obj, cell_label="home_id", value=home.id, row=row, edit_flag=True
        )

        overwrite = "No" if home.id else "New"
        self.set_home__overwrite(
            sheet_obj=sheet_obj, cell_label="home__overwrite", value=overwrite, row=row
        )

        address = home.get_home_address_display_parts(
            include_city_state_zip=True, company=self.user.company
        )

        self.set_label_optional_value(
            sheet_obj=sheet_obj, cell_label="home__lot_number", value=address.lot_number, row=row
        )

        self.set_label_required_value(
            sheet_obj=sheet_obj,
            cell_label="home__street_line1",
            value=address.street_line1,
            row=row,
        )
        self.set_label_optional_value(
            sheet_obj=sheet_obj,
            cell_label="home__street_line2",
            value=address.street_line2,
            row=row,
        )
        self.set_multi_family(
            sheet_obj=sheet_obj,
            cell_label="home__is_multi_family",
            value=home.is_multi_family,
            row=row,
        )
        self.set_label_optional_value(
            sheet_obj=sheet_obj, cell_label="home__parcel", value="", row=row
        )

        city = home.city.name if home.city else None
        city_id = home.city.id if home.city else None
        self.set_label_required_value(
            sheet_obj=sheet_obj, cell_label="home__city__name", value=city, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home__city_id", value=city_id, row=row
        )

        county = home.city.county.name if home.city and home.city.county else None
        county_id = home.city.county.id if home.city and home.city.county else None
        self.set_county(
            sheet_obj=sheet_obj,
            cell_label="home__county__name",
            value=county,
            row=row,
            required=True,
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home__county_id", value=county_id, row=row
        )

        state = home.city.county.state if home.city and home.city.county else None
        self.set_state(sheet_obj=sheet_obj, cell_label="home__state__name", value=state, row=row)
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home__state", value=state, row=row
        )

        self.set_zipcode(
            sheet_obj=sheet_obj, cell_label="home__zipcode", value=address.zipcode, row=row
        )

        subdision = home.subdivision.name if home.subdivision else None
        subdision_id = home.subdivision.id if home.subdivision else None
        self.set_subdivision(
            sheet_obj=sheet_obj, cell_label="subdivision__name", value=subdision, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="subdivision_id", value=subdision_id, row=row
        )

        community = (
            home.subdivision.community.name
            if home.subdivision and home.subdivision.community
            else None
        )
        community_id = (
            home.subdivision.community.id
            if home.subdivision and home.subdivision.community
            else None
        )
        self.set_community(
            sheet_obj=sheet_obj, cell_label="community__name", value=community, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="community_id", value=community_id, row=row
        )

        if row is not None:
            return

        # Cosmetics
        start = self.get_cell_from_label(sheet_obj.title, "home__overwrite").row
        end = self.get_cell_from_label(sheet_obj.title, "subdivision_id").row
        for cell_range in sheet_obj.merged_cells.ranges:
            if cell_range in self.handled_merged_cells:
                continue
            cell_label = cell_range.coord.split(":")[0]
            row = int(re.sub(r"[A-Z]+", "", cell_label))
            if start <= row <= end:
                self.set_merged_bordered_style(sheet_obj, sheet_obj[cell_label], **self.basic_style)

        log.debug("Setting home border -- B{start}:N{end}".format(start=start, end=end))
        self.set_border_range("B{start}:N{end}".format(start=start, end=end), sheet_obj=sheet_obj)

    def write_home_status_values(self, home_status, eep_program, sheet_obj, row=None, **context):
        eep_program_id = eep_program.id

        overwrite = "No" if home_status.id else "New"
        self.set_home_status_overwrite(
            sheet_obj=sheet_obj, cell_label="home_status__overwrite", value=overwrite, row=row
        )

        status_id = home_status.id if home_status else None
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home_status_id", value=status_id, row=row
        )

        program = home_status.eep_program.name if home_status.eep_program_id else None
        program_id = home_status.eep_program.id if home_status.eep_program_id else None
        if eep_program_id:
            from axis.eep_program.models import EEPProgram

            _program = EEPProgram.objects.get(id=eep_program_id)
            program, program_id = _program.name, _program.id
        self.set_program(
            sheet_obj=sheet_obj,
            cell_label="home_status__program",
            value=program,
            row=row,
            eep_program_id=eep_program_id,
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home_status__program_id", value=program_id, row=row
        )

        company = home_status.company.name if home_status.company_id else None
        company_id = home_status.company.id if home_status.company_id else None
        if eep_program_id and self.user.company.company_type in ["rater", "provider"]:
            company, company_id = self.user.company.name, self.user.company.id
        self.set_home_status_company(
            sheet_obj=sheet_obj,
            cell_label="home_status__company",
            value=company,
            row=row,
            eep_program=eep_program,
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home_status__company_id", value=company_id, row=row
        )

        rater = (
            "{}, {}".format(
                home_status.rater_of_record.last_name, home_status.rater_of_record.first_name
            )
            if home_status and home_status.rater_of_record
            else None
        )
        rater_id = (
            home_status.rater_of_record.id if home_status and home_status.rater_of_record else None
        )
        self.set_home_status_rater(
            sheet_obj=sheet_obj,
            cell_label="home_status__rater",
            value=rater,
            row=row,
            company_id=company_id,
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home_status__rater_id", value=rater_id, row=row
        )

        certification_date = (
            home_status.certification_date
            if home_status and home_status.certification_date
            else None
        )
        self.set_certification_date(
            sheet_obj=sheet_obj, cell_label="certification_date", value=certification_date, row=row
        )

        has_floorplan = self.get_cell_from_label(
            sheet_obj.title, "home_status__floorplan", raise_on_error=False
        )
        if has_floorplan:
            floorplan = home_status.floorplan if home_status and home_status.floorplan else None
            floorplan_id = floorplan.id if floorplan else None
            self.set_floorplan(
                sheet_obj=sheet_obj, cell_label="home_status__floorplan", value=floorplan, row=row
            )
            self.set_label_white_value(
                sheet_obj=sheet_obj,
                cell_label="home_status__floorplan_id",
                value=floorplan_id,
                row=row,
            )

        has_simulation = self.get_cell_from_label(
            sheet_obj.title, "home_status__simulation", raise_on_error=False
        )
        if has_simulation:
            simulation = (
                home_status.floorplan.remrate_target
                if home_status and home_status.floorplan and home_status.floorplan.remrate_target
                else None
            )
            simulation_id = simulation.id if simulation else None
            self.set_simulation(
                sheet_obj=sheet_obj, cell_label="home_status__simulation", value=simulation, row=row
            )
            self.set_label_white_value(
                sheet_obj=sheet_obj,
                cell_label="home_status__simulation_id",
                value=simulation_id,
                row=row,
            )

        if row is not None:
            return

        # Cosmetics
        start = self.get_cell_from_label(sheet_obj.title, "home_status__overwrite").row
        last = "home_status__simulation_id" if has_simulation else "certification_date"
        end = self.get_cell_from_label(sheet_obj.title, last).row
        for cell_range in sheet_obj.merged_cells.ranges:
            if cell_range in self.handled_merged_cells:
                continue
            cell_label = cell_range.coord.split(":")[0]
            row = int(re.sub(r"[A-Z]+", "", cell_label))
            if start <= row <= end:
                self.set_merged_bordered_style(sheet_obj, sheet_obj[cell_label], **self.basic_style)

        log.debug("Setting status border -- B{start}:N{end}".format(start=start, end=end))
        self.set_border_range("B{start}:N{end}".format(start=start, end=end), sheet_obj=sheet_obj)

    def write_field_qa_values(self, home, home_status, eep_program, sheet_obj, row=None, **context):
        has_field_qa = self.get_cell_from_label(
            sheet_obj.title, "field_qa_overwrite", raise_on_error=False
        )

        qa_program = eep_program
        rater_program = qa_program.get_rater_program()

        qa_company = home_status.company if home_status.company_id else self.user.company

        try:
            QARequirement.objects.get(
                eep_program=rater_program, type="field", qa_company=qa_company
            )
        except QARequirement.DoesNotExist:
            log.info("QA Requirement does not exist")
            return

        if not has_field_qa:
            return

        overwrite = "No" if home_status.id else "New"
        self.set_home_status_overwrite(
            sheet_obj=sheet_obj, cell_label="field_qa_overwrite", value=overwrite, row=row
        )

        self.set_qa_state(
            sheet_obj=sheet_obj, cell_label="field_qa__new_state", value=None, row=row
        )
        self.set_label_required_value(
            sheet_obj=sheet_obj, cell_label="field_qa__new_notes", value=None, row=row
        )

        # Cosmetics
        start = self.get_cell_from_label(sheet_obj.title, "field_qa_overwrite").row
        end = self.get_cell_from_label(sheet_obj.title, "field_qa__new_notes").row
        for cell_range in sheet_obj.merged_cells.ranges:
            if cell_range in self.handled_merged_cells:
                continue
            cell_label = cell_range.coord.split(":")[0]
            row = int(re.sub(r"[A-Z]+", "", cell_label))
            if start <= row <= end:
                self.set_merged_bordered_style(sheet_obj, sheet_obj[cell_label], **self.basic_style)

        log.debug("Setting Field QA border -- B{start}:N{end}".format(start=start, end=end))
        self.set_border_range("B{start}:N{end}".format(start=start, end=end), sheet_obj=sheet_obj)

    def write_company_values(self, home, home_status, eep_program, sheet_obj, row=None, **context):
        eep_program_id = eep_program.id

        builder = home.get_builder() if home.id else None
        builder_id = builder.id if builder else None
        overwrite = "No" if home.id and builder else "New"
        self.set_company_type(
            sheet_obj=sheet_obj,
            cell_label="home_builder",
            value=builder,
            row=row,
            target_column="T",
            company_type="builder",
        )
        self.set_company_overwrite(
            sheet_obj=sheet_obj, cell_label="home_builder_overwrite", value=overwrite, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="home_builder_id", value=builder_id, row=row
        )

        e_utility = home.get_electric_company() if home.id else None
        e_utility_id = e_utility.id if e_utility else None
        overwrite = "No" if home.id and e_utility else "New"
        self.set_company_type(
            sheet_obj=sheet_obj,
            cell_label="electric_utility",
            value=e_utility,
            row=row,
            target_column="U",
            company_type="utility",
            utility_type="electric",
        )
        self.set_company_overwrite(
            sheet_obj=sheet_obj, cell_label="electric_utility_overwrite", value=overwrite, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="electric_utility_id", value=e_utility_id, row=row
        )

        g_utility = home.get_gas_company() if home.id else None
        g_utility_id = g_utility.id if g_utility else None
        overwrite = "No" if home.id and g_utility else "New"
        self.set_company_type(
            sheet_obj=sheet_obj,
            cell_label="gas_utility",
            value=g_utility,
            row=row,
            target_column="V",
            company_type="utility",
            utility_type="gas",
        )
        self.set_company_overwrite(
            sheet_obj=sheet_obj, cell_label="gas_utility_overwrite", value=overwrite, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="gas_utility_id", value=g_utility_id, row=row
        )

        hvac = home.get_hvac_company() if home.id else None
        hvac_id = hvac.id if hvac else None
        overwrite = "No" if home.id and hvac else "New"
        self.set_company_type(
            sheet_obj=sheet_obj,
            cell_label="hvac_company",
            value=hvac,
            row=row,
            target_column="W",
            company_type="hvac",
        )
        self.set_company_overwrite(
            sheet_obj=sheet_obj, cell_label="hvac_company_overwrite", value=overwrite, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="hvac_company_id", value=hvac_id, row=row
        )

        provider = home_status.get_provider() if home_status.id else None
        provider_id = provider.id if provider else None
        if (
            eep_program_id
            and hasattr(self, "user")
            and self.user.company.company_type == "provider"
        ):
            provider, provider_id = self.user.company, self.user.company_id
        overwrite = "No" if home.id and provider else "New"
        self.set_company_type(
            sheet_obj=sheet_obj,
            cell_label="provider_organization",
            value=provider,
            target_column="X",
            row=row,
            company_type="provider",
        )
        self.set_company_overwrite(
            sheet_obj=sheet_obj,
            cell_label="provider_organization_overwrite",
            value=overwrite,
            row=row,
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="provider_organization_id", value=provider_id, row=row
        )

        if eep_program.is_qa_program:
            rater = None
            company_type = "rater"
            if home_status:
                try:
                    rater_hs = EEPProgramHomeStatus.objects.get(
                        eep_program=eep_program.get_rater_program(), home=home_status.home
                    )
                    rater = rater_hs.company
                    company_type = rater_hs.company.company_type if rater_hs.id else "rater"
                except Exception:
                    pass
        else:
            rater = home_status.company if home_status.id else None
            company_type = home_status.company.company_type if home_status.id else "rater"
            if (
                eep_program_id
                and hasattr(self, "user")
                and self.user.company.company_type == "rater"
            ):
                overwrite = "No" if home.id and rater else "New"
            rater, rater_id = self.user.company, self.user.company_id

        rater_id = rater.id if rater else None
        self.set_company_type(
            sheet_obj=sheet_obj,
            cell_label="rating_company",
            value=rater,
            row=row,
            target_column="Y",
            company_type=company_type,
        )
        self.set_company_overwrite(
            sheet_obj=sheet_obj, cell_label="rating_company_overwrite", value=overwrite, row=row
        )
        self.set_label_white_value(
            sheet_obj=sheet_obj, cell_label="rating_company_id", value=rater_id, row=row
        )

        if row is not None:
            return

        # Cosmetics
        start = self.get_cell_from_label(sheet_obj.title, "home_builder").row - 1
        end = self.get_cell_from_label(sheet_obj.title, "rating_company_id").row
        for cell_range in sheet_obj.merged_cells.ranges:
            if cell_range in self.handled_merged_cells:
                continue
            cell_label = cell_range.coord.split(":")[0]
            row = int(re.sub(r"[A-Z]+", "", cell_label))
            if start <= row <= end:
                self.set_merged_bordered_style(sheet_obj, sheet_obj[cell_label], **self.basic_style)

        log.debug("Setting status border -- B{start}:N{end}".format(start=start, end=end))
        self.set_border_range("B{start}:N{end}".format(start=start, end=end), sheet_obj=sheet_obj)

    def get_validation(self, annotation_type):
        return ExcelChecklistCollector(None).type_validations.get(annotation_type.data_type)

    def set_annotation_input(self, annotation_type, annotation, cell_obj, sheet_obj):
        # Validation
        info = self.get_validation(annotation_type)
        if info:
            type_name = info["type"].capitalize()
            self.add_validation(
                sheet_obj,
                cell_obj,
                DataValidation(
                    **{
                        "type": info["type"],
                        "operator": info.get("operator"),
                        "formula1": info.get("formula1"),
                        "formula2": info.get("formula2"),
                        "errorTitle": info.get("error_title", "Invalid {}".format(type_name)),
                        "error": info.get("error", "Entry must be a {}".format(type_name)),
                    }
                ),
            )

        # Style
        style = self.basic_style.copy()
        if annotation_type.is_required:
            style.update(self.basic_required_style)
        else:
            style.update(self.basic_optional_style)

        style.update(self.basic_required_style)
        style.update(
            **{
                # Forced styles
                "alignment": Alignment(horizontal="left", wrap_text=True),
            }
        )
        self.set_merged_bordered_style(sheet_obj, cell_obj, **style)

        # Choices
        choices = annotation_type.get_valid_multiplechoice_values()
        if choices:
            self.set_valid_choices(sheet_obj, cell_obj, choices=choices)

        # Current value
        if annotation:
            cell_obj.value = annotation.content
        else:
            cell_obj.value = None

    def write_annotations(self, home_status, eep_program, sheet_obj, **context):
        if not eep_program.required_annotation_types.count():
            return

        start_row = self.set_section_header(
            sheet_obj, home_status, "Annotations", "annotations__overwrite"
        )

        row = start_row
        row += 1

        annotations = {}
        if home_status and home_status.pk:
            annotations = home_status.get_annotations_breakdown()

        for annotation_type in eep_program.required_annotation_types.all().order_by("id"):
            height, cell_obj = self.set_generic_check_list_annotation(
                row=row,
                sheet_obj=sheet_obj,
                named_cell="annotation_{}".format(annotation_type.slug.replace("-", "_")),
                text=annotation_type.name,
                description=annotation_type.description,
                data_type=annotation_type.data_type,
            )
            self.set_annotation_input(
                annotation_type=annotation_type,
                annotation=annotations.get(annotation_type),
                cell_obj=cell_obj,
                sheet_obj=sheet_obj,
            )

            # Style
            box_border_range = "B{start}:N{stop}".format(start=row, stop=row + height - 1)
            self.set_border_range(box_border_range, sheet_obj=sheet_obj)

            row += height

    def write_checklist(self, **context):
        if not context["eep_program"].collection_request_id:
            return self.write_legacy_checklist(**context)
        return self.write_input_collection(**context)

    def set_section_header(
        self,
        sheet_obj,
        home_status=None,
        section_value="Checklist Questions",
        named_result="answers__overwrite",
    ):
        last_row = self.get_last_data_row_number(sheet_obj)
        sheet_obj.row_dimensions[last_row + 1].height = 7.0

        start_row = row = last_row + 2
        sheet_obj.row_dimensions[row].height = 14.0
        sheet_obj.merge_cells("B{row}:K{row}".format(row=start_row))
        cell_obj = sheet_obj["B{row}".format(row=start_row)]
        self.set_style(
            sheet_obj,
            cell_obj,
            **{
                "font": Font(name="Calibri", size=14, bold=True, color=Color(rgb=WHITE_FONT_COLOR)),
                "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color=MEDIUM_GREY),
                "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                "border": Border(
                    bottom=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                    top=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                    left=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                    right=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                ),
            },
        )
        cell_obj.value = section_value

        sheet_obj.merge_cells("L{row}:M{row}".format(row=start_row))
        cell_obj = sheet_obj["L{row}".format(row=start_row)]
        self.set_style(
            sheet_obj,
            cell_obj,
            **{
                "font": Font(
                    name="Calibri", size=12, italic=True, color=Color(rgb=WHITE_FONT_COLOR)
                ),
                "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color=MEDIUM_GREY),
                "alignment": Alignment(horizontal="right", vertical="center", wrap_text=True),
                "border": Border(
                    bottom=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                    top=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                    left=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                    right=Side(border_style="thin", color=Color(rgb=MEDIUM_GREY)),
                ),
            },
        )
        cell_obj.value = "Overwrite"

        cell_obj = sheet_obj["N{row}".format(row=start_row)]
        # Style
        style = self.basic_style.copy()
        style.update(self.basic_required_style)
        style.update(
            **{
                # Forced styles
                "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                "border": self.set_thin_bordered_style(color="FF000000"),
            }
        )
        style.update(self.set_thin_bordered_style())
        self.set_style(sheet_obj, cell_obj, **style)

        # Create a named range - Note: this was from a deprecated function create_named_range
        defn = DefinedName(name=named_result)
        defn.value = "{0}!{1}".format(quote_sheetname(sheet_obj.title), f"$N${row}")
        self.workbook.defined_names.append(defn)

        self.set_valid_choices(sheet_obj, cell_obj, choices=["No", "Yes", "New"], formula=None)
        cell_obj.value = "No" if home_status.id else "Yes"

        box_border_range = f"B{row}:N{row}"
        self.set_border_range(box_border_range, sheet_obj=sheet_obj)

        return row

    def write_input_collection(
        self, home_status, collection_request, sheet_obj, row=None, **context
    ):
        if not hasattr(self, "user"):
            return

        # Init
        collection_context = {
            "user": self.user,
        }
        collector = ExcelChecklistCollector(collection_request, **collection_context)
        collector.context["user_role"] = collector.get_user_role(self.user)

        if row is None:
            row = self.set_section_header(
                sheet_obj,
                home_status,
                section_value="Checklist Questions",
                named_result="answers__overwrite",
            )
        start_row = row
        row += 1

        # Write checklist
        queryset = collector.get_xls_instruments()
        group_info = collector.get_breakdown(queryset, "group_id")
        for group_id, instruments in group_info.items():
            if group_id not in ["default", None]:
                self.merge_and_set(
                    "B{row}:N{row}".format(row=row),
                    group_id,
                    **dict(
                        self.basic_style,
                        **{
                            "style": "header",
                            "sheet_obj": sheet_obj,
                        },
                    ),
                )
                row += 1
            for instrument in instruments:
                if instrument.measure_id == "insulation-company":
                    if context.get("eep_program") and context["eep_program"].slug == "eto-2020-qa":
                        continue
                kwargs = {
                    "collector": collector,
                    "instrument_obj": instrument,
                    "collected_input": collector.get_inputs(instrument).first(),
                    "sheet_obj": sheet_obj,
                }
                row += self.set_instrument(row=row, long=True, **kwargs)

                # Style
                box_border_range = "B{start}:N{stop}".format(start=start_row, stop=row - 1)
                self.set_border_range(box_border_range, sheet_obj=sheet_obj)

        return row

    def set_generic_check_list_annotation(
        self, named_cell, text, description, data_type, sheet_obj, row, long=True
    ):
        height = 2
        question_range = "B{row}:L{row}".format(row=row + 0)
        input_range = "M{row}:N{row}".format(row=row + 0)
        input_ref = "$M${row}".format(row=row + 0)
        description_range = "B{row}:L{row}".format(row=row + 1)
        type_range = "M{row}:N{row}".format(row=row + 1)

        if long:
            height = 3
            question_range = "B{row}:N{row}".format(row=row + 0)
            description_range = "B{row}:L{row}".format(row=row + 1)
            type_range = "M{row}:N{row}".format(row=row + 1)
            input_range = "B{row}:N{row}".format(row=row + 2)
            input_ref = "$B${row}".format(row=row + 2)

        input_name = named_cell

        self.merge_and_set(
            question_range,
            value=text,
            **{
                "sheet_obj": sheet_obj,
            },
        )
        self.merge_and_set(
            description_range,
            value=description or None,
            **dict(
                self.basic_style,
                **{
                    "sheet_obj": sheet_obj,
                    "font": Font(
                        name="Calibri", size=8, italic=True, color=Color(rgb=DARK_GREY_FONT_COLOR)
                    ),
                },
            ),
        )
        self.merge_and_set(
            type_range,
            value=data_type,
            **dict(
                self.basic_style,
                **{
                    "sheet_obj": sheet_obj,
                    "font": Font(
                        name="Calibri", size=8, italic=False, color=Color(rgb=GREY_FONT_COLOR)
                    ),
                    "alignment": Alignment(horizontal="right", wrap_text=False),
                },
            ),
        )

        # # Prototype of using a simple getter for collection data, too
        # rendered_data = collector.get_data_display(instrument_obj, measure)
        # self.merge_and_set(input_range, value=rendered_data, **dict(self.basic_style, **{
        #     'sheet_obj': sheet_obj,
        #     'alignment': Alignment(horizontal="left", wrap_text=True),
        # }))

        # Input area
        sheet_obj.merge_cells(input_range)

        # Create a named range - Note: this was from a deprecated function create_named_range
        defn = DefinedName(name=input_name)
        defn.value = "{0}!{1}".format(quote_sheetname(sheet_obj.title), input_ref)
        self.workbook.defined_names.append(defn)

        cell_obj = sheet_obj[input_range.split(":")[0]]

        # Record for external uses
        if hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)

        return height, cell_obj

    def set_instrument(self, instrument_obj, row, long=True, **context):
        height, cell_obj = self.set_generic_check_list_annotation(
            row=row,
            sheet_obj=context["sheet_obj"],
            named_cell="measure_{}".format(instrument_obj.measure_id.replace("-", "_")),
            text=instrument_obj.text,
            description=instrument_obj.description,
            data_type=instrument_obj.type_id,
            long=long,
        )
        self.set_collected_input(instrument_obj=instrument_obj, cell_obj=cell_obj, **context)
        return height

    def set_collected_input(self, instrument_obj, collected_input, collector, cell_obj, **context):
        sheet_obj = context["sheet_obj"]

        # Validation
        info = collector.get_xls_validation_info(instrument_obj)
        if info:
            type_name = info["type"].capitalize()
            self.add_validation(
                sheet_obj,
                cell_obj,
                DataValidation(
                    **{
                        "type": info["type"],
                        "operator": info.get("operator"),
                        "formula1": info.get("formula1"),
                        "formula2": info.get("formula2"),
                        "errorTitle": info.get("error_title", "Invalid {}".format(type_name)),
                        "error": info.get("error", "Entry must be a {}".format(type_name)),
                    }
                ),
            )

        # Style
        style = self.basic_style.copy()
        if instrument_obj.response_policy.required:
            style.update(self.basic_required_style)
        else:
            style.update(self.basic_optional_style)
        style.update(
            **{
                # Forced styles
                "alignment": Alignment(horizontal="left", wrap_text=True),
            }
        )
        self.set_merged_bordered_style(sheet_obj, cell_obj, **style)

        # Choices
        choices = collector.get_instrument_choices(instrument_obj)
        if choices:
            self.set_valid_choices(sheet_obj, cell_obj, choices=choices)

        # Current value
        if collected_input:
            cell_obj.value = collector.get_data_display(instrument_obj)
            comment = collected_input.data.get("comment")
            if comment:
                cell_obj.comment = Comment(comment, collected_input.user.get_full_name())
        else:
            cell_obj.value = None

    def write_legacy_checklist(self, home, home_status, eep_program, sheet_obj, **context):
        if not hasattr(self, "user"):
            return

        if not home_status and not home:
            return

        from axis.checklist.models import Question, Answer, CheckList

        if home_status.id:
            try:
                questions = Question.objects.filter_by_home_status(home_status)
                answers = Answer.objects.filter_by_home_status(home_status)
            except FieldDoesNotExist as err:
                # This appears to be triggered only when running from command line.  I "think" this has to do with this
                # 1.7 bug - https://code.djangoproject.com/ticket/24513
                self.log.warning("Bug handled - %r", err)
                from axis.eep_program.models import EEPProgram

                RequiredChecklistThrough = EEPProgram._meta.get_field(
                    "required_checklists"
                ).rel.through
                checklist_ids = RequiredChecklistThrough.objects.filter(
                    eepprogram_id=home_status.eep_program.id
                )
                checklist_ids = checklist_ids.values_list("checklist_id", flat=True)

                QuestionThrough = CheckList._meta.get_field("questions").rel.through
                questions = QuestionThrough.objects.filter(checklist_id__in=checklist_ids)
                questions = Question.objects.filter(
                    id__in=questions.values_list("question_id", flat=True)
                )

                answers = Answer.objects.filter(home=home_status.home)
        elif eep_program.id:
            try:
                questions = Question.objects.filter_by_eep(eep_program)
                answers = Answer.objects.none()
            except FieldDoesNotExist as err:
                # This appears to be triggered only when running from command line.  I "think" this has to do with this
                # 1.7 bug - https://code.djangoproject.com/ticket/24513
                self.log.warning("Bug handled - %r", err)
                from axis.eep_program.models import EEPProgram

                RequiredChecklistThrough = EEPProgram._meta.get_field(
                    "required_checklists"
                ).rel.through
                checklist_ids = RequiredChecklistThrough.objects.filter(
                    eepprogram_id=eep_program.id
                )
                checklist_ids = checklist_ids.values_list("checklist_id", flat=True)

                QuestionThrough = CheckList._meta.get_field("questions").rel.through
                questions = QuestionThrough.objects.filter(checklist_id__in=checklist_ids)
                questions = Question.objects.filter(
                    id__in=questions.values_list("question_id", flat=True)
                )
                answers = Answer.objects.none()
        elif home.id:
            questions = Question.objects.filter_by_home(home)
            answers = Answer.objects.filter_by_home(home)

        if not questions.count():
            return

        start_row = row = 35

        overwrite = "No" if home_status.id else "Yes"
        self.set_answer__overwrite(
            sheet_obj=sheet_obj, cell_label="answers__overwrite", value=overwrite, row=row
        )

        row += 1
        for row_idx, question in enumerate(questions):
            try:
                answer = answers.get(question_id=question.id)
            except Answer.DoesNotExist:
                answer = None
            self.set_question(sheet_obj=sheet_obj, question_obj=question, row=row + row_idx * 2)
            self.set_answer(
                sheet_obj=sheet_obj, question_obj=question, answer_obj=answer, row=row + row_idx * 2
            )

        row += row_idx * 2 + 1
        self.set_border_range(
            "B{start_row}:N{row}".format(start_row=start_row, row=row), sheet_obj=sheet_obj
        )

    def set_answer(self, sheet_obj, question_obj, answer_obj, row):
        answer_column = "K{row}:N{row_plus}"
        base_style = self.basic_style.copy()
        base_style["alignment"] = Alignment(horizontal="left", wrap_text=True)
        sheet_obj.merge_cells(answer_column.format(row=row, row_plus=row + 1))

        # Create a named range - Note: this was from a deprecated function create_named_range
        defn = DefinedName(name=f"question_{question_obj.pk}")
        defn.value = "{0}!{1}".format(quote_sheetname(sheet_obj.title), f"$K${row}")
        self.workbook.defined_names.append(defn)

        cell_obj = sheet_obj["K{row}".format(row=row)]

        style = base_style.copy()
        if question_obj.is_optional:
            style.update(self.basic_optional_style)
        else:
            style.update(self.basic_required_style)
        self.set_merged_bordered_style(sheet_obj, cell_obj, **style)

        value = answer_obj.answer if answer_obj else None

        if question_obj.type == "multiple-choice":
            try:
                choices = question_obj.question_choice.all()
            except FieldDoesNotExist as err:
                self.log.warning("Bug handled - %r", err)
                from axis.checklist.models import Question, QuestionChoice

                ChoiceThrough = Question._meta.get_field("question_choice").rel.through
                choice_ids = ChoiceThrough.objects.filter(question_id=question_obj.id).values_list(
                    "questionchoice_id", flat=True
                )
                choices = QuestionChoice.objects.filter(id__in=choice_ids)
            self.set_valid_choices(sheet_obj, cell_obj, choices=[x.choice for x in choices])
        elif question_obj.type == "open":
            data_validation = DataValidation(
                type="textLength",
                operator="lessThan",
                formula1="256",
                error="Entry must be less than 255 characters.",
            )
            self.add_validation(sheet_obj, cell_obj, data_validation)
        elif question_obj.type == "date":
            data_validation = DataValidation(
                type="date",
                operator="between",
                formula1="=TODAY()-365",
                ErrorTitle="Invalid date",
                error="Date must be a year from today",
                formula2="=TODAY()+365",
            )
            self.add_validation(sheet_obj, cell_obj, data_validation)
            cell_obj.number_format = "mm/dd/yyyy"
        elif question_obj.type == "datetime":
            data_validation = DataValidation(
                type="date",
                operator="between",
                formula1="=NOW()-365",
                ErrorTitle="Invalid datetime",
                error="Datetime must be a year from today",
                formula2="=NOW()+365",
            )
            self.add_validation(sheet_obj, cell_obj, data_validation)
            cell_obj.number_format = "mm/dd/yyyy hh:mm"
        elif question_obj.type in ["integer", "int", "float"]:
            vtype = "decimal" if question_obj.type == "float" else "whole"
            operator, formula1, formula2, error, ErrorTitle = None, None, None, None, None
            if answer_obj:
                value = float(value) if question_obj.type == "float" else int(value)
                ErrorTitle = "Invalid {}".format(vtype.capitalize())
                error = "Entry must be a {}".format(question_obj.get_type_display())
            if question_obj.minimum_value is not None and question_obj.maximum_value is None:
                operator = "greaterThanOrEqual"
                formula1 = question_obj.minimum_value
                ErrorTitle = "Invalid {}".format(vtype.capitalize())
                val1 = (
                    question_obj.minimum_value
                    if question_obj.type == "float"
                    else int(question_obj.minimum_value)
                )
                error = "{} must be greater than {}".format(question_obj.get_type_display(), val1)

            elif question_obj.maximum_value is not None and question_obj.minimum_value is None:
                operator = "lessThanOrEqual"
                formula1 = question_obj.maximum_value
                ErrorTitle = "Invalid {}".format(vtype.capitalize())
                val1 = (
                    question_obj.maximum_value
                    if question_obj.type == "float"
                    else int(question_obj.maximum_value)
                )
                error = "{} must be less than {}".format(question_obj.get_type_display(), val1)

            elif question_obj.maximum_value is not None and question_obj.minimum_value is not None:
                operator = "between"
                formula1 = question_obj.minimum_value
                formula2 = question_obj.maximum_value
                ErrorTitle = "Invalid {}".format(question_obj.get_type_display())
                val1 = (
                    question_obj.minimum_value
                    if question_obj.type == "float"
                    else int(question_obj.minimum_value)
                )
                val2 = (
                    question_obj.maximum_value
                    if question_obj.type == "float"
                    else int(question_obj.maximum_value)
                )
                error = "{} must be between {}-{}".format(vtype.capitalize(), val1, val2)
            data_validation = DataValidation(
                type=vtype,
                operator=operator,
                formula1=formula1,
                formula2=formula2,
                errorTitle=ErrorTitle,
                error=error,
            )
            self.add_validation(sheet_obj, cell_obj, data_validation)
        elif question_obj.type == "csv":
            pass

        elif question_obj.type == "kvfloatcsv":
            pass

        cell_obj.value = value
        if answer_obj and answer_obj.comment:
            cell_obj.comment = Comment(answer_obj.comment, answer_obj.user.get_full_name())

        if hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)

        self.set_border_range(
            "B{row}:N{row_plus}".format(row=row, row_plus=row + 1), sheet_obj=sheet_obj
        )

    def set_question(self, sheet_obj, question_obj, row):
        question_column = "B{row}:J{row}"
        base_style = self.basic_style.copy()
        base_style["alignment"] = Alignment(horizontal="left", wrap_text=True)

        sheet_obj.merge_cells(question_column.format(row=row))
        cell_obj = sheet_obj["B{row}".format(row=row)]
        self.set_merged_bordered_style(sheet_obj, cell_obj, **base_style)
        cell_obj.value = question_obj.question.strip()

        description_column = "B{row}:G{row}"
        sheet_obj.merge_cells(description_column.format(row=row + 1))
        cell_obj = sheet_obj["B{row}".format(row=row + 1)]
        base_style["font"] = Font(
            name="Calibri", size=8, italic=True, color=Color(rgb=GREY_FONT_COLOR)
        )
        self.set_merged_bordered_style(sheet_obj, cell_obj, **base_style)
        cell_obj.value = question_obj.description.strip() if question_obj.description else None

        type_column = "H{row}:J{row}"
        sheet_obj.merge_cells(type_column.format(row=row + 1))
        cell_obj = sheet_obj["H{row}".format(row=row + 1)]
        base_style["alignment"] = Alignment(horizontal="right", wrap_text=False)
        base_style["font"] = Font(
            name="Calibri", size=8, italic=False, color=Color(rgb=GREY_FONT_COLOR)
        )
        self.set_merged_bordered_style(sheet_obj, cell_obj, **base_style)
        cell_obj.value = "{}".format(question_obj.get_type_display())

    def write_extra(self, **context):
        if context["eep_program"].slug == "eto-2020-qa":
            self.write_generic_data_eto2020_qa(**context)

    def write_generic_data_eto2020_qa(self, sheet_obj, **context):
        # make measure_insulation_company editable
        cell_obj = self.get_cell_from_label(sheet_obj.title, "measure_insulation_company", None)
        self.set_style(sheet_obj, cell_obj, **self.basic_optional_style)
        if hasattr(self, "editable_cells"):
            self.editable_cells.append(cell_obj.coordinate)
