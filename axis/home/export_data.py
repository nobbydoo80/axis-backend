"""export_data.py: Django home"""

import datetime
import inspect
import itertools
import logging
import numbers
import operator
import pprint
import re
import time
from collections import OrderedDict, namedtuple
from decimal import Decimal
from functools import partial
from itertools import zip_longest

from django.apps import apps
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import FieldDoesNotExist
from django.db.models import Max, Sum, Case, When, F
from django.db.models.functions import Coalesce
from django.db.models.query import QuerySet
from django.http import QueryDict
from django.utils import formats
from django.utils.encoding import force_str
from hashid_field import Hashid
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.comments import Comment
from openpyxl.styles import Font, numbers as excel_numbers
from openpyxl.utils import get_column_letter

from axis.checklist.xls_checklist import XLSChecklist
from axis.customer_eto.utils import get_eto_region_name_for_zipcode
from axis.customer_hirl.models import HIRLProjectRegistration, HIRLProject
from axis.customer_neea.utils import NEEA_BPA_SLUGS
from axis.home.models import EEPProgramHomeStatus
from axis.home.views import HomeStatusReportMixin, HomeStatusView
from axis.invoicing.models import InvoiceItemGroup
from axis.qa.models import Observation, QANote
from axis.qa.models import QARequirement, QAStatus
from axis.qa.state_machine import QAStateMachine

__author__ = "Steven Klass"
__date__ = "1/23/14 9:45 AM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven Klass",
]

log = logging.getLogger(__name__)
customer_hirl_app = apps.get_app_config("customer_hirl")


CellParser = namedtuple(
    "CellParser",
    [
        "attr",
        "pretty_name",
        "clean_method",
        "section",
    ],
)
CellObject = namedtuple(
    "CellObj", ["attr", "pretty_name", "clean_method", "section", "raw_value", "value"]
)

ReportOn = {
    "status": "Energy Efficiency Programs",
    "eep_program": "Programs",
    "ipp": "Incentive Payments",
    "invoicing": "Invoicing",
    "qa": "QA Data",
    "customer_aps": "APS Metersets",
    "simulation_basic": "Simulation Basic",
    "simulation_advanced": "Simulation Advanced",
    "checklist_answers": "Program Questions / Answers",
    "relationships": "Associations",
    "customer_eto": "ETO Data",
    "hes_data": "HES Data",
    "ngbs_data": "NGBS Data",
    "retain_empty": "Retain Empty Columns",
}


class BaseHomeStatusDataDump(object):
    currency_format = excel_numbers.FORMAT_CURRENCY_USD_SIMPLE
    percentage_format = "0.00%"
    number_format = excel_numbers.FORMAT_NUMBER
    float_format = excel_numbers.FORMAT_NUMBER_00
    date_format = "mm/dd/yy"
    datetime_format = "mm/dd/yy h:mm AM/PM"

    def __init__(self, *args, **kwargs):
        self.user_id = kwargs.pop("user_id")
        if self.user_id:
            self.user = get_user_model().objects.get(id=self.user_id)

        self.log = kwargs.get("log")
        if self.log is None:
            self.log = logging.getLogger("BaseHomeStatusDataDump")
        if kwargs.get("loglevel"):
            self.log.setLevel(kwargs.pop("loglevel"))
        self.stored_data_file = None

        self.report_on = kwargs.pop("report_on", None)
        self.home_status_ids = kwargs.get("home_status_ids")

        # This defines the column order..
        if self.report_on is None:
            log.info("Setting default report on")
            self.report_on = [
                "home",
                "status",
                "subdivision",
                "community",
                "relationships",
                "eep_program",
                "ipp",
                "invoicing",
                "sampleset",
                "floorplan",
                "simulation_basic",
                "annotations",
                "checklist_answers",
                "qa",
                "customer_aps",
                "customer_eto",
                "hes_data",
            ]

        else:
            if "status" not in self.report_on:
                if len(self.report_on) and self.report_on[0] == "home":
                    self.report_on = [self.report_on[0], "status"] + self.report_on[1:]
                else:
                    self.report_on = ["status"] + self.report_on

        # Add Calculator values for supers and NEEA when exporting with v2 Program.
        report_on_standard_protocol_calculator = any(
            [
                self.user.is_superuser,
                self.user.company.slug in ("neea",),
                self.user.company.sponsors.filter(slug="neea").exists(),
            ]
        )
        if report_on_standard_protocol_calculator:
            if self.get_queryset().filter(eep_program__slug__in=NEEA_BPA_SLUGS).exists():
                self.report_on.append("neea_standard_protocol_calculator")

        # Make sure Simulation information is always the last to go.
        if "simulation_basic" in self.report_on:
            self.report_on.append(self.report_on.pop(self.report_on.index("simulation_basic")))
        elif "simulation_advanced" in self.report_on:
            self.report_on.append(self.report_on.pop(self.report_on.index("simulation_advanced")))

        self._dataset = None
        self._columns = []
        self.specified_columns = kwargs.get("specified_columns", [])
        self._data = []
        self.max_num = kwargs.get("max_num", None)
        self.reuse_storage = kwargs.get("reuse_storage", True)

        self.default_none = "-"
        self.default_true = "Yes"
        self.default_false = "No"
        self.join_resulting_values = True

        self.output_ids = False

        self.task = kwargs.get("task")
        self.retain_empty = kwargs.get("retain_empty", False)

        self.default_col_width = kwargs.get("default_col_width", 25)
        self.include_home_status_data = True
        self.include_builder_agreement_data = True

    def update_task_progress(
        self, state="STARTED", steps=None, step=0, current=1, total=1, **kwargs
    ):
        """
        Update an asynchronous tasks state. Primarily used for send data to the client.
        :param state: string of custom state to set Task to.
        :param steps: list of processing steps
        :param step: int current step. 0 based
        :param current: int current progress point in total
        :param total: int total number of progress points
        """
        if self.task:
            steps = steps or ["processing", "writing"]

            current_step = steps[step]

            meta = {}
            for s in steps:
                if s == current_step:
                    meta[s] = {"current": current, "total": total}
                elif steps.index(s) > steps.index(current_step):
                    # takes place after current step
                    meta[s] = {"current": 0, "total": 1}
                else:
                    # takes place before current step
                    meta[s] = {"current": 1, "total": 1}

            meta.update(kwargs)

            self.task.update_state(state=state, meta=meta)

    def get_queryset(self, *args, **kwargs):
        if not self.home_status_ids:
            raise NotImplementedError(
                "You need to either override get_queryset or provide "
                "home_status_ids to the constructor."
            )
        if not hasattr(self, "_queryset"):
            qs = EEPProgramHomeStatus.objects.filter(id__in=self.home_status_ids)
            self._queryset = qs

        return self._queryset

    def rec_getattr(self, obj, attr, default=None):
        """Get object's attribute. May use dot notation."""
        if "." not in attr:
            if "()" in attr:
                attr = attr.split("(")[0]
                return getattr(obj, attr, default)()
            return getattr(obj, attr, default)
        else:
            split_attr = attr.split(".")
            return self.rec_getattr(getattr(obj, split_attr[0]), ".".join(split_attr[1:]), default)

    def set_cell_integer_value(self, cell, value, style_args):
        try:
            cell.value = int(value)
            style_args["number_format"] = self.number_format
        except ValueError:
            self.set_cell_explicit_value(cell, value)

    def set_cell_float_value(self, cell, value, style_args):
        try:
            cell.value = float(value)
            style_args["number_format"] = self.float_format
        except ValueError:
            self.set_cell_explicit_value(cell, value)

    def set_cell_decimal_value(self, cell, value, style_args):
        try:
            cell.value = str(value)
            exps = -1 * value.as_tuple().exponent
            style_args["number_format"] = "0." + "0" * exps
        except ValueError:
            self.set_cell_explicit_value(cell, value)

    def set_cell_currency_value(self, cell, value, style_args):
        try:
            cell.value = float(value[1:])
            style_args["number_format"] = self.currency_format
        except ValueError:
            self.set_cell_explicit_value(cell, value)

    def set_cell_percentage_value(self, cell, value, style_args):
        try:
            cell.value = float(value[:-1]) / 100
            style_args["number_format"] = self.percentage_format
        except ValueError:
            self.set_cell_explicit_value(cell, value)

    def set_cell_date_value(self, cell, value, style_args):
        cell.value = value
        style_args["number_format"] = self.date_format

    def set_cell_datetime_value(self, cell, value, style_args):
        cell.value = value
        style_args["number_format"] = self.datetime_format

    def set_cell_explicit_value(self, cell, value):
        cell.value = value

    def get_formatted_datetime(self, value):
        """Return the string components - as TZ Aware issue exist when setting the format the
        writer"""
        tz = self.user.timezone_preference
        try:
            return formats.date_format(value.astimezone(tz), "SHORT_DATETIME_FORMAT")
        except AttributeError:
            return self.default_none

    def get_formatted_date(self, value):
        """Return the string components - as TZ Aware issue exist when setting the format the
        writer"""
        if not value:
            return self.default_none
        if isinstance(value, datetime.datetime):
            tz = self.user.timezone_preference
            return formats.date_format(value.astimezone(tz).date(), "SHORT_DATE_FORMAT")

        # Plain date obj
        return formats.date_format(value, "SHORT_DATE_FORMAT")

    def get_formatted_boolean(self, value):
        return self.default_true if value else self.default_false

    def get_formatted_nullboolean(self, value):
        if value is None:
            return self.default_none
        return self.default_true if value else self.default_false

    def get_formatted_decimal(self, value):
        return "{}".format(value) if value else self.default_none

    def get_field_and_clean_method(self, ModelObject, field_name):
        """Returns a field and a clean method for the field"""

        try:
            field = ModelObject._meta.get_field(field_name.split("__")[0])
        except FieldDoesNotExist:
            raise

        if "__" not in field_name:
            if field.__class__.__name__ == "DateTimeField":
                return field, "get_formatted_datetime"
            elif field.__class__.__name__ == "DateField":
                return field, "get_formatted_date"
            elif field.__class__.__name__ == "BooleanField":
                return field, "get_formatted_boolean"
            elif field.__class__.__name__ == "NullBooleanField":
                return field, "get_formatted_nullboolean"
            elif field.__class__.__name__ == "DecimalField":
                return field, "get_formatted_decimal"
            elif field.__class__.__name__ == "ForeignKey":
                log.warning("We need a real field name for %s to dead-end into", field_name)
            elif field.__class__.__name__ not in [
                "AutoField",
                "CharField",
                "IntegerField",
                "UUIDField",
                "FloatField",
                "TextField",
                "USStateField",
                "PositiveIntegerField",
                "StateField",
                "URLField",
                "SlugField",
                "ManyToOneRel",
                "BigAutoField",
            ]:
                log.info(
                    "We might not know how to handle %s for field %s",
                    field_name,
                    field.__class__.__name__,
                )
            return field, None
        elif field.related_model:
            return self.get_field_and_clean_method(
                field.related_model, "__".join(field_name.split("__")[1:])
            )
        else:
            raise ValueError(
                "{0} ({1}) is not a valid field.".format(
                    ModelObject.__name__, field.__class__.__name__
                )
            )

    def assign_basic(
        self,
        ModelObject,
        include=None,
        exclude=None,
        section="home",
        drop_prefix=None,
        replace_dict=None,
        clean_dict=None,
    ):
        """This will return a generic mapping of attributes to a pretty name with a default method"""

        fields = [x.name for x in ModelObject._meta.fields]

        _err = "%s must be defined as list you gave %s"
        if include:
            try:
                _foo = iter(include)
            except TypeError:
                raise TypeError(_err % ("include", type(include)))
            fields = include
        if exclude:
            try:
                _foo = iter(exclude)
            except TypeError:
                raise TypeError(_err % ("exclude", type(exclude)))
            fields = [f for f in fields if f not in exclude]

        if all([include, exclude]) is None:
            fields = [
                f.name
                for f in ModelObject._meta.fields
                if f.__class__.__name__ not in ["ForeignKey"]
            ]

        value_tuples = []
        for field_name in fields:
            if field_name.split("__")[0] not in [f.name for f in ModelObject._meta.get_fields()]:
                continue

            field, clean_function = self.get_field_and_clean_method(ModelObject, field_name)
            if clean_dict and field_name in clean_dict:
                clean_function = clean_dict[field_name]

            subnames = []
            if replace_dict and field_name in replace_dict.keys():
                subnames = [replace_dict[field_name]]
            else:
                if len(field_name.split("__")) > 1:
                    _fname = field_name
                    if drop_prefix and not field_name.endswith("__id"):
                        _fname = re.sub(drop_prefix, "", _fname)
                    for x in _fname.split("__")[:-1]:
                        if replace_dict and x in replace_dict:
                            subnames.append(replace_dict[x])
                        else:
                            subnames.append(x.capitalize())
                if field_name == "id" or field_name.endswith("__id"):
                    subnames.append("ID")
                else:
                    subnames.append(field.verbose_name.capitalize())

            pretty_name = " ".join(subnames)
            value_tuples.append(CellParser(field_name, pretty_name, clean_function, section))

        return value_tuples

    def assign_json(
        self,
        ModelObject,
        include=None,
        section="home",
        drop_prefix=None,
        replace_dict=None,
        clean_dict=None,
    ):
        """This will return a generic mapping of attributes to a pretty name with a default method"""

        assert isinstance(include, (list, tuple)), "include must be defined as a list"
        fields = include

        value_tuples = []
        for field_name in fields:
            clean_function = None
            if clean_dict and field_name in clean_dict:
                clean_function = clean_dict[field_name]

            subnames = []
            if replace_dict and field_name in replace_dict.keys():
                subnames = [replace_dict[field_name]]
            else:
                if len(field_name.split("__")) > 1:
                    _fname = field_name
                    if drop_prefix and not field_name.endswith("__id"):
                        _fname = re.sub(drop_prefix, "", _fname)
                    for x in _fname.split("__")[:-1]:
                        if replace_dict and x in replace_dict:
                            subnames.append(replace_dict[x])
                        else:
                            subnames.append(x.capitalize())
                if field_name == "id" or field_name.endswith("__id"):
                    subnames.append("ID")
                else:
                    subnames.append(field_name)  # FIXME: 'camelCase' -> 'Camel case'

            pretty_name = " ".join(subnames)
            value_tuples.append(CellParser(field_name, pretty_name, clean_function, section))

        return value_tuples

    def merge_results(self, baseline_results, structure, key="id"):
        """
        baseline_results = [[1,2,3],[4,5,6],[7,8,9]]
        structure = [
            (attr, pretty_name, clean_method, section),
            (attr, pretty_name, clean_method, section),
            (attr, pretty_name, clean_method, section)]

        output:[
            [ (attr, pretty_name, clean_method, section, raw_value, value=1),
              (attr, pretty_name, clean_method, section, raw_value, value=2),
              (attr, pretty_name, clean_method, section, raw_value, value=3)],

            [ (attr, pretty_name, clean_method, section, raw_value, value=4),
              (attr, pretty_name, clean_method, section, raw_value, value=5),
              (attr, pretty_name, clean_method, section, raw_value, value=6)],

            [ (attr, pretty_name, clean_method, section, raw_value, value=7),
              (attr, pretty_name, clean_method, section, raw_value, value=8),
              (attr, pretty_name, clean_method, section, raw_value, value=9)],

        """
        # log.debug("Received {} results".format(len(baseline_results)))
        results = OrderedDict()
        for result_vals in baseline_results:
            # log.debug(" - Sub result conains {} entries".format(len(result_vals)))
            # Get the index number from the structure for the index, then get the value
            key_idx = structure.index(next((item for item in structure if item.attr == key)))
            key_val = result_vals[key_idx]
            if key_val not in results.keys():
                # This will add two place holders for lists of values at the end.
                results[key_val] = [list(v) + [[], []] for v in structure[:]]
                # log.debug(" - Added new key {}".format(key_val))
            result = results[key_val]

            assert len(result_vals) == len(result), "Mismatch in the number of values vs results.."

            for idx, raw_value in enumerate(result_vals):
                clean_value = raw_value
                result[idx][-2].append(raw_value)
                clean_method = result[idx][2]
                # log.debug("  - Inserting raw value {} into index {}".format(raw_value, idx))
                if clean_method:
                    if callable(clean_method):
                        clean_value = clean_method(raw_value)
                    else:
                        clean_value = getattr(self, clean_method)(raw_value)
                clean_value = clean_value if clean_value is not None else self.default_none
                # if clean_value != raw_value:
                #     log.debug("   - Adding clean value {} into index {}".format(clean_value, idx))
                result[idx][-1].append(clean_value)

        # Now though these into our final structure.
        # log.debug("Building Final Structure")
        for dict_key, source_result_set in results.items():
            for idx, result in enumerate(source_result_set):
                # Handle the primary key - it should only have one..
                if result[0] == key:
                    assert (
                        len(set(result[-2])) == 1
                    ), "Primary Key Raw result is not unique {}".format(result[-2])
                    assert len(set(result[-1])) == 1, "Primary Key result is not unique {}".format(
                        result[-1]
                    )
                    result[-2] = result[-2][0]
                    result[-1] = result[-1][0]
                    # log.debug(" - Trimming Primary Key - {}".format(result[-1]))
                elif len(result[-1]) > 1 and self.join_resulting_values:
                    result[-1] = ", ".join(["{}".format(v) for v in result[-1]])
                    # log.debug(" - Multiple values found - {}".format(result[-1]))
                elif len(result[-1]) == 1:
                    result[-1] = result[-1][0]
                    # log.debug(" - Single values found - {}".format(result[-1]))
                else:
                    result[-1] = self.default_none
                    # log.debug(" - No value found - {}".format(result[-1]))
                source_result_set[idx] = CellObject(*result)
        return list(results.values())

    def get_home_data(self):
        include_eto_region = (
            self.user.company.slug in ("eto", "peci", "csg-qa") or self.user.is_superuser
        )

        display_raw_addresses = False
        if getattr(self.user, "company", None) and self.user.company.display_raw_addresses:
            display_raw_addresses = True

        def get_climate_zone(ident):
            if not hasattr(self, "climate_zone_dict"):
                log.debug("Looking up Climate Zone data")
                from axis.geographic.models import ClimateZone

                self.climate_zone_dict = {}
                cz_data = ClimateZone.objects.all().values_list("id", "zone", "moisture_regime")
                for _id, zone, moisture in cz_data:
                    self.climate_zone_dict[_id] = "{}{}".format(zone, moisture if moisture else "")
            return self.climate_zone_dict.get(ident, self.default_none)

        def map_zipcode_to_eto_region(zipcode):
            return get_eto_region_name_for_zipcode(zipcode) or self.default_none

        field_aliases = {
            "eto_region": "_zipcode",
        }

        section_name = "home"

        structure = [
            CellParser("id", "Status ID", None, section_name),
            CellParser("home__id", "AXIS ID", None, section_name),
            CellParser("home__lot_number", "Lot #", None, section_name),
            CellParser("_street_line1", "Street Line 1", None, section_name),
            CellParser("_street_line2", "Street Line 2", None, section_name),
            CellParser(
                "home__is_multi_family", "Multi-Family", self.get_formatted_boolean, section_name
            ),
            CellParser("home__alt_name", "Alt ID", None, section_name),
            CellParser("_city_name", "City", None, "home"),
            CellParser("_city_latitude", "City Latitude", None, section_name),
            CellParser("_city_longitude", "City Longitude", None, section_name),
            CellParser(
                "home__subdivision__latitude",
                "Subdivision/MF Development Latitude",
                None,
                section_name,
            ),
            CellParser(
                "home__subdivision__longitude",
                "Subdivision/MF Development Longitude",
                None,
                section_name,
            ),
            CellParser("_state", "State", None, section_name),
            CellParser("_zipcode", "ZIP Code", None, section_name),
            CellParser("home__metro__name", "Metro", None, section_name),
            CellParser("_county_name", "County", None, section_name),
            CellParser("_county_latitude", "County Latitude", None, section_name),
            CellParser("_county_longitude", "County Longitude", None, section_name),
            CellParser("home__latitude", "Latitude", None, section_name),
            CellParser("home__longitude", "Longitude", None, section_name),
            CellParser("home__confirmed_address", "Confirmed Address", None, section_name),
            CellParser("home__address_override", "Address Override", None, section_name),
            CellParser("home__climate_zone__id", "Climate Zone", get_climate_zone, section_name),
            CellParser(
                "home__created_date", "Created Date", self.get_formatted_datetime, section_name
            ),
        ]

        if include_eto_region:
            eto_region_cell = CellParser(
                "eto_region", "ETO Region", map_zipcode_to_eto_region, "home"
            )
            structure.insert(17, eto_region_cell)  # insert just after zipcode

        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]

        if display_raw_addresses:
            objects = (
                self.get_queryset()
                .select_related(*select_related)
                .annotate(
                    _street_line1=Coalesce(
                        F("home__geocode_response__geocode__raw_street_line1"),
                        F("home__street_line1"),
                    ),
                    _street_line2=Coalesce(
                        F("home__geocode_response__geocode__raw_street_line2"),
                        F("home__street_line2"),
                    ),
                    _city_name=Coalesce(
                        F("home__geocode_response__geocode__raw_city__name"),
                        F("home__city__name"),
                    ),
                    _city_latitude=Coalesce(
                        F("home__geocode_response__geocode__raw_city__latitude"),
                        F("home__city__latitude"),
                    ),
                    _city_longitude=Coalesce(
                        F("home__geocode_response__geocode__raw_city__longitude"),
                        F("home__city__longitude"),
                    ),
                    _state=Coalesce(
                        F("home__geocode_response__geocode__raw_city__county__state"),
                        F("home__state"),
                    ),
                    _zipcode=Coalesce(
                        F("home__geocode_response__geocode__raw_zipcode"),
                        F("home__zipcode"),
                    ),
                    _county_name=Coalesce(
                        F("home__geocode_response__geocode__raw_city__county__name"),
                        F("home__county__name"),
                    ),
                    _county_latitude=Coalesce(
                        F("home__geocode_response__geocode__raw_city__county__latitude"),
                        F("home__county__latitude"),
                    ),
                    _county_longitude=Coalesce(
                        F("home__geocode_response__geocode__raw_city__county__longitude"),
                        F("home__county__longitude"),
                    ),
                )
            )
        else:
            objects = (
                self.get_queryset()
                .select_related(*select_related)
                .annotate(
                    _street_line1=F("home__street_line1"),
                    _street_line2=F("home__street_line2"),
                    _city_name=F("home__city__name"),
                    _city_latitude=F("home__city__latitude"),
                    _city_longitude=F("home__city__longitude"),
                    _state=F("home__state"),
                    _zipcode=F("home__zipcode"),
                    _county_name=F("home__county__name"),
                    _county_latitude=F("home__county__latitude"),
                    _county_longitude=F("home__county__longitude"),
                )
            )

        attrs_order = [x.attr for x in structure if x.attr not in field_aliases]
        values = list(objects.values_list(*attrs_order))
        if include_eto_region:
            for i, row in enumerate(values):
                for alias, field_name in field_aliases.items():
                    value_i = attrs_order.index(field_name)
                    row = list(row)
                    values[i] = row
                    row.insert(17, row[value_i])

        data = self.merge_results(values, structure)

        return data

    def get_status_data(self):
        from axis.home.state_machine import HomeStatusStateMachine

        replace_dict = OrderedDict(
            [
                ("id", "Status ID"),
                ("state", "Project Status"),
                ("rater_of_record__id", "Rater of Record"),
                ("energy_modeler__id", "Energy Modeler"),
                ("field_inspectors__id", "Field Inspectors"),
                ("customer_hirl_rough_verifier__id", "Rough Verifier"),
                ("customer_hirl_final_verifier__id", "Final Verifier"),
                ("pct_complete", "% Complete"),
                ("samplesethomestatus__is_test_home", "Rating Type"),
                ("certification_date", "Certification Date"),
            ]
        )

        # remove unnecessary and duplicate columns
        if "ngbs_data" in self.report_on:
            replace_dict = OrderedDict(
                [
                    ("id", "Status ID"),
                    ("state", "Project Status"),
                    ("customer_hirl_rough_verifier__id", "Rough Verifier"),
                    ("customer_hirl_final_verifier__id", "Final Verifier"),
                    ("certification_date", "Certification Date"),
                ]
            )

        def get_state_name(ident):
            return dict(HomeStatusStateMachine.get_state_choices())[ident]

        def get_pct_complete(ident):
            return "{}%".format(ident)

        def get_rating_type(ident):
            if ident:
                return "Sample Test"
            elif ident is None:
                return "Confirmed"
            return "Sampled"

        def get_user_full_name(user_id):
            if user_id:
                if isinstance(user_id, (str, int)):
                    user_id = [user_id]
                vals = [str(x) for x in get_user_model().objects.filter(id__in=user_id)]
                return ", ".join(vals)
            return "-"

        clean_dict = {
            "state": get_state_name,
            "pct_complete": get_pct_complete,
            "rater_of_record__id": get_user_full_name,
            "field_inspectors__id": get_user_full_name,
            "energy_modeler__id": get_user_full_name,
            "customer_hirl_rough_verifier__id": get_user_full_name,
            "customer_hirl_final_verifier__id": get_user_full_name,
            "samplesethomestatus__is_test_home": get_rating_type,
        }

        objects = self.get_queryset().distinct()

        # Only add this column if there are homes that can be submitted to the registry
        if objects.filter(eep_program__require_input_data=True).exists():
            key = "floorplan__simulation__project__resnet_registry_identification"
            replace_dict[key] = "RESNET Registry ID"

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="status",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        # For home statuses we only have one entry.
        results = []
        for _row in data:
            items = []
            for item in _row:
                if isinstance(item.raw_value, list):
                    item = item._replace(
                        raw_value=list(set(item.raw_value)),
                        value=", ".join(list(set(item.value.split(", ")))),
                    )
                items.append(item)
            results.append(items)
        return results

    def get_historical_certifier_data(self):
        """This will attempt to get the certifier by one of three methods

        1. The person who transitioned the state model
        2. The historical user on eepprogram home status model
        3. The company of the
        """

        target_ids = (
            self.get_queryset()
            .filter(certification_date__isnull=False)
            .values_list("id", flat=True)
        )
        section_name = "state"
        result = OrderedDict()

        StateLog = EEPProgramHomeStatus._state_log_model
        objects = StateLog.objects.filter(id__in=target_ids, to_state="complete")
        objects = objects.select_related("user", "user__company")
        objects = objects.values_list(
            "on_id", "id", "user__first_name", "user__last_name", "user__company__name"
        )
        t_found = []
        for _id, _h, user_f, user_l, comp in objects.order_by("id"):
            if _id in t_found:
                continue
            t_found.append(_id)
            if comp:
                result[_id] = "{} {}".format(user_f, user_l), comp

        state_log = len(result.keys())
        log.debug("State Log found %s/%s", state_log, target_ids.count())

        remaining = list(set(target_ids) - set(result.keys()))
        if remaining:
            History = EEPProgramHomeStatus.history.model
            history = History.objects.filter(
                id__in=remaining, certification_date__isnull=True, state="complete"
            )
            history = history.select_related("user", "user__company")
            history = history.values_list(
                "id",
                "certification_date",
                "history_id",
                "history_user__first_name",
                "history_user__last_name",
                "history_user__company__name",
            )
            t_found = []
            for _id, cd, _h, user_f, user_l, comp in history.order_by("history_id"):
                if _id in t_found:
                    continue
                t_found.append(_id)
                if comp:
                    result[_id] = "{} {}".format(user_f, user_l), comp

            hist_found = len(result.keys()) - state_log
            log.debug("History found %s/%s", hist_found, target_ids.count())

        remaining = list(set(target_ids) - set(result.keys()))
        if remaining:
            vals = self.get_queryset().filter(id__in=remaining)
            vals = vals.select_related("company")
            vals = vals.values_list("id", "company__company_type", "company__name")
            reldata = self.get_relationships_data()
            for _id in remaining:
                _, co_type, company = next((x for x in vals if x[0] == _id))
                if co_type != "rater":
                    result[_id] = self.default_none, company
                rel = next((x for x in reldata if x[0][-1] == _id), [])
                provider = next((x for x in rel if x[0] == "provider"), None)
                if provider:
                    result[_id] = self.default_none, provider.value

        # Final Glue..
        final = []
        for _id in self.get_queryset().values_list("id", flat=True):
            user, comp = result.get(_id, (self.default_none, self.default_none))
            final.append(
                [
                    CellObject("id", "ID", None, section_name, _id, _id),
                    CellObject(
                        "certifying_user", "Certifying User", None, section_name, user, user
                    ),
                    CellObject(
                        "certifying_company", "Certifying Company", None, section_name, comp, comp
                    ),
                ]
            )

        return final

    def get_subdivision_data(self):
        """Pull all of the Sudivision Data"""

        replace_dict = OrderedDict(
            [
                ("id", "ID"),
                ("home__subdivision__name", "Subdivision/MF Development Name"),
                ("home__subdivision__slug", "Subdivision/MF Development Slug"),
                ("home__subdivision__builder_org__name", "Subdivision/MF Development Builder Name"),
                ("home__subdivision__cross_roads", "Subdivision/MF Development Cross Roads"),
            ]
        )

        clean_dict = None

        if self.user.is_superuser or self.user.company.slug == "aps":
            from axis.customer_aps.models import SMART_TSTAT_MODELS, SMART_TSTAT_ELIGIBILITY

            def clean_eligibility(ident):
                """Gets the pretty name"""
                return dict(SMART_TSTAT_ELIGIBILITY).get(ident, "-")

            def clean_models(ident):
                """Gets the pretty name"""
                if ident is None:
                    return "-"
                if ident and not isinstance(ident, set):
                    ident = eval(ident)
                return ",".join([dict(SMART_TSTAT_MODELS).get(x, "-") for x in ident])

            key = "home__subdivision__aps_thermostat_option__models"
            replace_dict[key] = "Smart Thermostat Eligible"
            key = "home__subdivision__aps_thermostat_option__eligibility"
            replace_dict[key] = "Smart Thermostat Eligible"

            clean_dict = {
                "home__subdivision__aps_thermostat_option__models": clean_models,
                "home__subdivision__aps_thermostat_option__eligibility": clean_eligibility,
            }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            drop_prefix="home__",
            section="subdivision",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)

        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        return data

    def get_builder_agreement_data(self):
        return []

    def get_community_data(self):
        include = [
            "id",
            "home__subdivision__community__name",
            "home__subdivision__community__slug",
            "home__subdivision__community__cross_roads",
            "home__subdivision__community__website",
        ]
        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=include,
            drop_prefix="home__subdivision__",
            section="community",
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    # flake8: noqa: C901
    def get_relationships_data(self):
        from axis.relationship.models import Relationship
        from axis.home.models import Home
        from axis.company.strings import COMPANY_TYPES

        if hasattr(self, "relationship_data"):
            return self.relationship_data

        # stat_to_home = dict(self.get_queryset().values_list('home_id', 'id'))
        home_to_stat = dict(self.get_queryset().values_list("id", "home_id"))

        include = [
            "object_id",
            "company__id",
            "company__name",
            "company__slug",
            "company__company_type",
            "utilitysettings__is_electric_utility",
            "utilitysettings__is_gas_utility",
            "company__water_provider",
        ]

        structure = self.assign_basic(
            Relationship, include=include, drop_prefix="company__", section="associations"
        )

        objects = Relationship.objects.filter(content_type=ContentType.objects.get_for_model(Home))
        objects = objects.filter(object_id__in=list(home_to_stat.values()))
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related).distinct()
        data = objects.values_list(*[x.attr for x in structure])

        utility_types = []
        has_electricity = any([x[5] for x in data if x])
        if has_electricity:
            utility_types.append("electric")
        has_gas = any([x[6] for x in data if x])
        if has_gas:
            utility_types.append("gas")
        has_water = any([x[7] for x in data if x])
        if has_water:
            utility_types.append("water")
        has_utility = not any([y for x in data for y in x[5:]])
        if has_utility:
            utility_types.append("default")

        section_name = "relationships"
        result_dict = OrderedDict()
        company_types = set(x[4] for x in data)
        for item in data:
            home_id, co_id, co_name, co_slug, co_type, electric, gas, water = item
            for stat_id in [x for x in home_to_stat if home_to_stat[x] == home_id]:
                # Build the baseline structure
                if stat_id not in result_dict.keys():
                    result_dict[stat_id] = [["id", "ID", None, section_name, stat_id, stat_id]]
                    for company_type, pretty_name in COMPANY_TYPES:
                        if company_type not in company_types:
                            continue
                        if company_type == "utility":
                            for item in utility_types:
                                data_package = [
                                    company_type,
                                    pretty_name,
                                    None,
                                    section_name,
                                    set(),
                                    set(),
                                ]
                                slug_package = [
                                    "{}_slug".format(company_type),
                                    company_type,
                                    None,
                                    section_name,
                                    set(),
                                    set(),
                                ]
                                if item != "default":
                                    pretty_name = "{} Utility".format(item.capitalize())
                                    c_type = "{}_{}".format(item, company_type)
                                    data_package = [
                                        c_type,
                                        pretty_name,
                                        None,
                                        section_name,
                                        set(),
                                        set(),
                                    ]
                                    slug_package = [
                                        "{}_slug".format(c_type),
                                        c_type,
                                        None,
                                        section_name,
                                        set(),
                                        set(),
                                    ]
                                result_dict[stat_id].append(data_package)
                                result_dict[stat_id].append(slug_package)
                        else:
                            data_package = [
                                company_type,
                                pretty_name,
                                None,
                                section_name,
                                set(),
                                set(),
                            ]
                            slug_package = [
                                "{}_slug".format(company_type),
                                company_type,
                                None,
                                section_name,
                                set(),
                                set(),
                            ]
                            result_dict[stat_id].append(data_package)
                            result_dict[stat_id].append(slug_package)
                # Add our data into the structure
                result_set = result_dict[stat_id]
                if co_type == "utility":
                    if electric:
                        result = next((x for x in result_set if x[0] == "electric_utility"))
                        result[-2].add(co_name)
                        result = next((x for x in result_set if x[0] == "electric_utility_slug"))
                        result[-2].add(co_slug)
                    if gas:
                        result = next((x for x in result_set if x[0] == "gas_utility"))
                        result[-2].add(co_name)
                        result = next((x for x in result_set if x[0] == "gas_utility_slug"))
                        result[-2].add(co_slug)
                    if water:
                        result = next((x for x in result_set if x[0] == "water_utility"))
                        result[-2].add(co_name)
                        result = next((x for x in result_set if x[0] == "water_utility_slug"))
                        result[-2].add(co_slug)
                    if not any([electric, gas, water]):
                        result = next((x for x in result_set if x[0] == "utility"), None)
                        if result:
                            result[-2].add(co_name)
                        result = next((x for x in result_set if x[0] == "utility_slug"), None)
                        if result:
                            result[-2].add(co_slug)
                else:
                    result = next((x for x in result_set if x[0] == co_type))
                    result[-2].add(co_name)
                    result = next((x for x in result_set if x[0] == "{}_slug".format(co_type)))
                    result[-2].add(co_slug)

        # Finally translate this over to our default structure.
        result = OrderedDict()
        for key, values in result_dict.items():
            result[key] = []
            for v in values:
                v[-1] = v[-2]
                if isinstance(v[-2], set):
                    v[-1] = list(v[-2])
                if isinstance(v[-1], (list, tuple)):
                    if len(v[-1]) > 1 and self.join_resulting_values:
                        v[-1] = ", ".join(v[-1])
                    elif len(v[-1]) == 1:
                        v[-1] = v[-1][0]
                result[key].append(CellObject(*v))

        self.relationship_data = list(result.values())
        return self.relationship_data

    def get_eep_program_data(self):
        replace_dict = OrderedDict(
            [
                ("id", "Status ID"),
                ("eep_program__name", "Program Name"),
                ("eep_program__slug", "Program Slug"),
                ("eep_program__owner__name", "Program Owner"),
                ("eep_program__owner__slug", "Program Owner Slug"),
                ("eep_program__min_hers_score", "Program Minimum HERS Score"),
                ("eep_program__max_hers_score", "Program Maximum HERS Score"),
            ]
        )
        replace_dict["eep_program__builder_incentive_dollar_value"] = "Program Builder Incentive"
        replace_dict["eep_program__rater_incentive_dollar_value"] = "Program Rater Incentive"
        replace_dict["eep_program__program_start_date"] = "Program Start Date"
        replace_dict["eep_program__program_end_date"] = "Program End Date"

        def get_incentive_payment(ident):
            return "${}".format(ident) if ident else None

        clean_dict = {
            "eep_program__builder_incentive_dollar_value": get_incentive_payment,
            "eep_program__rater_incentive_dollar_value": get_incentive_payment,
        }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="eep_program",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_floorplan_data(self):
        include = [
            "id",
            "floorplan__name",
            "floorplan__number",
            "floorplan__square_footage",
            "floorplan__is_custom_home",
            "floorplan__owner__name",
            "floorplan__comment",
        ]

        structure = self.assign_basic(EEPProgramHomeStatus, include=include, section="floorplan")
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_basic_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import FOUNDATION_TYPES
        from axis.remrate_data.strings import HOME_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("building__project__name", "Project Name"),
                ("buildinginfo__conditioned_area", "Area of Conditioned Space (sq. ft.)"),
                ("buildinginfo__volume", "Volume of Conditioned Space (cu. ft.)"),
                ("buildinginfo__number_stories", "Floors on or Above-Grade"),
                ("buildinginfo__number_bedrooms", "Number of Bedrooms"),
                ("buildinginfo__num_units", "Number of Units"),
                ("buildinginfo__type", "Building Type"),
                ("buildinginfo__foundation_type", "Foundation Type"),
                ("hers__score", "HERS Score"),
                ("hers__passes_2005_epact_tax_credit", "EPACT Tax Credit"),
                (
                    "energystar__energy_star_v2p5_pv_score",
                    "ENERGY STAR v2.5 photo voltaic adjusted HERS score",
                ),
                (
                    "energystar__energy_star_v2p5_hers_score",
                    "ENERGY STAR v2.5 Reference Design HERS score",
                ),
                (
                    "energystar__energy_star_v2p5_hers_saf",
                    "ENERGY STAR v2.5 size adjustment factor",
                ),
                (
                    "energystar__energy_star_v3_pv_score",
                    "ENERGY STAR v3 photo voltaic adjusted HERS score",
                ),
                (
                    "energystar__energy_star_v3_hers_score",
                    "ENERGY STAR v3 Reference Design HERS score",
                ),
                ("energystar__energy_star_v3_hers_saf", "ENERGY STAR v3 size adjustment factor"),
                (
                    "energystar__energy_star_v3p1_pv_score",
                    "ENERGY STAR v3.1 photo voltaic adjusted HERS score",
                ),
                (
                    "energystar__energy_star_v3p1_hers_score",
                    "ENERGY STAR v3.1 Reference Design HERS score",
                ),
                (
                    "energystar__energy_star_v3p1_hers_saf",
                    "ENERGY STAR v3.1 size adjustment factor",
                ),
                ("energystar__passes_doe_zero", "Passes DOE Zero Energy Rated Home"),
                ("iecc__passes_iecc06_code", "Passes IECC 2006 Code"),
                ("iecc__passes_iecc09_code", "Passes IECC 2009 Code"),
                ("iecc__passes_iecc12_code", "Passes IECC 2012 Code"),
                ("iecc__passes_iecc15_code", "Passes IECC 2015 Code"),
                ("iecc__passes_iecc18_code", "Passes IECC 2018 Code"),
                ("version", "REM/Rate Version"),
                ("flavor", "REM/Rate Flavor"),
                ("simulation_date", "REM/Rate Simulation Date"),
                ("remrate_user__username", "REM/Rate Username"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_bldg_type(ident):
            return dict(HOME_TYPES).get(ident, self.default_none)

        def get_foundation_type(ident):
            return dict(FOUNDATION_TYPES).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "buildinginfo__type": get_bldg_type,
            "buildinginfo__foundation_type": get_foundation_type,
            "buildinginfo__conditioned_area": lambda value: value and int(value),
            "buildinginfo__volume": lambda value: value and int(value),
        }

        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="simulation_basic",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]

        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_building_shell_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("building__ceiling_attic_area", "Building Shell Ceiling with Attic"),
                ("building__ceiling_cathedral_area", "Building Shell Vaulted Ceiling"),
                (
                    "building__above_ground_wall_conditioned_to_outdoor_area",
                    "Building Shell Above Grade Walls (Conditioned)",
                ),
                (
                    "building__above_ground_wall_buffer_to_outdoor_area",
                    "Building Shell Above Grade Walls (Buffer)",
                ),
                (
                    "building__foundation_walls_conditioned_to_outdoor_area",
                    "Building Shell Foundation Walls (Conditioned)",
                ),
                (
                    "building__foundation_walls_buffer_to_outdoor_area",
                    "Building Shell Foundation Walls (Buffer)",
                ),
                (
                    "building__frame_floor_conditioned_to_outdoor_area",
                    "Building Shell Frame Floors (Conditioned)",
                ),
                ("building__window_wall_ratio", "Building Shell Window/Wall Ratio"),
                ("building__window_floor_ratio", "Window / Floor Ratio"),
            ]
        )

        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        clean_dict = {"id": get_ident}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_building_shell",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_framefloor_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("framefloor__area", "Frame Floor Area (sq.ft.)"),
                ("framefloor__type__continuous_insulation", "Frame Floor Continuous Insulation"),
                ("framefloor__type__cavity_insulation", "Frame Floor Framed Cavity Insulation"),
                ("framefloor__type__composite_type__name", "Frame Floor Component Name"),
                ("framefloor__type__composite_type__u_value", "Frame Floor U-Value"),
            ]
        )

        def round_3(ident):
            try:
                return Decimal("%.3f" % float(ident))
            except Exception:
                return ident

        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        clean_dict = {"id": get_ident, "framefloor__type__composite_type__u_value": round_3}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_framefloor",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_joist_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import JOIST_ABOVE_GRADE_WALL_LOCATIONS

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("joist__area", "Rim and Band Joist Area (sq.ft.)"),
                (
                    "joist__continuous_insulation_r_value",
                    "Rim and Band Joist Continuous Insulation",
                ),
                ("joist__cavity_insulation_r_value", "Rim and Band Joist Framed Cavity Insulation"),
                ("joist__location", "Rim and Band Joist Location"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_joist__location(ident):
            return dict(JOIST_ABOVE_GRADE_WALL_LOCATIONS).get(ident, self.default_none)

        clean_dict = {"id": get_ident, "joist__location": get_joist__location}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_joist",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_abovegradewall_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("abovegradewall__gross_area", "Above-Grade Wall Total gross Area (sq.ft.)"),
                (
                    "abovegradewall__type__continuous_insulation",
                    "Above-Grade Wall Total Continuous R-value",
                ),
                (
                    "abovegradewall__type__cavity_insulation",
                    "Above-Grade Wall Total Cavity Insulation",
                ),
                ("abovegradewall__type__composite_type__name", "Above-Grade Wall Component Name"),
                ("abovegradewall__type__composite_type__u_value", "Above-Grade Wall U-Value"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def round_3(ident):
            try:
                return Decimal("%.3f" % float(ident))
            except Exception:
                return ident

        clean_dict = {"id": get_ident, "abovegradewall__type__composite_type__u_value": round_3}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_abovegradewall",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_window_door_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("window__area", "Window Information Glazing area"),
                ("window__type__shgc", "Window Information SHGC"),
                ("window__type__u_value", "Window Information U-Value"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        clean_dict = {"id": get_ident}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_window",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_door_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("door__area", "Door Information Opaque Area (sq.ft.)"),
                ("door__type__r_value", "Door Information R-value of Opaque Area"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        clean_dict = {"id": get_ident}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_door",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_roof_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import COLORS, ROOF_STYLE

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("roof__area", "Roof Information Gross Area (sq.ft.)"),
                ("roof__sealed_attic_roof_area", "Roof Information Attic Exterior (sq.ft.)"),
                ("roof__color", "Roof Information Color"),
                ("roof__radiant_barrier", "Roof Information Radiant Barrier"),
                ("roof__style", "Roof Information Type (Attic)"),
                ("roof__clay_or_concrete", "Roof Information Clay or Cement Tiles"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_color(ident):
            return dict(COLORS).get(ident, self.default_none)

        def get_true_false(ident):
            if ident == 1:
                return self.default_true
            return self.default_false

        def get_style(ident):
            return dict(ROOF_STYLE).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "roof__color": get_color,
            "roof__radiant_barrier": get_true_false,
            "roof__style": get_style,
            "roof__clay_or_concrete": get_true_false,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_roof",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_ceiling_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                (
                    "ceilingtype__continuous_insulation",
                    "Ceiling Information Continuous Insulation (R-value)",
                ),
                (
                    "ceilingtype__cavity_insulation",
                    "Ceiling Information Cavity Insulation (R-value)",
                ),
                ("ceilingtype__composite_type__name", "Ceiling Component Name"),
                ("ceilingtype__composite_type__u_value", "Ceiling Component U-Value"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def round_3(ident):
            try:
                return Decimal("%.3f" % float(ident))
            except Exception:
                return ident

        clean_dict = {"id": get_ident, "ceilingtype__composite_type__u_value": round_3}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_ceiling",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_general_mechaical_data(self, object_map):
        from axis.remrate_data.models import GeneralMechanicalEquipment

        return GeneralMechanicalEquipment.objects.get_home_status_export_data(
            simulation_ids=list(object_map.values()),
            object_map=object_map,
            default_null=self.default_none,
        )

    def get_remrate_air_conditioning_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import COOLING_TYPES, COOLING_EFF_UNITS, FUEL_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("airconditioner__type", "Air Conditioning System Type"),
                ("airconditioner__fuel_type", "Air Conditioning Fuel Type"),
                ("airconditioner__output_capacity", "Air Conditioning Rated Output Capacity"),
                ("airconditioner__efficiency", "Air Conditioning Seasonal Equipment Efficiency"),
                ("airconditioner__efficiency_unit", "Air Conditioning Efficiency Units"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_cooling_type(ident):
            return dict(COOLING_TYPES).get(ident, self.default_none)

        def get_fuel_type(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        def get_cooling_efficiency_unit(ident):
            return dict(COOLING_EFF_UNITS).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "airconditioner__type": get_cooling_type,
            "airconditioner__fuel_type": get_fuel_type,
            "airconditioner__efficiency_unit": get_cooling_efficiency_unit,
        }

        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_airconditioner",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_heater_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import HEATER_TYPES, FUEL_TYPES, HEATING_EFF_UNITS

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("heater__type", "Heating System Type"),
                ("heater__fuel_type", "Heating Fuel Type"),
                ("heater__output_capacity", "Heating Rated Output Capacity"),
                ("heater__efficiency", "Heating Seasonal Equipment Efficiency"),
                ("heater__efficiency_unit", "Heating Efficiency Units"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_fuel_type(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        def get_heating_type(ident):
            return dict(HEATER_TYPES).get(ident, self.default_none)

        def get_heating_efficiency_unit(ident):
            return dict(HEATING_EFF_UNITS).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "heater__type": get_heating_type,
            "heater__fuel_type": get_fuel_type,
            "heater__efficiency_unit": get_heating_efficiency_unit,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_heater",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_ashp_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import COOLING_EFF_UNITS, FUEL_TYPES, HEATING_EFF_UNITS

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("airsourceheatpump__name", "Air Source Heat Pump System Type"),
                ("airsourceheatpump__fuel_type", "Air Source Heat Pump Fuel Type"),
                ("airsourceheatpump__heating_capacity", "Air Source Heat Pump Heating Capacity"),
                ("airsourceheatpump__heating_capacity", "Air Source Heat Pump Heating Capacity"),
                (
                    "airsourceheatpump__heating_efficiency",
                    "Air Source Heat Pump Heating Efficiency",
                ),
                (
                    "airsourceheatpump__heating_efficiency_units",
                    "Air Source Heat Pump Heating Efficiency Units",
                ),
                ("airsourceheatpump__cooling_capacity", "Air Source Heat Pump Cooling Capacity"),
                (
                    "airsourceheatpump__cooling_efficiency",
                    "Air Source Heat Pump Cooling Efficiency",
                ),
                (
                    "airsourceheatpump__cooling_efficiency_units",
                    "Air Source Heat Pump Cooling Efficiency Units",
                ),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_fuel_type(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        def get_cooling_efficiency_unit(ident):
            return dict(COOLING_EFF_UNITS).get(ident, self.default_none)

        def get_heating_efficiency_unit(ident):
            return dict(HEATING_EFF_UNITS).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "airsourceheatpump__fuel_type": get_fuel_type,
            "airsourceheatpump__heating_efficiency_units": get_heating_efficiency_unit,
            "airsourceheatpump__cooling_efficiency_units": get_cooling_efficiency_unit,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_airsourceheatpump",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_gshp_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import FUEL_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("groundsourceheatpump__name", "Ground Source Heat Pump System Type"),
                ("groundsourceheatpump__fuel_type", "Ground Source Heat Pump Fuel Type"),
                (
                    "groundsourceheatpump__heating_coefficient_of_performance_70f",
                    "Ground Source Heat Pump Heating COP at 70°F",
                ),
                (
                    "groundsourceheatpump__heating_capacity_70f",
                    "Ground Source Heat Pump Heating Capacity at 70°F",
                ),
                (
                    "groundsourceheatpump__cooling_energy_efficiency_ratio_70f",
                    "Ground Source Heat Pump Cooling EER at 70°F",
                ),
                (
                    "groundsourceheatpump__cooling_capacity_70f",
                    "Ground Source Heat Pump Cooling Capacity at 70°F",
                ),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_fuel_type(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "groundsourceheatpump__fuel_type": get_fuel_type,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_groundsourceheatpump",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_iswh_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import (
            FUEL_TYPES,
            FAN_CONTROL_TYPES,
            ISWH_DIST_TYPES,
            ISWH_TYPES,
        )

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("integratedspacewaterheater__name", "Integrated Space Water Heater Name"),
                (
                    "integratedspacewaterheater__fuel_type",
                    "Integrated Space Water Heater Fuel Type",
                ),
                ("integratedspacewaterheater__type", "Integrated Space Water Heater System Type"),
                (
                    "integratedspacewaterheater__distribution_type",
                    "Integrated Space Water Heater Distribution Type",
                ),
                (
                    "integratedspacewaterheater__output_capacity",
                    "Integrated Space Water Heater Output Capacity",
                ),
                (
                    "integratedspacewaterheater__space_heating_efficiency",
                    "Integrated Space Water Heater Heating Efficiency",
                ),
                (
                    "integratedspacewaterheater__water_heating_energy_factor",
                    "Integrated Space Water Heater Water Heater Energy Factor",
                ),
                (
                    "integratedspacewaterheater__tank_size",
                    "Integrated Space Water Heater Water Heater Tank Size",
                ),
                (
                    "integratedspacewaterheater__fan_control_type",
                    "Integrated Space Water Heater Fan Control Type",
                ),
                (
                    "integratedspacewaterheater__fan_high_speed",
                    "Integrated Space Water Heater Fan High Speed",
                ),
                (
                    "integratedspacewaterheater__fan_low_speed",
                    "Integrated Space Water Heater Fan Low Speed",
                ),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_fuel_type(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        def get_sys_type(ident):
            return dict(ISWH_TYPES).get(ident, self.default_none)

        def get_dist_type(ident):
            return dict(ISWH_DIST_TYPES).get(ident, self.default_none)

        def get_fan_type(ident):
            return dict(FAN_CONTROL_TYPES).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "integratedspacewaterheater__fuel_type": get_fuel_type,
            "integratedspacewaterheater__type": get_sys_type,
            "integratedspacewaterheater__distribution_type": get_dist_type,
            "integratedspacewaterheater__fan_control_type": get_fan_type,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_integratedspacewaterheater",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_water_heater_part1_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import H2O_HEATER_TYPES, FUEL_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("hotwaterheater__type", "Water Heating Equipment Water Heater Type"),
                ("hotwaterheater__fuel_type", "Water Heating Equipment Fuel Type"),
                ("hotwaterheater__energy_factor", "Water Heating Equipment Energy Factor"),
                ("hotwaterheater__tank_size", "Water Heating Equipment Water Tank Size (gallons)"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_type(ident):
            return dict(H2O_HEATER_TYPES).get(ident, self.default_none)

        def get_fuel_type(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "hotwaterheater__type": get_type,
            "hotwaterheater__fuel_type": get_fuel_type,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_hot_water",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_water_heater_part2_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import MECHANICAL_EQUIP_LOCATIONS

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("installedequipment__location", "Water Heating Equipment Location"),
                (
                    "installedequipment__hot_water_heater_load_served_pct",
                    "Water Heating Equipment Percent Load Served",
                ),
                (
                    "installedequipment__performance_adjustment_pct",
                    "Water Heating Equipment Performance Adjustment",
                ),
                ("installedequipment__qty_installed", "Water Heating Equipment Number of Units"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_location(ident):
            return dict(MECHANICAL_EQUIP_LOCATIONS).get(ident, self.default_none)

        clean_dict = {"id": get_ident, "installedequipment__location": get_location}
        objects = Simulation.objects.filter(
            id__in=list(object_map.values()), installedequipment__hot_water_heater__isnull=False
        )
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_hot_water",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_ductsystem_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import INFILTRATION_UNITS

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("ductsystem__supply_leakage", "Supply Duct Leakage Outside"),
                ("ductsystem__return_leakage", "Return Duct Leakage Outside"),
                ("ductsystem__total_leakage", "Duct Leakage Outside"),
                ("ductsystem__leakage_unit", "Duct Leakage Outside Units"),
                ("ductsystem__supply_area", "Duct System Supply Area (sq.ft.)"),
                ("ductsystem__return_area", "Duct System Return Area (sq.ft.)"),
                (
                    "ductsystem__total_leakage",
                    "Duct System Duct Leakage (supply and return) to outside",
                ),
                ("ductsystem__leakage_unit", "Duct System Duct Leakage to outside Units"),
                ("ductsystem__total_real_leakage", "Total Duct Leakage"),
                ("ductsystem__total_real_leakage_unit", "Total Duct Units"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def leakage_unit(ident):
            return dict(INFILTRATION_UNITS).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "ductsystem__leakage_unit": leakage_unit,
            "ductsystem__total_real_leakage_unit": leakage_unit,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_ductsystem",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_duct_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import DUCT_LOCATIONS

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("duct__area", "Duct Information Percent Area"),
                ("duct__r_value", "Duct Information R-Value"),
                ("duct__location", "Duct Information Location"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def location(ident):
            return dict(DUCT_LOCATIONS).get(ident, self.default_none)

        clean_dict = {"id": get_ident, "duct__location": location}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_ductsystem",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_infiltration_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import INFILTRATION_UNITS
        from axis.remrate_data.strings import INFILTRATION_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("infiltration__heating_value", "Infiltration Heating Season Infiltration Value"),
                ("infiltration__cooling_value", "Infiltration Cooling Season Infiltration Value"),
                ("infiltration__units", "Infiltration units"),
                ("infiltration__mechanical_vent_type", "Mechanical Ventilation Type"),
                ("infiltration__mechanical_vent_cfm", "Mechanical Ventilation Rate (cfm)"),
                ("infiltration__hours_per_day", "Mechanical Ventilation Hours per day"),
                ("infiltration__mechanical_vent_power", "Mechanical Ventilation Fan Power (watts)"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_units(ident):
            return dict(INFILTRATION_UNITS).get(ident, self.default_none)

        def get_vent_type(ident):
            return dict(INFILTRATION_TYPES).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "infiltration__units": get_units,
            "infiltration__mechanical_vent_type": get_vent_type,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_infiltration",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_solarsystem_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import SOLAR_TYPES, LOOP_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("solarsystem__type", "Active Solar Array Solar Type System"),
                ("solarsystem__collector_loop_type", "Active Solar Array Collector loop type"),
                ("solarsystem__collector_area", "Active Solar Array Active Solar Area"),
                ("solarsystem__storage_volume", "Active Solar Array Active Solar Storage volume"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_solar_type(ident):
            return dict(SOLAR_TYPES).get(ident, self.default_none)

        def get_loop_type(ident):
            return dict(LOOP_TYPES).get(ident, self.default_none)

        clean_dict = {
            "id": get_ident,
            "solarsystem__type": get_solar_type,
            "solarsystem__collector_loop_type": get_loop_type,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_solarsystem",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_pv_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import PV_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("photovoltaic__type", "PV Array Photovoltaic Type"),
                ("photovoltaic__area", "PV Array Photovolatic Area"),
                ("photovoltaic__orientation", "PV Array Photovolatic Orientation"),
                ("photovoltaic__peak_power", "PV Array Photovolatic Peak Power"),
                ("photovoltaic__tilt", "PV Array Photovolatic Tilt"),
                ("photovoltaic__inverter_efficiency", "PV Array Photovolatic Inverter Efficiency"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_pv_type(ident):
            return dict(PV_TYPES).get(ident, self.default_none)

        clean_dict = {"id": get_ident, "photovoltaic__type": get_pv_type}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_solarsystem",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_misc_data(self, object_map):
        from axis.remrate_data.models import Simulation

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("building__project__name", "Performance Summary Building name"),
                ("building__project__property_owner", "Performance Summary Owners Name"),
                ("building__project__property_owner", "Performance Summary Owners Name"),
                ("building__project__property_address", "Performance Summary Property Address"),
                ("building__project__property_city", "Performance Summary Property City"),
                ("building__project__builder_name", "Performance Summary Builder Name"),
                ("site__climate_zone", "Performance Summary Weather Site"),
                ("building__filename", "Performance Summary File Name"),
                ("buildinginfo__rating_number", "Performance Summary Rating No."),
                ("building__project__rating_organization", "Performance Summary Rating Org"),
                ("building__project__rater_name", "Performance Summary Rater's Name"),
                ("building__project__rater_id", "Performance Summary Raters's No"),
                ("building__project__rating_type", "Performance Summary Rating Type"),
                ("building__project__resnet_registry_id", "Performance Summary RESNET Registry ID"),
                ("building__project__rating_date", "Performance Summary Rating Date"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        clean_dict = {"id": get_ident}
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_misc",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_fuelsummary_data(self, object_map):
        from axis.remrate_data.models import FuelSummary

        return FuelSummary.objects.get_home_status_export_data(
            simulation_ids=list(object_map.values()),
            object_map=object_map,
            default_null=self.default_none,
        )

    def get_dominant_equipment_data(self, object_map):
        from axis.remrate_data.models import InstalledEquipment

        return InstalledEquipment.objects.get_home_status_export_data(
            simulation_ids=list(object_map.values()),
            object_map=object_map,
            default_null=self.default_none,
        )

    def get_remrate_results_data(self, object_map):
        from axis.remrate_data.models import Simulation, Results

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("results__heating_system_efficiency", None),
                ("results__cooling_system_efficiency", None),
                ("results__hot_water_system_efficiency", None),
                ("results__roof_heating_load", None),
                ("results__roof_cooling_load", None),
                ("results__joist_heating_load", None),
                ("results__joist_cooling_load", None),
                ("results__frame_floor_heating_load", None),
                ("results__frame_floor_cooling_load", None),
                ("results__above_ground_walls_heating_load", None),
                ("results__above_ground_walls_cooling_load", None),
                ("results__foundation_wall_heating_load", None),
                ("results__foundation_wall_cooling_load", None),
                ("results__crawl_heating_load", None),
                ("results__crawl_cooling_load", None),
                ("results__slab_heating_load", None),
                ("results__slab_cooling_load", None),
                ("results__windows_skylights_heating_load", None),
                ("results__windows_skylights_cooling_load", None),
                ("results__door_heating_load", None),
                ("results__door_cooling_load", None),
                ("results__sunspace_heating_load", None),
                ("results__sunspace_cooling_load", None),
                ("results__internal_gains_heating_load", None),
                ("results__internal_gains_cooling_load", None),
                ("results__active_solar_heating_load", None),
                ("results__active_solar_cooling_load", None),
                ("results__mechanical_ventilation_heating_load", None),
                ("results__mechanical_ventilation_cooling_load", None),
                ("results__infiltration_heating_load", None),
                ("results__infiltration_cooling_load", None),
                ("results__duct_heating_load", None),
                ("results__duct_cooling_load", None),
                ("results__whole_house_fan_heating_load", None),
                ("results__whole_house_fan_cooling_load", None),
                ("results__total_heating_load", None),
                ("results__total_cooling_load", None),
                ("results__heating_design_load", None),
                ("results__calculated_sensible_cooling_load", None),
                ("results__calculated_latent_cooling_load", None),
                ("results__calculated_sensible_latent_cooling_load", None),
                ("results__heating_consumption", None),
                ("results__cooling_consumption", None),
                ("results__hot_water_consumption", None),
                ("results__hot_water_heating_consumption", None),
                ("results__refrigerator_consumption", None),
                ("results__freezer_consumption", None),
                ("results__dryer_consumption", None),
                ("results__oven_consumption", None),
                ("results__lights_and_appliances_consumption", None),
                ("results__lights_hs_consumption", None),
                ("results__lights_cs_consumption", None),
                ("results__lights_and_appliances_total_consumption", None),
                ("results__photo_voltaic_consumption", None),
                ("results__unit_heating_cost", None),
                ("results__unit_cooling_cost", None),
                ("results__unit_hot_water_cost", None),
                ("results__heating_cost", None),
                ("results__cooling_cost", None),
                ("results__hot_water_cost", None),
                ("results__refrigerator_cost", None),
                ("results__freezer_cost", None),
                ("results__dryer_cost", None),
                ("results__oven_cost", None),
                ("results__lights_and_appliances_cost", None),
                ("results__lighting_cost", None),
                ("results__lights_and_appliances_total_cost", None),
                ("results__photo_voltaic_cost", None),
                ("results__solar_savings", None),
                ("results__service_cost", None),
                ("results__total_cost", None),
                ("results__shell_area", None),
                # ('results__heating_load_per_shell_area_hdd75', None),
                # ('results__cooling_load_per_shell_area_cdd74', None),
                # ('results__heating_design_load_per_shell_area_hdd75', None),
                # ('results__cooling_design_load_per_shell_area_cdd74', None),
                ("results__heating_natural_ach", None),
                ("results__cooling_natural_ach", None),
                # ('results__rating_number', None),
                # ('results__co2_total_emission', None),
                # ('results__s02_total_emission', None),
                # ('results__nox_total_emission', None),
                # ('results__co2_heating_emission', None),
                # ('results__co2_cooling_emission', None),
                # ('results__co2_hot_water_emission', None),
                # ('results__co2_lights_appliance_emission', None),
                # ('results__co2_photo_voltaic_emission', None),
                # ('results__so2_heating_emission', None),
                # ('results__so2_cooling_emission', None),
                # ('results__so2_hot_water_emission', None),
                # ('results__so2_lights_appliance_emission', None),
                # ('results__so2_photo_voltaic_emission', None),
                # ('results__nox_heating_emission', None),
                # ('results__nox_cooling_emission', None),
                # ('results__nox_hot_water_emission', None),
                # ('results__nox_lights_appliance_emission', None),
                # ('results__nox_photo_voltaic_emission', None),
                # ('results__co2_hers_emission_savings', None),
                # ('results__so2_hers_emission_savings', None),
                # ('results__nox_hers_emission_savings', None),
                # ('results__source_energy_heating', None),
                # ('results__source_energy_cooling', None),
                # ('results__source_energy_hot_water', None),
                # ('results__source_energy_lights_appliance', None),
                # ('results__source_energy_photo_voltaic', None),
            ]
        )

        for k in replace_dict:
            f = Results._meta.get_field(k.replace("results__", ""))
            replace_dict[k] = f.verbose_name

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_cost(ident):
            return "${:.2f}".format(ident) if ident else self.default_none

        clean_dict = {
            "id": get_ident,
            "results__heating_cost": get_cost,
            "results__cooling_cost": get_cost,
            "results__hot_water_cost": get_cost,
            "results__lights_and_appliances_cost": get_cost,
            "results__photo_voltaic_cost": get_cost,
            "results__service_cost": get_cost,
            "results__total_cost": get_cost,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_results",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_utility_rate_data(self, object_map):
        from axis.remrate_data.models import UtilityRate

        return UtilityRate.objects.get_home_status_export_data(
            simulation_ids=list(object_map.values()),
            object_map=object_map,
            default_null=self.default_none,
        )

    def get_remrate_lights_appl_data(self, object_map):
        from axis.remrate_data.models import Simulation
        from axis.remrate_data.strings import FUEL_TYPES

        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                # ('lightsandappliance__pct_florescent', 'Lights and Appliances Percent Florescent - Pin Based'),
                (
                    "lightsandappliance__pct_interior_cfl",
                    "Lights and Appliances Percent Florescent Interior - CFL",
                ),
                (
                    "lightsandappliance__pct_exterior_cfl",
                    "Lights and Appliances Percent Florescent Exterior - CFL",
                ),
                (
                    "lightsandappliance__pct_garage_cfl",
                    "Lights and Appliances Percent Florescent Garage - CFL",
                ),
                (
                    "lightsandappliance__pct_interior_led",
                    "Lights and Appliances Percent Florescent Interior - LED",
                ),
                (
                    "lightsandappliance__pct_exterior_led",
                    "Lights and Appliances Percent Florescent Exterior - LED",
                ),
                (
                    "lightsandappliance__pct_garage_led",
                    "Lights and Appliances Percent Florescent Garage - LED",
                ),
                (
                    "lightsandappliance__refrigerator_kw_yr",
                    "Lights and Appliances Refrigerator (kW/yr)",
                ),
                (
                    "lightsandappliance__dishwasher_energy_factor",
                    "Lights and Appliances Dishwasher Energy Factor",
                ),
                (
                    "lightsandappliance__ceiling_fan_cfm_watt",
                    "Lights and Appliances Ceiling Fan (cfm)",
                ),
                ("lightsandappliance__clothes_dryer_fuel", "Lights and Appliances Dryer Fuel"),
                (
                    "lightsandappliance__clothes_dryer_energy_factor",
                    "Lights and Appliances Dryer Efficiency",
                ),
                (
                    "lightsandappliance__clothes_dryer_moisture_sensing",
                    "Lights and Appliances Dryer Moisture Sensing",
                ),
                (
                    "lightsandappliance__clothes_washer_label_energy_rating",
                    "Lights and Appliances Clothes Washer LER",
                ),
                (
                    "lightsandappliance__clothes_washer_electric_rate",
                    "Lights and Appliances Clothes Washer Elec Rate",
                ),
                (
                    "lightsandappliance__clothes_washer_gas_rate",
                    "Lights and Appliances Clothes Gas Rate",
                ),
                (
                    "lightsandappliance__clothes_washer_gas_cost",
                    "Lights and Appliances Clothes Annual Gas Cost",
                ),
                (
                    "lightsandappliance__clothes_washer_capacity",
                    "Lights and Appliances Clothes Washer Capacity",
                ),
                ("lightsandappliance__oven_fuel", "Lights and Appliances Range/Oven Fuel"),
            ]
        )

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def get_fuel(ident):
            return dict(FUEL_TYPES).get(ident, self.default_none)

        def get_yes_no(ident):
            return "Yes" if ident is not None else "No"

        clean_dict = {
            "id": get_ident,
            "lightsandappliance__clothes_dryer_fuel": get_fuel,
            "lightsandappliance__clothes_dryer_moisture_sensing": get_yes_no,
            "lightsandappliance__oven_fuel": get_fuel,
        }
        objects = Simulation.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_basic(
            Simulation,
            include=list(replace_dict.keys()),
            section="remrate_lights_appl",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_remrate_data(self, advanced=False):
        object_map = dict(self.get_queryset().values_list("id", "floorplan__remrate_target__id"))

        if not len(object_map.values()) and not len(object_map.keys()):
            return []

        results = OrderedDict([(x, []) for x in self.get_queryset().values_list("id", flat=True)])

        data = self.get_remrate_basic_data(object_map)
        results = self.munge_data(results, data)

        if advanced:
            self.log.debug("Gathering REM/Rate Advanced Building data")
            data = self.get_building_shell_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Frame Floor data")
            data = self.get_remrate_framefloor_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Above Grade data")
            data = self.get_remrate_abovegradewall_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Joist data")
            data = self.get_remrate_joist_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Window data")
            data = self.get_window_door_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Door data")
            data = self.get_door_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Roof data")
            data = self.get_remrate_roof_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Ceiling data")
            data = self.get_remrate_ceiling_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Mechanical data")
            data = self.get_remrate_general_mechaical_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Dominant Equipment data")
            data = self.get_dominant_equipment_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced A/C data")
            data = self.get_remrate_air_conditioning_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Heater data")
            data = self.get_remrate_heater_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced ASHP data")
            data = self.get_remrate_ashp_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced GSHP data")
            data = self.get_remrate_gshp_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced IWSH data")
            data = self.get_remrate_iswh_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Water Heater (pt 1) data")
            data = self.get_remrate_water_heater_part1_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Water Heater (pt 2) data")
            data = self.get_remrate_water_heater_part2_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Duct System data")
            data = self.get_remrate_ductsystem_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Duct data")
            data = self.get_remrate_duct_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Infiltration data")
            data = self.get_remrate_infiltration_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Solar data")
            data = self.get_remrate_solarsystem_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced PV data")
            data = self.get_remrate_pv_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Misc data")
            data = self.get_remrate_misc_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Fuel Summary data")
            data = self.get_remrate_fuelsummary_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Result data")
            data = self.get_remrate_results_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced Utility data")
            data = self.get_remrate_utility_rate_data(object_map)
            results = self.munge_data(results, data)

            self.log.debug("Gathering REM/Rate Advanced L&A data")
            data = self.get_remrate_lights_appl_data(object_map)
            results = self.munge_data(results, data)

        final = OrderedDict()
        field = "simulation_advanced" if advanced else "simulation_basic"
        for _id, simulation_id in object_map.items():
            cell_obj = CellObject("id", "ID", None, field, _id, _id)
            found = False
            for value_set in results.values():
                # if not len(value_set) and not self.retain_empty:
                if not len(value_set):
                    continue
                ident = next((v for v in value_set if v.attr == "id")).raw_value
                if ident == simulation_id:
                    final[_id] = [cell_obj] + value_set[1:]
                    found = True
                    break
            if not found:
                final[_id] = [cell_obj]
        return list(final.values())

    # flake8: noqa: C901
    def _get_from_json(self, paths, data, sep="__"):
        """
        Reads a collection of paths from a data dictionary, where paths are bits joined by 'sep'.
        """

        # This can handle a single list somehwere in the chain of 'path' bits.  It will silently do
        # very wrong things if there are nested lists in the path.  It might also raise IndexErrors.

        start_value = data
        multi_total = None

        line_values = []
        # print('=================')

        for attr in paths:
            # print('---', repr(attr))

            first = True
            multi = None
            column_value = None

            # Obtain a value for 'column_value'
            # This is easy if we just follow the chain of attr1__attr2__attr3, but in the case of
            # a middle element being a list, we have to loop through each position and finish the
            # lookups to the end of the chain.  Those multiple values go in a list and then
            # collapsed.
            # Finally that value can be appended to our row data.
            while first or multi:
                first = False
                value = start_value
                bits = attr.split(sep)
                for bit in bits:
                    if value is None:  # Exhausted chain as far as data allows
                        break
                    if isinstance(value, list) and len(value):
                        if multi is None:
                            # print('    *** starting 0')
                            multi = 0
                            multi_total = len(value)
                        # else:
                        #     print('    *** continuing', multi)

                        value = value[multi]
                    if hasattr(value, bit):
                        value = getattr(value, bit)
                    elif isinstance(value, dict):
                        value = value.get(bit, None)
                    elif hasattr(value, "data"):  # Look in JSON data instead
                        value = value.data.get(bit, None)
                    else:
                        value = None

                        # print('   ', repr(bit), '(%s)' % (type(value)))

                # if multi is not None:
                #     multi_str = "%d/%d" % (multi + 1, multi_total)
                # else:
                #     multi_str = ""
                # print('... Collected value:', repr(value), multi_str)

                if multi is None:
                    column_value = value
                else:
                    if column_value is None:
                        column_value = []
                    column_value.append(value)

                    # print('    Adding to multi list')
                    multi += 1
                    if multi == multi_total:
                        # Reached the end of the list, disable multi mode, outer loop will
                        # immediately exit and we're done.
                        multi = None
                        #     print('    Loop stopping')
                        # else:
                        #     print('    Looping')

            def _make_string(v):
                if v is None:
                    return self.default_none
                if type(v) == bool:
                    if v:
                        return self.default_true
                    return self.default_false
                return "{}".format(v)

            if isinstance(column_value, list):
                column_value = ", ".join(map(_make_string, column_value))
            line_values.append(column_value)

        return line_values

    def get_ekotrope_basic_data(self, object_map):
        from axis.ekotrope.models import HousePlan

        # This will map back the simulation id to the real ID
        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def _get_nullboolean_or_warning(value):
            if value == "warn":
                return "Warn"
            return self.get_formatted_nullboolean(value)

        items = OrderedDict(
            [
                (
                    "id",
                    {
                        "label": "House Plan ID",
                        "clean": get_ident,
                    },
                ),
                ("project__id", {"label": "Project ID"}),
                ("project__algorithmVersion", {"label": "Ekotrope Version"}),
                ("building__project__name", {"label": "Project Name"}),
                (
                    "houseplan__thermalEnvelope__summary__conditionedArea",
                    "Area of Conditioned Space (sq.ft.)",
                ),
                (
                    "houseplan__thermalEnvelope__summary__conditionedVolume",
                    "Volume of Conditioned Space",
                ),
                ("houseplan__details__numberOfFloorsOnOrAboveGrade", "Floors on or Above-Grade"),
                ("houseplan__details__bedrooms", "Number of Bedrooms"),
                # ('buildinginfo__num_units', 'Number of Units'),
                ("houseplan__details__residenceType", "Building Type"),
                ("houseplan__thermalEnvelope__foundationType", "Foundation Type"),
                ("analysis__hersScore", "HERS Score"),
                (
                    "analysis__compliances__IECC2018ERI",
                    {
                        "label": "Passes IECC 2018 ERI",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2018Prescriptive",
                    {
                        "label": "Passes IECC 2018 Prescriptive",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2018Performance",
                    {
                        "label": "Passes IECC 2018 Performance",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2015ERI",
                    {
                        "label": "Passes IECC 2015 ERI",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2015Prescriptive",
                    {
                        "label": "Passes IECC 2015 Prescriptive",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2015Performance",
                    {
                        "label": "Passes IECC 2015 Performance",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2012Prescriptive",
                    {
                        "label": "Passes IECC 2012 Prescriptive",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2012Performance",
                    {
                        "label": "Passes IECC 2012 Performance",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2009Prescriptive",
                    {
                        "label": "Passes IECC 2009 Prescriptive",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2009Performance",
                    {
                        "label": "Passes IECC 2009 Performance",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2006Prescriptive",
                    {
                        "label": "Passes IECC 2006 Prescriptive",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__IECC2006Performance",
                    {
                        "label": "Passes IECC 2006 Performance",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__EnergyStarV3",
                    {
                        "label": "Passes EnergyStar V3",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__EnergyStarV31",
                    {
                        "label": "Passes EnergyStar V3.1",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                (
                    "analysis__compliances__TaxCredit45L",
                    {
                        "label": "Passes TaxCredit 45L",
                        "clean": _get_nullboolean_or_warning,
                    },
                ),
                # ('project__version', 'REM/Rate Version'),
                # ('flavor', 'REM/Rate Flavor'),
                (
                    "analysis__created_date",
                    {
                        "label": "Ekotrope Simulation Date",
                        "clean": self.get_formatted_datetime,
                    },
                ),
                # ('remrate_user__username', 'REM/Rate Username'),
            ]
        )

        replace_dict = OrderedDict(
            [k, (v["label"] if isinstance(v, dict) else v)] for k, v in items.items()
        )
        clean_dict = OrderedDict([k, v["clean"]] for k, v in items.items() if "clean" in v)

        objects = HousePlan.objects.filter(id__in=list(object_map.values()))
        structure = self.assign_json(
            HousePlan,
            include=list(replace_dict.keys()),
            section="simulation_basic",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        # This ruins the flow of the deferred values_list query, but as a JSON field with lots of
        # these values inside, it's not clear how to play nice.

        objects = objects.select_related("project")
        values = []
        for houseplan in objects:
            ekotrope_objects = {
                "id": houseplan.id,
                "project": houseplan.project,
                "houseplan": houseplan,
                "analysis": houseplan.analysis,
            }
            values.append(self._get_from_json(list(replace_dict.keys()), ekotrope_objects))
        data = self.merge_results(values, structure)
        return data

    def get_ekotrope_advanced_data(self, object_map):
        from axis.ekotrope.models import HousePlan

        def get_ident(ident):
            return next((key for key, value in object_map.items() if value == ident))

        def invert_uFactor(u):
            if u is None:
                return self.default_none
            u_factors = [float(bit.strip()) for bit in u.split(",")]
            return ", ".join(["{}".format(1.0 / u) for u in u_factors])

        def round_3(ident):
            if ident is None:
                return self.default_none
            values = [Decimal("%.3f" % float(bit)) for bit in ident.split(",")]
            return ", ".join(["{}".format(value) for value in values])

        items = OrderedDict(
            [
                ## HOUSEPLAN INFO
                (
                    "ekotrope_framedfloors",
                    OrderedDict(
                        [
                            ("houseplan__thermalEnvelope__framedFloors__name", "Frame Floor Name"),
                            (
                                "houseplan__thermalEnvelope__framedFloors__surfaceArea",
                                "Frame Floor Area (sq.ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__framedFloors__betweenInteriorAnd",
                                "Frame Floor Location",
                            ),
                            (
                                "houseplan__thermalEnvelope__framedFloors__type__assemblyDetails__continuousR",
                                "Frame Floor Continuous Insulation (R-value)",
                            ),
                            (
                                "houseplan__thermalEnvelope__framedFloors__type__assemblyDetails__cavityR",
                                "Frame Floor Framed Cavity Insulation (R-value)",
                            ),
                            (
                                "houseplan__thermalEnvelope__framedFloors__type__name",
                                "Frame Floor Component Name ",
                            ),
                            (
                                "houseplan__thermalEnvelope__framedFloors__type__uFactor",
                                {
                                    "label": "Frame Floor Component U-Value",
                                    "clean": round_3,
                                },
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_rimjoists",
                    OrderedDict(
                        [
                            (
                                "houseplan__thermalEnvelope__rimJoists__name",
                                "Rim and Band Joist Name",
                            ),
                            (
                                "houseplan__thermalEnvelope__rimJoists__surfaceArea",
                                "Rim and Band Joist Area (sq.ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__rimJoists__betweenInteriorAnd",
                                "Rim and Band Joist Location",
                            ),
                            (
                                "houseplan__thermalEnvelope__rimJoists__type__name",
                                "Rim and Band Joist Type Name",
                            ),
                            (
                                "houseplan__thermalEnvelope__rimJoists__type__uFactor",
                                {
                                    "label": "Rim and Band Joist Type Composite Insulation (R-value)",
                                    "clean": invert_uFactor,
                                },
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_windows",
                    OrderedDict(
                        [
                            ("houseplan__thermalEnvelope__windows__name", "Window Name"),
                            (
                                "houseplan__thermalEnvelope__windows__surfaceArea",
                                "Window Glazing area",
                            ),
                            ("houseplan__thermalEnvelope__windows__type__name", "Window Type Name"),
                            ("houseplan__thermalEnvelope__windows__type__shgc", "Window Type SHGC"),
                            (
                                "houseplan__thermalEnvelope__windows__type__uFactor",
                                "Window Type U-Factor",
                            ),
                        ]
                    ),
                ),
                # ('ekotrope_skylights', OrderedDict([
                #     ('houseplan__thermalEnvelope__skylights__surfaceArea', 'Window Information Glazing area'),
                #     ('houseplan__thermalEnvelope__skylights__type__name', 'Window Information Type Name'),
                #     ('houseplan__thermalEnvelope__skylights__type__shgc', 'Window Information SHGC'),
                #     ('houseplan__thermalEnvelope__skylights__type__uFactor', 'Window Information Type U-Factor'),
                # ])),
                (
                    "ekotrope_walls",
                    OrderedDict(
                        [
                            ("houseplan__thermalEnvelope__walls__name", "Above-Grade Wall Name"),
                            (
                                "houseplan__thermalEnvelope__walls__surfaceArea",
                                "Above-Grade Wall Area (sq.ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__walls__betweenInteriorAnd",
                                "Above-Grade Wall Location",
                            ),
                            (
                                "houseplan__thermalEnvelope__walls__type__assemblyDetails__continuousR",
                                "Above-Grade Wall Continuous Insulation (R-value)",
                            ),
                            (
                                "houseplan__thermalEnvelope__walls__type__assemblyDetails__cavityR",
                                "Above-Grade Wall Framed Cavity Insulation (R-value)",
                            ),
                            (
                                "houseplan__thermalEnvelope__walls__type__name",
                                "Above-Grade Wall Component Name ",
                            ),
                            (
                                "houseplan__thermalEnvelope__walls__type__uFactor",
                                {
                                    "label": "Above-Grade Wall Component U-Value ",
                                    "clean": round_3,
                                },
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_ceilings",
                    OrderedDict(
                        [
                            ("houseplan__thermalEnvelope__ceilings__name", "Ceiling Name"),
                            (
                                "houseplan__thermalEnvelope__ceilings__surfaceArea",
                                "Ceiling Area (sq.ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__ceilings__betweenInteriorAnd",
                                "Ceiling Location",
                            ),
                            (
                                "houseplan__thermalEnvelope__ceilings__clayConcreteRoofTiles",
                                "Ceiling Clay Concrete Roof Tiles",
                            ),
                            (
                                "houseplan__thermalEnvelope__ceilings__type__assemblyDetails__continuousR",
                                "Ceiling Continuous Insulation (R-value)",
                            ),
                            (
                                "houseplan__thermalEnvelope__ceilings__type__assemblyDetails__cavityR",
                                "Ceiling Framed Cavity Insulation (R-value)",
                            ),
                            (
                                "houseplan__thermalEnvelope__ceilings__type__name",
                                "Ceiling Component Name ",
                            ),
                            (
                                "houseplan__thermalEnvelope__ceilings__type__uFactor",
                                {
                                    "label": "Ceiling Component U-Value",
                                    "clean": round_3,
                                },
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_foundationwalls",
                    OrderedDict(
                        [
                            (
                                "houseplan__thermalEnvelope__foundationWalls__name",
                                "Foundation Wall Name",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__surfaceArea",
                                "Foundation Wall Area (sq.ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__heightAboveGrade",
                                "Foundation Wall Height Above Grade (ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__depthBelowGrade",
                                "Foundation Wall Depth Below Grade (ft.)",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__perimeter",
                                "Foundation Wall Perimeter",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__encloses",
                                "Foundation Wall Encloses",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__type__name",
                                "Foundation Wall Type Name",
                            ),
                            (
                                "houseplan__thermalEnvelope__foundationWalls__type__uFactor",
                                "Foundation Wall Type U-Factor",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_doors",
                    OrderedDict(
                        [
                            ("houseplan__thermalEnvelope__doors__name", "Door Name"),
                            (
                                "houseplan__thermalEnvelope__doors__surfaceArea",
                                "Door Area (sq.ft.)",
                            ),
                            ("houseplan__thermalEnvelope__doors__type__name", "Door Type Name"),
                            (
                                "houseplan__thermalEnvelope__doors__type__uFactor",
                                "Door Type U-Factor",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_infiltration",
                    OrderedDict(
                        [
                            (
                                "houseplan__thermalEnvelope__infiltration__specificLeakageArea",
                                "Infiltration Specific Leakage Area",
                            ),
                            (
                                "houseplan__thermalEnvelope__infiltration__effectiveLeakageArea",
                                "Infiltration Effective Leakage Area",
                            ),
                            (
                                "houseplan__thermalEnvelope__infiltration__cfm50",
                                "Infiltration CFM50",
                            ),
                            (
                                "houseplan__thermalEnvelope__infiltration__ach50",
                                "Infiltration ACH50",
                            ),
                            (
                                "houseplan__thermalEnvelope__infiltration__coolingNaturalACH",
                                "Infiltration Cooling Natural ACH",
                            ),
                            (
                                "houseplan__thermalEnvelope__infiltration__heatingNaturalACH",
                                "Infiltration Heating Natural ACH",
                            ),
                        ]
                    ),
                ),
                # ('ekotrope_slabs', OrderedDict([
                #     ('houseplan__slabs__name', 'XXX Name'),
                #     ('houseplan__slabs__surfaceArea', 'XXX Area (sq.ft.)'),
                # ])),
                (
                    "ekotrope_mechanicalequipment",
                    OrderedDict(
                        [
                            (
                                "houseplan__mechanicals__mechanicalEquipment__name",
                                "Mechanical Equipment Name",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__heatingPercentLoad",
                                "Mechanical Equipment Heating Percent Load",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__hotWaterPercentLoad",
                                "Mechanical Equipment Hot Water Percent Load",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__coolingPercentLoad",
                                "Mechanical Equipment Cooling Percent Load",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__name",
                                "Mechanical Equipment Type Name",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__fuel",
                                "Mechanical Equipment Type Fuel",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__heating__efficiency",
                                "Mechanical Equipment Type Heating Efficiency",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__heating__capacity",
                                "Mechanical Equipment Type Heating Capacity",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__heating__efficiencyType",
                                "Mechanical Equipment Type Heating Efficiency Type",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__hotWater__efficiency",
                                "Mechanical Equipment Type Hot Water Efficiency",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__hotWater__tankSize",
                                "Mechanical Equipment Type Hot Water Tank Size",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__hotWater__efficiencyType",
                                "Mechanical Equipment Type Hot Water Efficiency Type",
                            ),
                            (
                                "houseplan__mechanicals__mechanicalEquipment__type__hotWater__isTankless",
                                "Mechanical Equipment Type Hot Water Is Tankless",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_equipment",
                    OrderedDict(
                        [
                            ("houseplan__mechanicals__equipment__equipmentType", "Equipment Type"),
                            (
                                "houseplan__mechanicals__equipment__heating__efficiency",
                                "Equipment Heating Efficiency",
                            ),
                            (
                                "houseplan__mechanicals__equipment__heating__efficiencyType",
                                "Equipment Heating Efficiency Type",
                            ),
                            (
                                "houseplan__mechanicals__equipment__heating__percentLoad",
                                "Equipment Heating Percent Load",
                            ),
                            (
                                "houseplan__mechanicals__equipment__waterHeating__efficiency",
                                "Equipment Water Heating Efficiency",
                            ),
                            (
                                "houseplan__mechanicals__equipment__waterHeating__efficiencyType",
                                "Equipment Water Heating Efficiency Type",
                            ),
                            (
                                "houseplan__mechanicals__equipment__waterHeating__percentLoad",
                                "Equipment Water Heating Percent Load",
                            ),
                            (
                                "houseplan__mechanicals__equipment__cooling__efficiency",
                                "Equipment Cooling Efficiency",
                            ),
                            (
                                "houseplan__mechanicals__equipment__cooling__efficiencyType",
                                "Equipment Cooling Efficiency Type",
                            ),
                            (
                                "houseplan__mechanicals__equipment__cooling__percentLoad",
                                "Equipment Cooling Percent Load",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_mechanicalventilation",
                    OrderedDict(
                        [
                            (
                                "houseplan__mechanicalVentilation__ventilationType",
                                "Mechanical Ventilation Type",
                            ),
                            (
                                "houseplan__mechanicalVentilation__ventilationRate",
                                "Mechanical Ventilation Rate (cfm)",
                            ),
                            (
                                "houseplan__mechanicalVentilation__operationalHoursPerDay",
                                "Mechanical Ventilation Hours per day",
                            ),
                            (
                                "houseplan__mechanicalVentilation__watts",
                                "Mechanical Ventilation Fan Power (watts)",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_distributionsystems",
                    OrderedDict(
                        [
                            (
                                "houseplan__mechanicals__distributionSystems__systemType",
                                "Distribution System Type",
                            ),
                            (
                                "houseplan__mechanicals__distributionSystems__isTested",
                                "Distribution System Is Tested",
                            ),
                            (
                                "houseplan__mechanicals__distributionSystems__untestedDetails__allEquipmentInConditionedSpace",
                                "Distribution System Untested Details All Equipment In Conditioned Space",
                            ),
                            (
                                "houseplan__mechanicals__distributionSystems__untestedDetails__ductSystemEfficiency",
                                "Distribution System Untested Details Duct System Efficiency",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_thermostats",
                    OrderedDict(
                        [
                            (
                                "houseplan__mechanicals__thermostats__thermostatType",
                                "Thermostat Type",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_appliances",
                    OrderedDict(
                        [
                            (
                                "houseplan__appliances__clothesDryer__utilizationFactor",
                                "Dryer Utilization factor",
                            ),
                            (
                                "houseplan__appliances__clothesWasher__integratedModifiedEnergyFactor",
                                "Washer IMEF",
                            ),
                            ("houseplan__appliances__clothesWasher__electricRate", "Electric Rate"),
                            ("houseplan__appliances__clothesWasher__gasRate", "Gas Rate"),
                            (
                                "houseplan__appliances__clothesWasher__annualGasCost",
                                "Natural Gas Operating Cost",
                            ),
                        ]
                    ),
                ),
                ## ANALYSIS INFO
                (
                    "ekotrope_hers",
                    OrderedDict(
                        [
                            ("analysis__hersScore", "HERS Score"),
                            ("analysis__hersScoreNoPv", "HERS Score No PV"),
                        ]
                    ),
                ),
                (
                    "ekotrope_breakdownbyfuel",
                    OrderedDict(
                        [
                            (
                                "analysis__energy__breakdown__byFuel__fuel",
                                "Energy Breakdown By Fuel Type",
                            ),
                            (
                                "analysis__energy__breakdown__byFuel__cost",
                                "Energy Breakdown By Fuel Cost",
                            ),
                            (
                                "analysis__energy__breakdown__byFuel__heatingConsumption",
                                "Energy Breakdown By Fuel Heating Consumption",
                            ),
                            (
                                "analysis__energy__breakdown__byFuel__waterHeatingConsumption",
                                "Energy Breakdown By Fuel Water Heating Consumption",
                            ),
                            (
                                "analysis__energy__breakdown__byFuel__coolingConsumption",
                                "Energy Breakdown By Fuel Cooling Consumption",
                            ),
                            (
                                "analysis__energy__breakdown__byFuel__lightingAndAppliancesConsumption",
                                "Energy Breakdown By Fuel Lighting And Appliances Consumption",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_breakdownbycomponent",
                    OrderedDict(
                        [
                            (
                                "analysis__energy__breakdown__byComponent__category",
                                "Energy Breakdown By Component Category",
                            ),
                            (
                                "analysis__energy__breakdown__byComponent__heatingLoad",
                                "Energy Breakdown By Component Heating Load",
                            ),
                            (
                                "analysis__energy__breakdown__byComponent__coolingLoad",
                                "Energy Breakdown By Component Cooling Load",
                            ),
                        ]
                    ),
                ),
                (
                    "ekotrope_analysissummary",
                    OrderedDict(
                        [
                            (
                                "analysis__energy__summary__coolingConsumption",
                                "Energy Summary Cooling Consumption",
                            ),
                            (
                                "analysis__energy__summary__heatingConsumption",
                                "Energy Summary Heating Consumption",
                            ),
                            (
                                "analysis__energy__summary__lightingAndAppliancesConsumption",
                                "Energy Summary Lighting And Appliances Consumption",
                            ),
                            (
                                "analysis__energy__summary__waterHeatingConsumption",
                                "Energy Summary Water Heating Consumption",
                            ),
                            (
                                "analysis__energy__summary__solarGeneration",
                                "Energy Summary Solar Generation",
                            ),
                            (
                                "analysis__energy__summary__summerElectricPowerPeak",
                                "Energy Summary Summer Electric Power Peak",
                            ),
                            (
                                "analysis__energy__summary__winterElectricPowerPeak",
                                "Energy Summary Winter Electric Power Peak",
                            ),
                            (
                                "analysis__energy__summary__generationSavings",
                                "Energy Generation Savings",
                            ),
                            ("analysis__energy__summary__cost", "Energy Summary Cost"),
                        ]
                    ),
                ),
                ## ENERGY STAR V3 Reference ANALYSIS INFO
                (
                    "ekotrope_EnergyStarV31Reference",
                    OrderedDict(
                        [
                            (
                                "analysis__building_types__EnergyStarV3Reference__hersScore",
                                "ENERGY STAR V3 Reference HERS Score",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__hersScoreNoPv",
                                "ENERGY STAR V3 Reference HERS Score No PV",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byFuel__fuel",
                                "ENERGY STAR V3 Reference Energy Breakdown By Fuel Type",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byFuel__cost",
                                "ENERGY STAR V3 Reference Energy Breakdown By Fuel Cost",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byFuel__heatingConsumption",
                                "ENERGY STAR V3 Reference Energy Breakdown By Fuel Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byFuel__waterHeatingConsumption",
                                "ENERGY STAR V3 Reference Energy Breakdown By Fuel Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byFuel__coolingConsumption",
                                "ENERGY STAR V3 Reference Energy Breakdown By Fuel Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byFuel__lightingAndAppliancesConsumption",
                                "ENERGY STAR V3 Reference Energy Breakdown By Fuel Lighting And Appliances Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byComponent__category",
                                "ENERGY STAR V3 Reference Energy Breakdown By Component Category",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byComponent__heatingLoad",
                                "ENERGY STAR V3 Reference Energy Breakdown By Component Heating Load",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__breakdown__byComponent__coolingLoad",
                                "ENERGY STAR V3 Reference Energy Breakdown By Component Cooling Load",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__coolingConsumption",
                                "ENERGY STAR V3 Reference Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__heatingConsumption",
                                "ENERGY STAR V3 Reference Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__lightingAndAppliancesConsumption",
                                "ENERGY STAR V3 Reference Lights & Appliance Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__waterHeatingConsumption",
                                "ENERGY STAR V3 Reference Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__solarGeneration",
                                "ENERGY STAR V3 Reference Energy Summary Solar Generation",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__summerElectricPowerPeak",
                                "ENERGY STAR V3 Reference Energy Summary Summer Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__winterElectricPowerPeak",
                                "ENERGY STAR V3 Reference Energy Summary Winter Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__generationSavings",
                                "ENERGY STAR V3 Reference Energy Generation Savings",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3Reference__energy__summary__cost",
                                "ENERGY STAR V3 Reference Cost",
                            ),
                        ]
                    ),
                ),
                ## ANALYSIS INFO
                (
                    "ekotrope_EnergyStarV31Reference",
                    OrderedDict(
                        [
                            (
                                "analysis__building_types__EnergyStarV31Reference__hersScore",
                                "ENERGY STAR V3.1 Reference HERS Score",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__hersScoreNoPv",
                                "ENERGY STAR V3.1 Reference HERS Score No PV",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byFuel__fuel",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Fuel Type",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byFuel__cost",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Fuel Cost",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byFuel__heatingConsumption",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Fuel Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byFuel__waterHeatingConsumption",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Fuel Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byFuel__coolingConsumption",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Fuel Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byFuel__lightingAndAppliancesConsumption",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Fuel Lighting And Appliances Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byComponent__category",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Component Category",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byComponent__heatingLoad",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Component Heating Load",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__breakdown__byComponent__coolingLoad",
                                "ENERGY STAR V3.1 Reference Energy Breakdown By Component Cooling Load",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__coolingConsumption",
                                "ENERGY STAR V3.1 Reference Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__heatingConsumption",
                                "ENERGY STAR V3.1 Reference Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__lightingAndAppliancesConsumption",
                                "ENERGY STAR V3.1 Reference Lights & Appliance Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__waterHeatingConsumption",
                                "ENERGY STAR V3.1 Reference Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__solarGeneration",
                                "ENERGY STAR V3.1 Reference Energy Summary Solar Generation",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__summerElectricPowerPeak",
                                "ENERGY STAR V3.1 Reference Energy Summary Summer Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__winterElectricPowerPeak",
                                "ENERGY STAR V3.1 Reference Energy Summary Winter Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__EnergyStarV3R1eference__energy__summary__generationSavings",
                                "ENERGY STAR V3.1 Reference Energy Generation Savings",
                            ),
                            (
                                "analysis__building_types__EnergyStarV31Reference__energy__summary__cost",
                                "ENERGY STAR V3.1 Reference Cost",
                            ),
                        ]
                    ),
                ),
                ## ANALYSIS INFO
                (
                    "ekotrope_IECC2018Reference",
                    OrderedDict(
                        [
                            (
                                "analysis__building_types__IECC2018Reference__hersScore",
                                "IECC 2018 Reference HERS Score",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__hersScoreNoPv",
                                "IECC 2018 Reference HERS Score No PV",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byFuel__fuel",
                                "IECC 2018 Reference Energy Breakdown By Fuel Type",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byFuel__cost",
                                "IECC 2018 Reference Energy Breakdown By Fuel Cost",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byFuel__heatingConsumption",
                                "IECC 2018 Reference Energy Breakdown By Fuel Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byFuel__waterHeatingConsumption",
                                "IECC 2018 Reference Energy Breakdown By Fuel Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byFuel__coolingConsumption",
                                "IECC 2018 Reference Energy Breakdown By Fuel Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byFuel__lightingAndAppliancesConsumption",
                                "IECC 2018 Reference Energy Breakdown By Fuel Lighting And Appliances Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byComponent__category",
                                "IECC 2018 Reference Energy Breakdown By Component Category",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byComponent__heatingLoad",
                                "IECC 2018 Reference Energy Breakdown By Component Heating Load",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__breakdown__byComponent__coolingLoad",
                                "IECC 2018 Reference Energy Breakdown By Component Cooling Load",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__coolingConsumption",
                                "IECC 2018 Reference Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__heatingConsumption",
                                "IECC 2018 Reference Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__lightingAndAppliancesConsumption",
                                "IECC 2018 Reference Lights & Appliance Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__waterHeatingConsumption",
                                "IECC 2018 Reference Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__solarGeneration",
                                "IECC 2018 Reference Energy Summary Solar Generation",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__summerElectricPowerPeak",
                                "IECC 2018 Reference Energy Summary Summer Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__winterElectricPowerPeak",
                                "IECC 2018 Reference Energy Summary Winter Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__generationSavings",
                                "IECC 2018 Reference Energy Generation Savings",
                            ),
                            (
                                "analysis__building_types__IECC2018Reference__energy__summary__cost",
                                "IECC 2018 Reference Cost",
                            ),
                        ]
                    ),
                ),
                ## ANALYSIS INFO
                (
                    "ekotrope_IECC2018Proposed",
                    OrderedDict(
                        [
                            (
                                "analysis__building_types__IECC2018Proposed__hersScore",
                                "IECC 2018 Proposed HERS Score",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__hersScoreNoPv",
                                "IECC 2018 Proposed HERS Score No PV",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byFuel__fuel",
                                "IECC 2018 Proposed Energy Breakdown By Fuel Type",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byFuel__cost",
                                "IECC 2018 Proposed Energy Breakdown By Fuel Cost",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byFuel__heatingConsumption",
                                "IECC 2018 Proposed Energy Breakdown By Fuel Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byFuel__waterHeatingConsumption",
                                "IECC 2018 Proposed Energy Breakdown By Fuel Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byFuel__coolingConsumption",
                                "IECC 2018 Proposed Energy Breakdown By Fuel Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byFuel__lightingAndAppliancesConsumption",
                                "IECC 2018 Proposed Energy Breakdown By Fuel Lighting And Appliances Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byComponent__category",
                                "IECC 2018 Proposed Energy Breakdown By Component Category",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byComponent__heatingLoad",
                                "IECC 2018 Proposed Energy Breakdown By Component Heating Load",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__breakdown__byComponent__coolingLoad",
                                "IECC 2018 Proposed Energy Breakdown By Component Cooling Load",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__coolingConsumption",
                                "IECC 2018 Proposed Cooling Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__heatingConsumption",
                                "IECC 2018 Proposed Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__lightingAndAppliancesConsumption",
                                "IECC 2018 Proposed Lights & Appliance Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__waterHeatingConsumption",
                                "IECC 2018 Proposed Water Heating Consumption",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__solarGeneration",
                                "IECC 2018 Proposed Energy Summary Solar Generation",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__summerElectricPowerPeak",
                                "IECC 2018 Proposed Energy Summary Summer Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__winterElectricPowerPeak",
                                "IECC 2018 Proposed Energy Summary Winter Electric Power Peak",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__generationSavings",
                                "IECC 2018 Proposed Energy Generation Savings",
                            ),
                            (
                                "analysis__building_types__IECC2018Proposed__energy__summary__cost",
                                "IECC 2018 Proposed Cost",
                            ),
                        ]
                    ),
                ),
            ]
        )

        data_list = []
        for section, section_dict in items.items():
            replace_dict = OrderedDict(
                [["id", "House Plan ID"]]
                + [[k, v["label"] if isinstance(v, dict) else v] for k, v in section_dict.items()]
            )
            clean_dict = OrderedDict(
                [["id", get_ident]]
                + [[k, v["clean"]] for k, v in section_dict.items() if "clean" in v]
            )

            objects = HousePlan.objects.filter(id__in=list(object_map.values()))
            structure = self.assign_json(
                HousePlan,
                include=list(replace_dict.keys()),
                section=section,
                replace_dict=replace_dict,
                clean_dict=clean_dict,
            )

            objects = objects.select_related("project")
            values = []
            for houseplan in objects:
                ekotrope_objects = {
                    "id": houseplan.id,
                    "project": houseplan.project,
                    "houseplan": houseplan,
                    "analysis": houseplan.analysis,
                }
                values.append(self._get_from_json(list(replace_dict.keys()), ekotrope_objects))
            data = self.merge_results(values, structure)
            data_list.append(data)

        return data_list

    def get_ekotrope_data(self, advanced=False):
        object_map = dict(
            self.get_queryset().values_list("id", "floorplan__ekotrope_houseplan__id")
        )
        # assert set(object_map.values()) != set(object_map.keys()), "Hmmm"

        results = OrderedDict([(x, []) for x in self.get_queryset().values_list("id", flat=True)])

        self.log.debug("Gathering Ekotrope Basic data")
        data = self.get_ekotrope_basic_data(object_map)
        results = self.munge_data(results, data)

        if advanced:
            self.log.debug("Gathering Ekotrope Advanced data")
            data_list = self.get_ekotrope_advanced_data(object_map)
            for data in data_list:
                results = self.munge_data(results, data)

        final = OrderedDict()
        field = "simulation_advanced" if advanced else "simulation_basic"
        for _id, houseplan_id in object_map.items():
            cell_obj = CellObject("id", "ID", None, field, _id, _id)
            found = False
            for value_set in results.values():
                # if not len(value_set) and not self.retain_empty:
                if not len(value_set):
                    continue
                ident = next((v for v in value_set if v.attr == "id")).raw_value
                if ident == houseplan_id:
                    final[_id] = [cell_obj] + value_set[1:]
                    found = True
                    break
            if not found:
                final[_id] = [cell_obj]
        return list(final.values())

    def get_sampleset_data(self):
        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("samplesethomestatus__sampleset__uuid", "Sample Set ID"),
                ("samplesethomestatus__sampleset__alt_name", "Sample Set Alt Name"),
                ("samplesethomestatus__sampleset__confirm_date", "Sample Set Confirm Date"),
            ]
        )

        def trim_slug(ident):
            try:
                return ident[:7]
            except TypeError:
                return self.default_none

        clean_dict = {"samplesethomestatus__sampleset__uuid": trim_slug}

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="sampleset",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related).distinct()
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_annotation_data(self):
        from axis.eep_program.models import EEPProgram

        section_name = "annotations"
        used_program_ids = list(self.get_queryset().values_list("eep_program", flat=True))
        eeps = EEPProgram.objects.filter_by_user(self.user).filter(id__in=used_program_ids)

        base_ann_types = eeps.values_list(
            "id",
            "required_annotation_types__id",
            "required_annotation_types__name",
            "required_annotation_types__slug",
        )

        ann_type_ids = list(set([a_id for _ep, a_id, _n, _s in base_ann_types if a_id]))

        objects = self.get_queryset().filter(annotations__type__id__in=ann_type_ids)

        # remove unnecessary and duplicate columns
        if "ngbs_data" in self.report_on:
            objects = objects.exclude(
                annotations__type__slug__in=[
                    "certified-nat-gbs",
                    "certification-standard",
                    "certification-date",
                    "certification-number",
                    "certification-record-id",
                    "project-id",
                    "unit-count",
                    "hers-score",
                    "hud-disaster-case-number",
                ]
            )

        include = ["id", "eep_program__id", "annotations__type__slug", "annotations__content"]
        structure = self.assign_basic(EEPProgramHomeStatus, include=include, section=section_name)
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        results = []
        for stat_id, eep_id in self.get_queryset().values_list("id", "eep_program_id"):
            # Add in all of the Annotations to a single set of results
            result = [["id", "ID", None, section_name, stat_id, stat_id]]
            for eep_id, at_id, at_name, at_slug in base_ann_types:
                if not at_id:
                    continue
                result.append(["ann-type-{}".format(at_slug), at_name, None, section_name, [], []])

            # Join any results we have for this home to the annotations
            home_stat_anns = next((x for x in data if x[0][-1] == stat_id), None)
            if home_stat_anns:
                home_stat_type_slugs = home_stat_anns[-2][-2]
                home_stat_type_contents = home_stat_anns[-1][-2]
                assert len(home_stat_type_slugs) == len(home_stat_type_contents)
                for slug, value in dict(zip(home_stat_type_slugs, home_stat_type_contents)).items():
                    result_item = next(
                        (x for x in result if x[0] == "ann-type-{}".format(slug)), None
                    )
                    if result_item:
                        result[result.index(result_item)][-2].append(value)

            # Clean up and prune it
            final = []
            for item in result:
                if item[-1]:
                    pass
                elif len(item[-2]) > 1 and self.join_resulting_values:
                    item[-1] = ", ".join(["{}".format(v) for v in item[-1]])
                elif len(item[-2]) == 1:
                    item[-1] = item[-2][0]
                else:
                    item[-1] = item[-2] = self.default_none
                final.append(CellObject(*item))
            results.append(final)

        return results

    def get_collection_answer_data(self, only_latest_answer=True):
        from axis.eep_program.models import EEPProgram

        section_name = "checklist_answers"

        eeps = list(
            self.get_queryset()
            .values_list("eep_program__id", flat=True)
            .order_by("eep_program__id")
        )
        eeps = EEPProgram.objects.filter(id__in=eeps, collection_request__isnull=False).order_by(
            "id"
        )

        base_instruments = eeps.values_list(
            "pk",
            "collection_request__collectioninstrument__text",
            "collection_request__collectioninstrument__measure_id",
            "collection_request__collectioninstrument__order",
        ).order_by("pk")

        # Order the questions based on the program, order, text
        _sorted_base_instruments = list(sorted(base_instruments, key=lambda x: (x[0], x[-1], x[1])))

        # if the measure and the text is the same then we can consider them the same
        # for a column in excel
        sorted_base_instruments = []
        for k, v in itertools.groupby(_sorted_base_instruments, lambda d: (d[1], d[2])):
            sorted_base_instruments.append(list(v)[0])

        inputs = self.get_queryset().values_list(
            "pk",
            "eep_program_id",
            "collection_request__collectedinput_set__instrument__measure_id",
            "collection_request__collectedinput_set__data",
        )

        from django_input_collection.models import CollectionRequest
        from axis.checklist.collection import ChecklistCollector

        collection_request_ids = list(
            inputs.values_list("collection_request", flat=True).distinct()
        )
        context = {
            "user": self.user,
            "user_role": "rater",
        }
        collectors = {
            (
                cr.eepprogramhomestatus.id,
                cr.eepprogramhomestatus.eep_program_id,
            ): ChecklistCollector(cr, **context)
            for cr in CollectionRequest.objects.filter(id__in=collection_request_ids)
        }

        results = []
        for (stat_id, eep_id), collector in collectors.items():
            result = [CellObject("id", "ID", None, section_name, stat_id, stat_id)]
            for pk, text, measure, _order in sorted_base_instruments:
                collector_method = collector.get_method(measure=measure)
                raw_answers = [
                    ans
                    for _pk, _eep_id, _measure, ans in inputs
                    if _pk == stat_id and _eep_id == eep_id and _measure == measure
                ]
                comment = None
                if raw_answers == [None]:
                    continue
                if not len(raw_answers):
                    answer = self.default_none
                elif len(raw_answers) == 1:
                    _answer = raw_answers[0]
                    answer = collector_method.get_data_display(_answer["input"])
                    comment = _answer.get("comment") or None  # Avoid '' as a db value
                elif len(raw_answers) > 1:
                    if only_latest_answer:
                        _answer = raw_answers[-1]
                        answer = collector_method.get_data_display(_answer["input"])
                        comment = _answer.get("comment") or None  # Avoid '' as a db value
                    else:
                        log.warning(
                            "Stat: %d Measure: %r Collected Inputs: %r",
                            stat_id,
                            measure,
                            raw_answers,
                        )
                        answer = ", ".join(
                            [collector_method.get_data_display(x["input"]) for x in raw_answers]
                        )
                        comment = ", ".join([(x.get("comment") or "-") for x in raw_answers])

                result.append(
                    CellObject(measure, text, None, section_name, raw_answers, (answer, comment))
                )
            results.append(result)
        return results

    # flake8: noqa: C901
    def get_checklist_answer_data(self, include_sampleset_unbounded=True):
        from axis.eep_program.models import EEPProgram

        section_name = "checklist_answers"

        eeps = list(self.get_queryset().values_list("eep_program__id", flat=True))
        eeps = EEPProgram.objects.filter(
            id__in=eeps, required_checklists__isnull=False, collection_request__isnull=True
        )
        base_questions = eeps.values_list(
            "id",
            "required_checklists__questions__id",
            "required_checklists__questions__question",
            "required_checklists__questions__slug",
            "required_checklists__id",
            "required_checklists__questions__priority",
        )

        # Order the questions based on the checklist id, question priority, question text
        sorted_base_questions = sorted(base_questions, key=lambda x: (x[-2], x[-1], x[1]))

        # Work through the questions and get a good list
        question_ids = []
        for _e_id, q_id, _qq, _qs, _c_id, _q_p in sorted_base_questions:
            if q_id and q_id not in question_ids:
                question_ids.append(q_id)

        ss_qs = (
            self.get_queryset()
            .in_sampleset()
            .values_list("id", "samplesethomestatus__sampleset__id")
            .annotate(n=Max("samplesethomestatus__revision"))
        )
        ss_dict = {}
        if ss_qs:
            for _id, ss, _ in ss_qs:
                if ss not in ss_dict.keys():
                    ss_dict[ss] = []
                ss_dict[ss].append(_id)

        # Get all of the answers - even those outside of my view..
        objects = self.get_queryset().filter(home__answer__question__id__in=question_ids)
        o_ids = list(objects.values_list("id", flat=True))
        o_ids += list(self.get_queryset().in_sampleset().values_list("id", flat=True))

        objects = (
            EEPProgramHomeStatus.objects.filter(id__in=o_ids)
            .distinct()
            .filter(home__answer__question__id__in=question_ids)
        )

        log.debug("Merging answers..")
        include = [
            "id",
            "eep_program__id",
            "home__answer__question__slug",
            "home__answer__answer",
            "home__answer__comment",
        ]
        structure = self.assign_basic(EEPProgramHomeStatus, include=include, section=section_name)
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        log.debug("Done merging answers..")
        results = []
        _query_set = self.get_queryset().values_list(
            "id",
            "eep_program_id",
            "samplesethomestatus__revision",
            "samplesethomestatus__is_test_home",
            "samplesethomestatus__sampleset__revision",
            "samplesethomestatus__sampleset__id",
            "samplesethomestatus__sampleset__uuid",
        )

        handled_ids = set()
        stat_ids = list(self.get_queryset().values_list("id", flat=True))
        my_stat_ids = list(
            EEPProgramHomeStatus.objects.filter_by_user(self.user).values_list("id", flat=True)
        )
        # The _query_set variable will have duplicate entries for the same id if multiple samplesets
        # have participated in its lifetime.  We will force "continue" duplicates
        for stat_id, eep_id, sshs_rev, is_test_home, ss_rev, ss_id, ss_slug in _query_set:
            if stat_id in handled_ids:
                continue
            if sshs_rev not in (None, ss_rev):
                continue

            handled_ids.add(stat_id)

            # Add in all of the questions to a single set of results
            result = [CellObject("id", "ID", None, section_name, stat_id, stat_id)]

            # Now work through the answers and lets fill them in.
            homes_stat_answers = next((x for x in data if x[0][-1] == stat_id), None)
            homes_stat_answers = [homes_stat_answers] if homes_stat_answers else []

            if ss_id:  # Sampled
                homes_stat_answers += [
                    x for x in data if x[0][-1] in ss_dict.get(ss_id, []) if x[0][-1] != stat_id
                ]

            # Create a dictionary to hold all of answers which contribute to this status.
            answers_dict = {}
            has_outside = False
            for home_stat_answers in homes_stat_answers:
                active = home_stat_answers[0][-1] == stat_id
                via_ss = home_stat_answers[0][-1] != stat_id  # These answers came via the SS
                outside_queryset = (
                    home_stat_answers[0][-1] not in stat_ids
                )  # These came outside the QS
                outside_my_view = home_stat_answers[0][-1] not in my_stat_ids
                home_stat_question_slugs = home_stat_answers[-3][-2]
                home_stat_question_answers = home_stat_answers[-2][-2]
                home_stat_question_answers_comments = home_stat_answers[-1][-2]
                answer_comments = zip(
                    home_stat_question_answers, home_stat_question_answers_comments
                )
                for slug, value in dict(zip(home_stat_question_slugs, answer_comments)).items():
                    comments = []
                    if isinstance(value, tuple):
                        value, _comment = value
                        if _comment not in ["", None, "None"]:
                            comments.append(_comment)
                    if slug not in answers_dict.keys():
                        answers_dict[slug] = []
                    if active:
                        answers_dict[slug].append((value, "\n".join(comments)))
                    else:
                        # Only add the answer if we don't have a direct contributor
                        if not len(answers_dict[slug]):
                            _scomment = ["Note: Answer originated from"]
                            if via_ss:
                                _scomment.append("Sample Set {}".format(ss_slug[:8]))
                            if outside_queryset:
                                _scomment.append(
                                    "{}outside of basic filters".format(
                                        "and " if len(_scomment) > 1 else ""
                                    )
                                )
                                has_outside = True
                            if len(_scomment) != 1:
                                comments.append(" ".join(_scomment))
                                answers_dict[slug].append((value, "\n".join(comments)))

            # Iterate over the sorted questions and add in the answers..
            for _e_id, _q_id, q_question, q_slug, _c_id, _q_p in sorted_base_questions:
                if not q_id:
                    continue
                raw_answers = answers_dict.get(q_slug, [])
                answer = self.default_none
                if len(raw_answers) > 1 and self.join_resulting_values:
                    answer = ", ".join(["{}".format(v) for v in raw_answers])
                elif len(raw_answers) == 1:
                    raw_answers = raw_answers[0]
                    answer = raw_answers
                else:
                    raw_answers, answer = None, self.default_none

                result.append(
                    CellObject(q_slug, q_question, None, section_name, raw_answers, answer)
                )

            if has_outside:
                log.info("Outside answers found!!")

            results.append(result)

        return results

    def get_incentive_data(self):
        from axis.incentive_payment.models import IncentiveDistribution
        from axis.incentive_payment.state_machine import IPPItemStateMachine
        from axis.company.strings import COMPANY_TYPES_MAPPING

        section_name = "ipp"

        # This section runs in two blocks, first for the incentivepaymentstatus state, then another
        # pass for the dynamically-sized dataset of ippitem results.

        queryset = self.get_queryset()

        final_data = []
        state_labels = dict(IPPItemStateMachine.get_state_choices())

        # First process the homestatus->incentivepaymentstatus data
        # Side note: I think this concise grouping of parallel data structures is repetitive enough
        # to stand on its own as an example for why the structure/merging/munging strategy could be
        # simplified.
        replace_dict = OrderedDict(
            [
                ("id", "Homestatus Id"),
                ("incentivepaymentstatus__state", "Incentive Payment Status"),
                ("ippitem__cost", "Incentive Total"),
            ]
        )
        include = [
            "id",
            "incentivepaymentstatus__state",
            "ippitem__cost",
        ]

        def get_cost(ident):
            return "${:.2f}".format(ident) if ident else self.default_none

        clean_dict = {
            "incentivepaymentstatus__state": lambda state: state_labels.get(
                state, self.default_none
            ),
            "ippitem__cost": get_cost,
        }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=include,
            section="ipp",
            clean_dict=clean_dict,
            replace_dict=replace_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = queryset.select_related(*select_related).annotate(Sum("ippitem__cost"))
        data = self.merge_results(
            objects.values_list(
                # Drop one-to-many 'ippitem__cost' value in favor of the annotation Sum() field
                *[x.attr for x in structure[:-1]]
                + ["ippitem__cost__sum"]
            ),
            structure,
        )
        final_data.extend(data)

        # Add IPPItem and associated IncentiveDistribution data
        # NOTE: From here to the end, I can't read anything it's doing or why it's doing it. The
        # only wisdom I have to share is that it's having to process a plural dataset and can't
        # send it directly to assign_basic() for a fixed structure.
        ipp_ids = IncentiveDistribution.objects.filter_by_user(self.user).values_list(
            "id", flat=True
        )
        objects = self.get_queryset().filter(ippitem__incentive_distribution__id__in=ipp_ids)
        include = [
            "id",
            "ippitem__incentive_distribution__status",
            "ippitem__incentive_distribution__slug",
            "ippitem__incentive_distribution__company__name",
            "ippitem__incentive_distribution__customer__name",
            "ippitem__incentive_distribution__customer__company_type",
            "ippitem__incentive_distribution__check_requested_date",
            "ippitem__incentive_distribution__paid_date",
            "ippitem__incentive_distribution__check_number",
            "ippitem__cost",
        ]

        select_related = [
            "__".join(item.split("__")[:-1]) for item in include if len(item.split("__")) > 1
        ]
        objects = objects.select_related(*select_related)
        data = objects.values_list(*include)

        cust_types = list(sorted(set(x[5] for x in data if x[5])))

        result = OrderedDict()
        for item in data:
            id, status, slug, co_name, cust_name, cust_type, req_date, pd_date, chk_n, cost = item
            if id not in result:
                result[id] = [
                    CellObject("id", "ID", None, section_name, id, id),
                    CellObject(
                        "incentive_distribution__company__name",
                        "Payee Company",
                        None,
                        section_name,
                        co_name,
                        co_name,
                    ),
                ]
                for _cust_type in cust_types:
                    short = "{}_incentive_distribution__check_requested_date".format(_cust_type)
                    label = "{} Check Request Date".format(COMPANY_TYPES_MAPPING[_cust_type])
                    result[id].append(CellObject(short, label, None, section_name, "", ""))

                    short = "{}_incentive_distribution__paid_date".format(_cust_type)
                    label = "{} Check Paid Date".format(COMPANY_TYPES_MAPPING[_cust_type])
                    result[id].append(CellObject(short, label, None, section_name, "", ""))

                    short = "{}_incentive_distribution__check_number".format(_cust_type)
                    label = "{} Check Number".format(COMPANY_TYPES_MAPPING[_cust_type])
                    result[id].append(CellObject(short, label, None, section_name, "", ""))

                    short = "{}_incentive_distribution__slug".format(_cust_type)
                    label = "{} Invoice".format(COMPANY_TYPES_MAPPING[_cust_type])
                    result[id].append(CellObject(short, label, None, section_name, "", ""))

            try:
                req_date = formats.date_format(req_date, "SHORT_DATE_FORMAT")
            except:
                req_date = self.default_none
            try:
                pd_date = formats.date_format(pd_date, "SHORT_DATE_FORMAT")
            except:
                pd_date = self.default_none

            for idx, _item in enumerate(result[id]):
                if _item[0] == "{}_incentive_distribution__slug".format(cust_type):
                    short = "{}_incentive_distribution__slug".format(cust_type)
                    label = "{} Invoice".format(COMPANY_TYPES_MAPPING[cust_type])
                    value = slug[:7]
                    result[id][idx] = CellObject(short, label, None, section_name, slug, value)
                elif _item[0] == "{}_incentive_distribution__check_requested_date".format(
                    cust_type
                ):
                    short = "{}_incentive_distribution__check_requested_date".format(cust_type)
                    label = "{} Check Requested Date".format(COMPANY_TYPES_MAPPING[cust_type])
                    result[id][idx] = CellObject(short, label, None, section_name, slug, req_date)
                elif _item[0] == "{}_incentive_distribution__paid_date".format(cust_type):
                    short = "{}_incentive_distribution__paid_date".format(cust_type)
                    label = "{} Check Paid Date".format(COMPANY_TYPES_MAPPING[cust_type])
                    result[id][idx] = CellObject(short, label, None, section_name, slug, pd_date)
                elif _item[0] == "{}_incentive_distribution__check_number".format(cust_type):
                    short = "{}_incentive_distribution__check_number".format(cust_type)
                    label = "{} Check Number".format(COMPANY_TYPES_MAPPING[cust_type])
                    result[id][idx] = CellObject(short, label, None, section_name, slug, chk_n)
                elif _item[0] == "{}_incentive_distribution__cost".format(cust_type):
                    short = "{}_incentive_distribution__cost".format(cust_type)
                    label = "{} Incentive".format(COMPANY_TYPES_MAPPING[cust_type])
                    result[id][idx] = CellObject(
                        short, label, None, section_name, cost, "${}".format(cost)
                    )

        # Fold the IPP data into the existing IncentivePaymentStatus-primed data
        for cell_list in final_data:
            # log.error("MAPPING (%d) %r", len(cell_list), cell_list)
            if len(cell_list):  # Contains stubbed paymentstatus id and state, or is empty
                id_cell = cell_list[0]
                cell_list.extend(result.get(id_cell.value, []))
                # log.error("INTO (%d) %r: %r", len(cell_list), id_cell.value, result.get(id_cell.value))

        return final_data

    def get_invoicing_data(self):
        """Invoicing info"""
        section_name = "invoicing"

        # We can't get the total paid balance - join on the field is not permitted
        home_status_ids = list(self.get_queryset().values_list("id", flat=True))
        objects = (
            InvoiceItemGroup.objects.filter_by_user(self.user)
            .filter(home_status_id__in=home_status_ids)
            .select_related(
                "invoice__issuer", "invoice__customer", "home_status__customer_hirl_project"
            )
        )

        project_data = list(
            HIRLProject.objects.filter(home_status_id__in=home_status_ids)
            .annotate_billing_info()
            .values_list("home_status_id", "billing_state", "most_recent_notice_sent")
        )

        def trim_slug(ident):
            try:
                return str(ident)[:8]
            except TypeError:
                return self.default_none

        def get_cost(ident):
            return "${}".format(ident) if ident else "$0.00"

        from axis.invoicing.models import Invoice

        def get_state(ident):
            return dict(Invoice.STATE_CHOICES).get(ident)

        def get_billing_state(ident):
            project = next((x for x in project_data if x[0] == ident), None)
            if project:
                return HIRLProject.BILLING_STATE_DISPLAY[project[1]]

        def clean_date(ident):
            project = next((x for x in project_data if x[0] == ident), None)
            if project and project[2]:
                return formats.date_format(project[2], "SHORT_DATE_FORMAT")
            return self.default_none

        # We have to manually assign this b/c the totals are calculated on the qs
        structure = [
            CellParser("home_status_id", "Project Status", None, section_name),
            CellParser("total", "Total", get_cost, section_name),
            CellParser("total_paid", "Total Paid", get_cost, section_name),
            CellParser("current_balance", "Balance", get_cost, section_name),
            CellParser("invoice_id", "Invoice Number", trim_slug, section_name),
            CellParser("invoice__issuer__name", "Invoice Issuer", None, section_name),
            CellParser("invoice__customer__name", "Invoice Customer", None, section_name),
            CellParser("invoice__state", "Invoice Status", get_state, section_name),
        ]
        if self.user.company.slug == customer_hirl_app.CUSTOMER_SLUG:
            structure += [
                CellParser(
                    "home_status__customer_hirl_project__h_number", "H-Number", None, section_name
                ),
                CellParser("home_status_id", "Billing State", get_billing_state, section_name),
                CellParser(
                    "home_status__customer_hirl_project__is_jamis_milestoned",
                    "JAMIS Milestoned",
                    "get_formatted_boolean",
                    section_name,
                ),
                CellParser("home_status_id", "Most Recent Notice Sent", clean_date, section_name),
            ]

        data = self.merge_results(
            objects.values_list(*[x.attr for x in structure]), structure, key="home_status_id"
        )

        # Munge back our billing state and notices
        results = []
        for _row in data:
            items = []
            for item in _row:
                if item.pretty_name == "Billing State":
                    if isinstance(item.raw_value, list):
                        raw_value = [x[1] for x in project_data if x[0] in item.raw_value]
                    else:
                        raw_value = [x[1] for x in project_data if x[0] == item.raw_value][0]
                    item = item._replace(
                        attr="home_status__customer_hirl_project__billing_state",
                        raw_value=raw_value,
                    )
                elif item.pretty_name == "Most Recent Notice Sent":
                    if isinstance(item.raw_value, list):
                        raw_value = [x[2] for x in project_data if x[0] in item.raw_value]
                    else:
                        raw_value = [x[2] for x in project_data if x[0] == item.raw_value][0]
                    item = item._replace(
                        attr="home_status__customer_hirl_project__most_recent_notice_sent",
                        raw_value=raw_value,
                    )
                items.append(item)
            results.append(items)
        return results

    def get_customer_aps_data(self):
        include = ["id", "home__apshome__premise_id", "home__apshome__meterset_date"]
        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=include,
            drop_prefix="home__apshome__",
            section="customer_aps",
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_customer_eto_data(self):
        replace_dict = OrderedDict(
            [
                ("id", "Simulation ID"),
                ("fasttracksubmission__project_id", "ENH PT ID"),
                ("fasttracksubmission__solar_project_id", "SLE PT ID"),
                ("fasttracksubmission__eps_score", "EPS Score"),
                ("fasttracksubmission__eps_score_built_to_code_score", "EPS Score (Code Build)"),
                ("fasttracksubmission__percent_improvement", "Percent Improvement"),
                ("fasttracksubmission__builder_incentive", "Total Builder Incentive"),
                ("fasttracksubmission__original_builder_incentive", "Builder Incentive Override"),
                ("fasttracksubmission__builder_electric_incentive", "Builder Electric Incentive"),
                ("fasttracksubmission__builder_gas_incentive", "Builder Gas Incentive"),
                ("fasttracksubmission__rater_incentive", "Total Verifier Incentive"),
                ("fasttracksubmission__original_rater_incentive", "Verifier Incentive Override"),
                ("fasttracksubmission__rater_electric_incentive", "Verifier Electric Incentive"),
                ("fasttracksubmission__rater_gas_incentive", "Verifier Gas Incentive"),
                ("fasttracksubmission__carbon_score", "Carbon Score"),
                ("fasttracksubmission__carbon_built_to_code_score", "Carbon Score (Code Build)"),
                ("fasttracksubmission__estimated_annual_energy_costs", "Annual Energy Costs"),
                ("fasttracksubmission__estimated_monthly_energy_costs", "Monthly Energy Costs"),
                ("fasttracksubmission__similar_size_eps_score", "Similar sized EPS Score"),
                ("fasttracksubmission__similar_size_carbon_score", "Similar sized Carbon Score"),
                ("fasttracksubmission__therm_savings", "Savings (therms)"),
                ("fasttracksubmission__kwh_savings", "Savings (KWH)"),
                ("fasttracksubmission__mbtu_savings", "Savings (MBtu)"),
                ("fasttracksubmission__net_zero_eps_incentive", "Net Zero EPS Incentive"),
                # (
                #     "fasttracksubmission__energy_smart_homes_eps_incentive",
                #     "Energy Smart Homes EPS Incentive",
                # ),
                # (
                #     "fasttracksubmission__energy_smart_homes_solar_incentive",
                #     "Energy Smart Homes Solar Incentive",
                # ),
                ("fasttracksubmission__net_zero_solar_incentive", "Net Zero Solar Incentive"),
                (
                    "fasttracksubmission__solar_ready_builder_incentive",
                    "Solar Ready Builder Incentive",
                ),
                (
                    "fasttracksubmission__solar_ready_verifier_incentive",
                    "Solar Ready Verifier Incentive",
                ),
                ("fasttracksubmission__ev_ready_builder_incentive", "ESH: EV Ready Incentive"),
                (
                    "fasttracksubmission__solar_storage_builder_incentive",
                    "ESH: Solar + Storage Incentive",
                ),
                (
                    "fasttracksubmission__triple_pane_window_incentive",
                    "Fire Rebuild Triple Pane Windows Bonus Incentive",
                ),
                (
                    "fasttracksubmission__rigid_insulation_incentive",
                    "Fire Rebuild Exterior Rigid Insulation Bonus Incentive",
                ),
                (
                    "fasttracksubmission__sealed_attic_incentive",
                    "Fire Rebuild Sealed Attic Bonus Incentive",
                ),
                (
                    "fasttracksubmission__heat_pump_water_heater_incentive",
                    "Heat Pump Water Heater Deduction",
                ),
                ("fasttracksubmission__cobid_builder_incentive", "DEI Builder Incentive"),
                ("fasttracksubmission__cobid_verifier_incentive", "DEI Rater Incentive"),
                (
                    "fasttracksubmission__required_credits_to_meet_code",
                    "WA Code Credit Required Code Credits",
                ),
                (
                    "fasttracksubmission__achieved_total_credits",
                    "WA Code Credit Selected Credits",
                ),
                (
                    "fasttracksubmission__eligible_gas_points",
                    "WA Code Credit Eligible Above Code Credits",
                ),
                (
                    "fasttracksubmission__code_credit_incentive",
                    "WA Code Credit Code Incentive",
                ),
                (
                    "fasttracksubmission__thermostat_incentive",
                    "WA Code Credit Thermostat Incentive",
                ),
                (
                    "fasttracksubmission__fireplace_incentive",
                    "WA Code Credit Fireplace Incentive",
                ),
                (
                    "home__gbr__gbr_id",
                    "Green Building Registry ID",
                ),
            ]
        )

        def get_incentive_payment(ident):
            return "${}".format(ident) if ident else None

        def get_percent(ident):
            return "{:.2%}".format(ident) if ident else None

        def get_incentive_update(ident):
            return "Yes" if ident is not None else "No"

        clean_dict = {
            "fasttracksubmission__builder_incentive": get_incentive_payment,
            "fasttracksubmission__builder_electric_incentive": get_incentive_payment,
            "fasttracksubmission__builder_gas_incentive": get_incentive_payment,
            "fasttracksubmission__rater_incentive": get_incentive_payment,
            "fasttracksubmission__rater_electric_incentive": get_incentive_payment,
            "fasttracksubmission__rater_gas_incentive": get_incentive_payment,
            "fasttracksubmission__estimated_annual_energy_costs": get_incentive_payment,
            "fasttracksubmission__estimated_monthly_energy_costs": get_incentive_payment,
            "fasttracksubmission__percent_improvement": get_percent,
            "fasttracksubmission__original_builder_incentive": get_incentive_update,
            # "fasttracksubmission__net_zero_eps_incentive": get_incentive_payment,
            # "fasttracksubmission__energy_smart_homes_eps_incentive": get_incentive_payment,
            "fasttracksubmission__net_zero_solar_incentive": get_incentive_payment,
            "fasttracksubmission__energy_smart_homes_solar_incentive": get_incentive_payment,
            "fasttracksubmission__solar_ready_builder_incentive": get_incentive_payment,
            "fasttracksubmission__solar_ready_verifier_incentive": get_incentive_payment,
            "fasttracksubmission__ev_ready_builder_incentive": get_incentive_payment,
            "fasttracksubmission__solar_storage_builder_incentive": get_incentive_payment,
            "fasttracksubmission__heat_pump_water_heater_incentive": get_incentive_payment,
            "fasttracksubmission__cobid_builder_incentive": get_incentive_payment,
            "fasttracksubmission__cobid_verifier_incentive": get_incentive_payment,
            "fasttracksubmission__code_credit_incentive": get_incentive_payment,
            "fasttracksubmission__thermostat_incentive": get_incentive_payment,
            "fasttracksubmission__fireplace_incentive": get_incentive_payment,
        }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="customer_eto",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_customer_eto_permit_data(self):
        replace_dict = OrderedDict(
            [
                ("id", "Status ID"),
                (
                    "home__permitandoccupancysettings__signed_building_permit__id",
                    "City of Hillsboro Building Permit Compliance Report",
                ),
                (
                    "home__permitandoccupancysettings__signed_certificate_of_occupancy__id",
                    "City of Hillsboro Certificate of Occupancy Compliance Report",
                ),
            ]
        )

        def get_yes_no(ident):
            return "Yes" if ident is not None else "No"

        clean_dict = {
            "home__permitandoccupancysettings__signed_building_permit__id": get_yes_no,
            "home__permitandoccupancysettings__signed_certificate_of_occupancy__id": get_yes_no,
        }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="customer_eto",
            replace_dict=replace_dict,
            drop_prefix="home__permitandoccupancysettings__",
            clean_dict=clean_dict,
        )

        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)
        return data

    def get_hes_data(self):
        replace_dict = OrderedDict(
            [
                ("id", "HES Score Status"),
                ("hes_score_statuses__worst_case_simulation__base_score", "HES Score"),
                ("hes_score_statuses__worst_case_simulation__hescore_version", "HES Version"),
                (
                    "hes_score_statuses__worst_case_simulation__assessment_type",
                    "HES Assessment Type",
                ),
                (
                    "hes_score_statuses__worst_case_simulation__assessment_date",
                    "HES Assessment Date",
                ),
            ]
        )

        def clean_date(x):
            if x is None:
                return self.default_none
            return formats.date_format(x, "SHORT_DATE_FORMAT")

        clean_dict = {"hes_score_statuses__worst_case_simulation__assessment_date": clean_date}

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="hes_data",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        return data

    def get_ngbs_data(self):
        section_name = "ngbs_data"

        def clean_project_type(project_type):
            choices_dict = dict(HIRLProjectRegistration.PROJECT_TYPE_CHOICES)
            # force_str() to coerce lazy strings.
            return force_str(choices_dict.get(project_type, project_type), strings_only=True)

        def clean_project_state(state):
            choices_dict = dict(HIRLProjectRegistration.STATE_CHOICES)
            # force_str() to coerce lazy strings.
            return force_str(choices_dict.get(state, state), strings_only=True)

        structure = []

        structure.append(CellParser("id", "ID", None, section_name))

        structure.append(
            CellParser("customer_hirl_project__id", "NGBS Project ID", None, section_name)
        )

        structure.append(
            CellParser("hirl_legacy_project_id", "Legacy Project ID", None, section_name)
        )

        structure.append(
            CellParser("eep_program__name", "Certification Standard", None, section_name)
        )

        structure.append(
            CellParser(
                "customer_hirl_project__registration__project_type",
                "NGBS Project Type",
                clean_project_type,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__is_accessory_structure",
                "Accessory Structure",
                self.get_formatted_boolean,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__is_accessory_dwelling_unit",
                "Accessory Dwelling Unit",
                self.get_formatted_boolean,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__commercial_space_type",
                "Commercial Space",
                self.get_formatted_boolean,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__registration__state",
                "NGBS Project State",
                clean_project_state,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__green_energy_badges__name",
                "Green Energy Badges",
                None,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__story_count",
                "Story Count",
                None,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__registration__is_build_to_rent",
                "Is this project Build-To-Rent?",
                None,
                section_name,
            )
        )

        structure.append(
            CellParser(
                "customer_hirl_project__is_appeals_project",
                "Is this an appeals project?",
                None,
                section_name,
            )
        )

        structure.append(
            CellParser("unit_count", "Unit Count", None, section_name),
        )
        structure.append(
            CellParser("client_ca_status", "Client CA Status", None, section_name),
        )
        structure.append(
            CellParser(
                "client_coi_status",
                "Client COI Status",
                lambda coi: coi.capitalize() if coi else "",
                section_name,
            ),
        )
        structure.append(
            CellParser(
                "certification_level",
                "Certification Level",
                lambda lvl: lvl.capitalize() if lvl else "",
                section_name,
            ),
        )
        structure.append(
            CellParser("certification_number", "Certification Number", None, section_name),
        )
        structure.append(
            CellParser("certification_date", "Certification Date", None, section_name),
        )

        select_related = [
            "customer_hirl_project",
            "customer_hirl_project__registration",
            "customer_hirl_project__registration__builder_organization",
            "customer_hirl_project__registration__architect_organization",
            "customer_hirl_project__registration__developer_organization",
            "customer_hirl_project__registration__community_owner_organization",
        ]

        objects = (
            self.get_queryset()
            .select_related(*select_related)
            .annotate_customer_hirl_legacy_project_id()
            .annotate_customer_hirl_unit_count()
            .annotate_customer_hirl_certification_level()
            .annotate_customer_hirl_client_ca_status()
            .annotate_customer_hirl_certification_number()
            .annotate_customer_hirl_client_coi_expiration_status()
            .prefetch_related(
                "customer_hirl_project__registration__builder_organization__customer_hirl_enrolled_agreements",
                "customer_hirl_project__registration__architect_organization__customer_hirl_enrolled_agreements",
                "customer_hirl_project__registration__developer_organization__customer_hirl_enrolled_agreements",
                "customer_hirl_project__registration__community_owner_organization__customer_hirl_enrolled_agreements",
                "customer_hirl_project__green_energy_badges",
            )
            .order_by("customer_hirl_project__id")
        )
        data = self.merge_results(
            objects.values_list(*[x.attr for x in structure]).distinct(),
            structure,
        )

        return data

    def get_qa_cycle_time_data(self):
        qa_types_map = dict(QARequirement.QA_REQUIREMENT_TYPES)
        allowed_types = list(dict(QARequirement.QA_REQUIREMENT_TYPES).keys())

        def clean_duration(duration):
            days = duration.days + ((float(duration.seconds) / 3600) / 24)
            return "{0:.2f} days".format(days)

        def get_qa_types_handler(value):
            return ", ".join([qa_types_map[qa_type] for qa_type in value["qa_types"]])

        def get_counts_handler(state):
            def counts_handler(value):
                temp = []
                s = "{count}"
                for qa_type, count in value["counts"][state].items():
                    temp.append(s.format(type=qa_types_map[qa_type], count=count))
                return ", ".join(temp)

            return counts_handler

        def get_duration_handler(state):
            def duration_handler(value):
                temp = []
                s = "{duration}"
                for qa_type, duration in value["durations"][state].items():
                    temp.append(
                        s.format(type=qa_types_map[qa_type], duration=clean_duration(duration))
                    )

                return ", ".join(temp)

            return duration_handler

        def get_cycle_time_handler(value):
            temp = []
            s = "{duration}"
            for qa_type, states in value["durations"].items():
                if qa_type not in allowed_types:
                    continue
                tdelta = sum(list(states.values()), datetime.timedelta())
                temp.append(s.format(type=qa_types_map[qa_type], duration=clean_duration(tdelta)))

            return ", ".join(temp)

        def completed_date_handler(value):
            temp = []
            for transitions in value["transitions"].values():
                for transition in transitions:
                    _, start_time, _, _, _, to_state, _ = transition
                    if to_state == "complete":
                        temp.append(self.get_formatted_datetime(start_time))

            # Try to return a datetime object directly
            if len(temp) == 1:
                return temp[0]

            # Format into a list of datetimes if necessary.  This breaks the column type but it's
            # our only option in this framework.
            return ", ".join(temp)

        ids = list(self.get_queryset().values_list("id", flat=True))
        qa_metrics = QAStatus.objects.get_qa_metrics_for_home_statuses(ids)

        structure = [
            CellParser("id", "ID", None, "qa"),
            CellParser("qa_types", "QA Types", get_qa_types_handler, "qa"),
        ]

        for state, pretty in QAStatus.get_state_choices():
            if state in ["received", "in_progress", "complete"]:
                continue
            structure.append(
                CellParser(state + "_count", pretty + " Count", get_counts_handler(state), "qa")
            )
            structure.append(
                CellParser(
                    state + "_duration", pretty + " Duration", get_duration_handler(state), "qa"
                )
            )

        structure.append(
            CellParser("qa_completed_date", "QA Status Completed", completed_date_handler, "qa")
        )
        structure.append(
            CellParser("total_cycle_time", "Total QA Cycle Time", get_cycle_time_handler, "qa")
        )

        data = []

        for key, val in qa_metrics.items():
            if not isinstance(key, int):
                continue

            args = (key,) + ((val,) * (len(structure) - 1))
            data.append(args)

        data = self.merge_results(data, structure)

        return data

    def get_qa_data(self):
        include = [
            "id",
            "qastatus__id",
            "qastatus__created_on",
            "qastatus__owner__name",
            "qastatus__qa_designee__id",
            "qastatus__state",
            "qastatus__result",
            "qastatus__observation",
            "qastatus__qanote",
        ]

        note_string = "{user}({last_update}): {note}"

        def reverse_foreign_key_getter(source, cleaner, sort=None):
            def reverse_foreign_key(ident):
                if ident is None:
                    return self.default_none
                elif isinstance(ident, (list, set)):
                    sources = [source[i] for i in filter(None, ident)]
                    if sort:
                        sources = sort(sources)
                    return ", ".join(map(cleaner, sources))
                else:
                    return cleaner(source[ident])

            return reverse_foreign_key

        def get_note_string(note):
            return note_string.format(
                **{
                    "user": note.user.get_full_name(),
                    "last_update": self.get_formatted_datetime(note.last_update),
                    "note": note.note,
                }
            )

        def get_user_full_name(user_id):
            if user_id:
                if isinstance(user_id, (str, int)):
                    user_id = [user_id]
                vals = [str(x) for x in get_user_model().objects.filter(id__in=user_id)]
                return ", ".join(vals)
            return "-"

        # Get QA data relevant to this dataset (not just filter_by_user())
        homestatus_id_list = list(self.get_queryset().values_list("id", flat=True))
        qaobservations_id_list = list(
            Observation.objects.filter(
                qa_status__home_status__id__in=homestatus_id_list
            ).values_list("id", flat=True)
        )
        qa_observations = Observation.objects.in_bulk(qaobservations_id_list)
        qanote_id_list = list(
            QANote.objects.filter(qa_status__home_status__id__in=homestatus_id_list).values_list(
                "id", flat=True
            )
        )
        qa_notes = QANote.objects.in_bulk(qanote_id_list)

        qa_note_sorter = lambda sources: sorted(sources, key=operator.attrgetter("last_update"))
        clean_dict = {
            "qastatus__result": lambda ident: dict(QAStatus.STATUS).get(ident, self.default_none),
            "qastatus__state": lambda ident: dict(QAStateMachine.get_state_choices()).get(
                ident, self.default_none
            ),
            "qastatus__observation": reverse_foreign_key_getter(qa_observations, str),
            "qastatus__qanote": reverse_foreign_key_getter(
                qa_notes, get_note_string, sort=qa_note_sorter
            ),
            "qastatus__qa_designee__id": get_user_full_name,
        }
        name_dict = {"qastatus__observation": "QA Observations", "qastatus__qanote": "QA Notes"}

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=include,
            section="qa",
            clean_dict=clean_dict,
            replace_dict=name_dict,
        )
        select_related = [
            "__".join(item.attr.split("__")[:-1])
            for item in structure
            if len(item.attr.split("__")) > 1
        ]
        objects = self.get_queryset().select_related(*select_related)
        keys = [x.attr for x in structure]
        reverse_foreign_keys = ["qastatus__qanote", "qastatus__observation"]

        def compress_reverse_foreign_keys(source):
            compressed_values = {}
            for instance in source:
                instance_id = instance["id"]
                if instance_id not in compressed_values:
                    compressed_values[instance_id] = instance
                else:
                    for reverse_key in reverse_foreign_keys:
                        # With multiple reversed FKs in a values_list,
                        # there will be duplicates because of pairing.
                        if not isinstance(compressed_values[instance_id][reverse_key], set):
                            compressed_values[instance_id][reverse_key] = set(
                                [compressed_values[instance_id][reverse_key]]
                            )
                        else:
                            compressed_values[instance_id][reverse_key].add(instance[reverse_key])
            return list(compressed_values.values())

        def dict_to_list(source):
            final = []
            for instance in source:
                temp = []
                for key in keys:
                    temp.append(instance[key])
                final.append(temp)
            return final

        data = objects.values(*keys)  # get initial data
        data = compress_reverse_foreign_keys(data)  # clean reverse FKs
        data = dict_to_list(data)  # list of dicts -> list of lists

        state_model = EEPProgramHomeStatus._state_log_model
        qa_pending_transition_qs = state_model.objects.filter(to_state="qa_pending").values_list(
            "on_id", "start_time"
        )

        # Make sure we're always getting the latest transition of there are multiple
        qa_pending_transition = {}
        for on_id, start_time in qa_pending_transition_qs:
            if on_id not in qa_pending_transition:
                qa_pending_transition[on_id] = start_time
            elif qa_pending_transition[on_id] < start_time:
                qa_pending_transition[on_id] = start_time

        def get_qa_pending_entered(ident):
            value = qa_pending_transition.get(ident, self.default_none)
            if value != self.default_none:
                return self.get_formatted_date(value)
            return value

        structure.append(
            CellParser("qa_pending_entered", "Pending QA Entered", get_qa_pending_entered, "qa")
        )

        # Add extra id to end of data.
        # Used to look up when a stat entered pending qa
        temp = []
        for row in data:
            temp.append(row + [row[0]])

        data = self.merge_results(temp, structure)
        return data

    def get_customer_neea_standard_protocol_calculator_data(self):
        from axis.home.models import EEPProgramHomeStatus

        replace_dict = OrderedDict(
            [
                ("id", "Homestatus ID"),
                (
                    "standardprotocolcalculator__code_total_consumption_mmbtu",
                    "Total State Code Reference Home MMBtu Consumption",
                ),
                (
                    "standardprotocolcalculator__improved_total_consumption_mmbtu",
                    "Total Site MMBtu Consumption",
                ),
                (
                    "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings",
                    "Total Site MMBtu Consumption (with measure savings)",
                ),
                ("standardprotocolcalculator__total_kwh_savings", "Estimated Annual Savings (kWh)"),
                (
                    "standardprotocolcalculator__total_therm_savings",
                    "Estimated Annual Savings (Therms)",
                ),
                ("standardprotocolcalculator__heating_kwh_savings", "Heating kWh Savings"),
                ("standardprotocolcalculator__heating_therm_savings", "Heating Therm Savings"),
                ("standardprotocolcalculator__cooling_kwh_savings", "Cooling kWh Savings"),
                ("standardprotocolcalculator__cooling_therm_savings", "Cooling Therm Savings"),
                (
                    "standardprotocolcalculator__smart_thermostat_kwh_savings",
                    "Smart Thermostat kWh Savings",
                ),
                (
                    "standardprotocolcalculator__smart_thermostat_therm_savings",
                    "Smart Thermostat Therm Savings",
                ),
                (
                    "standardprotocolcalculator__water_heater_kwh_savings",
                    "Water Heater kWh Savings",
                ),
                (
                    "standardprotocolcalculator__water_heater_therm_savings",
                    "Water Heater Therm Savings",
                ),
                (
                    "standardprotocolcalculator__showerhead_kwh_savings",
                    "Low Flow Shower Heads kWh Savings",
                ),
                (
                    "standardprotocolcalculator__showerhead_therm_savings",
                    "Low Flow Shower Heads Therm Savings",
                ),
                ("standardprotocolcalculator__lighting_kwh_savings", "Lighting kWh Savings"),
                ("standardprotocolcalculator__appliance_kwh_savings", "Appliances kWh Savings"),
                ("standardprotocolcalculator__percent_improvement", "Percent Improvement"),
                (
                    "standardprotocolcalculator__revised_percent_improvement",
                    "Alternate Percent Improvement",
                ),
                (
                    "standardprotocolcalculator__pct_improvement_method",
                    "Percent Improvement Method",
                ),
                ("standardprotocolcalculator__builder_incentive", "Builder Incentive"),
            ]
        )

        def round_two(value):
            if isinstance(value, numbers.Number):
                return "{:.2f}".format(value)

        def percent(value):
            if isinstance(value, numbers.Number):
                return "{:.2f}%".format(value * 100)

        def payment(value):
            if isinstance(value, numbers.Number):
                return "${:.2f}".format(value)

        def capitalize(value):
            return "{}".format(value).capitalize()

        clean_dict = {
            "standardprotocolcalculator__reference_home_kwh": round_two,
            "standardprotocolcalculator__busbar_consumption": round_two,
            "standardprotocolcalculator__busbar_savings": round_two,
            "standardprotocolcalculator__heating_kwh_savings": round_two,
            "standardprotocolcalculator__heating_therm_savings": round_two,
            "standardprotocolcalculator__cooling_kwh_savings": round_two,
            "standardprotocolcalculator__cooling_therm_savings": round_two,
            "standardprotocolcalculator__smart_thermostat_kwh_savings": round_two,
            "standardprotocolcalculator__smart_thermostat_therm_savings": round_two,
            "standardprotocolcalculator__water_heater_kwh_savings": round_two,
            "standardprotocolcalculator__water_heater_therm_savings": round_two,
            "standardprotocolcalculator__showerhead_kwh_savings": round_two,
            "standardprotocolcalculator__showerhead_therm_savings": round_two,
            "standardprotocolcalculator__lighting_kwh_savings": round_two,
            "standardprotocolcalculator__appliance_kwh_savings": round_two,
            "standardprotocolcalculator__percent_improvement": percent,
            "standardprotocolcalculator__revised_percent_improvement": percent,
            "standardprotocolcalculator__pct_improvement_method": capitalize,
            "standardprotocolcalculator__builder_incentive": payment,
        }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="neea_standard_protocol_calculator",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        objects = self.get_queryset()
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        # Unify Percent improvement
        if len(data):
            pct_imp_col = data[0].index(
                next(
                    (
                        x
                        for x in data[0]
                        if x.attr == "standardprotocolcalculator__percent_improvement"
                    )
                )
            )
            rev_pct_imp_col = data[0].index(
                next(
                    (
                        x
                        for x in data[0]
                        if x.attr == "standardprotocolcalculator__revised_percent_improvement"
                    )
                )
            )
            pct_imp_met_col = data[0].index(
                next(
                    (
                        x
                        for x in data[0]
                        if x.attr == "standardprotocolcalculator__pct_improvement_method"
                    )
                )
            )

            imp_no_sav_col = data[0].index(
                next(
                    (
                        x
                        for x in data[0]
                        if x.attr == "standardprotocolcalculator__improved_total_consumption_mmbtu"
                    )
                )
            )
            imp_with_sav_col = data[0].index(
                next(
                    (
                        x
                        for x in data[0]
                        if x.attr
                        == "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings"
                    )
                )
            )

            for item in data[:]:
                if item[pct_imp_met_col].raw_value == ["alternate"]:
                    _d = list(item[pct_imp_col])
                    _d[-2] = item[rev_pct_imp_col].raw_value
                    _d[-1] = item[rev_pct_imp_col].value
                    item[pct_imp_col] = CellObject(*_d)

                    _e = list(item[imp_no_sav_col])
                    _e[-2] = item[imp_with_sav_col].raw_value
                    _e[-1] = item[imp_with_sav_col].value
                    item[imp_no_sav_col] = CellObject(*_e)

                item.pop(
                    item.index(
                        next(
                            (
                                x
                                for x in item
                                if x.attr == "standardprotocolcalculator__pct_improvement_method"
                            )
                        )
                    )
                )
                item.pop(
                    item.index(
                        next(
                            (
                                x
                                for x in item
                                if x.attr
                                == "standardprotocolcalculator__revised_percent_improvement"
                            )
                        )
                    )
                )
                item.pop(
                    item.index(
                        next(
                            (
                                x
                                for x in item
                                if x.attr
                                == "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings"
                            )
                        )
                    )
                )

        return data

    def epa_builder_rater_data(self, object_id):
        pass

    def munge_data(self, result_dict, new_data, key="id", verbose=False):
        """
        Takes each list in ``new_data`` and extends the corresponding running data pile at
        ``result_dict[id]``, where the appropriate id is available in each ``new_data`` list.
        """

        if verbose:
            if len(new_data):
                log.debug(pprint.pformat(new_data[-1]))
            log.debug("Result Keys: %r", list(result_dict.keys()))
        for item in new_data:
            if not len(item):
                log.warning("Zero length item found..")
                continue
            _id = next((x for x in item if x.attr == key)).value
            if verbose and new_data.index(item) == len(new_data) - 1:
                log.debug("ID %r", _id)
                log.debug("Initial Length %r", len(result_dict[_id]))

            if _id not in result_dict:
                raise ValueError(
                    "Unexpected {id_key} {id!r} after '{caller_function}' gathered data for queryset. "
                    "(valid {id_key} list={id_list!r}) (data={data!r})".format(
                        **{
                            "id_key": key,
                            "id": _id,
                            "id_list": list(sorted(result_dict.keys())),
                            "data": item,
                            "caller_function": inspect.stack()[1][3],
                        }
                    )
                )

            result_dict[_id] += item

            if verbose and new_data.index(item) == len(new_data) - 1:
                log.debug("ID %r", _id)
                log.debug("Final Length %r", len(result_dict[_id]))
                log.debug("Result item: %s", pprint.pformat(result_dict[_id]))
        return result_dict

    def get_datasets(self, report=False):
        if self._dataset:
            return self._dataset

        start = time.time()
        self.log.debug("Gathering %s datasets", self.get_queryset(report=True).count())

        results = self.build_datasets(report=report)

        self.log.debug("All data gathered.. (%.2fs)", time.time() - start)
        self.update_task(current=15)

        self._dataset = list(results.values())
        return self._dataset

    def build_datasets(self, report=False):
        results = OrderedDict([(x, []) for x in self.get_queryset().values_list("id", flat=True)])

        if "home" in self.report_on:
            self.log.debug("Gathering Home data")
            self.update_task(current=1)
            data = self.get_home_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Home data")

        if self.include_home_status_data:
            data = self.get_status_data()
            results = self.munge_data(results, data)

            data = self.get_historical_certifier_data()
            results = self.munge_data(results, data, verbose=False)

        if "subdivision" in self.report_on:
            self.log.debug("Gathering Subdivision data")
            self.update_task(current=2)
            data = self.get_subdivision_data()
            results = self.munge_data(results, data)

            if "community" in self.report_on:
                self.log.debug("Gathering Community data")
                self.update_task(current=3)
                data = self.get_community_data()
                results = self.munge_data(results, data)

            if self.include_builder_agreement_data:
                self.log.debug("Gathering Builder Agreement data")
                data = self.get_builder_agreement_data()
                results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Subdivision / Community data")

        if "relationships" in self.report_on:
            self.log.debug("Gathering Association data")
            self.update_task(current=4)
            data = self.get_relationships_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Association data")

        if "eep_program" in self.report_on:
            self.log.debug("Gathering Program data")
            self.update_task(current=5)
            data = self.get_eep_program_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Program data")

        if "ipp" in self.report_on:
            self.log.debug("Gathering Incentive data")
            self.update_task(current=6)
            data = self.get_incentive_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Incentive data")

        if "invoicing" in self.report_on:
            self.log.debug("Gathering Invoicing data")
            self.update_task(current=7)
            data = self.get_invoicing_data()
            results = self.munge_data(results, data, key="home_status_id")
        else:
            self.log.debug("Excluding Incentive data")

        if "sampleset" in self.report_on:
            self.log.debug("Gathering Sampleset data")
            self.update_task(current=8)
            data = self.get_sampleset_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Sampleset data")

        if "annotations" in self.report_on:
            self.log.debug("Gathering Annotation data")
            self.update_task(current=9)
            data = self.get_annotation_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Annotation data")

        if "checklist_answers" in self.report_on:
            self.log.debug("Gathering Checklist Answer data")
            self.update_task(current=10)
            data = self.get_collection_answer_data()
            results = self.munge_data(results, data)

            data = self.get_checklist_answer_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Checklist Answer data")

        if "qa" in self.report_on:
            self.log.debug("Gathering QA data")
            self.update_task(current=11)
            data = self.get_qa_data()
            results = self.munge_data(results, data)

            data = self.get_qa_cycle_time_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding QA data")

        if "customer_aps" in self.report_on:
            self.log.debug("Gathering Customer APS data")
            self.update_task(current=12)
            data = self.get_customer_aps_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Customer APS data")

        if "customer_eto" in self.report_on:
            self.log.debug("Gathering Customer ETO data")
            self.update_task(current=13)
            data = self.get_customer_eto_data()
            results = self.munge_data(results, data)
            data = self.get_customer_eto_permit_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Customer APS data")

        if "floorplan" in self.report_on:
            self.log.debug("Gathering Floorplan data")
            self.update_task(current=6)
            data = self.get_floorplan_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding Floorplan data")

        if "simulation_basic" in self.report_on:
            self.log.debug("Gathering Simulation Basic data")
            self.update_task(current=14)
            data = self.get_remrate_data()
            results = self.munge_data(results, data)
            data = self.get_ekotrope_data()
            results = self.munge_data(results, data)
        elif "simulation_advanced" in self.report_on:
            self.log.debug("Gathering Simulation Advanced data")
            self.update_task(current=14)
            data = self.get_remrate_data(advanced=True)
            results = self.munge_data(results, data)
            data = self.get_ekotrope_data(advanced=True)
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding REM/Rate data")
            self.log.debug("Excluding Ekotrope data")

        if "hes_data" in self.report_on:
            self.log.debug("Gathering HES data")
            self.update_task(current=15)
            data = self.get_hes_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding HES data")

        if "neea_standard_protocol_calculator" in self.report_on:
            self.log.debug("Gathering NEEA Standard Protocol Calculator data")
            self.update_task(current=16)
            data = self.get_customer_neea_standard_protocol_calculator_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding NEEA Standard Protocol Calculator data")

        if "ngbs_data" in self.report_on:
            self.log.debug("Gathering NGBS data")
            self.update_task(current=17)
            data = self.get_ngbs_data()
            results = self.munge_data(results, data)
        else:
            self.log.debug("Excluding NGBS data")

        return results

    def get_columns(self):
        """You will typically override this"""
        columns = []
        for column in self.report_on:
            if column == "simulation_advanced":
                for sub_column in [
                    "simulation_basic",
                    "remrate_building_shell",
                    "remrate_framefloor",
                    "remrate_joist",
                    "remrate_abovegradewall",
                    "remrate_window",
                    "remrate_door",
                    "remrate_roof",
                    "remrate_ceiling",
                    "remrate_mechanical",
                    "remrate_dominant_values",
                    "remrate_airconditioner",
                    "remrate_heater",
                    "remrate_airsourceheatpump",
                    "remrate_groundsourceheatpump",
                    "remrate_integratedspacewaterheater",
                    "remrate_hot_water",
                    "remrate_ductsystem",
                    "remrate_infiltration",
                    "remrate_solarsystem",
                    "remrate_misc",
                    "fuel_summary",
                    "remrate_results",
                    "remrate_utility_rates",
                    "remrate_lights_appl",
                    "ekotrope_framedfloors",
                    "ekotrope_rimjoists",
                    "ekotrope_windows",
                    "ekotrope_walls",
                    "ekotrope_ceilings",
                    "ekotrope_foundationwalls",
                    "ekotrope_doors",
                    "ekotrope_infiltration",
                    "ekotrope_mechanicalequipment",
                    "ekotrope_equipment",
                    "ekotrope_distributionsystems",
                    "ekotrope_thermostats",
                    "ekotrope_appliances",
                    "ekotrope_hers",
                    "ekotrope_breakdownbyfuel",
                    "ekotrope_breakdownbycomponent",
                    "ekotrope_analysissummary",
                    "ekotrope_EnergyStarV3Reference",
                    "ekotrope_EnergyStarV31Reference",
                    "ekotrope_IECC2018Reference",
                    "ekotrope_IECC2018Proposed",
                ]:
                    columns.append(sub_column)
                continue
            columns.append(column)
        # self.log.debug("Columns: {}".format(", ".join(columns)))
        return columns

    def get_column_name_map(self):
        """Allows you to customize the column names

        If value is None it is not exported - but used by other applications
        """
        return {
            "home__apshome__premise_id": "APS Meterset ID",
            "home__county__latitude": None,
            "home__county__longitude": None,
            "home__city__latitude": None,
            "home__city__longitude": None,
            "home__subdivision__latitude": None,
            "home__subdivision__longitude": None,
            "home__subdivision__slug": None,
            "home__subdivision__community__slug": None,
            "eep_program__slug": None,
            "eep_program__owner__slug": None,
            "builder_slug": None,
            "eep_slug": None,
            "provider_slug": None,
            "rater_slug": None,
            "electric_utility_slug": None,
            "gas_utility_slug": None,
            "water_utility_slug": None,
            "hvac_slug": None,
            "qa_slug": None,
            "general_slug": None,
            "developer_slug": None,
            "communityowner_slug": None,
            "architect_slug": None,
        }

    @property
    def column_map(self):
        return self.get_column_name_map()

    def show_column(self, attr):
        if len(self.specified_columns) and attr not in self.specified_columns:
            return False

        try:
            pretty_name = self.get_column_name_map()[attr]
        except KeyError:
            if attr is not None:
                return True
        else:
            if pretty_name is not None:
                return True
            else:
                log.debug("Skipping column %s pretty name is None ", attr)
        return False

    def print_columns(self):
        section = None
        for column in self.columns:
            if column.section != section:
                section = column.section
                print("- Section: {}".format(section))
            print("  - {}".format(column.pretty_name))

    @property
    def columns(self):
        """Get the columns"""

        ColData = namedtuple("ColData", ["attr", "section", "pretty_name"])

        if self._columns:
            return self._columns

        self.log.debug("Gathering columns")

        data = self.get_datasets()
        stime = time.time()

        column_data = OrderedDict()

        current_step = 15

        for idx, row in enumerate(data):
            for section_name in self.get_columns():
                for raw_column in row:
                    if time.time() - stime >= 2:
                        current_step += 1
                        self.update_task(current=current_step)
                        self.log.debug(
                            "Processing %s Columns  %s/%s",
                            raw_column.pretty_name,
                            data.index(row),
                            len(data),
                        )
                        stime = time.time()

                    # Don't consider this if it's not the right column or if it's an "id"
                    if section_name != raw_column.section:
                        continue
                    if raw_column.attr == "id" and self.output_ids is False:
                        continue

                    if not self.show_column(raw_column.attr):
                        continue

                    if raw_column.section not in column_data:
                        column_data[raw_column.section] = OrderedDict()

                    # Get the pretty name..
                    pretty_name = raw_column.pretty_name
                    try:
                        pretty_name = self.get_column_name_map()[raw_column.attr]
                    except KeyError:
                        pass

                    column = ColData(raw_column.attr, raw_column.section, pretty_name)

                    if column not in column_data[raw_column.section].keys():
                        column_data[raw_column.section][column] = 1

                    # If it's actually been used consider it.
                    if not self.retain_empty and raw_column.value in [None, self.default_none, ""]:
                        continue

                    column_data[raw_column.section][column] += 1

        self._columns = []
        for section in column_data.keys():
            for column in column_data[section].keys():
                if column_data[section][column] > 1:
                    self._columns.append(column)

        # Regigger the column order if you gave us a specified column order
        if len(self.specified_columns):
            cols = []
            for col in self.specified_columns:
                f_col = next((x for x in self._columns if x.attr == col), None)
                if f_col:
                    cols.append(f_col)
            self._columns = cols

        ret = "" if self.retain_empty else "Not "
        self.log.debug("%sRetaining all columns", ret)
        self.log.debug("Done Gathering columns")

        # print("data = [")
        # for i in self._columns:
        #     print("    '{}',   # {}".format(i.attr, i.pretty_name))
        # print(']')

        # print("data = [")
        # for i in self._columns:
        #     print("    ('{}',   '{}'),   # {}".format(i.attr, i.attr, i.pretty_name))
        # print(']')

        return self._columns

    @property
    def data(self):
        """Fill in the holes"""
        if self._data:
            return self._data
        self.log.debug("Processing data")

        stime = time.time()

        datasets = self.get_datasets()

        current_step = 35

        for home in datasets:
            if time.time() - stime >= 2:
                current_step += 1
                self.update_task(current=current_step)
                self.log.debug("Processing Home %s/%s", datasets.index(home), len(datasets))
                stime = time.time()
            data = []
            for column in self.columns:
                cdata = next((coldata for coldata in home if coldata.attr == column.attr), None)
                if not self.show_column(column.attr):
                    continue
                data.append(cdata.value if cdata and cdata.value else self.default_none)
            try:
                if set(data) != {self.default_none}:
                    self._data.append(data)
            except TypeError:
                self._data.append(data)
        self.log.debug("Done Processing data")
        return self._data

    def print_sample(self, max_num=2, start_num=0, as_dict=False):
        self.max_num = max_num
        results = []
        for data in self.data[start_num : max_num + start_num]:
            result = {}
            for idx, coldata in enumerate(self.columns):
                label = self.column_map.get(coldata.attr, coldata.pretty_name)
                label = label if label else coldata.attr
                if not label:
                    continue
                if len(label.split(" ")) > 5:
                    label = " ".join(label.split(" ")[:5]) + "..."
                label = label[:32] + "..." if len(label) > 40 else label[:39]
                value = data[idx]
                comment = ""
                if isinstance(value, tuple):
                    value, comment = value
                    comment = comment if comment is not None else ""
                result[label] = value
                if not as_dict:
                    try:
                        print("{:<40} {:<50}{:20}".format(label, value, comment))
                    except TypeError:
                        print("{:<40} !! {}".format(label, data[idx]))
            results.append(result)
            if not as_dict:
                print("--" * 20)
        if as_dict:
            return results


class HomeDataXLSExport(BaseHomeStatusDataDump, XLSChecklist, HomeStatusReportMixin):
    def __init__(self, *args, **kwargs):
        self.kwargs = kwargs

        self.creator = kwargs.get("creator", "Axis")
        self.title = kwargs.get("title", "Project Export")
        self.subject = kwargs.get("subject", "Project Export")
        self.description = kwargs.get("description", "Axis Generated Document")
        self.max_num = kwargs.get("max_num", None)

        kwargs = {
            "user_id": kwargs.get("user_id"),
            "report_on": kwargs.get("report_on", kwargs.get("export_fields")),
            "log": kwargs.get("log"),
            "max_num": kwargs.get("max_num", None),
            "reuse_storage": kwargs.get("reuse_storage", True),
            "retain_empty": kwargs.get("retain_empty_field", False),
            "task": kwargs.get("task"),
            "specified_columns": kwargs.get("specified_columns", []),
        }
        super(HomeDataXLSExport, self).__init__(**kwargs)

    def get_queryset(self, report=False):
        if hasattr(self, "_queryset"):
            return self._queryset

        qs = EEPProgramHomeStatus.objects.filter_by_user(user=self.user)
        qs, filters = self.get_external_qs_filters(
            qs, user=self.user, return_filters=True, **self.kwargs
        )
        self.filters = filters

        if self.kwargs.get("id_list"):
            qs = qs.filter(id__in=self.kwargs["id_list"])

        search_term = self.kwargs.get("search_bar")
        if search_term:
            qs = self.get_datatable_filtered_qs(qs, search_term)

        _count = qs.count()
        if self.max_num:
            qs = qs.filter(id__in=list(qs.values_list("id", flat=True))[: self.max_num])

        if report:
            if search_term:
                _count = "{} further reduced by search of '{}' to {}".format(
                    _count, search_term, len(qs)
                )
            self.log.info(
                "%s Filters have been used resulted in %s results", len(self.filters), _count
            )
            for key, value in self.filters:
                self.log.info("&emsp;&emsp;{}&emsp;:&emsp;{}".format(key, value))
            if self.max_num:
                self.log.info("&emsp;&emsp;Capped to %s results", self.max_num)

        if len(self.report_on):
            self.log.info("Reporting on:")
            for item in zip_longest(*[iter(self.report_on)] * 5):
                msg = ", ".join([ReportOn.get(x, x.capitalize()) for x in item if x])
                self.log.info("&emsp;&emsp;{}".format(msg))

        if not isinstance(qs, QuerySet):
            qs = EEPProgramHomeStatus.objects.filter(id__in=[x.id for x in qs])

        self._queryset = qs

        return self._queryset

    def get_datatable_filtered_qs(self, queryset, search_term):
        self.kwargs["search[value]"] = search_term
        query_params = QueryDict(mutable=True)
        query_params.update(self.kwargs)
        Request = namedtuple("Request", ["user", "company", "method", "GET", "headers", "path"])
        _request = Request(
            self.user,
            self.user.company,
            "GET",
            query_params,
            {"x-requested-with": "XMLHttpRequest"},
            "/",
        )

        class NarrowQuerySetDTV(HomeStatusView):
            request = _request
            model = EEPProgramHomeStatus

            def get_datatable_kwargs(self, **kwargs):
                kwargs = super(NarrowQuerySetDTV, self).get_datatable_kwargs()
                kwargs["object_list"] = queryset
                return kwargs

        homestatus_view = NarrowQuerySetDTV(request=_request)
        datatable = homestatus_view.get_datatable(object_list=queryset)
        datatable.configure()
        datatable.populate_records()
        return datatable._records

    def add_filter_text(self, sheet, row, column):
        cell = sheet.cell(row=row, column=column)
        cell.value = "Filters used:"
        self.set_cell_large_style(cell)

        row += 1
        for row_1, row_2 in zip_longest(*(iter(self.filters),) * 2):
            t1, v1 = row_1
            sheet.merge_cells(
                start_row=row, start_column=column + 1, end_row=row, end_column=column + 3
            )
            cell = sheet.cell(row=row, column=column)
            cell.value = t1
            self.set_cell_default_style(cell)
            cell.font = Font(name="Arial", size=12, bold=True)
            cell = sheet.cell(row=row, column=column + 1)
            cell.value = v1
            self.set_cell_default_style(cell)
            if row_2:
                t2, v2 = row_2
                sheet.merge_cells(
                    start_row=row + 1,
                    start_column=column + 1,
                    end_row=row + 1,
                    end_column=column + 3,
                )
                cell = sheet.cell(row=row + 1, column=column)
                cell.value = t2
                self.set_cell_default_style(cell)
                cell.font = Font(name="Arial", size=12, bold=True)
                cell = sheet.cell(row=row + 1, column=column + 1)
                cell.value = v2
                self.set_cell_default_style(cell)
            column += 4
        return row + 2, 1

    def write(self, return_workbook=False, output="HomesStatusExport.xlsx"):
        self.update_task = partial(self.update_task_progress, total=50)
        self.update_task(current=1)

        self.get_datasets(report=True)
        self.log.debug("Starting to write export")
        workbook = Workbook()
        today = formats.date_format(datetime.date.today(), "SHORT_DATE_FORMAT")
        sheet = workbook.create_sheet(index=0, title=self.subject[:31])

        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        cell = sheet.cell(row=1, column=2)
        cell.value = self.title
        self.set_cell_title_style(cell)

        cell = sheet.cell(row=2, column=2)
        cell.value = "Ran by user: {} on {}".format(self.user.get_full_name(), today)
        self.set_cell_italic_small_style(cell)

        row, column = 4, 1
        if len(self.filters):
            row, column = self.add_filter_text(sheet, row, column)

        row += 1
        for coldata in self.columns:
            label = self.column_map.get(coldata.attr, coldata.pretty_name)
            if not label:
                log.info("No header for cell %s skipping", cell)
                continue
            self.set_cell_header_style(sheet.cell(row=row, column=column, value=label))
            column += 1
        row += 1

        stime = time.time()
        data_length = len(self.data)

        self.update_task = partial(self.update_task_progress, step=1, total=data_length + 4)

        for row_data in self.data:
            column = 1
            if time.time() - stime >= 2:
                current = self.data.index(row_data)
                self.update_task(current=current)
                self.log.debug("Writing %s/%s", current, data_length)
                stime = time.time()

            for value in list(row_data):
                cell = sheet.cell(row=row, column=column)
                # PyXL gets a bit too fancy.. uuid can be interpretted as a big number.
                comment = None
                style_args = {}
                if isinstance(value, tuple):
                    value, comment = value
                if isinstance(value, float):
                    self.set_cell_float_value(cell, value, style_args)
                elif isinstance(value, int):
                    self.set_cell_integer_value(cell, value, style_args)
                elif isinstance(value, datetime.datetime):
                    self.set_cell_datetime_value(cell, value, style_args)
                elif isinstance(value, datetime.date):
                    self.set_cell_date_value(cell, value, style_args)
                elif isinstance(value, Decimal):
                    self.set_cell_decimal_value(cell, value, style_args)
                elif isinstance(value, str):
                    # Do unicode escape for strings
                    # to avoid openpyxl.utils.exceptions.IllegalCharacterError
                    value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)
                    if value.isdigit() or value.lstrip("-").isdigit():
                        self.set_cell_integer_value(cell, value, style_args)
                    elif (
                        value.replace(".", "", 1).isdigit()
                        or value.lstrip("-").replace(".", "", 1).isdigit()
                    ):
                        self.set_cell_float_value(cell, value, style_args)
                    elif value.startswith("$"):
                        self.set_cell_currency_value(cell, value, style_args)
                    elif value.endswith("%"):
                        self.set_cell_percentage_value(cell, value, style_args)
                    else:
                        self.set_cell_explicit_value(cell, value)

                    if comment:
                        cell.comment = Comment(comment, "Axis")
                elif isinstance(value, Hashid):
                    self.set_cell_explicit_value(cell, str(value))
                else:
                    cell.value = value
                self.set_cell_default_style(cell, **style_args)
                sheet.column_dimensions[get_column_letter(column)].width = self.default_col_width
                column += 1
            row += 1

        self.log.debug("Adding properties and logo")
        self.update_task(current=data_length + 1)

        workbook.properties = self.properties()
        self.update_task(current=data_length + 2)

        self.add_logo(workbook, sheet)
        self.update_task(current=data_length + 3)

        if return_workbook:
            self.log.debug("Done to building export")
            self.update_task(current=data_length + 4)
            return workbook

        self.log.debug("Saving export")
        workbook.save(output)
        self.update_task(current=data_length + 4)
        self.log.debug("Done to saving export")
        return output


def get_bulk_items(**kwargs):
    obj = HomeDataXLSExport(**kwargs)
    total = obj.get_queryset().count()

    count_by, start = 100, 0
    while True:
        kwargs["max_num"] = start + count_by
        if kwargs["max_num"] > total:
            del kwargs["max_num"]
        obj = HomeDataXLSExport(**kwargs)
        obj.get_datasets()
        if not kwargs.get("max_num"):
            obj.write()
            break
        start += count_by
        print("5 Seconds cleared to kill: ")
        time.sleep(5)


if __name__ == "__main__":
    # user = get_user_model().objects.get(username="rjohnson")
    from django.apps import apps as django_app

    if not django_app.apps_ready:
        import django

        django.setup()

    user = get_user_model().objects.get(username="sklass")
    # user = get_user_model().objects.get(id=355) # Terra Bell
    # user = get_user_model().objects.get(first_name__istartswith="gavin")
    # get_bulk_items(user_id=user.id)
    # user = get_user_model().objects.get(first_name__istartswith="helen")

    # kwargs = {'user_id': user.id,
    #           'eep_program_id': 6,
    #           'max_num': 200,
    #           'report_on': ['home', 'status', 'subdivision', 'community', 'relationships'
    #                         'eep_program', 'floorplan', 'sampleset', 'remrate', 'annotations',
    #                         'checklist_answers', 'ipp', 'qa', 'customer_aps']
    # }

    # kwargs = {'user_id': user.id,
    #           'search_bar': '3 HRII',
    #           # 'max_num': 2000,
    #           # 'report_on': ['home', 'status', 'subdivision', 'community', 'relationships',
    #           #               'eep_program', 'floorplan', 'sampleset', 'remrate', 'annotations',
    #           #               'checklist_answers', 'ipp', 'qa', 'customer_aps']
    #           'report_on': ['home', 'status', 'relationships', 'annotations', 'eep_program']
    # }

    # kwargs = {'user_id': user.id,
    #           # 'subdivision_id': 419,
    #           # 'eep_program_id': 6,
    #           'retain_empty': True,
    #           'search_bar': '622 drager',
    #           # 'max_num': 2,
    #           # 'report_on': ['home', 'status', 'subdivision', 'community', 'relationships',
    #           #               'eep_program', 'floorplan', 'sampleset', 'remrate', 'annotations',
    #           #               'checklist_answers', 'ipp', 'qa', 'customer_aps']
    #           # 'report_on': ['home', 'status', 'sampleset', 'checklist_answers',]
    #           # 'report_on': ['home', 'status', 'relationships', 'annotations', 'eep_program']
    #           'report_on': ['relationships', 'annotations', 'remrate_advanced' ]
    #           # 'report_on': ['relationships', 'annotations', 'remrate_advanced', 'remrate_basic' ]
    # }

    # report_on = ReportOn.keys()
    # report_on.pop(report_on.index('remrate_basic'))

    kwargs = {
        "user_id": user.id,
        "search_bar": "2155 SE 12th Ave",
        # 'eep_program_id': [83, 36],
        # 'report_on': ['checklist_answers'],
        "report_on": ["checklist_answers"],
        # 'report_on': ['home', 'status', 'subdivision', 'community', 'relationships',
        #               'eep_program', 'floorplan', 'sampleset', 'remrate_advanced',
        #               'annotations', 'checklist_answers', 'ipp', 'qa', 'customer_aps',
        #               'customer_eto', 'hes_data'],
        # 'max_num': 100,
    }

    obj = HomeDataXLSExport(**kwargs)

    obj.write()

    obj.print_sample(max_num=2)

    # obj.print_sample(max_num=5)
    # obj.print_columns()
    # obj.write()
    # # print obj.report_on
    print(len(obj.data), "Rows", len(obj.columns), "Columns")
