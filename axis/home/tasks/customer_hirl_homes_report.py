"""customer_hirl_homes_report.py: """

__author__ = "Artem Hruzd"
__date__ = "06/12/2022 18:58"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import io

from celery import shared_task
from django.apps import apps
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.files.base import ContentFile
from django.db.models import OuterRef, Subquery
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from axis.core.reports import AxisReportFormatter


User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


@shared_task(bind=True, time_limit=60 * 60 * 3)
def customer_hirl_homes_report_task(self):
    """
    Generate Customer HIRL special report once a day
    :param self:
    :return:
    """
    from axis.filehandling.models import AsynchronousProcessedDocument

    workbook = Workbook()
    _certified_sheet(workbook=workbook)
    _in_progress_sheet(workbook=workbook)

    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    outfile = virtual_workbook.getvalue()

    today = timezone.now().today()

    filename = "NGBS-Green-Certification-Activity-{}.xlsx".format(today.strftime("%Y%m%d"))

    customer_hirl_company = customer_hirl_app.get_customer_hirl_provider_organization()

    customer_document, created = AsynchronousProcessedDocument.objects.get_or_create(
        company=customer_hirl_company,
        task_name=f"customer_hirl_homes_report_task",
        download=True,
        created_date__day=today.day,
        created_date__month=today.month,
        created_date__year=today.year,
        defaults={"final_status": "SUCCESS", "task_id": ""},
    )

    customer_document.document.save(filename, ContentFile(outfile, name=filename))
    customer_document.save()


def _certified_sheet(workbook):
    from axis.home.models import EEPProgramHomeStatus
    from axis.company.models import Company
    from axis.home.models import Home

    sheet = workbook.create_sheet(index=0, title="NGBS Green Certified")
    axis_report_formatter = AxisReportFormatter()

    row = 1
    column = 1
    labels = [
        {"text": "Company", "col_width": 35},
        {"text": "Project Name", "col_width": 35},
        {"text": "Unit Count", "col_width": 30},
        {"text": "Address", "col_width": 25},
        {"text": "City", "col_width": 35},
        {"text": "County", "col_width": 35},
        {"text": "State", "col_width": 35},
        {"text": "Zip", "col_width": 35},
        {"text": "Certification Level", "col_width": 35},
        {"text": "Certified Date", "col_width": 35},
        {"text": "Scoring path", "col_width": 35},
    ]

    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    row = 2

    builder_relationship_query = Company.objects.filter(
        company_type="builder",
        relationships__content_type=ContentType.objects.get_for_model(Home),
        relationships__object_id=OuterRef("home__pk"),
    )
    certified_home_statuses = (
        EEPProgramHomeStatus.objects.filter(
            eep_program__slug__in=customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS
            + customer_hirl_app.HIRL_PROJECT_LEGACY_EEP_PROGRAM_SLUGS
        )
        .select_related(
            "home",
            "eep_program",
            "customer_hirl_project",
            "home__subdivision",
            "home__city",
            "home__city__county",
        )
        .annotate(builder_name=Subquery(builder_relationship_query.values("name")[:1]))
        .annotate_customer_hirl_certification_level()
        .annotate_customer_hirl_unit_count()
        .exclude(certification_level=None)
        .order_by("-created_date")
    )

    for home_status in certified_home_statuses:
        # company
        sheet.cell(row=row, column=1).value = home_status.builder_name

        # project name
        if getattr(home_status.home, "subdivision", None):
            sheet.cell(row=row, column=2).value = home_status.home.subdivision.name

        # unit count
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=3), home_status.unit_count)

        # address
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=4), home_status.home.street_line1
        )

        if getattr(home_status.home, "city", None):
            # city
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=5), home_status.home.city.name
            )

            if getattr(home_status.home.city, "county", None):
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=6), home_status.home.city.county.name
                )
                # state
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=7), home_status.home.city.county.state
                )

        # zip
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=8), home_status.home.zipcode
        )

        # certification level
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=9), home_status.certification_level.title()
        )

        # certified date
        if home_status.certification_date:
            sheet.cell(row=row, column=10).value = home_status.certification_date

        # scoring path
        sheet.cell(row=row, column=11).value = home_status.eep_program.name

        row = row + 1

    sheet.auto_filter.ref = "A1:K99999"
    sheet.auto_filter.add_sort_condition("J2:J99999")


def _in_progress_sheet(workbook):
    from axis.home.models import EEPProgramHomeStatus
    from axis.company.models import Company
    from axis.home.models import Home

    sheet = workbook.create_sheet(index=1, title="NGBS GreenIn-process")
    axis_report_formatter = AxisReportFormatter()

    row = 1
    column = 1
    labels = [
        {"text": "Company", "col_width": 35},
        {"text": "Project Name", "col_width": 35},
        {"text": "Unit Count", "col_width": 30},
        {"text": "Address", "col_width": 25},
        {"text": "City", "col_width": 35},
        {"text": "County", "col_width": 35},
        {"text": "State", "col_width": 35},
        {"text": "Zip", "col_width": 35},
        {"text": "Scoring path", "col_width": 35},
    ]

    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    row = 2

    builder_relationship_query = Company.objects.filter(
        company_type="builder",
        relationships__content_type=ContentType.objects.get_for_model(Home),
        relationships__object_id=OuterRef("home__pk"),
    )
    in_progress_home_statuses = (
        EEPProgramHomeStatus.objects.filter(
            eep_program__slug__in=customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS
            + customer_hirl_app.HIRL_PROJECT_LEGACY_EEP_PROGRAM_SLUGS
        )
        .select_related(
            "home",
            "eep_program",
            "customer_hirl_project",
            "home__subdivision",
            "home__city",
            "home__city__county",
        )
        .annotate(builder_name=Subquery(builder_relationship_query.values("name")[:1]))
        .annotate_customer_hirl_certification_level()
        .annotate_customer_hirl_unit_count()
        .filter(certification_level=None)
        .order_by("-created_date")
    )

    for home_status in in_progress_home_statuses:
        # company
        sheet.cell(row=row, column=1).value = home_status.builder_name

        # project name
        if getattr(home_status.home, "subdivision", None):
            sheet.cell(row=row, column=2).value = home_status.home.subdivision.name

        # unit count
        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=3), home_status.unit_count)

        # address
        sheet.cell(row=row, column=4).value = home_status.home.street_line1

        if getattr(home_status.home, "city", None):
            # city
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=5), home_status.home.city.name
            )

            if getattr(home_status.home.city, "county", None):
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=6), home_status.home.city.county.name
                )
                # state
                axis_report_formatter.format_str_cell(
                    sheet.cell(row=row, column=7), home_status.home.city.county.state
                )

        # zip
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=8), home_status.home.zipcode
        )

        # scoring path
        sheet.cell(row=row, column=9).value = home_status.eep_program.name

        row = row + 1

    sheet.auto_filter.ref = "A1:K99999"
    sheet.auto_filter.add_sort_condition("J2:J99999")
