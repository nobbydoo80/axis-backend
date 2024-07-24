"""test_export_data.py - Axis"""

import logging
import tempfile
from decimal import Decimal

from django.apps import apps
from django.core import management

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from axis.checklist.tests.mixins import CollectedInputMixin
from axis.core.models import User
from axis.core.tests.test_views import DevNull
from axis.core.tests.testcases import AxisTestCase
from axis.customer_hirl.models import HIRLProject, HIRLProjectRegistration
from axis.customer_hirl.tests.factories import hirl_green_energy_badge_factory
from axis.eep_program.models import EEPProgram
from axis.geographic.tests.factories import real_city_factory
from axis.home.export_data import HomeDataXLSExport, ReportOn
from axis.home.models import EEPProgramHomeStatus
from axis.invoicing.models import Invoice, InvoiceItemGroup
from axis.invoicing.tests.factories import invoice_factory
from axis.relationship.utils import create_or_update_spanning_relationships

log = logging.getLogger(__name__)
customer_hirl_app = apps.get_app_config("customer_hirl")
GreenBuildingRegistry = apps.get_model("gbr", "GreenBuildingRegistry")

__author__ = "Steven K"
__date__ = "4/21/21 10:42"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven K",
]


class ProjectStatusReportMixin:
    """This will flex out the various components of the HSR / Project Report.  The testing of
    tasks is covered in the task stuff this hits the core of this report."""

    @classmethod
    def setUpTestData(cls):
        from axis.core.tests.factories import (
            eep_admin_factory,
            rater_admin_factory,
            provider_admin_factory,
            qa_admin_factory,
            utility_admin_factory,
            builder_admin_factory,
        )
        from axis.company.tests.factories import (
            communityowner_organization_factory,
            architect_organization_factory,
            developer_organization_factory,
        )
        from axis.eep_program.models import EEPProgram
        from axis.floorplan.tests.factories import floorplan_with_simulation_factory
        from axis.home.tests.factories import (
            eep_program_home_status_factory,
            home_factory,
        )
        from axis.relationship.models import Relationship
        from axis.checklist.tests.mixins import CollectedInputMixin
        from axis.customer_hirl.tests.factories import (
            hirl_project_registration_factory,
            hirl_project_factory,
            hirl_green_energy_badge_factory,
        )
        from axis.customer_hirl.models import HIRLProjectRegistration
        from axis.invoicing.tests.factories import invoice_factory
        from axis.invoicing.models import Invoice, InvoiceItemGroup
        import time

        start = time.time()
        customer_hirl_app = apps.get_app_config("customer_hirl")
        city = real_city_factory("Gilbert", "AZ")

        # Create Users
        eep_user = eep_admin_factory(company__name="ETO", company__slug="eto", company__city=city)
        qa_user = qa_admin_factory(company__name="QA1", company__slug="peci", company__city=city)

        aps_user = utility_admin_factory(
            company__name="APS",
            company__slug="aps",
            company__is_eep_sponsor=True,
            company__city=city,
            company__electricity_provider=True,
            company__gas_provider=False,
            company__water_provider=False,
        )

        ngbs_user = provider_admin_factory(
            company__name="HIRL",
            company__slug=customer_hirl_app.CUSTOMER_SLUG,
            company__is_eep_sponsor=True,
            company__city=city,
        )

        cls.rater_user = rater_admin_factory(company__name="Rater1", company__city=city)
        rater_user_2 = rater_admin_factory(company=cls.rater_user.company, is_company_admin=False)

        cls.provider_user = provider_admin_factory(company__name="Provider1", company__city=city)
        builder_user = builder_admin_factory(company__name="Provider1", company__city=city)

        # Create mutual relationships
        eto_companies = [
            eep_user.company,
            cls.rater_user.company,
            cls.provider_user.company,
            qa_user.company,
            builder_user.company,
        ]
        Relationship.objects.create_mutual_relationships(*eto_companies)

        aps_companies = [
            aps_user.company,
            cls.rater_user.company,
            cls.provider_user.company,
            qa_user.company,
            builder_user.company,
        ]
        Relationship.objects.create_mutual_relationships(*aps_companies)

        community_owner_company = communityowner_organization_factory(city=city)
        architect_company = architect_organization_factory(city=city)
        developer_company = developer_organization_factory(city=city)

        ngbs_companies = [
            ngbs_user.company,
            cls.rater_user.company,
            qa_user.company,
            builder_user.company,
            community_owner_company,
            architect_company,
            developer_company,
        ]
        Relationship.objects.create_mutual_relationships(*aps_companies)

        # Create EEPPrograms
        management.call_command(
            "build_program", "-p", "eto-2022", "--no_close_dates", stdout=DevNull()
        )
        eto_program = EEPProgram.objects.get(slug="eto-2022")
        EEPProgram.objects.filter(id=eto_program.id).update(
            program_close_date=None,
            program_submit_date=None,
            program_end_date=None,
        )

        management.call_command(
            "build_program", "-p", "aps-energy-star-v3-hers-60-2018", stdout=DevNull()
        )
        aps_program = EEPProgram.objects.get(slug="aps-energy-star-v3-hers-60-2018")

        management.call_command(
            "build_program", "-p", "ngbs-sf-new-construction-2020-new", stdout=DevNull()
        )
        ngbs_program = EEPProgram.objects.get(slug="ngbs-sf-new-construction-2020-new")

        # Create Home stats
        cls.home = home_factory(
            city=city,
            subdivision__city=city,
            street_line1="256 W Oxford Ln",
            zipcode="85233",
            subdivision__builder_org=builder_user.company,
        )
        eep_program_home_status_factory(
            company=cls.rater_user.company, eep_program=eto_program, home=cls.home
        )

        home_status = eep_program_home_status_factory(
            company=cls.provider_user.company, eep_program=aps_program, home=cls.home
        )
        CollectedInputMixin().satisfy_collection_request(home_status, user=cls.provider_user)

        # NGBS Projects
        ngbs_stat = eep_program_home_status_factory(
            company=cls.rater_user.company,
            eep_program=ngbs_program,
            home=cls.home,
            field_inspectors=[cls.rater_user, rater_user_2],
        )
        assert ngbs_stat.field_inspectors.count() == 2, "Missing Verifier"

        sf_registration = hirl_project_registration_factory(
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE,
            registration_user=cls.rater_user,
            builder_organization=builder_user.company,
            community_owner_organization=community_owner_company,
            architect_organization=architect_company,
            developer_organization=developer_company,
            eep_program=ngbs_program,
        )
        sf_project = hirl_project_factory(registration=sf_registration, home_status=ngbs_stat)

        energy_badge_1 = hirl_green_energy_badge_factory(name="Test Badge", cost=50)
        sf_project.green_energy_badges.add(energy_badge_1)
        energy_badge_2 = hirl_green_energy_badge_factory(name="Test Badge 2", cost=20)
        sf_project.green_energy_badges.add(energy_badge_2)

        sf_project.save()  # This will create the invoice items and the InvoiceItemGroup

        invoice_factory(
            invoice_type=Invoice.HIRL_PROJECT_INVOICE_TYPE,
            issuer=ngbs_user.company,
            customer=sf_registration.get_company_responsible_for_payment(),
            invoice_item_groups=InvoiceItemGroup.objects.all(),
        )

        for company in aps_companies + eto_companies + ngbs_companies:
            Relationship.objects.validate_or_create_relations_to_entity(cls.home, company)

        cls.home2 = home_factory(
            city=city,
            subdivision__city=city,
            street_line1="321 W Desert Ave",
            zipcode="85233",
            subdivision__builder_org=builder_user.company,
        )
        cls.floorplan2 = floorplan_with_simulation_factory(
            owner=cls.rater_user.company,
            simulation__project__resnet_registry_identification=12231,
        )
        eep_program_home_status_factory(
            company=cls.rater_user.company,
            eep_program=eto_program,
            floorplan=cls.floorplan2,
            home=cls.home2,
        )

        # Create relationships to home
        for company in eto_companies:
            Relationship.objects.validate_or_create_relations_to_entity(cls.home2, company)
        #  print(elapsed_time(time.time() - start).long_fmt)

    @property
    def report_on(self):
        return list(ReportOn.keys())

    @property
    def default_kwargs(self):
        user = self.user_model.objects.filter(company__slug="eto").first()
        from axis.eep_program.models import EEPProgram

        return {
            "user_id": user.id,
            "eep_program_id": list(
                EEPProgram.objects.filter_by_user(user).values_list("pk", flat=True)
            ),
            "search_bar": None,
            "report_on": self.report_on,
            "retain_empty_field": True,
            "filename": tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name,
        }

    def get_results(self, filename, as_dict=True, max_rows=15):
        """Open the file and get the contents - Specifically the table data"""
        workbook = load_workbook(filename=filename)
        self.assertEqual(workbook.sheetnames[0], "Project Export")
        sheet = workbook[workbook.sheetnames[0]]
        results = []
        for row in range(7, max_rows):
            cell = sheet[f"A{row}"]
            if not cell.value:
                continue

            row_values = [sheet[f"{get_column_letter(col)}{row}"].value for col in range(1, 1000)]

            # Remove no values
            while row_values and row_values[-1] is None:
                del row_values[-1]
            if len(row_values) >= 2:
                if len(results):
                    self.assertEqual(len(results[-1]), len(row_values))
                results.append(row_values)

        if as_dict:
            return [dict(zip(results[0], row)) for row in results[1:]]

        return results

    @property
    def default_reported_keys(self):
        return {
            "Project Status",
            "Rater of Record",
            "% Complete",
            "Rating Type",
            "Certification Date",
            "RESNET Registry ID",
            "Final Verifier",
            "Field Inspectors",
            "Rough Verifier",
            "Energy Modeler",
        }


class ProjectStatusReportTests(ProjectStatusReportMixin, AxisTestCase):
    def test_basics(self):
        """This will test the basic export"""
        kwargs = self.default_kwargs

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        queryset_count = obj.get_queryset().count()

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(queryset_count, len(data))

    def test_report_on_status(self):
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["status"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(set(data[0].keys()), self.default_reported_keys)

    def test_search_bar(self):
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["status"]
        kwargs["search_bar"] = "Desert Ave"

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(len(data), 1)
        self.assertEqual(set(data[0].keys()), self.default_reported_keys)

    def test_report_on_status_no_retain(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["status"]
        kwargs["retain_empty_field"] = False

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertIn("Project Status", data[0].keys())
        self.assertIn("% Complete", data[0].keys())
        self.assertIn("Rating Type", data[0].keys())
        self.assertIn("Rater of Record", data[0].keys())
        self.assertIn("RESNET Registry ID", data[0].keys())

    def test_values_field_inspectors(self):
        """We have multiple verifiers on this address"""

        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )

        kwargs = self.default_kwargs
        kwargs["report_on"] = ["status"]
        kwargs["user_id"] = stat.company.users.first().id
        kwargs["eep_program_id"] = stat.eep_program_id

        self.assertEqual(stat.field_inspectors.count(), 2)

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), 1)
        self.assertIn(", ", data[0]["Field Inspectors"])
        self.assertNotIn(", ", data[0]["Project Status"])
        self.assertNotIn(", ", data[0]["Rater of Record"])


class ProjectStatusReportHomeTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_home(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["home"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())
        expected = {
            "Alt ID",
            "Created Date",
            "Latitude",
            "Confirmed Address",
            "Multi-Family",
            "Metro",
            "Longitude",
            "State",
            "AXIS ID",
            "Street Line 1",
            "Address Override",
            "Climate Zone",
            "County",
            "ZIP Code",
            "City",
            "Lot #",
            "Street Line 2",
            "ETO Region",
            "City Longitude",
            "County Longitude",
            "County Latitude",
            "City Latitude",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))


class ProjectStatusCollectedInputTests(ProjectStatusReportMixin, AxisTestCase):
    @classmethod
    def setUpTestData(cls):
        super(ProjectStatusCollectedInputTests, cls).setUpTestData()

        eto_program = EEPProgram.objects.get(slug="eto-2022")
        home_status = EEPProgramHomeStatus.objects.get(eep_program=eto_program, home=cls.home)
        CollectedInputMixin().satisfy_collection_request(home_status, user=cls.rater_user)

        aps_program = EEPProgram.objects.get(slug="aps-energy-star-v3-hers-60-2018")
        home_status = EEPProgramHomeStatus.objects.get(eep_program=aps_program)
        CollectedInputMixin().satisfy_collection_request(home_status, user=cls.provider_user)

    def test_report_on_checklist_answers_multiple_programs(self):
        """Make sure that two programs show data respectively"""
        from axis.eep_program.models import EEPProgram

        kwargs = self.default_kwargs
        kwargs["user_id"] = self.provider_user.id
        kwargs["report_on"] = ["eep_program", "checklist_answers"]
        kwargs["search_bar"] = "Oxford"

        programs = EEPProgram.objects.filter_by_user(self.provider_user).filter(is_qa_program=False)
        kwargs["eep_program_id"] = list(programs.values_list("pk", flat=True))

        self.assertEqual(len(kwargs["eep_program_id"]), 2)

        stats = EEPProgramHomeStatus.objects.filter(
            home=self.home, eep_program_id__in=kwargs["eep_program_id"]
        )
        self.assertEqual(stats.count(), 2)
        # Make sure that we have answers for the questions
        for stat in stats:
            self.assertEqual(len(stat.get_unanswered_questions()), 0)
            self.assertGreater(len(stat.get_answers_for_home()), 9)

        status_1 = stats.first()
        status_1_program = status_1.eep_program.name
        status_1_answers_raw = status_1.get_answers_for_home()
        status_1_answers = {x.instrument.text: x.data["input"] for x in status_1_answers_raw}
        status_2 = stats.last()
        status_2_program = status_2.eep_program.name
        status_2_answers_raw = status_2.get_answers_for_home()
        status_2_answers = {x.instrument.text: x.data["input"] for x in status_2_answers_raw}
        self.assertNotEqual(status_1, status_2)
        self.assertNotEqual(status_1_program, status_2_program)

        # print(status_1.id, status_1_answers)
        # print(status_2.id, status_2_answers)

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        xls_data = self.get_results(kwargs["filename"])

        # import pprint
        #
        # for row in xls_data:
        #     print("--" * 20)
        #     pprint.pprint(row)

        self.assertEqual(len(xls_data), 2)

        xls_status_1_data = next((x for x in xls_data if x["Program Name"] == status_1_program))
        for k, v in status_1_answers.items():
            if isinstance(v, dict) and "brand_name" in v.keys():  # Cascade Selects
                continue
            #  with self.subTest(f"{status_1_program} {k!r} -> {v!r} vs {xls_status_1_data[k]}"):
            self.assertEqual(str(xls_status_1_data[k]), str(v))

        xls_status_2_data = next((x for x in xls_data if x["Program Name"] == status_2_program))
        for k, v in status_2_answers.items():
            if isinstance(v, dict) and "brand_name" in v.keys():  # Cascade Selects
                continue
            #  with self.subTest(f"{status_1_program} {k!r} -> {v!r} vs {xls_status_2_data[k]}"):
            self.assertEqual(str(xls_status_2_data[k]), str(v))

        self.assertNotEqual(xls_status_1_data, xls_status_2_data)

    def test_report_on_checklist_answers_multiple_like_programs(self):
        """ETO 2021 and ETO 2020 have very similar measures we are seeing where measures dont
        show in the report not sure why but this should expose and test for it."""
        from axis.eep_program.models import EEPProgram

        self.assertEqual(self.home2.homestatuses.count(), 1)
        home_status = self.home2.homestatuses.get()
        CollectedInputMixin().remove_collected_input(home_status=home_status, measure_id="__all__")

        management.call_command("build_program", "-p", "eto-2021", stdout=DevNull())
        eto_program = EEPProgram.objects.get(slug="eto-2021")

        home_status.eep_program = eto_program
        home_status.save()

        CollectedInputMixin().satisfy_collection_request(home_status, user=self.rater_user)
        programs = EEPProgram.objects.filter(owner__slug="eto", is_qa_program=False)

        kwargs = self.default_kwargs
        kwargs["user_id"] = self.provider_user.id
        kwargs["report_on"] = ["eep_program", "checklist_answers"]
        kwargs["eep_program_id"] = list(programs.values_list("pk", flat=True))

        stats = EEPProgramHomeStatus.objects.filter(eep_program_id__in=kwargs["eep_program_id"])
        self.assertEqual(stats.count(), 2)

        status_1 = stats.first()
        status_1_program = status_1.eep_program.name
        status_1_answers_raw = status_1.get_answers_for_home()
        status_1_answers = {x.instrument.text: x.data["input"] for x in status_1_answers_raw}
        status_2 = stats.last()
        status_2_program = status_2.eep_program.name
        status_2_answers_raw = status_2.get_answers_for_home()
        status_2_answers = {x.instrument.text: x.data["input"] for x in status_2_answers_raw}
        self.assertNotEqual(status_1, status_2)
        self.assertNotEqual(status_1_program, status_2_program)

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        xls_data = self.get_results(kwargs["filename"])

        self.assertEqual(len(xls_data), 2)

        xls_status_1_data = next((x for x in xls_data if x["Program Name"] == status_1_program))
        for k, v in status_1_answers.items():
            if isinstance(v, dict) and "brand_name" in v.keys():  # Cascade Selects
                continue
            #  with self.subTest(f"{status_1_program} {k!r} -> {v!r} vs {xls_status_1_data[k]}"):
            self.assertEqual(str(xls_status_1_data[k]), str(v))

        xls_status_2_data = next((x for x in xls_data if x["Program Name"] == status_2_program))
        for k, v in status_2_answers.items():
            if isinstance(v, dict) and "brand_name" in v.keys():  # Cascade Selects
                continue
            #  with self.subTest(f"{status_1_program} {k!r} -> {v!r} vs {xls_status_2_data[k]}"):
            self.assertEqual(str(xls_status_2_data[k]), str(v))
        self.assertNotEqual(xls_status_1_data, xls_status_2_data)


class ProjectStatusReportSubdivisionCommunityTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_subdivision(self):
        """Verify the reported values from a subdivision"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["subdivision"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Subdivision/MF Development Name",
            "Subdivision/MF Development Builder Name",
            "Subdivision/MF Development Cross Roads",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))

    def test_report_on_community(self):
        """Verify the reported values from a community.  This will get reported if
        Subdivision is also in this"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["subdivision", "community"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Community Name",
            "Subdivision/MF Development Cross Roads",
            "Community Website",
            "Community Crossroads",
            "Subdivision/MF Development Builder Name",
            "Subdivision/MF Development Name",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))


class ProjectStatusReportRelationshipTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_relationships(self):
        """Verify the reported values from a relationships."""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["relationships"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Rating/Verification Company",
            "Builder",
            "Program Sponsor",
            "Owner Company",
            "Developer Company",
            "Architect Company",
            "QA/QC Company",
            "Certification Organization",
            "Electric Utility",
        }

        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))

    def test_report_on_relationships_new_company_types(self):
        """Verify the reported values from a relationships."""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["relationships"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        from axis.company.tests.factories import (
            architect_organization_factory,
            developer_organization_factory,
            communityowner_organization_factory,
            hvac_organization_factory,
            utility_organization_factory,
            general_organization_factory,
        )

        city = EEPProgramHomeStatus.objects.first().home.city
        architect = architect_organization_factory(city=city)
        developer = developer_organization_factory(city=city)
        communityowner = communityowner_organization_factory(city=city)
        general = general_organization_factory(city=city)
        hvac = hvac_organization_factory(city=city)
        ele = utility_organization_factory(
            name="ele",
            electricity_provider=True,
            gas_provider=False,
            water_provider=False,
            city=city,
        )
        gas = utility_organization_factory(
            name="gas",
            gas_provider=True,
            electricity_provider=False,
            city=city,
            water_provider=False,
        )
        water = utility_organization_factory(
            name="water",
            electricity_provider=False,
            gas_provider=False,
            water_provider=True,
            city=city,
        )

        companies = [architect, developer, communityowner, general, hvac, water]

        stat = EEPProgramHomeStatus.objects.first()
        create_or_update_spanning_relationships(companies, stat.home)
        rel_ele = create_or_update_spanning_relationships(ele, stat.home)[0][0]
        rel_gas = create_or_update_spanning_relationships(gas, stat.home)[0][0]
        stat.home._generate_utility_type_hints(rel_gas, rel_ele)

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Electric Utility",
            "Builder",
            "Rating/Verification Company",
            "Water Utility",
            "QA/QC Company",
            "Certification Organization",
            "Owner Company",
            "Program Sponsor",
            "Architect Company",
            "Developer Company",
            "Gas Utility",
            "General Company",
            "HVAC Contractor",
            "Water Utility",
        }

        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))

        # Our first one should have values.
        for k, v in data[0].items():
            if k in self.default_reported_keys:
                continue
            self.assertNotEqual(v, "-", "Found %s" % k)


class ProjectStatusReportEEPProgramTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_eep_program(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["eep_program"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Program Builder Incentive",
            "Program Rater Incentive",
            "Program Owner",
            "Program Minimum HERS Score",
            "Program Maximum HERS Score",
            "Program Start Date",
            "Program End Date",
            "Program Name",
        }

        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))


class ProjectStatusReportFloorplanTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_floorplan(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["floorplan"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Floorplan Owner Name",
            "Floorplan Square footage",
            "Floorplan Is custom home",
            "Floorplan Name",
            "Floorplan Number",
            "Floorplan Comment",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))


class ProjectStatusReporIPPTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_ipp(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["floorplan"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Floorplan Owner Name",
            "Floorplan Square footage",
            "Floorplan Is custom home",
            "Floorplan Name",
            "Floorplan Number",
            "Floorplan Comment",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))


class ProjectStatusReportInvoicingTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_invoicing(self):
        """Verify that if we do not retain empty we have less that normal"""
        from axis.eep_program.models import EEPProgram

        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )

        kwargs = self.default_kwargs
        kwargs["report_on"] = ["invoicing"]
        kwargs["user_id"] = stat.company.users.first().id
        kwargs["eep_program_id"] = list(EEPProgram.objects.all().values_list("pk", flat=True))

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "Invoice Issuer",
            "Invoice Status",
            "Invoice Customer",
            "Total Paid",
            "Balance",
            "Invoice Number",
            "Total",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))

    def test_report_on_invoicing_values(self):
        """Verify that if we do not retain empty we have less that normal"""
        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )

        kwargs = self.default_kwargs
        kwargs["report_on"] = ["invoicing"]
        kwargs["user_id"] = stat.company.users.first().id
        kwargs["eep_program_id"] = stat.eep_program_id

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(len(data), 1)

        data = data[0]

        self.assertEqual(data["Total"], InvoiceItemGroup.objects.get().total)
        self.assertEqual(data["Total Paid"], InvoiceItemGroup.objects.get().total_paid)
        self.assertEqual(data["Balance"], InvoiceItemGroup.objects.get().current_balance)

        invoice = Invoice.objects.get()
        self.assertEqual(data["Invoice Number"], str(invoice.id)[:8])
        self.assertEqual(data["Invoice Issuer"], str(invoice.issuer))
        self.assertEqual(data["Invoice Customer"], str(invoice.customer))
        self.assertEqual(data["Invoice Status"], dict(Invoice.STATE_CHOICES).get(invoice.state))

    def test_report_on_invoicing_values_ngbs_staff(self):
        """Verify that if we do not retain empty we have less that normal"""
        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )

        kwargs = self.default_kwargs
        kwargs["report_on"] = ["invoicing"]
        kwargs["user_id"] = (
            User.objects.filter(company__slug=customer_hirl_app.CUSTOMER_SLUG).first().id
        )
        kwargs["eep_program_id"] = stat.eep_program_id

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(len(data), 1)

        invoice_expected = {
            "Invoice Issuer",
            "Invoice Status",
            "Invoice Customer",
            "Total Paid",
            "Balance",
            "Invoice Number",
            "Total",
            "H-Number",
            "JAMIS Milestoned",
            "Most Recent Notice Sent",
            "Billing State",
        }
        expected = self.default_reported_keys.union(invoice_expected) - {"RESNET Registry ID"}

        self.assertEqual(set(data[0].keys()), expected)

        data = data[0]
        project = HIRLProject.objects.get()

        self.assertEqual(data["H-Number"], project.h_number)
        self.assertEqual(data["JAMIS Milestoned"], "No")
        self.assertEqual(data["JAMIS Milestoned"], "No")
        self.assertEqual(data["Billing State"], "Notice Sent")
        self.assertNotEqual(data["Most Recent Notice Sent"], "-")

    def add_invoice(self, home_status):
        from axis.invoicing.models import Invoice, InvoiceItemGroup

        self.assertEqual(Invoice.objects.count(), 1)
        self.assertEqual(InvoiceItemGroup.objects.count(), 1)
        self.assertEqual(HIRLProject.objects.count(), 1)
        project = HIRLProject.objects.get()
        registration = HIRLProjectRegistration.objects.get()

        # Add a new badge and add a second invoice to this sucker.
        energy_badge = hirl_green_energy_badge_factory(name="New Badge", cost=200)
        project.green_energy_badges.add(energy_badge)

        project.save()  # This will create the invoice items and the InvoiceItemGroup

        self.assertEqual(InvoiceItemGroup.objects.count(), 2)

        invoice_factory(
            invoice_type=Invoice.HIRL_PROJECT_INVOICE_TYPE,
            issuer=home_status.eep_program.owner,
            customer=registration.get_company_responsible_for_payment(),
            invoice_item_groups=[InvoiceItemGroup.objects.last()],
        )

        self.assertEqual(Invoice.objects.count(), 2)

    def test_multiple_invoices(self):
        """Test multiple invoices will work itself out."""

        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )
        self.add_invoice(stat)

        # Now run the report
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["invoicing"]
        kwargs["user_id"] = stat.company.users.first().id
        kwargs["eep_program_id"] = stat.eep_program_id

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(len(data), 1)

        data = data[0]

        # Verify that we actually have multiples represented.
        self.assertIn(", ", data["Total"])
        self.assertIn("$0.00, $0.00", data["Total Paid"])
        self.assertIn(", ", data["Balance"])
        self.assertIn(", ", data["Invoice Number"])
        self.assertIn(", ", data["Invoice Issuer"])
        self.assertIn(", ", data["Invoice Status"])

    def test_multiple_invoices_nbgs(self):
        """Test multiple invoices will work itself out."""

        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )
        self.add_invoice(stat)

        # Now run the report
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["invoicing"]
        kwargs["user_id"] = (
            User.objects.filter(company__slug=customer_hirl_app.CUSTOMER_SLUG).first().id
        )
        kwargs["eep_program_id"] = stat.eep_program_id

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])

        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(len(data), 1)

        data = data[0]

        # Verify that we actually have multiples represented.
        self.assertIn(", ", data["Total"])
        self.assertIn("$0.00, $0.00", data["Total Paid"])
        self.assertIn(", ", data["Balance"])
        self.assertIn(", ", data["Invoice Number"])
        self.assertIn(", ", data["Invoice Issuer"])
        self.assertIn(", ", data["Invoice Status"])


class ProjectStatusSimulationBasicTests(ProjectStatusReportMixin, AxisTestCase):
    @property
    def rem_basic(self):
        return {
            "Area of Conditioned Space (sq. ft.)",
            "Building Type",
            "ENERGY STAR v2.5 Reference Design HERS score",
            "ENERGY STAR v2.5 photo voltaic adjusted HERS score",
            "ENERGY STAR v2.5 size adjustment factor",
            "ENERGY STAR v3 Reference Design HERS score",
            "ENERGY STAR v3 photo voltaic adjusted HERS score",
            "ENERGY STAR v3 size adjustment factor",
            "ENERGY STAR v3.1 Reference Design HERS score",
            "ENERGY STAR v3.1 photo voltaic adjusted HERS score",
            "ENERGY STAR v3.1 size adjustment factor",
            "EPACT Tax Credit",
            "Floors on or Above-Grade",
            "Foundation Type",
            "HERS Score",
            "Number of Bedrooms",
            "Number of Units",
            "Passes DOE Zero Energy Rated Home",
            "Passes IECC 2006 Code",
            "Passes IECC 2009 Code",
            "Passes IECC 2012 Code",
            "Passes IECC 2015 Code",
            "Passes IECC 2018 Code",
            "Project Name",
            "REM/Rate Flavor",
            "REM/Rate Simulation Date",
            "REM/Rate Username",
            "REM/Rate Version",
            "Volume of Conditioned Space (cu. ft.)",
        }

    def test_report_on_simulation_basic_remrate(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["simulation_basic"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        floorplan = EEPProgramHomeStatus.objects.first().floorplan
        from axis.remrate_data.tests.factories import simulation_factory

        floorplan.remrate_target = simulation_factory(company=floorplan.owner)
        floorplan.save()

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(self.rem_basic))

    @property
    def ekotrope_basic(self):
        return {
            "Area of Conditioned Space (sq.ft.)",
            "Building Type",
            "Ekotrope Simulation Date",
            "Ekotrope Version",
            "Floors on or Above-Grade",
            "Foundation Type",
            "HERS Score",
            "Number of Bedrooms",
            "Passes EnergyStar V3",
            "Passes EnergyStar V3.1",
            "Passes IECC 2006 Performance",
            "Passes IECC 2006 Prescriptive",
            "Passes IECC 2009 Performance",
            "Passes IECC 2009 Prescriptive",
            "Passes IECC 2012 Performance",
            "Passes IECC 2012 Prescriptive",
            "Passes IECC 2015 ERI",
            "Passes IECC 2015 Performance",
            "Passes IECC 2015 Prescriptive",
            "Passes IECC 2018 ERI",
            "Passes IECC 2018 Performance",
            "Passes IECC 2018 Prescriptive",
            "Passes TaxCredit 45L",
            "Project ID",
            "Project Name",
            "Volume of Conditioned Space",
        }

    def test_report_on_simulation_basic_ekotrope(self):
        """Verify that if we do not retain empty we have less that normal"""
        kwargs = self.default_kwargs
        kwargs["report_on"] = ["simulation_basic"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        floorplan = EEPProgramHomeStatus.objects.first().floorplan
        from axis.ekotrope.tests.factories import analysis_factory

        floorplan.ekotrope_houseplan = analysis_factory(company=floorplan.owner).houseplan
        floorplan.save()

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())
        # print(sorted(set(data[0].keys()) - self.default_reported_keys))

        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(self.ekotrope_basic))


class ProjectStatusReportNGBSDataTests(ProjectStatusReportMixin, AxisTestCase):
    def test_report_on_ngbs_data(self):
        """Verify the reported values from a subdivision"""
        stat = EEPProgramHomeStatus.objects.get(
            eep_program__owner__slug=customer_hirl_app.CUSTOMER_SLUG
        )

        kwargs = self.default_kwargs
        kwargs["report_on"] = ["ngbs_data"]
        kwargs["user_id"] = stat.company.users.first().id
        kwargs["eep_program_id"] = stat.eep_program_id

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        # import pprint
        # pprint.pprint(data)

        # print(sorted(set(data[0].keys()) - self.default_reported_keys))

        addl_expected = {
            "NGBS Project ID",
            "Legacy Project ID",
            "Certification Number",
            "NGBS Project Type",
            "NGBS Project State",
            "Green Energy Badges",
            "Story Count",
            "Unit Count",
            "Is this an appeals project?",
            "Is this project Build-To-Rent?",
            "Certification Level",
            "Client CA Status",
            "Client COI Status",
            "Project Status",
            "Final Verifier",
            "Rough Verifier",
            "Certification Date",
            "Certification Standard",
            "Accessory Dwelling Unit",
            "Commercial Space",
            "Accessory Structure",
        }
        self.assertEqual(set(data[0].keys()), addl_expected)


class ProjectStatusReportCustomerETOTests(ProjectStatusReportMixin, AxisTestCase):
    def setup_project_tracker(self):
        from axis.customer_eto.models import FastTrackSubmission

        defaults = {
            "project_id": "P6666",
            "solar_project_id": "P7777",
            "eps_score": 11,
            "eps_score_built_to_code_score": 12,
            "percent_improvement": 0.387,
            "builder_incentive": Decimal("2.50"),
            "original_builder_incentive": Decimal("2.51"),
            "builder_electric_incentive": Decimal("2.52"),
            "builder_gas_incentive": Decimal("2.53"),
            "rater_incentive": Decimal("2.54"),
            "original_rater_incentive": Decimal("2.55"),
            "rater_electric_incentive": Decimal("2.56"),
            "rater_gas_incentive": Decimal("2.57"),
            "carbon_score": Decimal("2.58"),
            "carbon_built_to_code_score": Decimal("2.59"),
            "estimated_annual_energy_costs": Decimal("2.60"),
            "estimated_monthly_energy_costs": Decimal("2.61"),
            "similar_size_eps_score": Decimal("2.62"),
            "similar_size_carbon_score": Decimal("2.63"),
            "therm_savings": Decimal("2.64"),
            "kwh_savings": Decimal("2.65"),
            "mbtu_savings": Decimal("2.66"),
            "net_zero_eps_incentive": Decimal("2.67"),
            "net_zero_solar_incentive": Decimal("2.68"),
            "energy_smart_homes_eps_incentive": Decimal("2.69"),
            "energy_smart_homes_solar_incentive": Decimal("2.70"),
            "solar_ready_builder_incentive": Decimal("2.71"),
            "solar_ready_verifier_incentive": Decimal("2.72"),
            "ev_ready_builder_incentive": Decimal("2.73"),
            "solar_storage_builder_incentive": Decimal("2.74"),
            "heat_pump_water_heater_incentive": Decimal("-2.75"),
            "cobid_builder_incentive": Decimal("2.76"),
            "required_credits_to_meet_code": Decimal("2.78"),
            "achieved_total_credits": Decimal("2.79"),
            "eligible_gas_points": Decimal("2.80"),
            "code_credit_incentive": Decimal("2.81"),
            "thermostat_incentive": Decimal("2.82"),
            "fireplace_incentive": Decimal("2.83"),
        }

        FastTrackSubmission.objects.update_or_create(
            home_status=EEPProgramHomeStatus.objects.filter(eep_program__slug="eto-2022").first(),
            defaults=defaults,
        )
        return defaults

    def test_report_on_ipp(self):
        """Verify that if we do not retain empty we have less that normal"""

        self.setup_project_tracker()
        GreenBuildingRegistry.objects.create(
            home=EEPProgramHomeStatus.objects.filter(eep_program__slug="eto-2022").first().home,
            gbr_id="Fuddy Duddy",
        )

        kwargs = self.default_kwargs
        kwargs["report_on"] = ["customer_eto"]

        obj = HomeDataXLSExport(**kwargs)
        obj.update_task = lambda: None

        obj.write(output=kwargs["filename"])
        data = self.get_results(kwargs["filename"])
        self.assertEqual(len(data), obj.get_queryset().count())

        expected = {
            "DEI Builder Incentive",
            "Annual Energy Costs",
            "Carbon Score (Code Build)",
            "Similar sized Carbon Score",
            "WA Code Credit Eligible Above Code Credits",
            "ENH PT ID",
            "EPS Score",
            "Solar Ready Verifier Incentive",
            "EPS Score (Code Build)",
            "Builder Incentive Override",
            "Verifier Electric Incentive",
            "Percent Improvement",
            "WA Code Credit Required Code Credits",
            "WA Code Credit Fireplace Incentive",
            "Builder Gas Incentive",
            "Verifier Gas Incentive",
            "Similar sized EPS Score",
            "Savings (KWH)",
            "WA Code Credit Selected Credits",
            "SLE PT ID",
            "Savings (MBtu)",
            "ESH: Solar + Storage Incentive",
            "Monthly Energy Costs",
            "Solar Ready Builder Incentive",
            "Total Verifier Incentive",
            "Verifier Incentive Override",
            "ESH: EV Ready Incentive",
            "Total Builder Incentive",
            "Builder Electric Incentive",
            "Heat Pump Water Heater Deduction",
            "WA Code Credit Thermostat Incentive",
            "WA Code Credit Code Incentive",
            # "Energy Smart Homes EPS Incentive",
            # "Energy Smart Homes Solar Incentive",
            "Net Zero EPS Incentive",
            "Fire Rebuild Triple Pane Windows Bonus Incentive",
            "Fire Rebuild Exterior Rigid Insulation Bonus Incentive",
            "Fire Rebuild Sealed Attic Bonus Incentive",
            "Net Zero Solar Incentive",
            "Carbon Score",
            "Savings (therms)",
            "DEI Rater Incentive",
            "City of Hillsboro Certificate of Occupancy Compliance Report",
            "City of Hillsboro Building Permit Compliance Report",
            "Green Building Registry ID",
        }
        self.assertEqual(set(data[0].keys()), self.default_reported_keys.union(expected))
        self.assertEqual(data[0]["ENH PT ID"], "P6666")
        self.assertEqual(data[0]["SLE PT ID"], "P7777")
        self.assertEqual(data[0]["EPS Score"], 11)
        self.assertEqual(data[0]["EPS Score (Code Build)"], 12)
        self.assertEqual(data[0]["Percent Improvement"], 0.387)
        self.assertEqual(data[0]["Total Builder Incentive"], 2.5)
        self.assertEqual(data[0]["Builder Incentive Override"], "Yes")
        self.assertEqual(data[0]["Builder Electric Incentive"], 2.52)
        self.assertEqual(data[0]["Builder Gas Incentive"], 2.53)
        self.assertEqual(data[0]["Total Verifier Incentive"], 2.54)
        self.assertEqual(data[0]["Verifier Incentive Override"], 2.55)
        self.assertEqual(data[0]["Verifier Electric Incentive"], 2.56)
        self.assertEqual(data[0]["Verifier Gas Incentive"], 2.57)
        self.assertEqual(data[0]["Carbon Score"], 2.58)
        self.assertEqual(data[0]["Carbon Score (Code Build)"], 2.59)
        self.assertEqual(data[0]["Annual Energy Costs"], 2.6)
        self.assertEqual(data[0]["Monthly Energy Costs"], 2.61)
        self.assertEqual(data[0]["Similar sized EPS Score"], 2)
        self.assertEqual(data[0]["Similar sized Carbon Score"], 2.63)
        self.assertEqual(data[0]["Savings (therms)"], 2.64)
        self.assertEqual(data[0]["Savings (KWH)"], 2.65)
        self.assertEqual(data[0]["Savings (MBtu)"], 2.66)
        # self.assertEqual(data[0]["Net Zero EPS Incentive"], 2.67)
        # self.assertEqual(data[0]["Energy Smart Homes EPS Incentive"], 2.69)
        # self.assertEqual(data[0]["Energy Smart Homes Solar Incentive"], 2.7)
        self.assertEqual(data[0]["Net Zero Solar Incentive"], 2.68)
        self.assertEqual(data[0]["Solar Ready Builder Incentive"], 2.71)
        self.assertEqual(data[0]["Solar Ready Verifier Incentive"], 2.72)
        self.assertEqual(data[0]["ESH: EV Ready Incentive"], 2.73)
        self.assertEqual(data[0]["ESH: Solar + Storage Incentive"], 2.74)
        self.assertEqual(data[0]["Heat Pump Water Heater Deduction"], -2.75)
        self.assertEqual(data[0]["DEI Builder Incentive"], 2.76)
        self.assertEqual(data[0]["DEI Rater Incentive"], "-")
        self.assertEqual(data[0]["WA Code Credit Required Code Credits"], 2.78)
        self.assertEqual(data[0]["WA Code Credit Selected Credits"], 2.79)
        self.assertEqual(data[0]["WA Code Credit Eligible Above Code Credits"], 2.8)
        self.assertEqual(data[0]["WA Code Credit Code Incentive"], 2.81)
        self.assertEqual(data[0]["WA Code Credit Thermostat Incentive"], 2.82)
        self.assertEqual(data[0]["WA Code Credit Fireplace Incentive"], 2.83)
        self.assertEqual(data[0]["City of Hillsboro Building Permit Compliance Report"], "No")
        self.assertEqual(
            data[0]["City of Hillsboro Certificate of Occupancy Compliance Report"],
            "No",
        )
        self.assertEqual(data[0]["Green Building Registry ID"], "Fuddy Duddy")
