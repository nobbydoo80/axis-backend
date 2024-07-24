"""project.py: """

__author__ = "Artem Hruzd"
__date__ = "12/08/2020 15:02"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import io
import os
import random
from unittest import mock

from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files import File
from django.core.files.base import ContentFile
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce
from django.urls import reverse_lazy
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status

from axis.company.tests.factories import (
    provider_organization_factory,
    builder_organization_factory,
    developer_organization_factory,
    architect_organization_factory,
    communityowner_organization_factory,
)
from axis.core.reports import AxisReportFormatter
from axis.core.tests.factories import provider_user_factory, SET_NULL
from axis.core.tests.testcases import ApiV3Tests
from axis.customer_hirl.models import HIRLProject, HIRLProjectRegistration, BuilderAgreement
from axis.customer_hirl.tests.factories import (
    hirl_project_factory,
    hirl_green_energy_badge_factory,
    hirl_project_registration_factory,
    builder_agreement_factory,
)
from axis.eep_program.tests.factories import basic_eep_program_factory
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.home.models import EEPProgramHomeStatus
from axis.invoicing.models import InvoiceItem, Invoice, InvoiceItemTransaction, InvoiceItemGroup

User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


class TestHIRLProjectViewSet(ApiV3Tests):
    def test_list(self):
        list_url = reverse_lazy("api_v3:hirl_projects-list")
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        sf_registration = hirl_project_registration_factory(
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE
        )
        sf_project = hirl_project_factory(
            registration=sf_registration, home_address_geocode_response=None
        )

        kwargs = dict(url=list_url, user=hirl_user)

        data = self.list(**kwargs)
        self.assertEqual(len(data), HIRLProject.objects.count())
        self.assertEqual(data[0]["id"], sf_project.id)

    def test_update_registration_for_project(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )

        eep_program = basic_eep_program_factory(
            name="SF New Construction 2020", slug="ngbs-sf-new-construction-2020-new"
        )

        builder_organization = builder_organization_factory(company_type="builder")
        developer_organization = developer_organization_factory(company_type="developer")
        architect_organization = architect_organization_factory(company_type="architect")
        community_owner_organization = communityowner_organization_factory(
            company_type="communityowner"
        )

        sf_registration = hirl_project_registration_factory(
            eep_program=eep_program,
            state=HIRLProjectRegistration.ACTIVE_STATE,
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE,
            builder_organization=builder_organization,
            developer_organization=developer_organization,
            architect_organization=architect_organization,
            community_owner_organization=community_owner_organization,
        )
        sf_project = hirl_project_factory(
            registration=sf_registration, home_status=SET_NULL, home_address_geocode_response=None
        )

        sf_registration2 = hirl_project_registration_factory(
            eep_program=eep_program,
            state=HIRLProjectRegistration.ACTIVE_STATE,
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE,
            builder_organization=builder_organization,
            developer_organization=developer_organization,
            architect_organization=architect_organization,
            community_owner_organization=community_owner_organization,
        )

        update_url = reverse_lazy("api_v3:hirl_projects-detail", args=(sf_project.id,))
        obj = self.update(
            user=hirl_user,
            url=update_url,
            partial=True,
            data=dict(registration=str(sf_registration2.id)),
        )
        self.assertEqual(obj["registration"], str(sf_registration2.id))

        with self.subTest("Test with different states"):
            sf_registration.state = HIRLProjectRegistration.REJECTED_STATE
            sf_registration.save()
            self.update(
                user=hirl_user,
                url=update_url,
                partial=True,
                data=dict(registration=str(sf_registration.id)),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )
            sf_registration.state = HIRLProjectRegistration.ACTIVE_STATE
            sf_registration.save()

        with self.subTest("Test with different eep program"):
            sf_registration.eep_program = None
            sf_registration.save()
            self.update(
                user=hirl_user,
                url=update_url,
                partial=True,
                data=dict(registration=str(sf_registration.id)),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )
            sf_registration.eep_program = eep_program
            sf_registration.save()

        with self.subTest("Test with different builder organization"):
            sf_registration.builder_organization = None
            sf_registration.save()
            self.update(
                user=hirl_user,
                url=update_url,
                partial=True,
                data=dict(registration=str(sf_registration.id)),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )
            sf_registration.builder_organization = builder_organization
            sf_registration.save()

        with self.subTest("Test with different developer organization"):
            sf_registration.developer_organization = None
            sf_registration.save()
            self.update(
                user=hirl_user,
                url=update_url,
                partial=True,
                data=dict(registration=str(sf_registration.id)),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )
            sf_registration.developer_organization = developer_organization
            sf_registration.save()

        with self.subTest("Test with different architect organization"):
            sf_registration.architect_organization = None
            sf_registration.save()
            self.update(
                user=hirl_user,
                url=update_url,
                partial=True,
                data=dict(registration=str(sf_registration.id)),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )
            sf_registration.architect_organization = architect_organization
            sf_registration.save()

        with self.subTest("Test with different community owner organization"):
            sf_registration.community_owner_organization = None
            sf_registration.save()
            self.update(
                user=hirl_user,
                url=update_url,
                partial=True,
                data=dict(registration=str(sf_registration.id)),
                expected_status=status.HTTP_400_BAD_REQUEST,
            )
            sf_registration.community_owner_organization = community_owner_organization
            sf_registration.save()

    def test_billing_rule_export_as_hirl_user(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        builder_organization = builder_organization_factory()
        architect_organization = architect_organization_factory()
        builder_agreement_factory(
            owner=hirl_company, company=architect_organization, state=BuilderAgreement.COUNTERSIGNED
        )

        mf_registration = hirl_project_registration_factory(
            eep_program__slug=random.choice(customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS),
            project_type=HIRLProjectRegistration.MULTI_FAMILY_PROJECT_TYPE,
            builder_organization=builder_organization,
            architect_organization=architect_organization,
            project_client=HIRLProjectRegistration.PROJECT_CLIENT_ARCHITECT,
        )
        mf_project = hirl_project_factory(
            registration=mf_registration,
            home_address_geocode_response=None,
            story_count=1,
        )

        # activate a project and create new home with program
        mf_project.registration.active()

        hirl_green_energy_badge = hirl_green_energy_badge_factory(name="Test Badge", cost=10)
        mf_project.green_energy_badges.add(hirl_green_energy_badge)

        total_fees = InvoiceItem.objects.filter(
            group__home_status=mf_project.home_status, protected=True
        ).aggregate(total_fees=Coalesce(Sum("cost"), 0, output_field=DecimalField()))["total_fees"]
        # certification fee * 2(because of badge story_count) + our two test badges
        expected_fees_cost = mf_project.calculate_certification_fees_cost() + 20
        self.assertEqual(total_fees, expected_fees_cost)

        # Create invoice for builder, so this means that he is ready to pay for this project
        invoice = Invoice.objects.create(issuer=hirl_company)

        for group in InvoiceItemGroup.objects.filter(home_status=mf_project.home_status):
            group.invoice = invoice
            group.save()

        billing_rule_export_url = reverse_lazy("api_v3:hirl_projects-billing-rule-export")

        self.client.force_authenticate(user=hirl_user)
        response = self.client.get(billing_rule_export_url, data={}, format="json")

        async_document = AsynchronousProcessedDocument.objects.get(id=response.data["id"])
        wb = load_workbook(ContentFile(async_document.document.read()), data_only=True)
        sheet = wb["BillingRule"]
        h_number = sheet.cell(row=2, column=1).value
        billing_rule_id = sheet.cell(row=2, column=2).value
        certification_fee = sheet.cell(row=2, column=3).value
        revenue_percentage = sheet.cell(row=2, column=4).value
        certification_status = sheet.cell(row=2, column=5).value
        responsible_company_id = sheet.cell(row=2, column=6).value
        job_id = sheet.cell(row=2, column=7).value
        invoice_rule_id = sheet.cell(row=2, column=8).value
        invoice_sent_date = sheet.cell(row=2, column=9).value

        self.assertEqual(h_number, mf_project.h_number)
        self.assertEqual(billing_rule_id, mf_project.get_billing_rule_id())
        self.assertEqual(billing_rule_id, mf_project.get_billing_rule_id())
        self.assertEqual(certification_fee, int(expected_fees_cost))
        self.assertEqual(revenue_percentage, "0")
        self.assertEqual(
            certification_status,
            HIRLProject.BILLING_STATE_DISPLAY[mf_project.NOTICE_SENT_BILLING_STATE],
        )
        self.assertEqual(
            responsible_company_id,
            f"B{mf_project.registration.architect_organization.hirlcompanyclient.id:05}",
        )
        self.assertEqual(job_id, mf_project.get_job_id())
        self.assertEqual(invoice_rule_id, mf_project.get_invoice_rule_id())

        most_recent_notice_date = InvoiceItemGroup.objects.first().created_at.strftime("GRN%m%Y")
        self.assertEqual(invoice_sent_date, most_recent_notice_date)

    @mock.patch(
        "axis.customer_hirl.tasks.projects.customer_hirl_app.get_accounting_users",
    )
    def test_green_payments_import_as_hirl_user(self, get_accounting_users):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS",
            last_name="Admin",
            email="test@gmail.com",
            company=hirl_company,
            is_company_admin=True,
        )

        get_accounting_users.return_value = User.objects.filter(id=hirl_user.id)

        sf_registration = hirl_project_registration_factory(
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE
        )
        sf_project = hirl_project_factory(
            registration=sf_registration, home_address_geocode_response=None
        )

        # activate a project and create new home with program
        sf_project.registration.active()

        hirl_green_energy_badge = hirl_green_energy_badge_factory(name="Test Badge", cost=10)
        sf_project.green_energy_badges.add(hirl_green_energy_badge)

        total_fees = InvoiceItem.objects.filter(
            group__home_status=sf_project.home_status,
        ).aggregate(total_fees=Coalesce(Sum("cost"), 0, output_field=DecimalField()))["total_fees"]
        # certification fee + our two test badges
        expected_fees_cost = sf_project.calculate_certification_fees_cost() + 10
        self.assertEqual(total_fees, expected_fees_cost)

        total_paid = InvoiceItemTransaction.objects.filter(
            item__group__home_status=sf_project.home_status,
        ).aggregate(total_paid=Coalesce(Sum("amount"), 0, output_field=DecimalField()))[
            "total_paid"
        ]

        self.assertEqual(total_paid, 0)

        file_path = os.path.join(
            settings.SITE_ROOT, "axis", "customer_hirl", "sources", "tests", "green_payments.xlsx"
        )
        with io.open(file_path, "rb") as f:
            green_payments_import_url = reverse_lazy("api_v3:hirl_projects-green-payments-import")

            self.client.force_authenticate(user=hirl_user)
            response = self.client.post(
                green_payments_import_url,
                data={"jamis_file": File(f, name=os.path.basename(file_path))},
                format="multipart",
            )
            self.assertEqual(response.status_code, 200)

        total_paid = InvoiceItemTransaction.objects.filter(
            item__group__home_status=sf_project.home_status,
        ).aggregate(total_paid=Coalesce(Sum("amount"), 0, output_field=DecimalField()))[
            "total_paid"
        ]

        self.assertEqual(total_paid, 200)

        # first email about registration change state
        # second about Project Fees update
        # third for ngbs admins green payment report
        self.assertEqual(len(mail.outbox), 3)
        self.assertEqual(mail.outbox[2].subject, "New Green Payments file has been imported")

    def test_project_billing_import_as_hirl_user(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )

        sf_registration = hirl_project_registration_factory(
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE
        )
        sf_project = hirl_project_factory(
            registration=sf_registration,
            home_address_geocode_response=None,
            home_status__state="complete",
            home_status__certification_date=timezone.now(),
        )

        self.assertFalse(sf_project.is_jamis_milestoned)

        file_path = os.path.join(
            settings.SITE_ROOT, "axis", "customer_hirl", "sources", "tests", "project_billing.xlsx"
        )

        with io.open(file_path, "rb") as f:
            project_billing_import_url = reverse_lazy("api_v3:hirl_projects-project-billing-import")

            self.client.force_authenticate(user=hirl_user)
            response = self.client.post(
                project_billing_import_url,
                data={"jamis_file": File(f, name=os.path.basename(file_path))},
                format="multipart",
            )
            self.assertEqual(response.status_code, 200)

        sf_project.refresh_from_db()
        self.assertTrue(sf_project.is_jamis_milestoned)

    def test_customer_hirl_all_projects_report(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )
        hirl_user = provider_user_factory(
            first_name="NGBS", last_name="Admin", company=hirl_company, is_company_admin=True
        )
        eep_program = basic_eep_program_factory(
            name="SF New Construction 2020",
            slug="ngbs-sf-new-construction-2020-new",
            owner=hirl_company,
        )

        # new ngbs home
        sf_registration = hirl_project_registration_factory(
            eep_program=eep_program, project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE
        )
        sf_project = hirl_project_factory(
            registration=sf_registration, home_address_geocode_response=None, home_status=SET_NULL
        )

        sf_project.registration.active()
        sf_project.registration.save()
        sf_project.create_home_status()

        all_project_report_url = reverse_lazy("api_v3:hirl_projects-all-projects-report")

        self.client.force_authenticate(user=hirl_user)
        response = self.client.get(
            all_project_report_url,
            data={"eep_program_slug": [eep_program.slug, "invalid_slug"], "invalid_param": "55"},
        )

        async_document = AsynchronousProcessedDocument.objects.get(id=response.data["id"])
        wb = load_workbook(ContentFile(async_document.document.read()), data_only=True)
        sheet = wb["NGBSCertficationRawData"]

        axis_report_formatter = AxisReportFormatter(user=hirl_user)

        street_line1 = sheet.cell(row=2, column=1).value
        certification_date = sheet.cell(row=2, column=2).value
        program_name = sheet.cell(row=2, column=3).value
        rough_receive_date = sheet.cell(row=2, column=4).value
        invoice_sent_date = sheet.cell(row=2, column=5).value
        payment_received_date = sheet.cell(row=2, column=6).value
        created_date = sheet.cell(row=2, column=7).value
        certification_status = sheet.cell(row=2, column=8).value
        invoice_number = sheet.cell(row=2, column=9).value
        total_fee = sheet.cell(row=2, column=10).value
        builder_id = sheet.cell(row=2, column=11).value

        self.assertEqual(street_line1, sf_project.home_status.home.street_line1)
        self.assertEqual(certification_date, None)
        self.assertEqual(program_name, eep_program.name)
        self.assertEqual(rough_receive_date, None)
        self.assertIsNotNone(invoice_sent_date)
        self.assertEqual(payment_received_date, "-")
        self.assertEqual(
            created_date,
            axis_report_formatter.get_formatted_datetime(sf_project.home_status.created_date),
        )
        self.assertEqual(certification_status, sf_project.NEW_NOTIFIED_BILLING_STATE)
        self.assertEqual(invoice_number, sf_project.h_number)
        self.assertEqual(total_fee, sf_project.calculate_certification_fees_cost())
        self.assertEqual(
            builder_id, sf_project.registration.builder_organization.hirlcompanyclient.id
        )

    def test_project_cycle_time_metrics(self):
        hirl_company = provider_organization_factory(
            name="Home Innovation Research Labs", slug=customer_hirl_app.CUSTOMER_SLUG
        )

        hirl_user = provider_user_factory(company=hirl_company)

        non_customer_hirl_user = provider_user_factory()

        sf_registration = hirl_project_registration_factory(
            project_type=HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE,
            eep_program__owner=hirl_company,
        )
        sf_project = hirl_project_factory(
            registration=sf_registration, home_address_geocode_response=None, home_status=SET_NULL
        )
        today = timezone.now().today()

        sf_project.create_home_status()
        sf_project.refresh_from_db()
        sf_project.home_status.state = EEPProgramHomeStatus.COMPLETE_STATE
        sf_project.home_status.certification_date = today
        sf_project.home_status.save()

        url = reverse_lazy("api_v3:hirl_projects-project-cycle-time-metrics")
        self.client.force_authenticate(user=hirl_user)
        response = self.client.get(url, format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data[0]["cycle_days"], 0)
        self.assertEqual(response.data[0]["projects_count"], 1)

        with self.subTest("Restrict access for non HIRL users"):
            self.client.force_authenticate(user=non_customer_hirl_user)
            response = self.client.get(url, format="json")
            self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
