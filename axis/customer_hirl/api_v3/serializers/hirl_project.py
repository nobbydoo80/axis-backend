"""hirl_project.py: """

__author__ = "Artem Hruzd"
__date__ = "04/16/2021 17:20"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import pytz
from django.apps import apps
from django.utils import timezone
from hashid_field.rest import HashidSerializerCharField
from openpyxl import load_workbook
from rest_framework import serializers

from axis.core.api_v3.serializers import UserInfoSerializer
from axis.customer_hirl.models import HIRLProject, HIRLProjectRegistration
from axis.customer_hirl.tasks import green_payments_import_task, project_billing_import_task
from axis.customer_hirl.utils import (
    hirl_project_address_is_unique,
    get_hirl_project_address_components,
)
from axis.eep_program.api_v3.serializers import EEPProgramInfoSerializer
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.geocoder.api_v3.serializers import (
    GeocodeInfoSerializer,
    GeocodeBrokeredResponseSerializer,
)
from axis.geographic.models import City
from axis.home.api_v3.serializers import EEPProgramHomeStatusInfoSerializer, HomeInfoSerializer
from axis.home.models import EEPProgramHomeStatus
from .green_energy_badge import HIRLGreenEnergyBadgeInfoSerializer

customer_hirl_app = apps.get_app_config("customer_hirl")


class HIRLProjectInfoSerializer(serializers.ModelSerializer):
    id = HashidSerializerCharField(source_field="customer_hirl.HIRLProject.id", read_only=True)
    home_status_info = EEPProgramHomeStatusInfoSerializer(source="home_status", read_only=True)
    home_address_geocode_info = GeocodeInfoSerializer(source="home_address_geocode", read_only=True)
    home_address_geocode_response_info = GeocodeBrokeredResponseSerializer(
        source="home_address_geocode_response", read_only=True
    )

    class Meta:
        model = HIRLProject
        fields = (
            "id",
            "home_status",
            "home_status_info",
            "home_address_geocode",
            "home_address_geocode_info",
            "home_address_geocode_response",
            "home_address_geocode_response_info",
            "is_accessory_structure",
            "accessory_structure_description",
            "is_accessory_dwelling_unit",
            "accessory_dwelling_unit_description",
            "h_number",
            "created_at",
        )


class HIRLProjectRegistrationProjectSerializerInfoSerializer(serializers.ModelSerializer):
    """
    Short info version for HIRLProjectSerializer that is not contain subdivision info
    """

    id = HashidSerializerCharField(
        source_field="customer_hirl.HIRLProjectRegistration.id", read_only=True
    )
    registration_user_info = UserInfoSerializer(source="registration_user", read_only=True)
    eep_program_info = EEPProgramInfoSerializer(source="eep_program", read_only=True)

    class Meta:
        model = HIRLProjectRegistration
        fields = (
            "id",
            "project_type",
            "state",
            "state_changed_at",
            "state_change_reason",
            "registration_user",
            "registration_user_info",
            "eep_program",
            "eep_program_info",
            "subdivision",
            "project_name",
        )


class HIRLProjectEEPProgramHomeStatusInfoSerializer(serializers.ModelSerializer):
    home_info = HomeInfoSerializer(source="home")

    class Meta:
        model = EEPProgramHomeStatus
        fields = (
            "id",
            "home",
            "home_info",
            "state",
            # Accounting
            "pct_complete",
            "is_billable",
            "modified_date",
            "created_date",
        )


class HIRLProjectSerializerMixin(metaclass=serializers.SerializerMetaclass):
    id = HashidSerializerCharField(source_field="customer_hirl.HIRLProject.id", read_only=True)
    registration = serializers.PrimaryKeyRelatedField(
        pk_field=HashidSerializerCharField(source_field="customer_hirl.HIRLProjectRegistration.id"),
        queryset=HIRLProjectRegistration.objects.all(),
    )
    commercial_space_parent = serializers.PrimaryKeyRelatedField(
        pk_field=HashidSerializerCharField(source_field="customer_hirl.HIRLProject.id"),
        queryset=HIRLProject.objects.all(),
        required=False,
    )
    home_status_info = HIRLProjectEEPProgramHomeStatusInfoSerializer(
        source="home_status", read_only=True
    )
    home_address_geocode_info = GeocodeInfoSerializer(source="home_address_geocode", read_only=True)
    home_address_geocode_response_info = GeocodeBrokeredResponseSerializer(
        source="home_address_geocode_response", read_only=True
    )
    green_energy_badges_info = HIRLGreenEnergyBadgeInfoSerializer(
        source="green_energy_badges", many=True, read_only=True
    )
    registration_info = HIRLProjectRegistrationProjectSerializerInfoSerializer(
        source="registration", read_only=True
    )

    def validate(self, data):
        instance = getattr(self, "instance", None)
        # project type is read only
        project_type = data.pop("project_type", None)
        registration = data.get("registration")

        if instance:
            project_type = instance.registration.project_type

            if registration and instance.registration != registration:
                if getattr(instance.registration, "state", None) != getattr(
                    registration, "state", None
                ):
                    raise serializers.ValidationError(
                        {
                            "registration": f"Current Registration {instance.registration.id} and "
                            f"provided Registration {registration.id} have different workflow states"
                        }
                    )
                if getattr(instance.registration, "eep_program", None) != getattr(
                    registration, "eep_program", None
                ):
                    raise serializers.ValidationError(
                        {
                            "registration": f"Current Registration {instance.registration.id} and "
                            f"provided Registration {registration.id} have different programs"
                        }
                    )
                if getattr(instance.registration, "builder_organization", None) != getattr(
                    registration, "builder_organization", None
                ):
                    raise serializers.ValidationError(
                        {
                            "registration": f"Current Registration {instance.registration.id} and "
                            f"provided Registration {registration.id} have different Builder Organizations"
                        }
                    )
                if getattr(instance.registration, "developer_organization", None) != getattr(
                    registration, "developer_organization", None
                ):
                    raise serializers.ValidationError(
                        {
                            "registration": f"Current Registration {instance.registration.id} and "
                            f"provided Registration {registration.id} have different Developer Organizations"
                        }
                    )
                if getattr(instance.registration, "architect_organization", None) != getattr(
                    registration, "architect_organization", None
                ):
                    raise serializers.ValidationError(
                        {
                            "registration": f"Current Registration {instance.registration.id} and "
                            f"provided Registration {registration.id} have different Architect Organizations"
                        }
                    )
                if getattr(instance.registration, "community_owner_organization", None) != getattr(
                    registration, "community_owner_organization", None
                ):
                    raise serializers.ValidationError(
                        {
                            "registration": f"Current Registration {instance.registration.id} and "
                            f"provided Registration {registration.id} have different Community Owner Organizations"
                        }
                    )

        is_accessory_structure = data.get("is_accessory_structure")

        if is_accessory_structure:
            if not data.get("accessory_structure_description"):
                raise serializers.ValidationError(
                    {"accessory_structure_description": "Description is required"}
                )

        is_accessory_dwelling_unit = data.get("is_accessory_dwelling_unit")

        if is_accessory_dwelling_unit:
            if not data.get("accessory_dwelling_unit_description"):
                raise serializers.ValidationError(
                    {"accessory_structure_description": "Description is required"}
                )
        if is_accessory_structure and is_accessory_dwelling_unit:
            raise serializers.ValidationError(
                {
                    "is_accessory_dwelling_unit": "Project can be only Accessory Structure "
                    "or Accessory Dwelling Unit. Not both"
                }
            )

        if project_type == HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE:
            required_fields = [
                "builder_organization",
            ]
        elif project_type == HIRLProjectRegistration.MULTI_FAMILY_PROJECT_TYPE:
            is_include_commercial_space = data.get("is_include_commercial_space")

            if is_include_commercial_space:
                if not data.get("commercial_space_type"):
                    raise serializers.ValidationError(
                        {
                            "commercial_space_type": [
                                "This field is required",
                            ]
                        }
                    )
                if not data.get("total_commercial_space"):
                    raise serializers.ValidationError(
                        {
                            "total_commercial_space": [
                                "This field is required",
                            ]
                        }
                    )

            required_fields = ["number_of_units", "story_count"]
        elif project_type == HIRLProjectRegistration.LAND_DEVELOPMENT_PROJECT_TYPE:
            required_fields = [
                "number_of_lots",
            ]
        else:
            raise serializers.ValidationError({"project_type": "Project Type field is required"})

        for required_field_name in required_fields:
            if required_field_name in data.keys() and data[required_field_name] is None:
                raise serializers.ValidationError(
                    {
                        required_field_name: [
                            "This field is required",
                        ]
                    }
                )

        home_address_geocode_response = data.get("home_address_geocode_response")

        # check provided address for uniqueness
        address_components = get_hirl_project_address_components(
            home_address_geocode=data.get("home_address_geocode"),
            home_address_geocode_response=home_address_geocode_response,
        )
        if address_components:
            is_unique = hirl_project_address_is_unique(
                street_line1=address_components["street_line1"],
                street_line2=address_components["street_line2"],
                city=address_components["city"],
                zipcode=address_components["zipcode"],
                project_type=project_type == HIRLProjectRegistration.MULTI_FAMILY_PROJECT_TYPE,
                project=instance,
                geocode_response=home_address_geocode_response,
            )

            if not is_unique:
                raise serializers.ValidationError("Project address is not unique")

        data["project_type"] = project_type
        return data


class HIRLProjectMeta:
    """
    Base Meta model for HIRLProject with common fields
    """

    model = HIRLProject
    fields = (
        "id",
        "registration",
        "registration_info",
        "home_status",
        "home_status_info",
        "is_accessory_structure",
        "accessory_structure_description",
        "green_energy_badges",
        "green_energy_badges_info",
        "lot_number",
        "home_address_geocode",
        "home_address_geocode_info",
        "home_address_geocode_response",
        "home_address_geocode_response_info",
        "hud_disaster_case_number",
        "h_number",
        "billing_state",
        "manual_billing_state",
        "is_require_water_sense_certification",
        "is_require_rough_inspection",
        "is_require_wri_certification",
        "is_appeals_project",
        # single family specific
        "is_accessory_dwelling_unit",
        "accessory_dwelling_unit_description",
        # multi family specific
        "building_number",
        "is_include_commercial_space",
        "commercial_space_type",
        "total_commercial_space",
        "story_count",
        "number_of_units",
        "commercial_space_parent",
        # land development specific
        "are_all_homes_in_ld_seeking_certification",
        "land_development_project_type",
        "land_development_phase_number",
        "number_of_lots",
    )
    read_only_fields = ("home_status", "h_number")


class BasicHIRLProjectSerializer(HIRLProjectSerializerMixin, serializers.ModelSerializer):
    """Basic control of HIRLProject instance."""

    class Meta(HIRLProjectMeta):
        read_only_fields = HIRLProjectMeta.read_only_fields + (
            "billing_state",
            "manual_billing_state",
        )


class HIRLProjectSerializer(HIRLProjectSerializerMixin, serializers.ModelSerializer):
    """Allows full control of HIRLProject instance."""

    class Meta(HIRLProjectMeta):
        pass


class HIRLProjectAddressIsUniqueRequestDataSerializer(serializers.Serializer):
    project = serializers.PrimaryKeyRelatedField(
        pk_field=HashidSerializerCharField(source_field="customer_hirl.HIRLProject.id"),
        queryset=HIRLProject.objects.all(),
        allow_null=True,
        required=False,
        default=None,
        help_text="If provided exclude this project instance from search",
    )
    street_line1 = serializers.CharField()
    street_line2 = serializers.CharField(
        allow_blank=True, allow_null=True, required=False, default=""
    )
    is_multi_family = serializers.BooleanField(default=False)
    city = serializers.PrimaryKeyRelatedField(queryset=City.objects.all())
    zipcode = serializers.CharField()
    geocode_response = serializers.IntegerField(allow_null=True, required=False)

    def create(self, validated_data):
        raise NotImplementedError

    def update(self, instance, validated_data):
        raise NotImplementedError


class GreenPaymentsImportLineSerializer(serializers.Serializer):
    """
    Green Payments per line data serialization
    """

    amount = serializers.DecimalField(decimal_places=2, default=0.0, max_digits=9)
    date_received = serializers.DateTimeField(default_timezone=pytz.timezone("US/Eastern"))
    h_number = serializers.IntegerField()

    def create(self, validated_data):
        raise NotImplementedError

    def update(self, instance, validated_data):
        raise NotImplementedError


class GreenPaymentsImportSerializer(serializers.Serializer):
    jamis_file = serializers.FileField()

    def create(self, validated_data):
        """
        :param validated_data:
        :return: HIRL projects that have been payed
        """
        user = self.context["request"].user

        async_document = AsynchronousProcessedDocument.objects.create(
            company=user.company,
            document=validated_data["jamis_file"],
            task_name=green_payments_import_task.name,
            task_id="",
            download=False,
        )

        result = green_payments_import_task.delay(
            company_id=user.company.id,
            user_id=user.id,
            result_object_id=async_document.id,
            rows=validated_data["validated_rows"],
        )

        async_document.task_id = result.task_id
        async_document.save()

        return async_document

    def update(self, instance, validated_data):
        raise NotImplementedError

    def validate(self, attrs):
        try:
            wb = load_workbook(attrs["jamis_file"], read_only=True, data_only=True)
        except Exception:
            raise serializers.ValidationError(
                {
                    "jamis_file": f'Cannot read provided file {attrs["jamis_file"].name}. '
                    f"Incorrect File Format or File is broken"
                }
            )
        ws = wb.active

        validated_rows = []
        for cells in ws.iter_rows(min_row=2):
            # skip empty rows
            cells_values = list(map(lambda cell: cell.value, cells))
            if not any(cells_values):
                continue

            amount, date_received, h_number, *_ = cells_values
            line_serializer = GreenPaymentsImportLineSerializer(
                data={"amount": amount, "date_received": date_received, "h_number": h_number}
            )

            line_serializer.is_valid(raise_exception=True)

            validated_rows.append(line_serializer.validated_data)

        if not validated_rows:
            raise serializers.ValidationError({"jamis_file": f"File is empty"})

        h_numbers = set([data["h_number"] for data in validated_rows])
        existing_projects = HIRLProject.objects.filter(h_number__in=h_numbers).values_list(
            "h_number", flat=True
        )

        diff_list = list(h_numbers - set(existing_projects))
        if diff_list:
            raise serializers.ValidationError(
                {"jamis_file": f"Projects with H-Numbers: {diff_list} do not exist"}
            )

        attrs["validated_rows"] = validated_rows
        return attrs


class ProjectBillingImportLineSerializer(serializers.Serializer):
    """
    Project Billing per line data serialization
    """

    h_number = serializers.IntegerField()

    def create(self, validated_data):
        raise NotImplementedError

    def update(self, instance, validated_data):
        raise NotImplementedError


class ProjectBillingImportSerializer(serializers.Serializer):
    jamis_file = serializers.FileField()

    def create(self, validated_data):
        """
        :param validated_data:
        :return: HIRL projects that have been updated
        """
        user = self.context["request"].user

        async_document = AsynchronousProcessedDocument.objects.create(
            company=user.company,
            document=validated_data["jamis_file"],
            task_name=project_billing_import_task.name,
            task_id="",
            download=True,
        )

        result = project_billing_import_task.delay(
            company_id=user.company.id,
            user_id=user.id,
            result_object_id=async_document.id,
            rows=validated_data["validated_rows"],
        )

        async_document.task_id = result.task_id
        async_document.save()

        return async_document

    def update(self, instance, validated_data):
        raise NotImplementedError

    def validate(self, attrs):
        try:
            wb = load_workbook(attrs["jamis_file"], read_only=True, data_only=True)
        except Exception:
            raise serializers.ValidationError(
                {
                    "jamis_file": f'Cannot read provided file {attrs["jamis_file"].name}. '
                    f"Incorrect File Format or File is broken"
                }
            )
        ws = wb.active

        validated_rows = []
        for cells in ws.iter_rows(min_row=2):
            # skip empty rows
            cells_values = list(map(lambda cell: cell.value, cells))
            if not any(cells_values):
                continue

            h_number = cells_values[0]
            line_serializer = ProjectBillingImportLineSerializer(data={"h_number": h_number})

            line_serializer.is_valid(raise_exception=True)

            validated_rows.append(line_serializer.validated_data)

        if not validated_rows:
            raise serializers.ValidationError({"jamis_file": f"File is empty"})

        h_numbers = set([data["h_number"] for data in validated_rows])
        existing_projects = HIRLProject.objects.filter(h_number__in=h_numbers).values_list(
            "h_number", flat=True
        )

        diff_list = list(h_numbers - set(existing_projects))
        if diff_list:
            raise serializers.ValidationError(
                {"jamis_file": f"Projects with H-Numbers: {diff_list} do not exist"}
            )

        attrs["validated_rows"] = validated_rows
        return attrs


class BillingRuleExportQueryParamsSerializer(serializers.Serializer):
    start_date = serializers.DateField(default=timezone.now() - timezone.timedelta(days=10))
    end_date = serializers.DateField(default=timezone.now())
    program_slugs = serializers.ListSerializer(
        child=serializers.CharField(), default=customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS
    )

    def create(self, validated_data):
        pass

    def update(self, instance, validated_data):
        pass


class MilestoneExportQueryParamsSerializer(serializers.Serializer):
    start_date = serializers.DateField(default=timezone.now() - timezone.timedelta(days=10))
    end_date = serializers.DateField(default=timezone.now())
    program_slugs = serializers.ListSerializer(
        child=serializers.CharField(), default=customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS
    )

    def create(self, validated_data):
        pass

    def update(self, instance, validated_data):
        pass
