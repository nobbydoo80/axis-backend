"""customer_hirl_all_projects_report.py: """

__author__ = "Artem Hruzd"
__date__ = "12/01/2021 8:36 PM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import io

from celery import shared_task
from django.apps import apps
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.db.models import (
    Max,
    Sum,
    Subquery,
    OuterRef,
)
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from axis.core.reports import AxisReportFormatter
from axis.customer_hirl.api_v3.filters import HIRLProjectFilter
from axis.customer_hirl.models import HIRLProject
from axis.filehandling.log_storage import LogStorage
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.invoicing.models import InvoiceItem

User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


@shared_task(bind=True, time_limit=60 * 60 * 3)
def customer_hirl_all_projects_report_task(self, user_id, result_object_id, query_params=None):
    if not query_params:
        query_params = {}

    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 1}}

    self.update_state(state="STARTED", meta=task_meta)

    app_log = LogStorage(model_id=result_object_id)

    user = User.objects.get(id=user_id)

    async_document = AsynchronousProcessedDocument.objects.get(id=result_object_id)

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="NGBSCertficationRawData")
    axis_report_formatter = AxisReportFormatter(user=user)

    row = 1
    column = 1
    labels = [
        {"text": "AddressL1", "col_width": 35},
        {"text": "CertificateSentDate", "col_width": 35},
        {"text": "ScoringPath", "col_width": 30},
        {"text": "RoughInRcvdDate", "col_width": 25},
        {"text": "InvoiceSentDate", "col_width": 35},
        {"text": "PaymentRcvdDate", "col_width": 35},
        {"text": "CreatedDate", "col_width": 35},
        {"text": "CertificationStatus", "col_width": 35},
        {"text": "InvoiceNumber", "col_width": 35},
        {"text": "TotalFee", "col_width": 35},
        {"text": "Builder ID", "col_width": 35},
    ]

    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    row = 2

    queryset = HIRLProject.objects.filter_by_user(user=user).filter(home_status__isnull=False)
    filter_cls = HIRLProjectFilter(data=query_params, queryset=queryset)
    queryset = filter_cls.qs

    queryset = queryset.select_related(
        "home_status",
        "home_status__home",
        "home_status__eep_program",
        "registration",
        "registration__builder_organization",
    ).annotate(
        most_recent_payment_received=Max(
            "home_status__invoiceitemgroup__invoiceitem__transactions__created_at"
        ),
        most_recent_rough_qa_received_date=Max("home_status__qastatus__created_on"),
        fee_total=Subquery(
            InvoiceItem.objects.filter(group__home_status__customer_hirl_project=OuterRef("pk"))
            .values("group__home_status__customer_hirl_project")
            .order_by("group__home_status__customer_hirl_project")
            .annotate(
                total=Sum("cost"),
            )
            .values("total")[:1]
        ),
    )

    for project in queryset:
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=1), project.home_status.home.street_line1
        )

        if project.home_status.certification_date:
            axis_report_formatter.format_datetime_cell(
                sheet.cell(row=row, column=2), project.home_status.certification_date
            )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=3), project.home_status.eep_program.name
        )

        if project.most_recent_rough_qa_received_date:
            axis_report_formatter.format_datetime_cell(
                sheet.cell(row=row, column=4),
                project.most_recent_rough_qa_received_date,
            )

        axis_report_formatter.format_datetime_cell(
            sheet.cell(row=row, column=5),
            project.most_recent_notice_sent,
        )
        axis_report_formatter.format_datetime_cell(
            sheet.cell(row=row, column=6), project.most_recent_payment_received
        )

        axis_report_formatter.format_datetime_cell(
            sheet.cell(row=row, column=7), project.home_status.created_date
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=8),
            project.billing_state,
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=9), project.home_status.customer_hirl_project.h_number
        )

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=10), project.fee_total)

        try:
            entity_responsible_for_payment = (
                project.registration.get_company_responsible_for_payment()
            )
        except ObjectDoesNotExist:
            entity_responsible_for_payment = None

        legacy_builder_id = None

        if entity_responsible_for_payment:
            hirlcompanyclient = getattr(entity_responsible_for_payment, "hirlcompanyclient", None)
            if hirlcompanyclient:
                legacy_builder_id = hirlcompanyclient.id

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=11), legacy_builder_id)

        row = row + 1

    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    outfile = virtual_workbook.getvalue()

    today = timezone.now().today().strftime("%Y%m%d")
    filename = "all_projects_{}.xlsx".format(today)

    async_document.document.save(filename, ContentFile(outfile, name=filename))
    async_document.task_id = self.request.id
    async_document.save()

    task_meta["writing"]["current"] = 1
    self.update_state(state="DONE", meta=task_meta)
