"""projects.py: """

__author__ = "Artem Hruzd"
__date__ = "06/01/2021 8:25 PM"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import decimal
import io
from datetime import datetime

import pytz
from celery import shared_task
from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.sites.models import Site
from django.core.exceptions import ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from axis.core.reports import AxisReportFormatter
from axis.customer_hirl.messages import GreenPaymentsImportAdminNotificationMessage
from axis.customer_hirl.models import HIRLProject, BuilderAgreement
from axis.filehandling.log_storage import LogStorage
from axis.filehandling.models import AsynchronousProcessedDocument
from axis.qa.models import QARequirement, QAStatus

User = get_user_model()
customer_hirl_app = apps.get_app_config("customer_hirl")


@shared_task(bind=True, time_limit=60 * 60)
def green_payments_import_task(self, company_id, user_id, result_object_id, rows):
    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 1}}

    self.update_state(state="STARTED", meta=task_meta)

    app_log = LogStorage(model_id=result_object_id)

    result_object = AsynchronousProcessedDocument.objects.get(id=result_object_id)
    user = User.objects.get(id=user_id)

    byte_file = result_object.document.read()
    load_workbook(ContentFile(byte_file), data_only=True, read_only=True)
    result_object.document.close()

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    # The e-mail contains a table with the following information: certification invoice number,
    # amounts previous paid, new amounts paid, the new final total amount paid, time of the update
    # and the new payment received date.
    admin_email_table_data = []

    for row in rows:
        amount = decimal.Decimal(row["amount"])
        if not isinstance(row["date_received"], datetime):
            date_paid = datetime.strptime(row["date_received"], "%Y-%m-%dT%H:%M:%S%z")
            date_paid = date_paid.replace(tzinfo=pytz.timezone("US/Eastern"))
        else:
            date_paid = row["date_received"]

        hirl_project = (
            HIRLProject.objects.filter(h_number=row["h_number"]).annotate_billing_info().first()
        )

        previous_paid = hirl_project.fee_total_paid
        _ = hirl_project.pay(
            amount=amount,
            paid_by=user,
            date_paid=date_paid,
            note="Automatically via Import",
        )
        app_log.info(
            f'Paid {row["amount"]} for Project '
            f'<a href="{hirl_project.get_absolute_url()}">#{hirl_project.id}</a>'
        )

        hirl_project = (
            HIRLProject.objects.filter(h_number=row["h_number"]).annotate_billing_info().first()
        )
        total_amount_paid = hirl_project.fee_total_paid
        admin_email_table_data.append(
            {
                "project": hirl_project,
                "previous_paid": previous_paid,
                "new_amounts_paid": amount,
                "total_amount_paid": total_amount_paid,
                "date_update": row["date_received"],
            }
        )

    ngbs_accounting_users = customer_hirl_app.get_accounting_users()

    base_url = Site.objects.get(id=settings.SITE_ID).domain

    GreenPaymentsImportAdminNotificationMessage().send(
        users=ngbs_accounting_users,
        context={
            "async_document_url": result_object.get_absolute_url(),
            "base_url": "https://%s" % base_url,
            "result_table_data": admin_email_table_data,
        },
    )

    app_log.info(f"Imported {len(rows)} Projects")
    task_meta["writing"]["current"] = 1
    self.update_state(state="DONE", meta=task_meta)


@shared_task(bind=True, time_limit=60 * 60)
def project_billing_import_task(self, company_id, user_id, result_object_id, rows):
    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 1}}

    self.update_state(state="STARTED", meta=task_meta)

    app_log = LogStorage(model_id=result_object_id)

    result_object = AsynchronousProcessedDocument.objects.get(id=result_object_id)
    User.objects.get(id=user_id)

    byte_file = result_object.document.read()
    load_workbook(ContentFile(byte_file), data_only=True, read_only=True)
    result_object.document.close()

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    for row in rows:
        hirl_project = HIRLProject.objects.get(h_number=row["h_number"])
        hirl_project.is_jamis_milestoned = True
        hirl_project.save()
        app_log.info(
            f"Project "
            f'<a href="{hirl_project.get_absolute_url()}">#{hirl_project.id}</a> '
            f"jamis milestoned updated"
        )

    app_log.info(f"Imported {len(rows)} Projects")
    task_meta["writing"]["current"] = 1
    self.update_state(state="DONE", meta=task_meta)


@shared_task(bind=True, time_limit=60 * 60)
def billing_rule_export_task(
    self, user_id, result_object_id, start_date=None, end_date=None, program_slugs=None
):
    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 1}}

    self.update_state(state="STARTED", meta=task_meta)

    app_log = LogStorage(model_id=result_object_id)

    user = User.objects.get(id=user_id)

    async_document = AsynchronousProcessedDocument.objects.get(id=result_object_id)

    if not program_slugs:
        program_slugs = customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS

    if start_date is None:
        start_date = timezone.now() - timezone.timedelta(days=10)
    elif isinstance(start_date, str):
        start_date = parse_datetime(start_date)

    if end_date is None:
        end_date = timezone.now()
    elif isinstance(end_date, str):
        end_date = parse_datetime(end_date)

    hirl_projects = (
        HIRLProject.objects.filter(registration__eep_program__slug__in=program_slugs)
        .annotate_billing_info()
        .annotate_client_ca_status()
        .filter(
            Q(
                billing_state__in=[
                    HIRLProject.NOTICE_SENT_BILLING_STATE,
                ]
            )
            | Q(
                billing_state=HIRLProject.COMPLETED_BILLING_STATE,
                home_status__certification_date__range=[
                    start_date,
                    end_date,
                ],
            ),
            Q(client_ca_status=BuilderAgreement.COUNTERSIGNED),
        )
        .select_related(
            "registration",
            "home_status",
        )
        .prefetch_related("hirllegacycertification_set", "home_status__qastatus_set")
        .distinct()
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="BillingRule")
    outfile = _fill_jamis_sheet(
        user=user, hirl_projects=hirl_projects, workbook=workbook, sheet=sheet, app_log=app_log
    )

    today = timezone.now().today().strftime("%Y%m%d")
    filename = "BillingRule{}.xlsx".format(today)

    async_document.document.save(filename, ContentFile(outfile, name=filename))
    async_document.task_id = self.request.id
    async_document.save()

    task_meta["writing"]["current"] = 1
    self.update_state(state="DONE", meta=task_meta)


@shared_task(bind=True, time_limit=60 * 60)
def milestone_export_task(
    self, user_id, result_object_id, start_date=None, end_date=None, program_slugs=None
):
    task_meta = {"processing": {"current": 0, "total": 1}, "writing": {"current": 0, "total": 1}}

    self.update_state(state="STARTED", meta=task_meta)

    app_log = LogStorage(model_id=result_object_id)

    user = User.objects.get(id=user_id)

    async_document = AsynchronousProcessedDocument.objects.get(id=result_object_id)

    if not program_slugs:
        program_slugs = customer_hirl_app.HIRL_PROJECT_EEP_PROGRAM_SLUGS

    if not start_date:
        start_date = timezone.now() - timezone.timedelta(days=10)
    else:
        start_date = parse_datetime(start_date)

    if not end_date:
        end_date = timezone.now()
    else:
        end_date = parse_datetime(end_date)

    hirl_projects = (
        HIRLProject.objects.filter(registration__eep_program__slug__in=program_slugs)
        .filter(is_jamis_milestoned=False)
        .annotate_billing_info()
        .annotate_client_ca_status()
        .filter(
            Q(
                billing_state__in=[
                    HIRLProject.NOTICE_SENT_BILLING_STATE,
                ]
            )
            | Q(
                billing_state=HIRLProject.COMPLETED_BILLING_STATE,
                home_status__certification_date__range=[
                    start_date,
                    end_date,
                ],
            ),
            Q(client_ca_status=BuilderAgreement.COUNTERSIGNED),
        )
        .select_related(
            "registration",
            "home_status",
        )
        .prefetch_related("hirllegacycertification_set", "home_status__qastatus_set")
        .distinct()
    )

    task_meta["processing"]["current"] = 1
    self.update_state(state="STARTED", meta=task_meta)

    workbook = Workbook()
    sheet = workbook.create_sheet(index=0, title="Milestones")
    outfile = _fill_jamis_sheet(
        user=user, hirl_projects=hirl_projects, workbook=workbook, sheet=sheet, app_log=app_log
    )

    today = timezone.now().today().strftime("%Y%m%d")
    filename = "Milestones{}.xlsx".format(today)

    async_document.document.save(filename, ContentFile(outfile, name=filename))
    async_document.task_id = self.request.id
    async_document.save()

    task_meta["writing"]["current"] = 1
    self.update_state(state="DONE", meta=task_meta)


def _fill_jamis_sheet(user, hirl_projects, workbook, sheet, app_log):
    axis_report_formatter = AxisReportFormatter(user=user)

    row = 1
    column = 1
    labels = [
        {"text": "ID", "col_width": 25},
        {"text": "BillingRuleID", "col_width": 35},
        {"text": "CertificationFee", "col_width": 30},
        {"text": "RevenuePercent", "col_width": 25},
        {"text": "CertificationStatus", "col_width": 35},
        {"text": "BuilderID", "col_width": 35},
        {"text": "Job ID", "col_width": 15},
        {"text": "InvoiceRuleID", "col_width": 15},
        {"text": "InvoiceSentDate", "col_width": 15},
    ]

    for label in labels:
        cell = sheet.cell(row=row, column=column, value=label["text"])
        axis_report_formatter.set_cell_header_style(cell)
        sheet.column_dimensions[get_column_letter(column)].width = label["col_width"]
        column += 1

    row = 2
    for hirl_project in hirl_projects:
        revenue_percentage = 0

        if (
            hirl_project.registration.eep_program.slug
            in customer_hirl_app.WRI_PROGRAM_LIST + customer_hirl_app.LAND_DEVELOPMENT_PROGRAM_LIST
        ):
            revenue_percentage = 0

            if hirl_project.billing_state == HIRLProject.COMPLETED_BILLING_STATE:
                revenue_percentage = 100
        else:
            if (
                hirl_project.billing_state == HIRLProject.NOTICE_SENT_BILLING_STATE
                and hirl_project.home_status
            ):
                rough_qa = hirl_project.home_status.qastatus_set.filter(
                    requirement__type=QARequirement.ROUGH_INSPECTION_QA_REQUIREMENT_TYPE
                ).first()

                final_qa = hirl_project.home_status.qastatus_set.filter(
                    requirement__type=QARequirement.FINAL_INSPECTION_QA_REQUIREMENT_TYPE
                ).first()

                if (not rough_qa and not final_qa) or (rough_qa and not final_qa):
                    legacy_certification = hirl_project.hirllegacycertification_set.all().first()
                    if legacy_certification:
                        # 4 - New State for NGBS legacy DB
                        if legacy_certification.data["fkCertificationStatus"] == 4:
                            revenue_percentage = 10

                if rough_qa:
                    if (
                        rough_qa.result == QAStatus.PASS_STATUS
                        or not hirl_project.is_require_rough_inspection
                    ):
                        revenue_percentage = 70

            if hirl_project.billing_state == HIRLProject.COMPLETED_BILLING_STATE:
                revenue_percentage = 100

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=1), hirl_project.h_number)
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=2), hirl_project.get_billing_rule_id()
        )
        # customer wants integer cost here
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=3), int(hirl_project.fee_total)
        )

        sheet.cell(row=row, column=4).value = f"{revenue_percentage}"
        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=5),
            HIRLProject.BILLING_STATE_DISPLAY[hirl_project.billing_state],
        )

        try:
            project_client_company = hirl_project.registration.get_project_client_company()
        except ObjectDoesNotExist:
            project_client_company = None

        legacy_id = None

        if project_client_company:
            hirlcompanyclient = getattr(project_client_company, "hirlcompanyclient", None)
            if hirlcompanyclient:
                legacy_id = hirlcompanyclient.id

        if legacy_id:
            axis_report_formatter.format_str_cell(
                sheet.cell(row=row, column=6),
                f"B{legacy_id:05}",
            )
        else:
            app_log.info(f"Project {hirl_project.id} do not have legacy NGBS ID")

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=7), hirl_project.get_job_id()
        )

        axis_report_formatter.format_str_cell(
            sheet.cell(row=row, column=8), hirl_project.get_invoice_rule_id()
        )

        initial_invoice_date = getattr(hirl_project, "initial_invoice_date", None)
        if initial_invoice_date:
            initial_notice_sent = initial_invoice_date.strftime("GRN%m%Y")
        else:
            initial_notice_sent = ""

        axis_report_formatter.format_str_cell(sheet.cell(row=row, column=9), initial_notice_sent)
        row += 1

    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    return virtual_workbook.getvalue()
