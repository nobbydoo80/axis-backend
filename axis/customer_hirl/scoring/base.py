"""base.py: """

__author__ = "Artem Hruzd"
__date__ = "06/19/2020 19:01"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Artem Hruzd",
    "Steven Klass",
]

import re

import dateutil
from django.apps import apps
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.functional import cached_property
from openpyxl.cell import ReadOnlyCell, Cell
from rest_framework.exceptions import ValidationError

from axis.annotation.models import Annotation, Type as AnnotationType
from axis.customer_hirl.messages import (
    HIRLScoringUploadNotificationMessage,
    HIRLScoringUploadFinalOutstandingFeeBalanceMessage,
    HIRLProjectInvoiceCantGeneratedWithoutClientAgreement,
)
from axis.customer_hirl.models import HIRLProject, HIRLProjectRegistration, BuilderAgreement
from axis.customer_hirl.verifier_agreements.states import VerifierAgreementStates
from axis.filehandling.models import CustomerDocument
from axis.home.models import EEPProgramHomeStatus
from axis.invoicing.messages import InvoiceCreatedNotificationMessage
from axis.invoicing.models import InvoiceItemGroup, Invoice
from axis.qa.models import QARequirement, QAStatus
from axis.user_management.models import Accreditation

customer_hirl_app = apps.get_app_config("customer_hirl")


class ScoringExtractionValidationException(Exception):
    """
    Raise when provided value in document is not exist or have invalid format
    """

    pass


class ScoringExtractionUnknownVersion(Exception):
    """
    Raising when we can't parse document version
    """

    pass


class ScoringExtractionVersionNotSupported(Exception):
    """
    Raise when document version is not supported for some action
    """

    def __init__(self, sheet_version=None, *args):
        self.sheet_version = sheet_version

    def __str__(self):
        if self.sheet_version:
            return (
                f"Version {self.sheet_version} not supported. "
                f"Make sure that you selected correct Verification Report Type"
            )
        else:
            return (
                "Scoring Extraction Version not supported. "
                "Make sure that you selected correct Verification Report Type"
            )


class ScoringExtractionRequirementsFailed(Exception):
    """
    Raise when provided value in document is not met certain requirements
    """

    pass


class BatchSubmissionCellConfig:
    """
    Contains settings and methods to get batch submission cell range
    """

    def __init__(self, project_id_cell_range, inspection_date_cell_range, sheet):
        self.project_id_cell_range = project_id_cell_range
        self.inspection_date_cell_range = inspection_date_cell_range
        self.sheet = sheet

    def get_cells(self):
        project_id_cells = self._get_cell_range(str_range=self.project_id_cell_range)
        inspection_date_cells = self._get_cell_range(str_range=self.inspection_date_cell_range)

        return zip(project_id_cells, inspection_date_cells)

    def _get_cell_range(self, str_range):
        r = re.compile("([a-zA-Z]+)([0-9]+)")
        start, end = str_range.split(":")

        m = r.match(start)
        start_col = m.group(1).upper()
        start_value = int(m.group(2))

        m2 = r.match(end)
        end_value = int(m2.group(2))
        return [f"{start_col}{index}" for index in range(start_value, end_value)]

    def __str__(self):
        return (
            f"Sheet: {self.sheet} "
            f"Project ID Cells: [{self.project_id_cell_range}] "
            f"Inspection Date Cells: [{self.inspection_date_cell_range}]"
        )


class SamplingConfig:
    """
    Contains settings and methods for sampling validation
    """

    def __init__(self, sheet_name, total_available_cell, total_inspected_cell, error_cell):
        self.sheet_name = sheet_name
        self.total_available_cell = total_available_cell
        self.total_inspected_cell = total_inspected_cell
        self.error_cell = error_cell


class BaseScoringExtraction(object):
    """
    This is a complex base class that provide the following actions:
    self.validate_project()
    self.validate_sampling()
    self.validate_verifier()
    self.populate_annotations()
    self.create_qa_programs()
    self.populate_home_status()
    self.modify_uploaded_worksheet()

    To start work with new program we need to declare annotation types for program,
    annotation map and serializer
    """

    # key and display are mandatory and is using by Registry class
    key = None
    display = None

    # Based on data type we select annotations
    # This prefix will be added to the end of annotations
    ROUGH_DATA_TYPE = "rough"
    FINAL_DATA_TYPE = "final"
    AVAILABLE_DATA_TYPES = [ROUGH_DATA_TYPE, FINAL_DATA_TYPE]

    # to validate annotation data from cells
    annotation_data_serializer_class = None

    # cells to read to perform batch_submission method. Expect BatchSubmissionCellConfig
    rough_submission_project_id_config = None
    final_submission_project_id_config = None

    # cells to read to perform sampling. Expect SamplingConfig
    rough_sampling_config = None
    final_sampling_config = None

    # destinations contains all named cells from document
    available_destinations = [
        "vsfProjectID",
    ]

    MAX_BATCH_SUBMISSION_PROJECTS = 50

    @cached_property
    def sheet_version(self):
        ws = self.workbook["Info & Intro"]
        sheet_version = ws["B7"].value

        if sheet_version:
            self.app_log.info("Detect version .xlsx {}".format(sheet_version))
        else:
            raise ScoringExtractionUnknownVersion
        return str(sheet_version)

    @cached_property
    def destinations(self):
        """
        Destinations - it is all named cells in spreadsheet(e.g vsfProjectID)
        :return: Dictionary
        """
        destinations_data = {}
        for defined_name in self.available_destinations:
            try:
                title, coord = next(self.workbook.defined_names[defined_name].destinations)
                value = self.workbook[title][coord].value
                destinations_data[defined_name] = value
            except KeyError:
                raise ScoringExtractionValidationException(
                    "{} cell name is required in uploaded file".format(defined_name)
                )
        return destinations_data

    def get_cell_by_defined_name(self, defined_name):
        title, coord = next(self.workbook.defined_names[defined_name].destinations)
        return self.workbook[title][coord]

    @cached_property
    def customer_hirl_provider_organization(self):
        return customer_hirl_app.get_customer_hirl_provider_organization()

    def get_project_id(self):
        hirl_project_id = self.destinations["vsfProjectID"]
        return self._clean_str(hirl_project_id)

    @cached_property
    def project(self):
        """
        Get HIRLProject in cell vsfProjectID and select/prefetch related data
        :raise ScoringExtractionRequirementsFailed:
        :return: HIRLProject
        """
        hirl_project_id = self.get_project_id()

        hirl_project = (
            HIRLProject.objects.filter_by_user(user=self.user)
            .filter(id=hirl_project_id)
            .annotate_fee_balance()
            .select_related("home_status", "registration")
            .first()
        )

        if not hirl_project:
            raise ScoringExtractionRequirementsFailed(
                f"Project with provided ID: "
                f"{hirl_project_id} not found on AXIS. "
                f"Please confirm that the correct project ID has been entered "
                f"in the Verification Report spreadsheet and that an "
                f"NGBS project with that ID has been registered."
            )
        if not hirl_project.home_status:
            raise ScoringExtractionRequirementsFailed(
                f"Project {hirl_project_id} is not approved yet by NGBS"
            )
        return hirl_project

    @cached_property
    def home_status(self):
        if not self.project.home_status:
            raise ScoringExtractionRequirementsFailed("Required programs does not exist for home")
        return self.project.home_status

    @cached_property
    def annotations_map(self):
        """
        Provide annotation map per version in following format:
        {
            self.to_annotation_name('energy-path'): {
                'sheet': 'Verification Report',
                'cell': 'W106'
            }
        }
        :return: Dict
        """
        raise NotImplementedError

    @cached_property
    def annotation_data(self):
        raw_data = {}
        for annotation_type_slug, value_location in self.annotations_map.items():
            if isinstance(value_location, (Cell, ReadOnlyCell)):
                value = value_location.value or ""
                self.app_log.debug(
                    "Validate ({} - {}) with value {}".format(
                        value_location.parent.title, value_location.coordinate, value
                    )
                )
            else:
                ws = self.workbook[value_location["sheet"]]
                value = ws[value_location["cell"]].value or ""
                self.app_log.debug(
                    "Validate ({} - {}) with value {}".format(
                        value_location["sheet"], value_location["cell"], value
                    )
                )
            try:
                AnnotationType.objects.get(slug=annotation_type_slug)
            except AnnotationType.DoesNotExist:
                self.app_log.debug(f"{annotation_type_slug} annotation does not exist")
                raise ScoringExtractionValidationException(
                    "NGBS Standard Version/Scoring Path Verification Report spreadsheet "
                    "does not match the NGBS program associated with the project. "
                    "Please review your project and/or VR upload selections"
                )
            raw_data[annotation_type_slug] = value

        return raw_data

    @cached_property
    def validated_annotation_data(self):
        serializer_class = self.get_annotation_data_serializer_class()
        data = {}
        # normalize data for serializer by removing program slug and type
        for annotation_type_slug, value in self.annotation_data.items():
            key = annotation_type_slug.replace(
                f"-{self.home_status.eep_program.slug}-{self.data_type}", ""
            )
            key = "_".join(key.split("-"))
            data[key] = value
        serializer = serializer_class(
            data=data, context=self.get_annotation_data_serializer_context()
        )
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as exc:
            formatted_errors = {}
            for key, value in exc.get_full_details().items():
                formatted_errors[key] = "\n".join(map(lambda error: error["message"], value))
            raise ScoringExtractionValidationException(formatted_errors)

        validated_data = {}
        for key, value in serializer.validated_data.items():
            validated_data[self.to_annotation_name("-".join(key.split("_")))] = value
        return validated_data

    def __init__(
        self,
        workbook,
        user,
        verifier,
        data_type,
        app_log,
        document,
        parent_project=None,
        *args,
        **kwargs,
    ):
        """
        :param workbook: excel document
        :param user: who uploading document
        :param verifier: selected by user
        :param data_type: rough or final
        :param app_log: AXIS app logger
        :param document: raw excel document
        :param parent_project: Parent HIRLProject
        :param args:
        :param kwargs: extra_data
        """
        self.workbook = workbook
        self.user = user
        self.verifier = verifier
        self.app_log = app_log
        self.document = document
        self.parent_project = parent_project
        self.extra_data = kwargs
        if data_type not in self.AVAILABLE_DATA_TYPES:
            raise ValueError("{} is not defined in Scoring DATA_TYPES".format(data_type))
        self.data_type = data_type

    def validate_verifier(self):
        if not self.verifier:
            raise ScoringExtractionRequirementsFailed("Verifier does not exists")

        self.app_log.debug(f"Validate verifier: {self.verifier}")
        # Verifier must have active accreditation for scoring path selected
        # Turns out, when HI accredits a verifier, they accredit them for an NGBS version (year)
        # and NOT on a per-program basis.
        # So for example, if a verifier is accredited for "NGBS 2020",
        # then they are accredited to submit for ALL
        # 2020 NGBS programs: SF NC, SF Remodel, MF NC, MF Remodel, SF Certified.
        eep_program = self.home_status.eep_program

        accreditation_exists = eep_program.hirl_program_have_accreditation_for_user(
            user=self.verifier
        )

        if not accreditation_exists:
            raise ScoringExtractionRequirementsFailed(
                f"{self.verifier} must have an Accreditation for program {eep_program}"
            )

        # Verifier must have active Verifier Agreement
        verifier_agreement_exists = self.verifier.customer_hirl_enrolled_verifier_agreements.filter(
            state=VerifierAgreementStates.COUNTERSIGNED
        ).exists()

        if not verifier_agreement_exists:
            raise ScoringExtractionRequirementsFailed(
                f"{self.verifier} do not have any countersigned Verifier Agreement"
            )

        #  Check expiration date of verifier COI’s
        active_cois_exists = self.verifier.company.coi_documents.active().exists()
        if not active_cois_exists:
            raise ScoringExtractionRequirementsFailed(
                f"{self.verifier} do not have any active Certificate of Insurance"
            )

    def validate_project(self):
        try:
            project_client = self.project.registration.get_project_client_company()
        except ObjectDoesNotExist:
            raise ScoringExtractionValidationException("Project Client Company is not set")

        countersigned_ca_exists = project_client.customer_hirl_enrolled_agreements.filter(
            state=BuilderAgreement.COUNTERSIGNED
        ).exists()

        if not countersigned_ca_exists:
            project_client_name = project_client.name
            project_client_url = project_client.get_absolute_url()
            ctx = {
                "project_client_url": project_client_url,
                "project_client_name": project_client_name,
                "project_url": self.project.get_absolute_url(),
                "project_name": self.project.id,
            }
            # raise same message to abort VR
            raise ScoringExtractionValidationException(
                HIRLProjectInvoiceCantGeneratedWithoutClientAgreement.content.format(**ctx)
            )

        if self.data_type == self.ROUGH_DATA_TYPE:
            if not self.project.is_require_rough_inspection:
                raise ScoringExtractionRequirementsFailed(
                    "Rough Inspection is disabled for this project"
                )
            if self.home_status.state in [
                EEPProgramHomeStatus.CUSTOMER_HIRL_PENDING_FINAL_QA_STATE,
                EEPProgramHomeStatus.COMPLETE_STATE,
                EEPProgramHomeStatus.FAILED_STATE,
                EEPProgramHomeStatus.ABANDONED_STATE,
            ]:
                raise ScoringExtractionRequirementsFailed("Rough Inspection already completed")

        if self.data_type == self.FINAL_DATA_TYPE:
            if self.home_status.state in [
                EEPProgramHomeStatus.COMPLETE_STATE,
                EEPProgramHomeStatus.FAILED_STATE,
                EEPProgramHomeStatus.ABANDONED_STATE,
            ]:
                raise ScoringExtractionRequirementsFailed("Final Inspection already completed")

            if (
                self.project.registration.project_type
                == HIRLProjectRegistration.MULTI_FAMILY_PROJECT_TYPE
            ):
                if self.project.registration.state == HIRLProjectRegistration.PENDING_STATE:
                    raise ScoringExtractionRequirementsFailed(
                        "A Final Verification Report may not be uploaded "
                        "for this project until the multifamily registration form is complete"
                    )

            if self.project.is_require_rough_inspection:
                qa_statuses = self.home_status.qastatus_set.filter(
                    requirement__type=QARequirement.ROUGH_INSPECTION_QA_REQUIREMENT_TYPE
                )
                if not qa_statuses:
                    raise ScoringExtractionRequirementsFailed(
                        "A Final Verification Report may not be "
                        "uploaded for this project prior to a Rough "
                        "Verification Report being uploaded"
                    )
                for qa_status in qa_statuses:
                    if qa_status.result != QAStatus.PASS_STATUS:
                        raise ScoringExtractionRequirementsFailed(
                            "Rough Inspection QA must be completed "
                            "prior to Final Verification Report upload"
                        )

    def validate_sampling(self):
        sampling_config = self.rough_sampling_config
        if self.data_type == self.FINAL_DATA_TYPE:
            sampling_config = self.final_sampling_config

        if not sampling_config:
            return

        ws = self.workbook[sampling_config.sheet_name]

        error_msg = ws[sampling_config.error_cell].value
        total_available = ws[sampling_config.total_available_cell].value or ""
        total_available = total_available.replace("TOTAL:", "")
        total_inspected = ws[sampling_config.total_inspected_cell].value or ""
        total_inspected = total_inspected.replace("TOTAL:", "")

        try:
            total_available = int(total_available)
        except (TypeError, ValueError):
            total_available = 0

        try:
            total_inspected = int(total_inspected)
        except (TypeError, ValueError):
            total_inspected = 0

        self.app_log.info(
            f"Project Sampling is set to {self.project.registration.get_sampling_display()}"
        )

        if self.data_type == self.ROUGH_DATA_TYPE and self.project.registration.sampling in [
            HIRLProjectRegistration.FINAL_ALL_SAMPLING,
            HIRLProjectRegistration.FINAL_TESTING_AND_PRACTICES_ONLY_SAMPLING,
        ]:
            self.app_log.debug(
                f"Skip sampling check, because Registration "
                f"sampling is set to <b>{self.project.registration.sampling}</b>"
            )
            return

        if self.data_type == self.FINAL_DATA_TYPE and self.project.registration.sampling in [
            HIRLProjectRegistration.ROUGH_ALL_SAMPLING,
            HIRLProjectRegistration.ROUGH_TESTING_AND_PRACTICES_ONLY_SAMPLING,
        ]:
            self.app_log.debug(
                f"Skip sampling check, because Registration "
                f"sampling is set to <b>{self.project.registration.sampling}</b>"
            )
            return

        if (
            not self.project.registration.sampling
            or self.project.registration.sampling == HIRLProjectRegistration.NO_SAMPLING
        ):
            if total_available or total_inspected:
                raise ScoringExtractionRequirementsFailed(
                    f"During project registration, sampling was not specified for this project "
                    f"yet sampling information was provided in the verification report. "
                    f"Please remove the data from the sampling section of the "
                    f"{self.data_type.capitalize()} "
                    f"signature page and re-upload or change the sampling specification for this "
                    f'<a href="{self.project.get_absolute_url()}">Project {self.project.id}</a> '
                    f"to “Yes”"
                )
            else:
                return

        if self.data_type == self.ROUGH_DATA_TYPE and self.project.registration.sampling in [
            HIRLProjectRegistration.FINAL_TESTING_AND_PRACTICES_ONLY_SAMPLING,
            HIRLProjectRegistration.FINAL_ALL_SAMPLING,
        ]:
            final_ws = self.workbook[self.final_sampling_config.sheet_name]
            _total_available = final_ws[self.final_sampling_config.total_available_cell].value or ""
            _total_available = _total_available.replace("TOTAL:", "")
            _total_inspected = final_ws[self.final_sampling_config.total_inspected_cell].value or ""
            _total_inspected = _total_inspected.replace("TOTAL:", "")
            try:
                _total_available = int(_total_available)
            except (TypeError, ValueError):
                _total_available = 0

            try:
                _total_inspected = int(_total_inspected)
            except (TypeError, ValueError):
                _total_inspected = 0

            if _total_available or _total_inspected:
                raise ScoringExtractionRequirementsFailed(
                    f"During project registration, sampling was specified for "
                    f"this project as <b>{self.project.registration.get_sampling_display()}</b>. "
                    f"You have provided sampling data for "
                    f"<b>{self.data_type.capitalize()}</b>, which is inconsistent with your project registration. "
                    f"Please remove the sampling section of the <b>{self.final_sampling_config.sheet_name}</b> "
                    f"page or change the sampling specification for "
                    f'<a href="{self.project.get_absolute_url()}">Project {self.project.id}</a> '
                )

        if self.data_type == self.FINAL_DATA_TYPE and self.project.registration.sampling in [
            HIRLProjectRegistration.ROUGH_TESTING_AND_PRACTICES_ONLY_SAMPLING,
            HIRLProjectRegistration.ROUGH_ALL_SAMPLING,
        ]:
            rough_ws = self.workbook[self.rough_sampling_config.sheet_name]
            _total_available = rough_ws[self.rough_sampling_config.total_available_cell].value or ""
            _total_available = _total_available.replace("TOTAL:", "")
            _total_inspected = rough_ws[self.rough_sampling_config.total_inspected_cell].value or ""
            _total_inspected = _total_inspected.replace("TOTAL:", "")
            try:
                _total_available = int(_total_available)
            except (TypeError, ValueError):
                _total_available = 0

            try:
                _total_inspected = int(_total_inspected)
            except (TypeError, ValueError):
                _total_inspected = 0
            if _total_available or _total_inspected:
                raise ScoringExtractionRequirementsFailed(
                    f"During project registration, sampling was specified for "
                    f"this project as <b>{self.project.registration.get_sampling_display()}</b>. "
                    f"You have provided sampling data for "
                    f"<b>{self.data_type.capitalize()}</b>, which is inconsistent with your project registration. "
                    f"Please remove the sampling section of the <b>{self.rough_sampling_config.sheet_name}</b> "
                    f"page or change the sampling specification for "
                    f'<a href="{self.project.get_absolute_url()}">Project {self.project.id}</a> '
                )

        # check that verifier has Master Verifier accreditation
        if self.project.registration.sampling in [
            HIRLProjectRegistration.ROUGH_ALL_SAMPLING,
            HIRLProjectRegistration.FINAL_ALL_SAMPLING,
            HIRLProjectRegistration.ALL_SAMPLING,
        ]:
            master_accreditation_exists = (
                self.verifier.accreditations.filter(name=Accreditation.MASTER_VERIFIER_NAME)
                .annotate_expiration_date()
                .filter(expiration_date__gt=timezone.now())
                .exclude(state=Accreditation.INACTIVE_STATE)
                .exists()
            )
            if not master_accreditation_exists:
                raise ScoringExtractionRequirementsFailed(
                    f"Make sure that Verifier {self.verifier} have an active Master Verifier accreditation"
                )

        if not total_available and not total_inspected:
            raise ScoringExtractionRequirementsFailed(
                f"During project registration, sampling was specified for "
                f"this project and no sampling "
                f"information was provided in the verification report. "
                f"Please review and complete the sampling section of the "
                f"{self.data_type.capitalize()} "
                f"signature page, upload a separate tracking file (Excel or PDF file preferred), "
                f"or change the sampling specification for this "
                f'<a href="{self.project.get_absolute_url()}">Project {self.project.id}</a> '
                f"to “No”."
            )

        if total_available and not total_inspected:
            formatted_error_msg = ""
            if error_msg:
                formatted_error_msg = f" Error: {error_msg}"

            raise ScoringExtractionValidationException(
                f"Please correct the Sampling section of the {self.data_type.capitalize()} "
                f"signature page and re-upload the spreadsheet.{formatted_error_msg}"
            )

        ratio = total_available / total_inspected

        if ratio > 7:
            formatted_error_msg = ""
            if error_msg:
                formatted_error_msg = f" Error: {error_msg}"

            raise ScoringExtractionValidationException(
                f"Please correct the Sampling section of the {self.data_type.capitalize()} "
                f"signature page and re-upload the spreadsheet.{formatted_error_msg}"
            )

    def validate_wri(self):
        self.app_log.info("Validate WRI")
        wri_value = self.validated_annotation_data.get(self.to_annotation_name("wri-score"))
        if self.project.is_require_wri_certification and not wri_value:
            raise ScoringExtractionValidationException(
                """
                Based on the registration for this project, WRI data is
                expected but none has been provided in the verification report
                """
            )

        if wri_value and not self.project.is_require_wri_certification:
            raise ScoringExtractionValidationException(
                """
                Based on the registration for this project,
                no WRI data was expected but data has been provided in the verification report
                """
            )

    def create_qa_programs(self):
        self.app_log.info("Checking QA")
        if self.data_type == self.ROUGH_DATA_TYPE:
            qa_requirement_rough = QARequirement.objects.get(
                qa_company=self.customer_hirl_provider_organization,
                type=QARequirement.ROUGH_INSPECTION_QA_REQUIREMENT_TYPE,
                eep_program=self.home_status.eep_program,
            )
            qa_status, created = QAStatus.objects.get_or_create(
                requirement=qa_requirement_rough,
                home_status=self.home_status,
                owner=self.customer_hirl_provider_organization,
                defaults=dict(
                    qa_designee=self.customer_hirl_provider_organization.users.filter(
                        is_company_admin=True
                    ).first(),
                ),
            )
            if created:
                self.app_log.info(
                    "Created {}".format(
                        qa_status,
                    )
                )
        elif self.data_type == self.FINAL_DATA_TYPE:
            qa_requirement_final = QARequirement.objects.get(
                qa_company=self.customer_hirl_provider_organization,
                type=QARequirement.FINAL_INSPECTION_QA_REQUIREMENT_TYPE,
                eep_program=self.home_status.eep_program,
            )
            qa_status, created = QAStatus.objects.get_or_create(
                requirement=qa_requirement_final,
                home_status=self.home_status,
                owner=self.customer_hirl_provider_organization,
                defaults=dict(
                    qa_designee=self.customer_hirl_provider_organization.users.filter(
                        is_company_admin=True
                    ).first(),
                ),
            )
            if created:
                self.app_log.info(
                    "Created {}".format(
                        qa_status,
                    )
                )

    def populate_annotations(self):
        home_status_ct = ContentType.objects.get_for_model(EEPProgramHomeStatus)

        self.app_log.info("Creating annotations for {}".format(self.home_status.eep_program))
        for annotation_type_slug, value in self.validated_annotation_data.items():
            annotation_type = AnnotationType.objects.get(slug=annotation_type_slug)
            annotation, created = Annotation.objects.update_or_create(
                type=annotation_type,
                content_type=home_status_ct,
                object_id=self.home_status.id,
                defaults={
                    "user": self.user,
                    "content": value,
                },
            )
            if created:
                self.app_log.debug("Created annotation {}".format(annotation))
            else:
                self.app_log.debug("Updated annotation {}".format(annotation))

    def populate_home_status(self):
        """
        Update home status attributes
        :return:
        """
        self.app_log.debug(f"Set {self.data_type} Verifier to {self.verifier}")
        verifier_attr = "customer_hirl_rough_verifier"

        if self.data_type == self.FINAL_DATA_TYPE:
            verifier_attr = "customer_hirl_final_verifier"

        setattr(self.home_status, verifier_attr, self.verifier)
        self.home_status.save()

    def create_attachments(self):
        """
        Use this method to create additional documents
        :return: list of CustomerDocument objects
        """
        self.app_log.info("Attaching document to Home")

        if self.data_type == self.ROUGH_DATA_TYPE:
            qa_status = self.home_status.qastatus_set.filter(
                requirement__type=QARequirement.ROUGH_INSPECTION_QA_REQUIREMENT_TYPE
            ).first()

            content_type = ContentType.objects.get_for_model(QAStatus)
            object_id = qa_status.id
        elif self.data_type == self.FINAL_DATA_TYPE:
            qa_status = self.home_status.qastatus_set.filter(
                requirement__type=QARequirement.FINAL_INSPECTION_QA_REQUIREMENT_TYPE
            ).first()

            content_type = ContentType.objects.get_for_model(QAStatus)
            object_id = qa_status.id
        else:
            content_type = ContentType.objects.get_for_model(QAStatus)
            object_id = self.project.home_status.home.id

        document = CustomerDocument.objects.create(
            company=self.user.company,
            description="Scoring upload {} data".format(self.data_type),
            document=self.document,
            content_type=content_type,
            object_id=object_id,
            is_public=True,
        )

        return [
            document,
        ]

    def create_invoices(self):
        """
        Once a project requiring Rough (has SUBMITTED a Rough VR) AND
        (an invoice has not already been generated)
        Auto-generate invoice
        Email invoice to client contact AND copy verifier
        (Note: “Rough Bypass” can be set for either SF or MF projects)
        If Rough VR is part of a batch submission, generate one invoice for all projects

        :return:
        """
        if not self.parent_project:
            batch_data = self.get_batch_submission_project_data()
            project_ids = set([batch_obj["project_id"] for batch_obj in batch_data])
            project_ids.add(str(self.project.id))

            projects = HIRLProject.objects.filter(
                id__in=project_ids, is_require_rough_inspection=True, home_status__isnull=False
            )

            try:
                project_client = self.project.registration.get_project_client_company()
            except ObjectDoesNotExist:
                raise ScoringExtractionValidationException("Project Client Company is not set")

            # create invoice only in case we have fees without invoice
            # or project has legacy ngbs invoice
            invoice_item_groups = InvoiceItemGroup.objects.filter(
                Q(home_status__customer_hirl_project__in=projects),
                Q(invoice__isnull=True),
                Q(
                    home_status__customer_hirl_project__hirllegacycertification__data__has_key="InvoiceSentDate",
                    home_status__customer_hirl_project__hirllegacycertification__data__InvoiceSentDate=None,
                )
                | ~Q(
                    home_status__customer_hirl_project__hirllegacycertification__data__has_key="InvoiceSentDate"
                ),
            )

            if not invoice_item_groups:
                return

            try:
                company_responsible_for_payment = (
                    self.project.registration.get_company_responsible_for_payment()
                )
            except ObjectDoesNotExist:
                raise ScoringExtractionValidationException("Project Invoicee Company is not set")

            invoice = Invoice.objects.create(
                invoice_type=Invoice.HIRL_PROJECT_INVOICE_TYPE,
                issuer=self.customer_hirl_provider_organization,
                customer=company_responsible_for_payment,
                note=f"Automatically generated after VR upload document for project {self.project}",
            )
            for invoice_item_group in invoice_item_groups:
                invoice_item_group.invoice = invoice
                invoice_item_group.save()

            url = invoice.get_absolute_url()
            invoice_created_context = {
                "invoice_detail_url": url,
                "customer": company_responsible_for_payment,
                "customer_url": company_responsible_for_payment.get_absolute_url(),
                "invoice_item_groups": invoice_item_groups,
                "invoice_id": invoice.id,
            }
            InvoiceCreatedNotificationMessage(
                url=url,
            ).send(
                company=self.customer_hirl_provider_organization,
                context=invoice_created_context,
            )
            InvoiceCreatedNotificationMessage(url=url).send(
                user=self.user, context=invoice_created_context
            )
            InvoiceCreatedNotificationMessage(url=url).send(
                company=company_responsible_for_payment, context=invoice_created_context
            )
            InvoiceCreatedNotificationMessage(url=url).send(
                company=project_client, context=invoice_created_context
            )

    def finalize_upload(self):
        """
        Use this method to call notifications and post messages for user after success upload
        :return:
        """

        if self.data_type == self.FINAL_DATA_TYPE:
            self.app_log.info(
                "<p>The Verification Report requires "
                "the upload of the following additional collateral:</p>"
                "<p><strong>- At least one Front Elevation photo</strong></p>"
                "<p><strong>- An Energy Analysis Document</strong></p>"
            )

            if self.project.fee_current_balance > 0:
                HIRLScoringUploadFinalOutstandingFeeBalanceMessage().send(
                    company=self.customer_hirl_provider_organization,
                    context={
                        "project_url": self.project.get_absolute_url(),
                        "project_id": self.project.id,
                        "home_url": reverse_lazy(
                            "home:view", kwargs={"pk": self.project.home_status.home.pk}
                        ),
                        "home_address": self.project.home_status.home.get_home_address_display(),
                    },
                )

        if not self.parent_project:
            self.app_log.info(
                f'<a href="{self.project.home_status.get_absolute_url()}'
                f'#/tabs/documents">Document created</a>'
                f"<p style='font-size: 23px; margin: 0 0 0px;'>"
                f"<a href='{self.project.home_status.get_absolute_url()}#/tabs/qa'>"
                f"Click here to upload photos and additional documentation</a></p>"
                f"<hr style='margin: 0'>"
            )

    def get_annotation_data_serializer_class(self):
        """
        Serializer class to validate data from document
        :return: Serializer class
        """
        return self.annotation_data_serializer_class

    def get_annotation_data_serializer_context(self):
        """
        Serializer context
        :return: Dict
        """
        return {"user": self.user, "data_type": self.data_type}

    def to_annotation_name(self, name):
        """
        Add a prefix to annotation name based on data_type
        {name}-{eep_program_slug}-{cls.ROUGH_DATA_TYPE}
        :param name: annotation name
        :return: string
        """
        if self.data_type == self.ROUGH_DATA_TYPE:
            return self.to_annotation_r_name(name, self.home_status.eep_program.slug)
        return self.to_annotation_f_name(name, self.home_status.eep_program.slug)

    @classmethod
    def to_annotation_r_name(cls, name, eep_program_slug):
        return f"{name}-{eep_program_slug}-{cls.ROUGH_DATA_TYPE}"

    @classmethod
    def to_annotation_f_name(cls, name, eep_program_slug):
        return f"{name}-{eep_program_slug}-{cls.FINAL_DATA_TYPE}"

    def get_batch_submission_project_data(self):
        """
        Extract batch submission table to list of dicts
        :return: list of dicts
        """
        batch_config = self.rough_submission_project_id_config
        if self.data_type == self.FINAL_DATA_TYPE:
            batch_config = self.final_submission_project_id_config

        self.app_log.debug(f"Reading batch config {batch_config}")
        if not batch_config:
            return []

        ws = self.workbook[batch_config.sheet]
        batch_data = []
        batch_cells = batch_config.get_cells()
        for project_id_cell, inspection_date_cell in batch_cells:
            try:
                project_id = self._clean_str(ws[project_id_cell].value)
            except (ValueError, TypeError):
                continue

            inspection_date = ""
            try:
                inspection_date = self._clean_str(ws[inspection_date_cell].value)
            except (ValueError, TypeError):
                pass

            if project_id:
                if not inspection_date:
                    raise ScoringExtractionValidationException(
                        f"Project {project_id} " f"Inspection date is not provided"
                    )

                try:
                    inspection_date = dateutil.parser.parse(inspection_date)
                except ValueError:
                    raise ScoringExtractionValidationException(
                        f"Project {project_id} have invalid format for"
                        f" Inspection date {inspection_date}. "
                        f"Must be MM/DD/YYYY"
                    )

                if not inspection_date:
                    raise ScoringExtractionValidationException(
                        f"Project {project_id} do not have Inspection date"
                    )

                inspection_date = timezone.make_aware(inspection_date)

                batch_data.append({"project_id": project_id, "inspection_date": inspection_date})

        if (
            self.project.registration.project_type
            == HIRLProjectRegistration.SINGLE_FAMILY_PROJECT_TYPE
        ):
            inspection_dates = [batch["inspection_date"] for batch in batch_data]
            if inspection_dates:
                min_date = min(inspection_dates)
                max_date = max(inspection_dates)
                if (max_date - min_date).days > 60:
                    raise ScoringExtractionValidationException(
                        "All projects inspection date must be within 60 days"
                    )
        return batch_data

    @cached_property
    def batch_submission_extraction_objects(self):
        """
        Get ready Scoring Extraction classes for batch Projects
        :return:
        """
        batch_data = self.get_batch_submission_project_data()
        submission_extraction_objects = []
        for batch_obj in batch_data:
            new_extraction_cls = type(self)(
                workbook=self.workbook,
                user=self.user,
                data_type=self.data_type,
                verifier=self.verifier,
                app_log=self.app_log,
                document=self.document,
                parent_project=self.project,
                **{"inspection_date": batch_obj["inspection_date"], **self.extra_data},
            )

            # Get new project ID and replace cached_property project
            project = (
                HIRLProject.objects.filter_by_user(user=self.user)
                .filter(id=batch_obj["project_id"])
                .annotate_fee_balance()
                .select_related("home_status", "registration")
                .first()
            )

            # ignore project if it is have same id as parent project
            if project == self.project:
                continue

            if not project:
                raise ScoringExtractionValidationException(
                    f"Batch submission project with ID: {batch_obj['project_id']} not found. "
                    f"Please correct ID or remove it from list to pass validation"
                )
            new_extraction_cls.__dict__["project"] = project

            submission_extraction_objects.append(new_extraction_cls)

        return submission_extraction_objects

    def validate_batch_project(self, parent_project=None):
        """
        Validate project for common requirements before batch submission
        :param parent_project: HIRL Project. Default self.parent_project
        :return:
        """
        if not parent_project:
            parent_project = self.parent_project

        if self.home_status.eep_program != parent_project.home_status.eep_program:
            raise ScoringExtractionValidationException(
                f"Projects {self.project.id} and {parent_project.id} have different Program "
            )

        if self.project.is_accessory_structure:
            raise ScoringExtractionValidationException(
                f"Project {self.project.id} is Accessory Structure. "
                f"Accessory Structures CANNOT be included in a rough or final batch"
            )

        if self.project.is_accessory_dwelling_unit:
            raise ScoringExtractionValidationException(
                f"Project {self.project.id} is Accessory Dwelling Unit. "
                f"Accessory Dwelling Units CANNOT be included in a rough or final batch"
            )

        if self.project.is_include_commercial_space:
            raise ScoringExtractionValidationException(
                f"Project {self.project.id} is Commercial Space. "
                f"Commercial Space buildings CANNOT be included in a rough or final batch"
            )

        if (
            self.project.registration.builder_organization
            != parent_project.registration.builder_organization
        ):
            raise ScoringExtractionValidationException(
                f"Projects {self.project.id} and {parent_project.id} "
                f"have different Builder Organization"
            )

        parent_city = getattr(parent_project.home_address_geocode, "raw_city", None)
        city = getattr(self.project.home_address_geocode, "raw_city", None)
        if not parent_city:
            raise ScoringExtractionValidationException(
                f"Project <a href='{parent_project.get_absolute_url()}'>{parent_project.id}</a> do not have City"
            )
        if not city:
            raise ScoringExtractionValidationException(
                f"Project <a href='{self.project.get_absolute_url()}'>{self.project.id}</a> do not have City"
            )
        if city.county != parent_city.county:
            raise ScoringExtractionValidationException(
                f"Projects <a href='{self.project.get_absolute_url()}'>{self.project.id}</a>"
                f" and <a href='{parent_project.get_absolute_url()}'>{parent_project.id}</a> have different Counties"
            )

    def validate_batch_submission(self):
        """
        Validate all batch extraction objects
        :return:
        """
        submission_extraction_objects = self.batch_submission_extraction_objects

        if not any(submission_extraction_objects):
            self.app_log.info("Batch submission not found")
            return

        self.app_log.info("<b>Batch Submission Detected</b>")
        self.app_log.info(f"Found {len(submission_extraction_objects)} additional projects")

        if len(submission_extraction_objects) > self.MAX_BATCH_SUBMISSION_PROJECTS:
            raise ScoringExtractionValidationException(
                f"Max batch submission projects count is: {self.MAX_BATCH_SUBMISSION_PROJECTS}"
            )

        if (
            self.project.registration.project_type
            == HIRLProjectRegistration.MULTI_FAMILY_PROJECT_TYPE
        ):
            if (
                not self.verifier.accreditations.filter(
                    approver__company__slug=customer_hirl_app.CUSTOMER_SLUG,
                    name=Accreditation.MASTER_VERIFIER_NAME,
                )
                .annotate_expiration_date()
                .filter(expiration_date__gt=timezone.now().date())
                .exists()
            ):
                raise ScoringExtractionValidationException(
                    f"To upload Multi Family Batch projects "
                    f"verifier {self.verifier} must have "
                    f"active Master Verifier Accreditation created by NGBS"
                )

        for new_extraction_cls in submission_extraction_objects:
            new_extraction_cls.validate_batch_project()

    def batch_submission(self):
        submission_extraction_objects = self.batch_submission_extraction_objects
        for new_extraction_cls in submission_extraction_objects:
            new_extraction_cls._process()

            if self.data_type == self.FINAL_DATA_TYPE:
                new_extraction_cls.project.vr_batch_submission_parent_final = self.project
            else:
                new_extraction_cls.project.vr_batch_submission_parent_rough = self.project
            new_extraction_cls.project.save()

    def _process(self):
        """
        Including all actions for batch submission
        :return:
        """
        try:
            self.app_log.info(f"Verification Report Type: {self.display} ({self.data_type})")
            self.app_log.info(
                'Found <a href="{}#/tabs/programs">Building ID {}</a>'.format(
                    self.project.home_status.get_absolute_url(), self.project.id
                )
            )
            self.app_log.info(
                '<a href="{}#/tabs/programs">{}</a>'.format(
                    self.project.home_status.get_absolute_url(),
                    self.project.home_status.home.get_home_address_display(),
                )
            )
            self.validate_project()
            if not self.parent_project:
                self.validate_batch_submission()
            self.validate_sampling()
            self.validate_verifier()
            self.validate_wri()
            self.populate_annotations()
            self.create_qa_programs()
            self.populate_home_status()
            self.create_attachments()
            self.create_invoices()
            self.finalize_upload()
        except (
            ScoringExtractionUnknownVersion,
            ScoringExtractionVersionNotSupported,
            ScoringExtractionValidationException,
            ScoringExtractionRequirementsFailed,
        ) as exc:
            self.app_log.error(exc)
            raise exc

    def process(self):
        self.app_log.info("Validate uploaded document")
        self._process()

        if not self.parent_project:
            self.batch_submission()

        self.app_log.update_model(throttle_seconds=None)
        self.app_log.info("Task completed")

        HIRLScoringUploadNotificationMessage(url=self.home_status.home.get_absolute_url()).send(
            company=self.customer_hirl_provider_organization,
            context={
                "verification_report_type": self.data_type.title(),
                "uploaded_user": self.user.get_full_name(),
                "home_url": self.home_status.home.get_absolute_url(),
                "uploaded_user_profile_url": reverse_lazy(
                    "profile:detail", kwargs={"pk": self.user.pk}
                ),
            },
        )

    def _clean_str(self, value):
        """
        Convert value to string and trim whitespaces. Convert None to empty string
        :param value: any value
        :return: string
        """
        if value is None:
            value = ""
        value = str(value)
        value = value.strip()
        return value
