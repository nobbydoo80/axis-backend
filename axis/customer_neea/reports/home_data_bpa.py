"""mixins.py: Django """


import numbers
import logging
from collections import OrderedDict

from openpyxl.styles import fills, Alignment

from axis.home.export_data import HomeDataXLSExport, CellObject
from .mixins import ExtraFiltersMixin, NEEAMixin

__author__ = "Steven K"
__date__ = "08/02/2019 11:18"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven K",
]

log = logging.getLogger(__name__)


class NEEAHomeDataBPAExport(ExtraFiltersMixin, NEEAMixin, HomeDataXLSExport):
    """NEEA BPA Report"""

    def __init__(self, *args, **kwargs):
        kwargs["subject"] = "Performance Path Calculator"  # ""BPA Utility Export"
        kwargs["title"] = "Performance Path Calculator Summary Report"  # ""BPA Utility Export"
        kwargs["specified_columns"] = self.get_specified_columns()
        kwargs["retain_empty"] = True
        kwargs["report_on"] = [
            "bpa_ues",
            "bpa_ues_measures",
            "status",
            "home",
            "relationships",
            "simulation_advanced",
            "neea_standard_protocol_calculator",
            "checklist_answers",
        ]
        super(NEEAHomeDataBPAExport, self).__init__(*args, **kwargs)

    def set_cell_default_style(self, cell, **kwargs):
        """Default Style"""
        # "Calculator" columns are yellow
        if cell.column in ["L", "M"]:
            kwargs["fill"] = fills.PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFFAA")

        # Hard coded columns are grey
        if cell.column in ["W", "Z", "AC", "AF"]:
            kwargs["fill"] = fills.PatternFill(fill_type=fills.FILL_SOLID, start_color="FFCCCCCC")

        super(NEEAHomeDataBPAExport, self).set_cell_default_style(cell, **kwargs)

    def add_logo(self, workbook, sheet):
        """Add Logo"""
        super(NEEAHomeDataBPAExport, self).add_logo(workbook, sheet)

        cell = sheet.cell(row=2, column=7)
        cell.value = (
            "A negative savings value within a measure is possible and may be a result of "
            "technologies working interactively in the new home. Negative savings within "
            "measures impact total savings and total payments."
        )
        self.set_cell_default_style(cell)
        sheet.merge_cells(start_row=2, start_column=7, end_row=4, end_column=13)
        cell.alignment = Alignment(wrap_text=True)

    def get_specified_columns(self):
        """Get the columns"""
        return [
            # BPA UES fields
            "funding_source",
            "reference_number",
            "quantity",
            "certification_date",  # Axis field
            "federal_facility",
            "unique_site_id",
            "site_name",
            # Axis Fields
            "home__street_line1",
            "home__city__name",
            "home__state",
            "home__zipcode",
            "standardprotocolcalculator__busbar_savings",  # Total Busbar kWh Savings
            "standardprotocolcalculator__total_incentive",  # Total Payment
            "rater",  # Rating Company
            "home__id",
            "buildinginfo__conditioned_area",  # Area of Conditioned Space (sq.ft.)
            "dominant_heating_fuel",  # Heating Fuel Type
            "neea-heating_source",  # From checklist question neea-heating_source
            "standardprotocolcalculator__code_total_consumption_mmbtu",
            "standardprotocolcalculator__improved_total_consumption_mmbtu",
            "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings",
            "standardprotocolcalculator__percent_improvement",  # Is Home 10% Above Code?
            "standardprotocolcalculator__revised_percent_improvement",
            "standardprotocolcalculator__pct_improvement_method",
            "standardprotocolcalculator__reported_shell_windows_kwh_savings",
            "windows_insulation_and_other_shell_rate_per_kwh",
            "standardprotocolcalculator__reported_shell_windows_incentive",
            "standardprotocolcalculator__reported_hvac_waterheater_kwh_savings",
            "hvac_measure_rate",
            "standardprotocolcalculator__reported_hvac_waterheater_incentive",
            "standardprotocolcalculator__bpa_appliance_kwh_savings",
            "appliance_and_other_electronics_rater_per_kwh",
            "standardprotocolcalculator__appliance_kwh_incentive",
            "standardprotocolcalculator__reported_lighting_showerhead_tstats_kwh_savings",
            "lighting_measure_rate",
            "standardprotocolcalculator__reported_lighting_showerhead_tstats_incentive",
        ]

    def get_column_name_map(self):
        """Allows you to customize the column names

        If value is None it is not exported - but used by other applications
        """
        return {
            "funding_source": "FUNDING SOURCE",
            "reference_number": "REFERENCE NUMBER",
            "quantity": "QUANTITY",
            "certification_date": "COMPLETION DATE",
            "federal_facility": "FEDERAL FACILITY",
            "unique_site_id": "UNIQUE SITE ID",
            "site_name": "SITE NAME",
            "home__street_line1": "STREET",
            "home__city__name": "CITY",
            "home__state": "STATE",
            "home__zipcode": "ZIP",
            "standardprotocolcalculator__busbar_savings": "CALCULATOR SAVINGS PER UNIT (Total Site kWh Savings)",
            "standardprotocolcalculator__total_incentive": "CALCULATOR REIMBURSEMENT PER UNIT (Total Payment)",
            "rater": "Rater Name",
            "buildinginfo__conditioned_area": "Home Square Footage",
            "dominant_heating_fuel": "Primary Heater Fuel",
            "dominant_cooling_type": "Primary HVAC Type",
        }

    def get_customer_neea_standard_protocol_calculator_data(self):
        """Get the data"""
        from axis.home.models import EEPProgramHomeStatus

        replace_dict = OrderedDict(
            [
                ("id", "Homestatus ID"),
                (
                    "standardprotocolcalculator__code_total_consumption_mmbtu",
                    "Total State Code Reference Home MMBtu Consumption",
                ),
                (
                    "standardprotocolcalculator__improved_total_consumption_mmbtu",
                    "Total Site MMBtu Consumption",
                ),
                (
                    "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings",
                    "Total Site MMBtu Consumption (with measure savings)",
                ),
                ("standardprotocolcalculator__total_kwh_savings", "Estimated Annual Savings (kWh)"),
                (
                    "standardprotocolcalculator__total_therm_savings",
                    "Estimated Annual Savings (Therms)",
                ),
                ("standardprotocolcalculator__busbar_savings", "Total Busbar kWh Savings"),
                ("standardprotocolcalculator__pct_improvement_method", "Percent Method Used"),
                ("standardprotocolcalculator__percent_improvement", "Percent Improvement"),
                (
                    "standardprotocolcalculator__revised_percent_improvement",
                    "Alternate Percent Improvement",
                ),
                ("standardprotocolcalculator__pct_improvement_method", "% Improvement method used"),
                ("standardprotocolcalculator__total_incentive", "Total Payment"),
                (
                    "standardprotocolcalculator__bpa_appliance_kwh_savings",
                    "Appliance Upgrades Site kWh Savings",
                ),
                (
                    "standardprotocolcalculator__appliance_kwh_incentive",
                    "Appliance Upgrades Payment",
                ),
                ("standardprotocolcalculator__builder_incentive", "Builder Incentive"),
                (
                    "standardprotocolcalculator__reported_shell_windows_kwh_savings",
                    "Shell Upgrades, incl. Windows Site kWh Savings",
                ),
                (
                    "standardprotocolcalculator__reported_shell_windows_incentive",
                    "Shell Upgrades, incl. Windows Payment",
                ),
                (
                    "standardprotocolcalculator__reported_hvac_waterheater_kwh_savings",
                    "HVAC and Water Heat Upgrades Site kWh Savings",
                ),
                (
                    "standardprotocolcalculator__reported_hvac_waterheater_incentive",
                    "HVAC and Water Heat Upgrades Payment",
                ),
                (
                    "standardprotocolcalculator__reported_lighting_showerhead_tstats_kwh_savings",
                    "Lighting, incl. Fixtures, Showerheads, and Smart Tstats Site kWh Savings",
                ),
                (
                    "standardprotocolcalculator__reported_lighting_showerhead_tstats_incentive",
                    "Lighting, incl. Fixtures, Showerheads, and Smart Tstats Payment",
                ),
            ]
        )

        def round_two(value):
            """Round to two places"""
            if isinstance(value, numbers.Number):
                return "{:.2f}".format(value)
            return "{}".format(value)

        def percent(value):
            """Percent"""
            if isinstance(value, numbers.Number):
                return "{:.2f}%".format(value * 100)
            return "{}".format(value)

        def payment(value):
            """Payment"""
            if isinstance(value, numbers.Number):
                return "${:.2f}".format(value)
            return "{}".format(value)

        clean_dict = {
            "standardprotocolcalculator__code_total_consumption_mmbtu": round_two,
            "standardprotocolcalculator__improved_total_consumption_mmbtu": round_two,
            "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings": round_two,
            "standardprotocolcalculator__busbar_savings": round_two,
            "standardprotocolcalculator__percent_improvement": percent,
            "standardprotocolcalculator__revised_percent_improvement": percent,
            "standardprotocolcalculator__total_incentive": payment,
            "standardprotocolcalculator__reported_shell_windows_kwh_savings": round_two,
            "standardprotocolcalculator__reported_shell_windows_incentive": payment,
            "standardprotocolcalculator__bpa_lighting_kwh_savings": round_two,
            "standardprotocolcalculator__lighting_kwh_incentive": payment,
            "standardprotocolcalculator__reported_lighting_showerhead_tstats_kwh_savings": round_two,
            "standardprotocolcalculator__reported_lighting_showerhead_tstats_incentive": payment,
            "standardprotocolcalculator__appliance_kwh_incentive": payment,
            "standardprotocolcalculator__bpa_appliance_kwh_savings": round_two,
            "standardprotocolcalculator__reported_hvac_waterheater_kwh_savings": round_two,
            "standardprotocolcalculator__reported_hvac_waterheater_incentive": payment,
            "standardprotocolcalculator__builder_incentive": payment,
        }

        structure = self.assign_basic(
            EEPProgramHomeStatus,
            include=list(replace_dict.keys()),
            section="neea_standard_protocol_calculator",
            replace_dict=replace_dict,
            clean_dict=clean_dict,
        )

        objects = self.get_queryset()
        data = self.merge_results(objects.values_list(*[x.attr for x in structure]), structure)

        # Unify Percent improvement
        if data:
            key = "standardprotocolcalculator__percent_improvement"
            pct_imp_col = data[0].index(next((x for x in data[0] if x.attr == key)))

            key = "standardprotocolcalculator__revised_percent_improvement"
            rev_pct_imp_col = data[0].index(next((x for x in data[0] if x.attr == key)))

            key = "standardprotocolcalculator__pct_improvement_method"
            pct_imp_met_col = data[0].index(next((x for x in data[0] if x.attr == key)))

            key = "standardprotocolcalculator__improved_total_consumption_mmbtu"
            imp_no_sav_col = data[0].index(next((x for x in data[0] if x.attr == key)))

            key = "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings"
            imp_with_sav_col = data[0].index(next((x for x in data[0] if x.attr == key)))

            for item in data[:]:
                if item[pct_imp_met_col].raw_value == ["alternate"]:
                    _d = list(item[pct_imp_col])
                    _d[-2] = item[rev_pct_imp_col].raw_value
                    _d[-1] = item[rev_pct_imp_col].value
                    item[pct_imp_col] = CellObject(*_d)

                    _e = list(item[imp_no_sav_col])
                    _e[-2] = item[imp_with_sav_col].raw_value
                    _e[-1] = item[imp_with_sav_col].value
                    item[imp_no_sav_col] = CellObject(*_e)

                key = "standardprotocolcalculator__pct_improvement_method"
                item.pop(item.index(next((x for x in item if x.attr == key))))

                key = "standardprotocolcalculator__revised_percent_improvement"
                item.pop(item.index(next((x for x in item if x.attr == key))))

                key = "standardprotocolcalculator__improved_total_consumption_mmbtu_with_savings"
                item.pop(item.index(next((x for x in item if x.attr == key))))

        return data
