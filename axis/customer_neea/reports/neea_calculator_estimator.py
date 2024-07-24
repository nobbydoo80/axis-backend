"""standard_protocol_estimator.py: Django """


import logging
import re
import tempfile

from openpyxl.drawing.image import Image
from openpyxl.styles import fills, Alignment, Font, PatternFill
from openpyxl.utils import range_boundaries
from simulation.enumerations import FuelType

from axis.checklist.xls_checklist import XLSChecklist
from axis.company.models import Company
from axis.customer_neea.rtf_calculator.constants.neea_v3 import (
    NEEA_CLOTHES_WASHER_CHOICE_MAP,
    NEEA_REFRIGERATOR_CHOICE_MAP,
)
from axis.geographic.models import County
from axis.home.single_home_checklist import AXIS_LOGO

__author__ = "Steven K"
__date__ = "08/02/2019 11:08"
__copyright__ = "Copyright 2011-2023 Pivotal Energy Solutions. All rights reserved."
__credits__ = [
    "Steven K",
]

log = logging.getLogger(__name__)


class NEEACalculatorV2EstimatorExport(XLSChecklist):
    """Used to download a simple representation of the calculator"""

    def __init__(self, *args, **kwargs):
        label = "BetterBuiltNW Performance Path Savings Estimator"
        kwargs["creator"] = kwargs.get("creator", "Axis")
        kwargs["title"] = kwargs.get("title", label)
        kwargs["subject"] = kwargs.get("subject", label)
        kwargs["description"] = kwargs.get("description", label)
        kwargs["sheet_name"] = kwargs.get("sheet_name", "Sheet 1")
        super(NEEACalculatorV2EstimatorExport, self).__init__(*args, **kwargs)
        self.start_row = 6

    def pre_save(self, workbook, sheet):
        """Add our logo"""
        image = Image(AXIS_LOGO)
        sheet.add_image(image, anchor=self.get_absolute_anchor(image, 0, 10, 100, 100))

    def get_item_details(self, cell, **kwargs):
        """Uniformly gets item details"""
        item = kwargs
        item["cell"] = cell
        item["name"] = "foo"
        return item

    @property
    def basic_style(self):
        """Basic Style"""
        return {
            "font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", shrinkToFit=True),
        }

    @property
    def thin_bordered_style(self):
        """Basic Bordered Style"""
        from openpyxl.styles import Color, Side, Border

        kwargs = self.basic_style
        kwargs["border"] = Border(
            bottom=Side(border_style="thin", color=Color(rgb="FF000000")),
            top=Side(border_style="thin", color=Color(rgb="FF000000")),
            left=Side(border_style="thin", color=Color(rgb="FF000000")),
            right=Side(border_style="thin", color=Color(rgb="FF000000")),
        )
        return kwargs

    @property
    def basic_header_style(self):
        """Basic Header Style"""
        return {
            "font": Font(name="Calibri", size=14, bold=True),
            "alignment": Alignment(horizontal="left"),
            "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
            "border": None,
        }

    def set_style(self, cell_obj, **attr):
        """Set our style"""
        style_kwargs = self.thin_bordered_style

        if attr.get("fill_color"):
            style_kwargs["fill"] = PatternFill(
                fill_type=fills.FILL_SOLID, start_color=attr.pop("fill_color")
            )
        else:
            style_kwargs["fill"] = PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFE1")

        style_kwargs.update(**attr)
        for k, v in style_kwargs.items():
            setattr(cell_obj, k, v)

    def set_merged_bordered_style(self, sheet_obj, cell_obj, **attrs):
        """Set a merged border"""

        def get_merge_range(cell_coordinate):  # pylint: disable=inconsistent-return-statements
            """Get the merge range"""
            for merge_range in sheet_obj.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
                for cell_list in sheet_obj.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    for cell_unit in cell_list:
                        if cell_unit.coordinate == cell_coordinate:
                            return merge_range

        cell_range = get_merge_range(cell_obj.coordinate)

        # Don't want to double this up.
        if hasattr(self, "handled_merged_cells"):
            if cell_range in self.handled_merged_cells:
                return cell_range
            self.handled_merged_cells.append(cell_range)

        min_col, min_row, max_col, max_row = range_boundaries(str(cell_range).upper())
        for cell_list in sheet_obj.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for _cell_obj in cell_list:
                self.set_style(_cell_obj, **attrs)
        return cell_range

    def _write_data(self, sheet, data, addr=0):
        """Write the data updating the rows"""
        final = []
        for idx, (key, value) in enumerate(data, start=self.row):
            match = re.match(r"([A-Z]+)(\d+)", key)
            column = match.group(1)
            row = int(match.group(2)) + addr
            if value.get("merge"):
                start, end = value.pop("merge").split(":")
                smatch = re.match(r"([A-Z]+)(\d+)", start)
                ematch = re.match(r"([A-Z]+)(\d+)", end)
                value["merge"] = "%s%d:%s%d" % (
                    smatch.group(1),
                    int(smatch.group(2)) + addr,
                    ematch.group(1),
                    int(ematch.group(2)) + addr,
                )
            final.append(("%s%d" % (column, row), value))

        for cell, item in final:
            cell_obj = sheet[cell]
            cell_obj.value = item.get("value")

            if item.get("merge"):
                sheet.merge_cells(item.get("merge"))
                self.set_merged_bordered_style(sheet, cell_obj, **item.get("style", {}))

            self.set_style(cell_obj, **item.get("style", {}))

    @property
    def header_title(self):
        return "BetterBuiltNW Performance Path Savings Estimator"

    def write_header(self, sheet):
        """Write out our header"""

        disclaimer = """Disclaimer: This Performance Path Savings Estimator tool provides estimated
        savings and may be used for preliminary reviews and design consulting with builders.
        Participating utilities must have their Performance Path program requirements configured
        in AXIS before homes can be submitted for incentives. Requirements vary by utility and
        may include space heating fuel and/or equipment, water heating fuel and/or equipment, %
        improvement over code, home certification, etc. The regional minimum % improvement over
        code is 10%, utilities may set higher minimums for their program. Incentive amounts vary
        by utility. Contact your utility for more specifics."""

        disclaimer = re.sub(r"\s+", " ", disclaimer)

        values = (
            ("A1", {"value": "", "style": self.basic_header_style, "merge": "A1:A6"}),
            (
                "B1",
                {
                    "value": self.header_title,
                    "style": self.basic_header_style,
                    "merge": "B1:H1",
                },
            ),
            (
                "B2",
                {
                    "value": disclaimer,
                    "style": {
                        "font": Font(name="Calibri", size=9, bold=False),
                        "alignment": Alignment(
                            horizontal="left", wrap_text=True, vertical="bottom"
                        ),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": "B2:H6",
                },
            ),
            ("A7", {"value": "", "style": self.basic_header_style, "merge": "A7:H7"}),
            ("A8", {"value": "Inputs:", "style": self.basic_header_style, "merge": "A8:E8"}),
            ("F8", {"value": "Results:", "style": self.basic_header_style, "merge": "F8:H8"}),
            ("A9", {"value": "", "style": self.basic_header_style, "merge": "A9:H9"}),
        )
        sheet.column_dimensions["A"].width = 15
        sheet.row_dimensions[7].height = 5
        sheet.row_dimensions[9].height = 5
        self._write_data(
            sheet,
            values,
        )

    def get_base_inputs(self, data):
        county = County.objects.get(id=data["county"])
        try:
            gas = Company.objects.get(id=data["gas_utility"])
        except ValueError:
            gas = "-"
        try:
            ele = Company.objects.get(id=data["electric_utility"])
        except ValueError:
            ele = "-"

        return [
            (
                "A1",
                {
                    "value": "Home Location and Size",
                    "style": {
                        "font": Font(name="Calibri", size=10, bold=True),
                        "alignment": Alignment(horizontal="left"),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": "A1:E1",
                },
            ),
            ("A2", {"value": "County", "style": {"fill_color": "FFFFFF"}}),
            ("A3", {"value": "{}, {}".format(county.name, county.state)}),
            ("B2", {"value": "Conditioned Area", "style": {"fill_color": "FFFFFF"}}),
            (
                "B3",
                {
                    "value": round(float(data["conditioned_area"]), 0),
                    "style": {"number_format": "#,##"},
                },
            ),
            ("A5", {"value": "Electric Utility", "style": {"fill_color": "FFFFFF"}}),
            ("A6", {"value": "{}".format(ele)}),
            ("B5", {"value": "Gas Utility", "style": {"fill_color": "FFFFFF"}}),
            ("B6", {"value": "{}".format(gas)}),
        ]

    def get_heating_cooling_inputs(self, data):
        return [
            (
                "A8",
                {
                    "value": "Heating and Cooling",
                    "style": {
                        "font": Font(name="Calibri", size=10, bold=True),
                        "alignment": Alignment(horizontal="left"),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": "A8:E8",
                },
            ),
            ("A9", {"value": "Heating Fuel", "style": {"fill_color": "FFFFFF"}}),
            ("A10", {"value": "{}".format(data["heating_fuel"].capitalize())}),
            (
                "B9",
                {"value": "Primary Heating", "style": {"fill_color": "FFFFFF"}, "merge": "B9:C9"},
            ),
            ("B10", {"value": "{}".format(data["primary_heating_type"]), "merge": "B10:C10"}),
            ("D9", {"value": "System Config", "style": {"fill_color": "FFFFFF"}}),
            ("D10", {"value": "{}".format(data["heating_system_config"].capitalize())}),
            ("A11", {"value": "Yes" if data.get("smart_thermostat_installed") == "true" else "No"}),
            (
                "B11",
                {
                    "value": "Smart thermostat installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": "B11:D11",
                },
            ),
            ("A13", {"value": "", "style": {"fill_color": "FFFFFF"}}),
            ("B13", {"value": "Reference Home", "style": {"fill_color": "FFFFFF"}}),
            ("C13", {"value": "Improved Home", "style": {"fill_color": "FFFFFF"}}),
            ("A14", {"value": "Heating (kWh)", "style": {"fill_color": "FFFFFF"}}),
            (
                "B14",
                {
                    "value": round(float(data["code_data_heating_kwh"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "C14",
                {
                    "value": round(float(data["improved_data_heating_kwh"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("A15", {"value": "Heating (Therms)", "style": {"fill_color": "FFFFFF"}}),
            (
                "B15",
                {
                    "value": round(float(data["code_data_heating_therms"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "C15",
                {
                    "value": round(float(data["improved_data_heating_therms"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("A16", {"value": "Cooling (kWh)", "style": {"fill_color": "FFFFFF"}}),
            (
                "B16",
                {
                    "value": round(float(data["code_data_cooling_kwh"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "C16",
                {
                    "value": round(float(data["improved_data_cooling_kwh"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("A17", {"value": "Total (kWh)", "style": {"fill_color": "FFFFFF"}}),
            (
                "B17",
                {
                    "value": round(float(data["code_data_total_consumption_kwh"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "C17",
                {
                    "value": round(float(data["improved_data_total_consumption_kwh"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("A18", {"value": "Total (Therms)", "style": {"fill_color": "FFFFFF"}}),
            (
                "B18",
                {
                    "value": round(float(data["code_data_total_consumption_therms"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "C18",
                {
                    "value": round(float(data["improved_data_total_consumption_therms"]), 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
        ]

    def get_water_heating_inputs(self, data):
        return [
            (
                "A20",
                {
                    "value": "Water Heating",
                    "style": {
                        "font": Font(name="Calibri", size=10, bold=True),
                        "alignment": Alignment(horizontal="left"),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": "A20:E20",
                },
            ),
            ("A21", {"value": "Water Heating", "style": {"fill_color": "FFFFFF"}}),
            ("A22", {"value": "{}".format(data["water_heater_tier"].capitalize())}),
        ]

    def get_shower_inputs(self, data):
        return [
            ("B21", {"value": "Qty 1.5 gpm Showerheads", "style": {"fill_color": "FFFFFF"}}),
            (
                "B22",
                {
                    "value": round(float(data["qty_shower_head_1p5"]), 0),
                    "style": {"number_format": "#,##0"},
                },
            ),
            ("C21", {"value": "Qty 1.75 gpm Showerheads", "style": {"fill_color": "FFFFFF"}}),
            (
                "C22",
                {
                    "value": round(float(data["qty_shower_head_1p75"]), 0),
                    "style": {"number_format": "#,##0"},
                },
            ),
        ]

    def get_lighting_inputs(self, data):
        return [
            (
                "A24",
                {
                    "value": "Lighting",
                    "style": {
                        "font": Font(name="Calibri", size=10, bold=True),
                        "alignment": Alignment(horizontal="left"),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": "A24:E24",
                },
            ),
            ("A25", {"value": "Qty CFL lamps", "style": {"fill_color": "FFFFFF"}}),
            (
                "A26",
                {
                    "value": round(float(data["cfl_installed"]), 0),
                    "style": {"number_format": "#,##0"},
                },
            ),
            ("B25", {"value": "Qty LED lamps", "style": {"fill_color": "FFFFFF"}}),
            (
                "B26",
                {
                    "value": round(float(data["led_installed"]), 0),
                    "style": {"number_format": "#,##0"},
                },
            ),
            ("C25", {"value": "Qty Total Lamps", "style": {"fill_color": "FFFFFF"}}),
            (
                "C26",
                {
                    "value": round(float(data["total_installed_lamps"]), 0),
                    "style": {"number_format": "#,##0"},
                },
            ),
        ]

    def get_appliance_inputs(self, data):
        return [
            (
                "F20",
                {
                    "value": "Appliances",
                    "style": {
                        "font": Font(name="Calibri", size=10, bold=True),
                        "alignment": Alignment(horizontal="left"),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": "F20:H20",
                },
            ),
            (
                "F21",
                {
                    "value": "Yes"
                    if data.get("estar_std_refrigerators_installed") == "true"
                    else "No"
                },
            ),
            (
                "G21",
                {
                    "value": "ENERGY STAR® Refrigerator Installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": "G21:H21",
                },
            ),
            ("F22", {"value": "Yes" if data.get("estar_dishwasher_installed") == "true" else "No"}),
            (
                "G22",
                {
                    "value": "ENERGY STAR® Dishwasher Installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": "G22:H22",
                },
            ),
            (
                "F23",
                {
                    "value": "Yes"
                    if data.get("estar_front_load_clothes_washer_installed") == "true"
                    else "No"
                },
            ),
            (
                "G23",
                {
                    "value": "ENERGY STAR® Front Load Clothes Washer Installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": "G23:H23",
                },
            ),
            ("F24", {"value": "{}".format(data["clothes_dryer_tier"].capitalize())}),
            (
                "G24",
                {
                    "value": "Clothes Dryer Tier",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": "G24:H24",
                },
            ),
        ]

    def get_annotation_inputs(self, data, start_row=26):
        return [
            (f"F{start_row}", {"value": "{}".format(data["certified_earth_advantage"])}),
            (
                f"G{start_row}",
                {
                    "value": "Earth Advantage Certification",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": f"G{start_row}:H{start_row}",
                },
            ),
        ]

    def get_input_values(self, data, sheet):
        result = self.get_base_inputs(data)
        result += self.get_heating_cooling_inputs(data)
        result += self.get_water_heating_inputs(data)
        result += self.get_shower_inputs(data)
        result += self.get_lighting_inputs(data)
        result += self.get_appliance_inputs(data)
        result += self.get_annotation_inputs(data)
        sheet.row_dimensions[28].height = 5
        return result

    def write_inputs(self, sheet, data):
        """Write the input data side"""

        values = self.get_input_values(data, sheet)

        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 3

        self._write_data(sheet, values, 9)

    def get_heating_cooling_output(self, data):
        return [
            ("F1", {"value": "", "style": {"fill_color": "FFFFFF"}}),
            ("G1", {"value": "kWh", "style": {"fill_color": "FFFFFF"}}),
            ("H1", {"value": "Therms", "style": {"fill_color": "FFFFFF"}}),
            ("F2", {"value": "Heating", "style": {"fill_color": "FFFFFF"}}),
            (
                "G2",
                {
                    "value": round(data["heating_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "H2",
                {
                    "value": round(data["heating_therm_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("F3", {"value": "Cooling", "style": {"fill_color": "FFFFFF"}}),
            (
                "G3",
                {
                    "value": round(data["cooling_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "H3",
                {
                    "value": round(data["cooling_therm_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("F4", {"value": "Smart Thermostat", "style": {"fill_color": "FFFFFF"}}),
            (
                "G4",
                {
                    "value": round(data["smart_thermostat_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "H4",
                {
                    "value": round(data["smart_thermostat_therm_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
        ]

    def get_water_heating_output(self, data):
        return [
            ("F5", {"value": "Water Heating", "style": {"fill_color": "FFFFFF"}}),
            (
                "G5",
                {
                    "value": round(data["water_heater_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "H5",
                {
                    "value": round(data["water_heater_therm_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
        ]

    def get_shower_outputs(self, data):
        return [
            ("F6", {"value": "Low Flow Shower Heads", "style": {"fill_color": "FFFFFF"}}),
            (
                "G6",
                {
                    "value": round(data["showerhead_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                "H6",
                {
                    "value": round(data["showerhead_therm_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
        ]

    def get_lighting_outputs(self, data):
        return [
            ("F7", {"value": "Lighting", "style": {"fill_color": "FFFFFF"}}),
            (
                "G7",
                {
                    "value": round(data["lighting_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("H7", {"value": "-"}),
        ]

    def get_appliance_outputs(self, data):
        return [
            ("F8", {"value": "Appliances", "style": {"fill_color": "FFFFFF"}}),
            (
                "G8",
                {
                    "value": round(data["appliance_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            ("H8", {"value": "-"}),
        ]

    def get_results_outputs(self, data, start_row=10):
        if data["pct_improvement_method"] == "alternate":
            percent_improvement = data["pretty_revised_percent_improvement"]
            improved_total_consumption = data["improved_total_consumption_mmbtu_with_savings"]
        else:
            percent_improvement = data["pretty_percent_improvement"]
            improved_total_consumption = data["improved_total_consumption_mmbtu"]

        return [
            (
                f"F{start_row}",
                {"value": "Reference Total Consumption (MBtu)", "style": {"fill_color": "FFFFFF"}},
            ),
            (
                f"G{start_row}",
                {
                    "value": round(data["code_total_consumption_mmbtu"], 2),
                    "style": {"number_format": "#,##0.00"},
                    "merge": f"G{start_row}:H{start_row}",
                },
            ),
            (
                f"F{start_row+1}",
                {"value": "As Built Total Consumption (MBtu)", "style": {"fill_color": "FFFFFF"}},
            ),
            (
                f"G{start_row+1}",
                {
                    "value": round(improved_total_consumption, 2),
                    "style": {"number_format": "#,##0.00"},
                    "merge": f"G{start_row+1}:H{start_row+1}",
                },
            ),
            (
                f"F{start_row+3}",
                {"value": "Total Annual Savings", "style": {"fill_color": "FFFFFF"}},
            ),
            (
                f"G{start_row+3}",
                {
                    "value": round(data["total_kwh_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                f"H{start_row+3}",
                {
                    "value": round(data["total_therm_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                },
            ),
            (
                f"F{start_row+4}",
                {"value": "Total Annual Savings (MBtu)", "style": {"fill_color": "FFFFFF"}},
            ),
            (
                f"G{start_row+4}",
                {
                    "value": round(data["total_mmbtu_savings"], 2),
                    "style": {"number_format": "#,##0.00"},
                    "merge": f"G{start_row+4}:H{start_row+4}",
                },
            ),
            (
                f"F{start_row+5}",
                {"value": "Percent Improvement", "style": {"fill_color": "FFFFFF"}},
            ),
            (
                f"G{start_row+5}",
                {"value": percent_improvement, "merge": f"G{start_row+5}:H{start_row+5}"},
            ),
            (
                f"F{start_row+7}",
                {"value": "Estimated Incentive", "style": {"fill_color": "FFFFFF"}},
            ),
            (
                f"G{start_row+7}",
                {
                    "value": data["pretty_builder_incentive"],
                    "merge": f"G{start_row+7}:H{start_row+7}",
                },
            ),
        ]

    def get_output_values(self, data):
        result = self.get_heating_cooling_output(data)
        result += self.get_water_heating_output(data)
        result += self.get_shower_outputs(data)
        result += self.get_appliance_outputs(data)
        result += self.get_results_outputs(data)
        return result

    def write_outputs(self, sheet, data):
        """Write the output data side"""
        values = self.get_output_values(data)
        sheet.column_dimensions["F"].width = 25
        sheet.column_dimensions["G"].width = 12
        sheet.column_dimensions["H"].width = 12
        self._write_data(sheet, values, 9)

    def write(self, input_data, result, output=None, return_workbook=False):
        """Write our summary"""
        from openpyxl import Workbook

        if output is None:
            _, output = tempfile.mkstemp(suffix=".xlsx", prefix=self.output_prefix)
            log.debug("Look here for the file that was just generated %s", output)

        assert output, "You need an output file"
        assert self.sheet_name is not None, "You need a valid sheet"

        workbook = Workbook()
        sheet = workbook.create_sheet(index=0, title=self.sheet_name)
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

        self.write_header(sheet)
        self.write_inputs(sheet, input_data)
        self.write_outputs(sheet, result)
        self.pre_save(workbook, sheet)

        if return_workbook:
            return workbook

        workbook.save(output)
        return output


class NEEACalculatorV3EstimatorExport(NEEACalculatorV2EstimatorExport):
    @property
    def header_title(self):
        return "BetterBuiltNW Performance Path Savings Estimator V3"

    def get_appliance_inputs(self, data, start_row=16):
        washer_labels = dict(NEEA_CLOTHES_WASHER_CHOICE_MAP)
        refrigerator_labels = dict(NEEA_REFRIGERATOR_CHOICE_MAP)
        fuel_labels = dict([(FuelType.ELECTRIC, "Electric"), (FuelType.NATURAL_GAS, "Gas")])
        return [
            (
                f"F{start_row}",
                {
                    "value": "Appliances",
                    "style": {
                        "font": Font(name="Calibri", size=10, bold=True),
                        "alignment": Alignment(horizontal="left"),
                        "fill": PatternFill(fill_type=fills.FILL_SOLID, start_color="FFFFFF"),
                        "border": None,
                    },
                    "merge": f"F{start_row}:H{start_row}",
                },
            ),
            (
                f"F{start_row+1}",
                {"value": refrigerator_labels.get(data.get("estar_std_refrigerators_installed"))},
            ),
            (
                f"G{start_row+1}",
                {
                    "value": "ENERGY STAR® Refrigerator Installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": f"G{start_row+1}:H{start_row+1}",
                },
            ),
            (
                f"F{start_row+2}",
                {"value": "Yes" if data.get("estar_dishwasher_installed") == "true" else "No"},
            ),
            (
                f"G{start_row+2}",
                {
                    "value": "ENERGY STAR® Dishwasher Installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": f"G{start_row+2}:H{start_row+2}",
                },
            ),
            (
                f"F{start_row+3}",
                {
                    "value": washer_labels.get(
                        data.get("estar_front_load_clothes_washer_installed")
                    ),
                },
            ),
            (
                f"G{start_row+3}",
                {
                    "value": "ENERGY STAR® Front Load Clothes Washer Installed",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": f"G{start_row+3}:H{start_row+3}",
                },
            ),
            (f"F{start_row + 4}", {"value": fuel_labels.get(data["clothes_dryer_fuel"])}),
            (
                f"G{start_row + 4}",
                {
                    "value": "Clothes Dryer Fuel",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": f"G{start_row+4}:H{start_row+4}",
                },
            ),
            (f"F{start_row+5}", {"value": "{}".format(data["clothes_dryer_tier"].capitalize())}),
            (
                f"G{start_row+5}",
                {
                    "value": "Clothes Dryer Tier",
                    "style": {"fill_color": "FFFFFF"},
                    "merge": f"G{start_row+5}:H{start_row+5}",
                },
            ),
        ]

    def get_input_values(self, data, sheet):
        result = self.get_base_inputs(data)
        result += self.get_heating_cooling_inputs(data)
        result += self.get_water_heating_inputs(data)
        result += self.get_appliance_inputs(data)
        result += self.get_annotation_inputs(data, start_row=23)
        return result

    def get_output_values(self, data):
        result = self.get_heating_cooling_output(data)
        result += self.get_water_heating_output(data)
        result += self.get_results_outputs(data, start_row=7)
        return result
